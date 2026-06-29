#!/usr/bin/env python3
"""LUTRON -> nama-website-kalian WooCommerce catalog generator v13, technical-spec-safe Indonesian WordPress output.

Fitur utama:
- menerima URL produk LUTRON dan/atau LUTRON dalam satu Excel;
- otomatis mendeteksi merek dari domain URL;
- membaca title, brand, model, deskripsi, seluruh konten produk, spesifikasi,
  kategori/breadcrumb, dan gambar utama;
- menyimpan seluruh konten resmi produk pada tab "Informasi Produk Resmi";
- memakai title/model/brand dari Excel sebagai fallback apabila elemen halaman
  produk tidak tersedia;
- output mengikuti kolom WooCommerce/nama-website-kalian seperti script LUTRON sebelumnya.

Contoh:
python uji_catalog_import_LUTRON_lutron.py --input "lutron.com.tw.xlsx" --output "uji_lutron_import_ready.xlsx"
python uji_catalog_import_LUTRON_lutron.py --input "produk_mixed.xlsx" --output "uji_mixed_import_ready.xlsx" --delay 0.6
python uji_catalog_import_LUTRON_lutron.py --input "lutron.com.tw.xlsx" --use-ai --deepseek-api-key "YOUR_KEY"
"""
from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import time
import shutil
import subprocess
import tempfile
from io import BytesIO
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple
from urllib.parse import quote, quote_plus, unquote, urljoin, urlparse, urlsplit, urlunsplit

try:
    import pandas as pd
except Exception as exc:
    raise SystemExit("pandas wajib tersedia. Install: pip install pandas openpyxl") from exc

try:
    import requests
except Exception as exc:
    raise SystemExit("requests wajib tersedia. Install: pip install requests") from exc

try:
    from bs4 import BeautifulSoup, Tag
except Exception as exc:
    raise SystemExit("beautifulsoup4 wajib tersedia. Install: pip install beautifulsoup4") from exc

try:
    from playwright.sync_api import sync_playwright
except Exception:
    sync_playwright = None

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:
    load_workbook = None


OUTPUT_COLS = [
    "Name", "Brand", "Product Description", "Product Short Description",
    "custom_tab_1_title", "custom_tab_1_content", "custom_tab_1_priority",
    "custom_tab_2_title", "custom_tab_2_content", "custom_tab_2_priority",
    "custom_tab_3_title", "custom_tab_3_content", "custom_tab_3_priority",
    "custom_tab_4_title", "custom_tab_4_content", "custom_tab_4_priority",
    "custom_tab_5_title", "custom_tab_5_content", "custom_tab_5_priority",
    "custom_tab_6_title", "custom_tab_6_content", "custom_tab_6_priority",
    "custom_tab_7_title", "custom_tab_7_content", "custom_tab_7_priority",
    "custom_tab_8_title", "custom_tab_8_content", "custom_tab_8_priority",
    "product categories", "product tags", "focus keyphrase", "meta description",
    "publication_date", "Processed", "Processing Time", "Content Quality",
    "image_url", "Source URL", "Website Validation", "Website Correction Log",
]

DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Upgrade-Insecure-Requests": "1",
}

UJI_CTA = (
    '<p>Untuk konsultasi produk, silakan kunjungi '
    '<strong><a href="https://nama-website-kalian/about-us-3/" target="_blank" rel="noopener">nama-website-kalian</a></strong> '
    'atau <strong><a href="https://nama-website-kalian/contact-us/" target="_blank" rel="noopener">hubungi kami</a></strong>. '
    'Telepon: +62 000000000000.</p>'
)

NOISE_TEXT = {
    "home", "products", "contact us", "about us", "search", "menu", "login", "register",
    "shopping cart", "privacy policy", "terms of use", "copyright", "all rights reserved",
    "facebook", "instagram", "youtube", "linkedin", "line", "language", "english", "中文",
    "download", "inquiry", "quotation", "please contact us", "back to top",
}

SPEC_HEADING_RE = re.compile(
    r"\b(specification|specifications|technical data|technical specification|product information|"
    r"product specification|measurement specification|features and specifications|details)\b", re.I
)

# Klasifikasi aksesori dibuat konservatif. Kata seperti "probe", "sensor",
# "transmitter", dan "receiver" dapat muncul pada alat utama sehingga tidak boleh
# langsung mengubah produk menjadi aksesori.
ACCESSORY_STRONG_KEYWORDS = re.compile(
    r"\b(replacement|spare(?:\s+part)?|accessor(?:y|ies)|optional\s+(?:cable|adapter|adaptor|charger|case|probe|sensor|electrode|printer)|"
    r"extension\s+cable|usb[-\s]?0?1|upcb[-\s]?0?2|carrying\s+case|protective\s+cover|"
    r"test\s+piece|eyepiece|desiccant|o[-\s]?ring|washer|mounting\s+kit|holder|stand|strap|pouch|"
    r"bottle|beaker|cap|lid|filter)\b", re.I
)
ACCESSORY_COMPONENT_TERMS = re.compile(
    r"\b(probe|sensor|electrode|cable|adapter|adaptor|printer|case|cover|battery|charger|"
    r"test\s+piece|eyepiece|desiccant|holder|stand|paper|replacement|spare|accessor(?:y|ies)|"
    r"pouch|strap|funnel|nozzle|needle|hose|lid|o[-\s]?ring|washer|extension)\b", re.I
)
PRIMARY_PRODUCT_TERMS = re.compile(
    r"\b(transmitter|receiver|meter|tester|analyzer|counter|recorder|logger|gauge|thermometer|"
    r"multimeter|clamp|controller|monitor|calibrator|balance|scale|wrench|detector|instrument)\b", re.I
)
PRIMARY_MODEL_PREFIX_RE = re.compile(
    r"\b(?:TR[-_]|HR[-_]|EMF[-_]|DW[-_]|FC[-_]|FR[-_]|TM[-_]|CM[-_]|DM[-_]|BWA[-_]|BPH[-_]|"
    r"BMG[-_]|BDO[-_]|BCT[-_]|GM[-_]|DT[-_]|VB[-_]|FG[-_]|DR[-_]|PDA[-_]|LCR[-_]|TQ[-_]|SC[-_])",
    re.I,
)

# Jangan pernah menggunakan aset halaman umum sebagai foto katalog produk.
IMAGE_REJECT_TOKENS = (
    "logo", "icon", "sprite", "flag", "payment", "blank", "placeholder", "tracking",
    "banner", "banners", "hero", "carousel", "slider", "homepage", "landing", "corporate",
    "company", "facebook", "youtube", "instagram", "linkedin", "favicon", "og-default",
)

FAMILY_RULES: List[Tuple[re.Pattern[str], Dict[str, Any]]] = [
    (re.compile(r"fruit\s+hardness|fruit\s+sclerometer|\bfr-\d", re.I), {
        "category": "Penguji Kekerasan Buah",
        "term": "penguji kekerasan buah",
        "function": "mengukur tingkat kekerasan buah untuk pengendalian mutu, penelitian, atau proses sortasi",
        "tags": ["Lutron", "penguji kekerasan buah", "fruit hardness tester", "alat ukur pertanian"],
    }),
    (re.compile(r"electromagnetic\s+field|\bemf\s*tester|\bemf-", re.I), {
        "category": "Pengukur Medan Elektromagnetik",
        "term": "pengukur medan elektromagnetik",
        "function": "mengukur tingkat medan elektromagnetik sesuai spesifikasi produk",
        "tags": ["Lutron", "EMF tester", "pengukur medan elektromagnetik", "alat ukur keselamatan"],
    }),
    (re.compile(r"pressure\s+transmitter|\btr-ps", re.I), {
        "category": "Transmitter Tekanan",
        "term": "transmitter tekanan",
        "function": "mengubah pengukuran tekanan menjadi sinyal keluaran untuk kebutuhan pemantauan atau otomasi",
        "tags": ["Lutron", "pressure transmitter", "transmitter tekanan", "instrumentasi industri"],
    }),
    (re.compile(r"power analyzer|power meter|wattmeter|energy meter|power quality|three phase|3 phase|dw-", re.I), {
        "category": "Power Analyzer dan Electrical Tester",
        "term": "power analyzer atau electrical tester",
        "function": "menganalisis parameter daya, energi, dan kelistrikan sesuai spesifikasi produk",
        "tags": ["Lutron", "power analyzer", "power meter", "electrical tester", "alat ukur listrik"],
    }),
    (re.compile(r"frequency counter|frequency meter|fc-", re.I), {
        "category": "Frequency Counter",
        "term": "frequency counter",
        "function": "mengukur frekuensi dan parameter sinyal elektronik sesuai spesifikasi produk",
        "tags": ["Lutron", "frequency counter", "frequency meter", "electronic instrument"],
    }),
    (re.compile(r"clamp meter|clamp|multimeter|digital multimeter|voltmeter|ammeter|current recorder|cm-", re.I), {
        "category": "Multimeter dan Clamp Meter",
        "term": "multimeter atau clamp meter",
        "function": "mengukur parameter kelistrikan sesuai spesifikasi produk",
        "tags": ["Lutron", "multimeter", "clamp meter", "alat ukur listrik"],
    }),
    (re.compile(r"thermometer|temperature|thermocouple|infrared thermometer|tm-", re.I), {
        "category": "Thermometer dan Temperature Meter",
        "term": "alat ukur suhu",
        "function": "mengukur suhu sampel, lingkungan, atau proses sesuai spesifikasi produk",
        "tags": ["Lutron", "thermometer", "temperature meter", "alat ukur suhu"],
    }),
    (re.compile(r"humidity|hygrometer|wet bulb|dew point|hr-", re.I), {
        "category": "Humidity Meter",
        "term": "humidity meter",
        "function": "mengukur kelembapan dan parameter lingkungan terkait sesuai spesifikasi produk",
        "tags": ["Lutron", "humidity meter", "hygrometer", "alat ukur kelembapan"],
    }),
    (re.compile(r"sound level|noise|decibel|sl-", re.I), {
        "category": "Sound Level Meter",
        "term": "sound level meter",
        "function": "mengukur tingkat kebisingan atau tekanan suara sesuai spesifikasi produk",
        "tags": ["Lutron", "sound level meter", "noise meter", "decibel meter"],
    }),
    (re.compile(r"light meter|lux meter|uv meter|solar power|lx-", re.I), {
        "category": "Light Meter dan Lux Meter",
        "term": "light meter atau lux meter",
        "function": "mengukur intensitas cahaya atau parameter radiasi terkait sesuai spesifikasi produk",
        "tags": ["Lutron", "light meter", "lux meter", "uv meter"],
    }),
    (re.compile(r"anemometer|air flow|air velocity|wind speed|am-", re.I), {
        "category": "Anemometer dan Air Flow Meter",
        "term": "anemometer atau air flow meter",
        "function": "mengukur kecepatan udara, aliran udara, atau parameter lingkungan terkait sesuai spesifikasi produk",
        "tags": ["Lutron", "anemometer", "air flow meter", "air velocity meter"],
    }),
    (re.compile(r"ph\s*meter|\bph\b|orp|tds|conductivity|dissolved oxygen|do meter|water quality", re.I), {
        "category": "Water Quality Meter",
        "term": "alat ukur kualitas air",
        "function": "mengukur parameter kualitas air atau larutan sesuai spesifikasi produk",
        "tags": ["Lutron", "water quality meter", "pH meter", "conductivity meter", "TDS meter"],
    }),
    (re.compile(r"moisture meter|moisture|wood moisture|grain moisture", re.I), {
        "category": "Moisture Meter",
        "term": "moisture meter",
        "function": "mengukur kadar kelembapan material sesuai spesifikasi produk",
        "tags": ["Lutron", "moisture meter", "wood moisture meter", "material tester"],
    }),
    (re.compile(r"tachometer|rpm|stroboscope", re.I), {
        "category": "Tachometer dan RPM Meter",
        "term": "tachometer atau RPM meter",
        "function": "mengukur putaran, kecepatan, atau parameter terkait sesuai spesifikasi produk",
        "tags": ["Lutron", "tachometer", "RPM meter", "speed meter"],
    }),
    (re.compile(r"force gauge|hardness|tension|coating thickness|thickness gauge", re.I), {
        "category": "Material Tester",
        "term": "alat uji material",
        "function": "mendukung pengukuran gaya, kekerasan, ketebalan, atau parameter material sesuai spesifikasi produk",
        "tags": ["Lutron", "material tester", "force gauge", "hardness tester", "thickness gauge"],
    }),
]


DEFAULT_FAMILY = {
    "category": "Alat Ukur Laboratorium dan Industri",
    "term": "alat ukur laboratorium dan industri",
    "function": "mendukung pengukuran, pemeriksaan, atau kontrol kualitas sesuai spesifikasi produk",
    "tags": ["Lutron", "alat ukur", "alat laboratorium", "alat industri", "quality control"],
}

ACCESSORY_FAMILY = {
    "category": "Aksesori dan Komponen",
    "term": "aksesori atau komponen pendukung",
    "function": "mendukung penggunaan unit utama yang kompatibel sesuai model dan spesifikasi produk",
    "tags": ["Lutron", "aksesori Lutron", "komponen Lutron", "spare part", "kelengkapan alat"],
}


@dataclass(frozen=True)
class SourceConfig:
    key: str
    brand: str
    domains: Tuple[str, ...]


LUTRON_CONFIG = SourceConfig("lutron", "Lutron", ("lutron.com.tw",))
SUPPORTED_SOURCES: Tuple[SourceConfig, ...] = (LUTRON_CONFIG,)


# ---------------------------------------------------------------------------
# Common helpers
# ---------------------------------------------------------------------------

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = html.unescape(str(value))
    s = s.replace("\xa0", " ").replace("\u3000", " ").replace("_x000D_", "\n")
    s = s.replace("\r", "\n")
    s = re.sub(r"[ \t\f\v]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()



# Karakter Han/Chinese tidak boleh masuk ke output WooCommerce.
# Nama model, satuan, angka, tanda teknis, dan istilah merek tetap dipertahankan.
CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
CJK_PUNCT_RE = re.compile(r"[，。；：「」『』（）【】《》、]")
GENERIC_LUTRON_PAGE_RE = re.compile(
    r"(?:lutron\s*(?:taiwan)?\s*official\s*(?:website|site)|"
    r"路昌電子官方網站|專業電子儀表|leading\s+brand\s+specializing\s+in\s+electronic\s+instruments)",
    re.I,
)
GENERIC_LUTRON_CLAUSE_RE = re.compile(
    r"(?:路昌電子官方網站[^。.\n]*(?:[。.]|$)|"
    r"lutron\s*taiwan\s*official\s*(?:website|site)[^.\n]*(?:[.]|$)|"
    r"the\s+leading\s+brand\s+specializing\s+in\s+electronic\s+instruments\s+and\s+meters[.!]?)",
    re.I,
)


def has_cjk(value: Any) -> bool:
    return bool(CJK_RE.search(clean_text(value)))


def strip_cjk(value: Any) -> str:
    """Hapus karakter Han/Chinese tanpa mengubah angka, simbol, model, atau satuan teknis."""
    s = clean_text(value)
    s = CJK_RE.sub(" ", s)
    s = CJK_PUNCT_RE.sub(" ", s)
    s = re.sub(r"\s{2,}", " ", s)
    return s.strip(" -–—:;,")


def is_generic_lutron_landing_text(value: Any) -> bool:
    s = clean_text(value)
    if not s:
        return False
    return bool(GENERIC_LUTRON_PAGE_RE.search(s))


def clean_catalog_source_text(value: Any) -> str:
    """Bersihkan teks sumber agar konten umum/berbahasa Chinese tidak masuk katalog."""
    raw = clean_text(value)
    if not raw:
        return ""
    # Halaman Lutron kadang menyisipkan slogan situs umum dalam satu node dengan
    # isi produk. Buang klausa umum itu, tetapi pertahankan kalimat produk yang
    # berada sesudahnya.
    raw = GENERIC_LUTRON_CLAUSE_RE.sub(" ", raw)
    s = strip_cjk(raw)
    if not s or (is_generic_lutron_landing_text(s) and len(s) < 180):
        return ""
    return s


def html_to_source_text(value: Any, max_chars: int = 8500) -> str:
    """Ubah HTML sumber menjadi teks aman untuk diterjemahkan AI."""
    raw = clean_text(value)
    if not raw:
        return ""
    try:
        soup = BeautifulSoup(raw, "html.parser")
        text = soup.get_text("\n", strip=True)
    except Exception:
        text = re.sub(r"<[^>]+>", " ", raw)
    return compact(clean_catalog_source_text(text), max_chars)


def esc(value: Any) -> str:
    return html.escape(clean_text(value), quote=False)


def compact(text: Any, max_chars: int = 900) -> str:
    s = clean_text(re.sub(r"\s+", " ", clean_text(text))).strip()
    if len(s) <= max_chars:
        return s
    return s[:max_chars].rsplit(" ", 1)[0].rstrip(" ,.;:") + "."


def dedupe_keep_order(items: Iterable[Any]) -> List[str]:
    out: List[str] = []
    seen = set()
    for item in items:
        v = clean_text(item)
        if not v:
            continue
        key = re.sub(r"\s+", " ", v).casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(v)
    return out


def normalize_url(value: Any) -> str:
    """Normalize a Lutron URL while preserving the host provided in Excel.

    Product URLs supplied by Lutron use ``www.lutron.com.tw``. The old script
    incorrectly forced those URLs to the non-www host. That host can return a
    minimal shell instead of the product document, which leaves image URL,
    description, full content, and specs blank.
    """
    url = clean_text(value)
    if not url:
        return ""
    if url.startswith("//"):
        url = "https:" + url
    elif not re.match(r"https?://", url, flags=re.I):
        url = "https://" + url.lstrip("/")
    parts = urlsplit(url)
    path = quote(unquote(parts.path), safe="/%+-_.,;~@()[]")
    query = quote(unquote(parts.query), safe="=&%+-_.,;~@()[]/:")
    fragment = quote(unquote(parts.fragment), safe="%+-_.,;~@()[]/:")
    return urlunsplit((parts.scheme or "https", parts.netloc, path, query, fragment))


def get_source_config(url: str) -> Optional[SourceConfig]:
    """Return configuration only for the official Lutron Taiwan domain."""
    host = (urlparse(normalize_url(url)).hostname or "").casefold()
    if any(host == domain or host.endswith("." + domain) for domain in LUTRON_CONFIG.domains):
        return LUTRON_CONFIG
    return None


def deterministic_index(text: str, modulo: int) -> int:
    if modulo <= 1:
        return 0
    h = hashlib.md5(clean_text(text).encode("utf-8")).hexdigest()
    return int(h[:8], 16) % modulo


def is_noise_text(text: str) -> bool:
    key = re.sub(r"\s+", " ", clean_text(text)).casefold().strip(" :|-")
    return not key or key in NOISE_TEXT


def normalize_label(label: str) -> str:
    label = clean_text(label).strip(" :-–—|")
    if not label:
        return ""
    label = re.sub(r"\s+", " ", label)
    # Pertahankan model, akronim, dan penulisan vendor seperti apa adanya.
    return label[:1].upper() + label[1:]


def table_html(rows: Sequence[Dict[str, str]]) -> str:
    body = []
    for row in rows:
        label, value = esc(row.get("label")), esc(row.get("value"))
        if label and value:
            body.append(f"<tr><th>{label}</th><td>{value}</td></tr>")
    return '<table class="uji-spec-table"><tbody>' + "".join(body) + "</tbody></table>" if body else ""


def _is_technical_spec_section(value: Any) -> bool:
    """True when a source section is a dedicated technical narrative block."""
    text = clean_text(value)
    return bool(re.search(r"\b(?:technical\s+specifications?|technical\s+data|measurement\s+specifications?)\b", text, re.I))


def specifications_tab_html(rows: Sequence[Dict[str, str]], already_deduplicated: bool = False) -> str:
    """Render one specification table.

    When DeepSeek AI deduplication is enabled, rows have already been selected
    from the full raw source list. Do not run the rule-based deduplicator again,
    otherwise it could remove a separate Meter/Probe or range-specific fact
    after the AI has deliberately retained it.
    """
    if already_deduplicated:
        cleaned_rows: List[Dict[str, str]] = []
        seen = set()
        for raw in rows or []:
            if not isinstance(raw, dict):
                continue
            label = clean_text(raw.get("label"))
            value = clean_text(raw.get("value"))
            if not label or not value:
                continue
            # Guard only against an exact repeated output row. Semantic duplicate
            # selection remains DeepSeek's responsibility in AI mode.
            key = (_spec_text_key(label), _spec_text_key(value))
            if key in seen:
                continue
            seen.add(key)
            cleaned_rows.append({"label": label, "value": value})
    else:
        cleaned_rows, _issues = canonicalize_specifications(rows)

    if not cleaned_rows:
        return "<h2>Spesifikasi</h2><p>Spesifikasi belum tersedia dalam bahasa Indonesia.</p>"
    return "<h2>Spesifikasi</h2>" + table_html(cleaned_rows)


# ---------------------------------------------------------------------------
# Excel input
# ---------------------------------------------------------------------------

def find_column(df: pd.DataFrame, aliases: Sequence[str]) -> str:
    cols = {str(c).strip().casefold(): c for c in df.columns}
    for alias in aliases:
        found = cols.get(alias.casefold())
        if found is not None:
            return str(found)
    return ""


def find_url_column(df: pd.DataFrame, requested: str = "") -> str:
    if requested:
        exact = {str(c).strip().casefold(): c for c in df.columns}.get(requested.strip().casefold())
        if exact is None:
            raise ValueError(f"Kolom URL '{requested}' tidak ditemukan. Kolom tersedia: {list(df.columns)}")
        return str(exact)
    result = find_column(df, ["url", "link", "source url", "product url", "product link", "product_link"])
    if result:
        return result
    for col in df.columns:
        vals = df[col].dropna().astype(str).head(30).tolist()
        if any("lutron.com.tw" in v.casefold() or v.casefold().startswith("http") for v in vals):
            return str(col)
    raise ValueError("Tidak menemukan kolom URL. Gunakan url/link/Source URL atau --url-column.")


def read_input_records(input_path: str, sheet: Any = 0, url_column: str = "") -> List[Dict[str, str]]:
    df = pd.read_excel(input_path, sheet_name=sheet)
    if df.empty:
        return []
    url_col = find_url_column(df, url_column)
    brand_col = find_column(df, ["brand", "merek"])
    title_col = find_column(df, ["title", "name", "product name", "nama", "nama produk"])
    model_col = find_column(df, ["model", "sku", "product model", "kode model"])

    records: List[Dict[str, str]] = []
    seen = set()
    for _, raw in df.iterrows():
        url = normalize_url(raw.get(url_col))
        if not url or url.casefold() in seen:
            continue
        seen.add(url.casefold())
        records.append({
            "url": url,
            "input_brand": clean_text(raw.get(brand_col)) if brand_col else "",
            "input_title": clean_text(raw.get(title_col)) if title_col else "",
            "input_model": clean_text(raw.get(model_col)) if model_col else "",
        })
    return records


# ---------------------------------------------------------------------------
# HTML parsing
# ---------------------------------------------------------------------------

def soup_text(node: Any) -> str:
    if node is None:
        return ""
    try:
        return clean_text(node.get_text(" ", strip=True))
    except Exception:
        return clean_text(node)


def clean_title(value: str) -> str:
    title = clean_text(value)
    title = re.sub(r"\s*[|｜]\s*(?:Lutron(?: Electronic)?(?: Taiwan)?).*?$", "", title, flags=re.I)
    title = re.sub(r"\s+", " ", title)
    return title.strip(" -|\t\n")


def jsonld_objects(soup: BeautifulSoup) -> Iterator[Dict[str, Any]]:
    def walk(v: Any) -> Iterator[Dict[str, Any]]:
        if isinstance(v, dict):
            yield v
            graph = v.get("@graph")
            if isinstance(graph, list):
                for x in graph:
                    yield from walk(x)
        elif isinstance(v, list):
            for x in v:
                yield from walk(x)

    for node in soup.select('script[type="application/ld+json"]'):
        raw = node.string or node.get_text() or ""
        raw = clean_text(raw)
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            # Situs tertentu menyisipkan komentar HTML atau beberapa JSON object.
            continue
        yield from walk(data)


def product_jsonld(soup: BeautifulSoup) -> List[Dict[str, Any]]:
    out = []
    for obj in jsonld_objects(soup):
        types = obj.get("@type", [])
        types = [types] if isinstance(types, str) else types
        if any(clean_text(t).casefold() == "product" for t in types if clean_text(t)):
            out.append(obj)
    return out


def first_meta(soup: BeautifulSoup, selectors: Sequence[str]) -> str:
    for sel in selectors:
        node = soup.select_one(sel)
        if node:
            value = clean_text(node.get("content") or node.get("href") or "")
            if value:
                return value
    return ""


def extract_title(soup: BeautifulSoup, fallback: str = "") -> str:
    for obj in product_jsonld(soup):
        title = clean_text(obj.get("name"))
        if title:
            return clean_title(title)
    for selector in ['h1', '[itemprop="name"]', '.product-title', '.product_name', '.product-name', '.productTitle']:
        node = soup.select_one(selector)
        title = soup_text(node)
        if title and len(title) <= 300:
            return clean_title(title)
    meta = first_meta(soup, ['meta[property="og:title"]', 'meta[name="twitter:title"]'])
    if meta:
        return clean_title(meta)
    if soup.title:
        title = clean_title(soup_text(soup.title))
        if title:
            return title
    return clean_title(fallback)


def valid_image_url(url: str) -> bool:
    """Return True only for a plausible product photograph URL.

    Lutron pages load site banners and corporate artwork together with product
    assets.  A generic visual is worse than an empty image field, so uncertain
    URLs are rejected before they can reach WooCommerce.
    """
    value = clean_text(url)
    if not value or value.startswith(("data:", "javascript:", "#")):
        return False
    low = value.casefold()
    if any(token in low for token in IMAGE_REJECT_TOKENS):
        return False
    # Common query parameters used by tracking pixels and social previews.
    if re.search(r"(?:[?&](?:utm_|fbclid|gclid|tracking)=|/pixel(?:[/?]|$))", low, re.I):
        return False
    return True


def _iter_image_values(value: Any) -> Iterator[str]:
    """Yield raw image URLs from scalar, mapping, or list values."""
    if isinstance(value, str):
        candidate = clean_text(value)
        if candidate:
            yield candidate
        return
    if isinstance(value, dict):
        for key in (
            "url", "contentUrl", "src", "srcUrl", "image", "imageUrl", "image_url",
            "large", "largeImage", "thumbnailUrl", "thumbnail", "path",
        ):
            if key in value:
                yield from _iter_image_values(value.get(key))
        return
    if isinstance(value, (list, tuple)):
        for item in value:
            yield from _iter_image_values(item)


def _srcset_urls(value: Any) -> Iterator[str]:
    srcset = clean_text(value)
    if not srcset:
        return
    for part in srcset.split(","):
        raw = clean_text(part.strip().split(" ")[0])
        if raw:
            yield raw


def _url_from_style(value: Any) -> Iterator[str]:
    style = clean_text(value)
    if not style:
        return
    for raw in re.findall(r"url\(\s*['\"]?([^'\")\s]+)['\"]?\s*\)", style, flags=re.I):
        raw = clean_text(raw)
        if raw:
            yield raw


def _image_score(url: str, source: str, model: str = "") -> int:
    """Rank likely product photos above generic interface images."""
    value = clean_text(url)
    low = value.casefold()
    score = {
        "jsonld": 190,
        "og": 170,
        "meta": 150,
        "product_attr": 145,
        "srcset": 125,
        "any_attr": 75,
        "style": 65,
        "script": 55,
        # Browser candidates are visually related to the product area but are
        # not automatically authoritative.  Generic banners used to receive
        # an excessive score here.
        "dom_primary": 140,
        "network": 85,
        "link": 45,
    }.get(source, 0)

    if re.search(r"(?:product|products|catalog|upload|uploads|storage|media|image|images|item|goods|photo)", low):
        score += 20
    if re.search(r"\.(?:jpe?g|png|webp|avif|gif)(?:[?#]|$)", low):
        score += 12
    if "thumbnail" in low or "thumb" in low:
        score -= 10

    normalized_model = re.sub(r"[^a-z0-9]+", "", clean_text(model).casefold())
    normalized_url = re.sub(r"[^a-z0-9]+", "", low)
    if normalized_model and len(normalized_model) >= 3 and normalized_model in normalized_url:
        score += 220
    # Images from anonymous assets are accepted only when the extraction source
    # is already strongly tied to the product DOM.
    if source in {"any_attr", "style", "script", "link", "network"} and not (
        normalized_model and normalized_model in normalized_url
    ):
        score -= 35
    return score


def extract_best_image_url(soup: BeautifulSoup, base_url: str, model: str = "", network_images: Optional[Sequence[str]] = None) -> str:
    """Extract a product image from Lutron metadata, lazy-load attributes, and page data.

    Lutron pages may expose the photo in different forms depending on the page
    template: Open Graph tags, JSON-LD, lazy-load attributes, source/srcset
    elements, inline style, or serialized page data. Every path is checked here.
    """
    candidates: List[Tuple[str, str]] = []

    def add(raw: Any, source: str) -> None:
        for value in _iter_image_values(raw):
            value = clean_text(value).replace("\\/", "/").strip(" '\"")
            if value:
                candidates.append((source, value))

    # JS-rendered Lutron pages sometimes expose a product photo only as a
    # network image request. Capture it as a candidate too.
    for network_image in (network_images or []):
        raw_network = clean_text(network_image)
        if raw_network.startswith("__LUTRON_PRIMARY__"):
            add(raw_network[len("__LUTRON_PRIMARY__"):], "dom_primary")
        else:
            add(raw_network, "network")

    # Product structured data can hold either a string, an object, or a list.
    for obj in product_jsonld(soup):
        add(obj.get("image"), "jsonld")
        add(obj.get("photo"), "jsonld")
        add(obj.get("thumbnailUrl"), "jsonld")

    # Primary social metadata generally points to the canonical product image.
    for selector in (
        'meta[property="og:image"]', 'meta[property="og:image:url"]',
        'meta[property="og:image:secure_url"]', 'meta[name="twitter:image"]',
        'meta[name="twitter:image:src"]', 'meta[name="image"]',
        'link[rel="image_src"]', 'link[rel="preload"][as="image"]',
    ):
        node = soup.select_one(selector)
        if node:
            add(node.get("content") or node.get("href"), "og" if "og:" in selector else "meta")

    product_selectors = (
        ".product img", ".product-detail img", ".product-detail-image img",
        ".product-image img", ".product-images img", ".gallery img", ".carousel img",
        "[class*='product'] img", "[class*='gallery'] img", "[id*='product'] img",
    )
    product_nodes = set()
    for selector in product_selectors:
        for node in soup.select(selector):
            product_nodes.add(id(node))
            for attr, raw in node.attrs.items():
                attr_key = clean_text(attr).casefold()
                if attr_key in {"srcset", "data-srcset"}:
                    for url in _srcset_urls(raw):
                        add(url, "srcset")
                elif any(token in attr_key for token in ("src", "image", "photo", "zoom", "original", "large", "lazy")):
                    add(raw, "product_attr")
            for url in _url_from_style(node.get("style")):
                add(url, "style")

    # Generic image tags and unusual lazy-load attribute names.
    for node in soup.find_all(["img", "source", "a", "div", "figure", "picture"]):
        source = "product_attr" if id(node) in product_nodes else "any_attr"
        for attr, raw in node.attrs.items():
            attr_key = clean_text(attr).casefold()
            if attr_key in {"srcset", "data-srcset"}:
                for url in _srcset_urls(raw):
                    add(url, "srcset")
            elif any(token in attr_key for token in ("src", "image", "photo", "zoom", "original", "large", "lazy", "background")):
                add(raw, source)
        for url in _url_from_style(node.get("style")):
            add(url, "style")

    # Some templates only expose the high-resolution photo as an <a href>.
    for node in soup.select("a[href]"):
        href = clean_text(node.get("href"))
        if re.search(r"\.(?:jpe?g|png|webp|avif|gif)(?:[?#]|$)", href, flags=re.I):
            add(href, "link")

    # Final fallback: serialized page state inside scripts. Do not require a file
    # extension because CMS/CDN endpoints may serve images from dynamic paths.
    raw_html = str(soup).replace("\\/", "/")
    patterns = (
        r'"(?:image|imageUrl|image_url|photo|photoUrl|thumbnail|thumbnailUrl|mainImage)"\s*:\s*"([^"]+)"',
        r'(?:(?:https?:)?//[^"\'<>\\\s]+(?:\.(?:jpe?g|png|webp|avif|gif)|/image/|/images/|/media/|/storage/)[^"\'<>\\\s]*)',
        r'(?:/(?:storage|uploads?|images?|media|assets)/[^"\'<>\\\s]+(?:\.(?:jpe?g|png|webp|avif|gif))?[^"\'<>\\\s]*)',
    )
    for pattern in patterns:
        for match in re.finditer(pattern, raw_html, flags=re.I):
            value = clean_text(match.group(1) if match.groups() else match.group(0))
            if value:
                add(value, "script")

    ranked: List[Tuple[int, str, str]] = []
    seen = set()
    for source, raw in candidates:
        full = urljoin(base_url, raw)
        full = clean_text(full).replace(" ", "%20")
        if not valid_image_url(full):
            continue
        key = full.casefold()
        if key in seen:
            continue
        seen.add(key)
        ranked.append((_image_score(full, source, model), full, source))

    if not ranked:
        return ""
    ranked.sort(key=lambda item: item[0], reverse=True)
    best_score, best_url, best_source = ranked[0]
    # Do not publish a weakly associated random asset.  JSON-LD, OG, a real
    # product container, or a browser candidate near the product title may use
    # a non-model filename. Other sources need the model token to be trusted.
    trusted_sources = {"jsonld", "og", "meta", "product_attr", "srcset", "dom_primary"}
    model_token = _model_token(model)
    url_token = _model_token(best_url)
    if best_source not in trusted_sources and not (model_token and model_token in url_token):
        return ""
    if best_score < 105:
        return ""
    return best_url


def extract_breadcrumb(soup: BeautifulSoup) -> str:
    selectors = ['nav[aria-label*="breadcrumb" i]', '.breadcrumb', '.breadcrumbs', '[class*="breadcrumb"]', '[id*="breadcrumb"]']
    for selector in selectors:
        node = soup.select_one(selector)
        text = soup_text(node)
        if text and len(text) <= 500:
            parts = [x.strip() for x in re.split(r"\s*(?:/|>|›|»|\||→)\s*", text) if x.strip()]
            parts = [p for p in parts if p.casefold() not in {"home", "products", "product"}]
            if parts:
                return " > ".join(parts[-3:])
    return ""


def extract_category(soup: BeautifulSoup, json_products: Sequence[Dict[str, Any]], fallback: str = "") -> str:
    for obj in json_products:
        cat = obj.get("category")
        if isinstance(cat, list):
            cat = " > ".join(clean_text(x) for x in cat if clean_text(x))
        cat = clean_text(cat)
        if cat:
            return cat
    breadcrumb = extract_breadcrumb(soup)
    if breadcrumb:
        return breadcrumb.split(" > ")[-1]
    for selector in ['[class*="category"] a', '[class*="category"]', '[itemprop="category"]']:
        node = soup.select_one(selector)
        text = soup_text(node)
        if text and len(text) <= 150 and not is_noise_text(text):
            return text
    return clean_text(fallback)


def choose_main_container(soup: BeautifulSoup) -> Tag:
    for selector in ['main', '[role="main"]', '#product', '.product-detail', '.product-details', '.product-page', '.product-info', '.product']:
        node = soup.select_one(selector)
        if isinstance(node, Tag):
            return node
    return soup.body if isinstance(soup.body, Tag) else soup


def strip_noise_nodes(root: Tag) -> None:
    selectors = [
        'script', 'style', 'noscript', 'svg', 'iframe', 'header', 'footer', 'nav', 'aside', 'form',
        '.navbar', '.navigation', '.menu', '.sidebar', '.footer', '.header', '.cookie', '.modal',
        '.social', '.share', '.related-products', '.breadcrumb', '.breadcrumbs', '[aria-label*="breadcrumb" i]',
    ]
    for selector in selectors:
        for node in root.select(selector):
            node.decompose()


def extract_description(soup: BeautifulSoup, products: Sequence[Dict[str, Any]], title: str) -> str:
    for obj in products:
        desc = clean_text(obj.get("description"))
        if desc:
            return compact(desc, 3500)
    for selector in [
        '[itemprop="description"]', '.product-description', '.product_desc', '.description',
        '[class*="description"]', '[id*="description"]', '.summary', '.product-summary',
    ]:
        nodes = soup.select(selector)
        for node in nodes:
            text = soup_text(node)
            if len(text) >= 40 and len(text) <= 6000 and clean_text(text).casefold() != clean_text(title).casefold():
                return compact(text, 3500)
    meta = first_meta(soup, ['meta[name="description"]', 'meta[property="og:description"]'])
    return compact(meta, 1500)


def normalize_content_line(text: str) -> str:
    text = clean_text(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" |\t\n")


def _node_text_lines(node: Tag) -> List[str]:
    """Return visible node text while retaining manual line breaks from <br> elements."""
    try:
        raw = node.get_text("\n", strip=True)
    except Exception:
        raw = soup_text(node)
    return [normalize_content_line(part) for part in clean_text(raw).splitlines() if normalize_content_line(part)]


def _looks_like_section_heading(node: Tag, value: str) -> bool:
    """Recognise semantic and div-based headings used by Lutron/Quasar pages."""
    text = normalize_content_line(value)
    heading_text = text.strip(" :–—-|")
    if not heading_text or len(heading_text) > 150:
        return False
    if node.name in {"h1", "h2", "h3", "h4", "h5", "h6"}:
        return not is_noise_text(heading_text)
    # Some Lutron templates use a styled div/span instead of an h2/h3.
    if SPEC_HEADING_RE.fullmatch(heading_text) or re.fullmatch(
        r"(?:technical\s+specifications?|technical\s+data|specifications?|"
        r"features|versatile\s+applications|applications|standard\s+accessories|"
        r"optional\s+accessories|data\s+output)",
        heading_text,
        flags=re.I,
    ):
        return True
    class_id = " ".join([clean_text(node.get("class")), clean_text(node.get("id"))]).casefold()
    return bool(re.search(r"(?:title|heading|section).*(?:spec|technical)|(?:spec|technical).*(?:title|heading|section)", class_id))


def extract_all_content(root: Tag, title: str) -> List[Tuple[str, str, str]]:
    """Return non-noise product content as ``(kind, section, text)``.

    This deliberately preserves line breaks inside paragraphs/divs while
    excluding table rows, which are parsed separately. On Lutron pages the
    ``Technical Specifications`` block can be a Quasar ``div`` with
    ``<br>`` separated facts rather than semantic list/table markup. The old
    parser ignored that shape or collapsed the full block into one sentence.
    """
    items: List[Tuple[str, str, str]] = []
    heading = "Informasi Produk"
    seen = set()
    max_items = 420
    target_names = ["h1", "h2", "h3", "h4", "h5", "h6", "p", "li", "dt", "dd", "div"]

    for node in root.find_all(target_names, recursive=True):
        if len(items) >= max_items or not isinstance(node, Tag):
            break
        # A div that contains a semantic child is only a layout wrapper. Reading
        # it would duplicate every line below. Leaf divs still matter because
        # Lutron often uses them as visual section headings or text blocks.
        if node.name == "div" and node.find(["h1", "h2", "h3", "h4", "h5", "h6", "p", "li", "dt", "dd", "tr", "div"], recursive=False):
            continue

        # Table rows belong exclusively to the structured-table extractor in
        # extract_specs(). Reading them again as narrative content causes the
        # same source fact to appear twice in the final WooCommerce table.
        lines = _node_text_lines(node)
        kind = "li" if node.name == "li" else "text"

        for line in lines:
            if _looks_like_section_heading(node, line):
                heading = line.strip(" :–—-|")[:180]
                continue
            if not line or is_noise_text(line) or len(line) < 2:
                continue
            if line.casefold() == clean_text(title).casefold():
                continue
            key = (heading.casefold(), line.casefold())
            if key in seen:
                continue
            seen.add(key)
            items.append((kind, heading, line))
    return items


def all_content_html(items: Sequence[Tuple[str, str, str]], title: str, max_chars: int = 16000) -> str:
    if not items:
        return ""
    blocks: List[str] = []
    current = ""
    bullets: List[str] = []
    used = 0

    def flush_bullets() -> None:
        nonlocal bullets
        if bullets:
            blocks.append("<ul>" + "".join(f"<li>{esc(x)}</li>" for x in bullets) + "</ul>")
            bullets = []

    for kind, heading, text in items:
        addition = len(heading) + len(text) + 30
        if used + addition > max_chars:
            break
        used += addition
        if heading != current:
            flush_bullets()
            if heading:
                blocks.append(f"<h3>{esc(heading)}</h3>")
            current = heading
        if kind == "li":
            bullets.append(text)
        else:
            flush_bullets()
            blocks.append(f"<p>{esc(text)}</p>")
    flush_bullets()
    return "".join(blocks)


def clean_spec_value(value: str) -> str:
    value = normalize_content_line(value)
    value = re.sub(r"\s*\|\s*", " / ", value)
    value = re.sub(r"\s{2,}", " ", value)
    return value.strip(" :–—|/")


def looks_like_spec_label(value: str) -> bool:
    value = clean_spec_value(value)
    if not (2 <= len(value) <= 120):
        return False
    if len(value.split()) > 11:
        return False
    return True


def _spec_text_key(value: Any) -> str:
    text = clean_spec_value(value).casefold()
    text = text.replace("℃", "c").replace("°", "")
    return re.sub(r"[^a-z0-9]+", "", text)


def _spec_numbers(value: Any) -> List[str]:
    return [token.replace(",", ".") for token in re.findall(r"[-+]?\d+(?:[.,]\d+)?", clean_text(value))]


def _spec_unit_signature(value: Any) -> Tuple[str, ...]:
    """Return units that make equal numeric values comparable safely."""
    text = clean_text(value).casefold().replace("℃", "c").replace("℉", "f")
    units = re.findall(r"(?:%\s*r\.?h\.?|%|(?:dc|ac)|ma|mv|kv|v|w|(?:kg|mg|lb|lbs|gram|g)|(?:mm|cm|inch|in\.?|m)|°?\s*[cf])", text, re.I)
    return tuple(sorted(set(re.sub(r"\s+", "", u).strip(".") for u in units)))


def _spec_values_equivalent(left: Any, right: Any) -> bool:
    """Compare values without treating changed technical numbers as duplicates.

    Lutron writes the same fact in several cosmetic forms, such as
    ``Less than 80% RH`` and ``< 80 % RH``. Equal numbers plus the same unit
    category are equivalent for humidity, temperature, power, dimensions, and
    weight. Values with different units are never merged merely because their
    numeric tokens match.
    """
    a, b = _spec_text_key(left), _spec_text_key(right)
    if not a or not b:
        return False
    if a == b:
        return True
    a_numbers, b_numbers = _spec_numbers(left), _spec_numbers(right)
    if a_numbers != b_numbers:
        return False
    a_units, b_units = _spec_unit_signature(left), _spec_unit_signature(right)
    # Different source wording around one physical fact is cosmetic when
    # numbers and unit category agree.
    if (_dimension_like(left) and _dimension_like(right)) or (_weight_like(left) and _weight_like(right)):
        return True
    if _humidity_like(left) and _humidity_like(right):
        return True
    if ((_temperature_like(left) and _temperature_like(right)) or (_power_like(left) and _power_like(right))) and a_units == b_units:
        return True
    short, long = min(a, b, key=len), max(a, b, key=len)
    return len(short) >= 10 and short in long


# Canonical keys express fact meaning. The rules accept English source labels
# and Indonesian labels produced by AI/fallback so the final audit can run both
# before and after translation.
_CANONICAL_LABEL_RULES: List[Tuple[re.Pattern[str], str, str]] = [
    (re.compile(r"\b(?:model|sku|mpn|catalog(?:ue)?\s*(?:no|number)?|part\s*(?:no|number)|item\s*(?:no|number))\b", re.I), "model", "Model"),
    (re.compile(r"\b(?:main\s*(?:instrument|unit)|instrumen\s*utama).*\b(?:dimension|dimensions|dimensi)\b|\b(?:main\s*(?:instrument|unit)|instrumen\s*utama)\b", re.I), "dimensions_main_unit", "Dimensi instrumen utama"),
    (re.compile(r"\b(?:meter|instrument|unit|alat)\s*(?:dimension|dimensions|dimensi)\b|\b(?:dimension|dimensions|dimensi)\s*(?:meter|instrument|unit|alat)\b", re.I), "dimensions_meter", "Dimensi meter"),
    (re.compile(r"\b(?:probe|sensor)\s*(?:dimension|dimensions|dimensi)\b|\b(?:dimension|dimensions|dimensi)\s*(?:probe|sensor)\b", re.I), "dimensions_probe", "Dimensi probe"),
    (re.compile(r"\b(?:dimension|dimensions|dimensi)\b", re.I), "dimensions", "Dimensi"),
    (re.compile(r"\b(?:meter|instrument|unit|alat)\s*(?:weight|berat)\b|\b(?:weight|berat)\s*(?:meter|instrument|unit|alat)\b", re.I), "weight_meter", "Berat meter"),
    (re.compile(r"\b(?:probe|sensor)\s*(?:weight|berat)\b|\b(?:weight|berat)\s*(?:probe|sensor)\b", re.I), "weight_probe", "Berat probe"),
    (re.compile(r"\b(?:weight|berat)\b", re.I), "weight", "Berat"),
    (re.compile(r"\b(?:operating|working|ambient)\s*(?:temperature|suhu)\b|\b(?:temperature|suhu)\s*(?:operasi|kerja|lingkungan)\b", re.I), "operating_temperature", "Suhu operasi"),
    (re.compile(r"\b(?:operating|working|ambient)\s*(?:humidity|kelembapan|kelembaban)\b|\b(?:humidity|kelembapan|kelembaban)\s*(?:operasi|kerja|lingkungan)\b", re.I), "operating_humidity", "Kelembapan operasi"),
    (re.compile(r"\b(?:measurement|measuring)\s*(?:range|rentang)|\b(?:range|rentang)\s*(?:measurement|pengukuran)\b", re.I), "measurement_range", "Rentang pengukuran"),
    (re.compile(r"\bfrequency\s+response\b|\brespons\s+frekuensi\b", re.I), "frequency_response", "Respons frekuensi"),
    (re.compile(r"\bfrequency\s+(?:range|band)\b|\brentang\s+frekuensi\b", re.I), "frequency_range", "Rentang frekuensi"),
    (re.compile(r"\bfrequency\s+accuracy\b|\bakurasi\s+frekuensi\b", re.I), "frequency_accuracy", "Akurasi frekuensi"),
    (re.compile(r"\b(?:data\s+error|error\s+no\.?|kesalahan\s+data)\b", re.I), "data_error_rate", "Kesalahan data"),
    (re.compile(r"\b(?:full\s+scale\s+deflection|defleksi\s+skala\s+penuh)\b", re.I), "full_scale_deflection", "Defleksi skala penuh"),
    (re.compile(r"\b(?:resolution|resolusi)\b", re.I), "resolution", "Resolusi"),
    (re.compile(r"\b(?:accuracy|akurasi|error|kesalahan)\b", re.I), "accuracy", "Akurasi"),
    (re.compile(r"\b(?:display\s+sampling\s+time|sampling\s+time\s+of\s+display|waktu\s+sampling\s+tampilan)\b", re.I), "display_sampling_time", "Waktu sampling tampilan"),
    (re.compile(r"\b(?:data\s*logger|datalogger|pencatat\s+data).*\b(?:sampling\s+time|waktu\s+sampling)\b", re.I), "datalogger_sampling_time", "Waktu sampling pencatat data"),
    (re.compile(r"\b(?:sampling\s+time|waktu\s+sampling)\b", re.I), "sampling_time", "Waktu sampling"),
    (re.compile(r"\b(?:power\s+supply|catu\s+daya|sumber\s+daya)\b", re.I), "power_supply", "Catu daya"),
    # Specific power modes must be evaluated before generic power draw.
    # Otherwise "Power Consumption - Normal Operation" and
    # "Power Consumption - Data Logging" collapse into the same key and are
    # falsely reported as conflicting values.
    (re.compile(r"\b(?:normal\s+operation|operasi\s+normal)\b", re.I), "power_draw_normal_operation", "Konsumsi daya operasi normal"),
    (re.compile(r"\b(?:data\s*(?:save|logging)|penyimpanan\s+data)\b", re.I), "power_draw_data_logging", "Konsumsi daya saat pencatatan data"),
    (re.compile(r"\b(?:power\s+consumption|power\s+current|arus\s+daya|konsumsi\s+daya)\b", re.I), "power_draw", "Konsumsi daya"),
    (re.compile(r"\b(?:accessories?\s+included|included\s+accessories?|aksesori\s+yang\s+disertakan)\b", re.I), "accessories_included", "Aksesori yang disertakan"),
    (re.compile(r"\b(?:standard\s+accessories?|aksesori\s+standar)\b", re.I), "accessories_standard", "Aksesori standar"),
    (re.compile(r"\b(?:optional\s+accessories?|aksesori\s+opsional)\b", re.I), "accessories_optional", "Aksesori opsional"),
    (re.compile(r"\b(?:data\s+output|output|interface|communication|communication\s+interface|keluaran\s+data|antarmuka)\b", re.I), "interface_output", "Antarmuka dan keluaran data"),
    (re.compile(r"\b(?:display\s+direction|arah\s+tampilan)\b", re.I), "display_direction", "Arah tampilan"),
    (re.compile(r"\b(?:display|tampilan|lcd)\b", re.I), "display", "Tampilan"),
    (re.compile(r"\b(?:memory\s+card|sd\s*card|kartu\s+memori)\b", re.I), "memory_card", "Kartu memori"),
    (re.compile(r"\b(?:operating\s+humidity|humidity|kelembapan|kelembaban)\b", re.I), "humidity", "Kelembapan"),
    (re.compile(r"\b(?:temperature|suhu)\b", re.I), "temperature", "Suhu"),
]


def _canonical_label_info(row: Dict[str, Any]) -> Tuple[str, str, str]:
    source_label = clean_text(row.get("source_label") or row.get("label"))
    display_label = clean_text(row.get("label"))
    probe = " ".join(x for x in (source_label, display_label) if x)
    for pattern, key, label in _CANONICAL_LABEL_RULES:
        if pattern.search(probe):
            return key, label, source_label
    # Preserve an unknown but valid field rather than forcing a wrong category.
    fallback_label = normalize_label(display_label or source_label) or "Rincian teknis"
    fallback_key = "other:" + _spec_text_key(source_label or fallback_label)[:80]
    return fallback_key, fallback_label, source_label


def _dimension_like(value: Any) -> bool:
    return bool(re.search(r"\b\d+(?:[.,]\d+)?\s*[x×]\s*\d+(?:[.,]\d+)?(?:\s*[x×]\s*\d+(?:[.,]\d+)?)?\s*(?:mm|cm|m|inch|in\.?)(?:\b|\s)", clean_text(value), re.I))


def _weight_like(value: Any) -> bool:
    return bool(re.search(r"\b\d+(?:[.,]\d+)?\s*(?:mg|g|kg|lb|lbs|gram)\b", clean_text(value), re.I))


def _humidity_like(value: Any) -> bool:
    return bool(re.search(r"\b(?:less\s+than|under|below|kurang\s+dari|≤|<)?\s*\d+(?:[.,]\d+)?\s*%\s*(?:r\.?h\.?)?\b", clean_text(value), re.I))


def _temperature_like(value: Any) -> bool:
    return bool(re.search(r"[-+]?\d+(?:[.,]\d+)?\s*°?\s*(?:c|f|℃|℉)\b", clean_text(value), re.I))


def _power_like(value: Any) -> bool:
    text = clean_text(value)
    return bool(re.search(r"\b(?:dc|ac)\s*\d|\b\d+(?:[.,]\d+)?\s*(?:v|volt|ma|a|w|watt)\b|\b(?:adapter|adaptor)\b", text, re.I))


def _accessory_like(value: Any) -> bool:
    # Battery wording appears in normal weight/power statements, so it is not
    # enough evidence by itself to classify a value as an accessory list.
    return bool(re.search(r"\b(?:manual|book|buku|cable|kabel|case|tas|software|clip|holder|stand|tip|penetrometer|electrode|electroda)\b", clean_text(value), re.I))


def _value_has_multiple_field_signatures(value: Any) -> List[str]:
    found = []
    if _dimension_like(value):
        found.append("dimensions")
    if _weight_like(value):
        found.append("weight")
    if _humidity_like(value):
        found.append("humidity")
    if _temperature_like(value):
        found.append("temperature")
    if _power_like(value):
        found.append("power")
    if _accessory_like(value):
        found.append("accessories")
    return found


def _dedupe_value_fragments(value: Any) -> str:
    """Remove repeated identical clauses inside a single scraped value.

    It only removes an exact cosmetic duplicate. It never performs a unit
    conversion or drops a clause that contains different numbers.
    """
    raw = clean_spec_value(value)
    if not raw:
        return ""
    # Normalise a repeated complete half: "A / B / A / B" -> "A / B".
    parts = [clean_spec_value(p) for p in re.split(r"\s*/\s*", raw) if clean_spec_value(p)]
    if len(parts) >= 4 and len(parts) % 2 == 0:
        half = len(parts) // 2
        if all(_spec_values_equivalent(parts[i], parts[i + half]) for i in range(half)):
            parts = parts[:half]
    # Drop exact duplicated slash clauses while keeping the first occurrence.
    deduped: List[str] = []
    seen = set()
    for part in parts:
        key = _spec_text_key(part)
        if key and key not in seen:
            seen.add(key)
            deduped.append(part)
    if len(deduped) >= 2:
        return " / ".join(deduped)
    return raw


def _canonical_base_key(key: str) -> str:
    if key.startswith("dimensions_"):
        return "dimensions"
    if key.startswith("weight_"):
        return "weight"
    if key in {"humidity", "operating_humidity"}:
        return "humidity"
    if key in {"temperature", "operating_temperature"}:
        return "temperature"
    return key


def _canonical_keys_can_compare(left: str, right: str) -> bool:
    return left == right or _canonical_base_key(left) == _canonical_base_key(right)


def _specificity_score(key: str, source_label: str, section: str) -> Tuple[int, int, int]:
    """Rank duplicate facts while preserving the user's source priority.

    When Technical Specifications and Specifications contain the same fact,
    Technical Specifications must be kept. A regular Specifications table is
    used only when the technical section has no equivalent fact.
    """
    component = int(key.startswith("dimensions_") or key.startswith("weight_"))
    section_text = clean_text(section)
    if _is_technical_spec_section(section_text):
        source = 3
    elif re.search(r"(?:structured|datasheet|specifications)", section_text, re.I):
        source = 2
    else:
        source = 1
    label = min(80, len(clean_text(source_label)))
    return (component, source, label)


def _issue(code: str, row: Dict[str, Any], detail: str, severity: str = "warning") -> Dict[str, str]:
    """Create a review record with an explicit severity.

    Cross-section wording differences are normal on Lutron pages. They should
    be visible for audit, but they must not suppress the entire catalog. Only a
    strong label/value mismatch is a publication blocker.
    """
    return {
        "Issue Code": code,
        "Severity": severity,
        "Detail": detail,
        "Canonical Key": clean_text(row.get("canonical_key")),
        "Display Label": clean_text(row.get("label")),
        "Source Label": clean_text(row.get("source_label")),
        "Value": clean_text(row.get("value")),
        "Source Section": clean_text(row.get("section")),
    }


def _is_blocking_spec_issue(issue: Dict[str, Any]) -> bool:
    """Return True only for evidence that a label is paired to the wrong value.

    A value conflict between Technical Specifications and Specifications may
    be a valid distinction such as meter versus probe data, or a manufacturer
    wording difference. It stays as a warning, while obvious cases such as an
    accessory label containing only dimensions are blocked.
    """
    if not isinstance(issue, dict):
        return False
    return clean_text(issue.get("Severity")).casefold() == "blocker" or (
        clean_text(issue.get("Issue Code")) == "label_value_mismatch"
    )


def _validate_canonical_row(row: Dict[str, Any]) -> List[Dict[str, str]]:
    """Flag only strong label/value mismatches. Unknown fields are allowed."""
    key = clean_text(row.get("canonical_key"))
    value = clean_text(row.get("value"))
    issues: List[Dict[str, str]] = []
    if not value:
        issues.append(_issue("empty_specification_value", row, "Nilai spesifikasi kosong.", severity="warning"))
        return issues
    signatures = _value_has_multiple_field_signatures(value)
    expected = {
        "dimensions": "dimensions", "dimensions_main_unit": "dimensions", "dimensions_meter": "dimensions", "dimensions_probe": "dimensions",
        "weight": "weight", "weight_meter": "weight", "weight_probe": "weight",
        "operating_humidity": "humidity", "operating_temperature": "temperature",
        "power_supply": "power", "power_draw": "power", "power_draw_normal_operation": "power", "power_draw_data_logging": "power",
        "accessories_included": "accessories", "accessories_standard": "accessories", "accessories_optional": "accessories",
    }.get(key)
    strong_pattern = {
        "dimensions": _dimension_like,
        "weight": _weight_like,
        "humidity": _humidity_like,
        "temperature": _temperature_like,
        "power": _power_like,
        "accessories": _accessory_like,
    }
    if expected and expected in strong_pattern and not strong_pattern[expected](value):
        # Do not flag generic free-text facts. Only obvious wrong pairings are
        # review-worthy. E.g. accessories label carrying a dimension value.
        other = [x for x in signatures if x != expected]
        if other:
            issues.append(_issue("label_value_mismatch", row, f"Label mengarah ke {expected}, tetapi nilai terlihat seperti {', '.join(other)}.", severity="blocker"))
    if expected and expected != "accessories" and len(signatures) >= 2 and expected in signatures:
        # Accessories routinely include cables, adapters, and memory cards.
        # Those are not mixed fields. For technical labels, retain only strong
        # incompatible signals such as humidity attached to a power entry.
        incompatible = [x for x in signatures if x != expected]
        if incompatible:
            issues.append(_issue("mixed_specification_fields", row, f"Satu nilai memuat beberapa jenis data: {', '.join(signatures)}.", severity="warning"))
    return issues


def _row_display_label(canonical_key: str, default_label: str, source_label: str) -> str:
    for _pat, key, label in _CANONICAL_LABEL_RULES:
        if key == canonical_key:
            return label
    # Source-specific labels such as "6 GHz" or "Pt 100 ohm" must stay
    # recognisable rather than becoming a generic field name.
    return normalize_label(default_label or source_label) or "Rincian teknis"


def canonicalize_specifications(rows: Sequence[Dict[str, Any]]) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    """Hapus spesifikasi yang benar-benar sama, lalu simpan salah satu saja.

    Prinsip v16 sengaja sederhana:
    - Technical Specifications dan Specifications boleh sama-sama dibaca.
    - Bila fakta yang sama muncul dua kali dengan nilai yang sama atau hanya
      berbeda format kecil, simpan satu baris.
    - Bila nilai berbeda, jangan dihapus dan jangan dipindahkan ke review.
      Nilai berbeda bisa memang mewakili probe, meter, kanal, atau rentang lain.
    - Tidak ada aturan yang menahan produk hanya karena spesifikasi.
    """
    prepared: List[Dict[str, Any]] = []
    for idx, raw in enumerate(rows or []):
        if not isinstance(raw, dict):
            continue
        value = _dedupe_value_fragments(raw.get("value", ""))
        if not value:
            continue
        key, standard_label, source_label = _canonical_label_info(raw)
        label = _row_display_label(key, clean_text(raw.get("label")) or standard_label, source_label)
        section = clean_text(raw.get("section")) or "Specifications"
        if _spec_text_key(label) == _spec_text_key(value):
            continue
        prepared.append({
            "label": label,
            "value": value,
            "section": section,
            "source_label": source_label or label,
            "canonical_key": key,
            "_index": idx,
        })

    # Pertahankan urutan sumber. Model tetap ditampilkan pertama.
    prepared.sort(key=lambda r: (0 if r["canonical_key"] == "model" else 1, r["_index"]))
    accepted: List[Dict[str, Any]] = []

    for row in prepared:
        duplicate_index: Optional[int] = None
        for pos, existing in enumerate(accepted):
            if not _canonical_keys_can_compare(
                clean_text(existing.get("canonical_key")), row["canonical_key"]
            ):
                continue
            if _spec_values_equivalent(existing.get("value"), row.get("value")):
                duplicate_index = pos
                break

        if duplicate_index is None:
            accepted.append(row)
            continue

        # Jika dua baris memang sama, pakai versi label yang lebih spesifik.
        old = accepted[duplicate_index]
        old_score = _specificity_score(
            clean_text(old.get("canonical_key")),
            clean_text(old.get("source_label")),
            clean_text(old.get("section")),
        )
        new_score = _specificity_score(
            row["canonical_key"], row["source_label"], row["section"]
        )
        if new_score > old_score or (
            new_score == old_score and len(row["value"]) > len(clean_text(old.get("value")))
        ):
            accepted[duplicate_index] = row

    output: List[Dict[str, str]] = []
    for row in accepted:
        output.append({
            "label": clean_text(row.get("label")),
            "value": clean_text(row.get("value")),
            "section": clean_text(row.get("section")),
            "source_label": clean_text(row.get("source_label")),
            "canonical_key": clean_text(row.get("canonical_key")),
        })
    return output, []


def dedupe_specs(rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    """Compatibility wrapper used across scraping, fallback, and AI output."""
    cleaned, _issues = canonicalize_specifications(rows)
    return cleaned



def _technical_heading_exact(value: Any) -> bool:
    text = clean_text(value).strip(" :–—|-")
    return bool(re.fullmatch(r"(?:technical\s+specifications?|technical\s+data|measurement\s+specifications?)", text, re.I))


def extract_specs(soup: BeautifulSoup, root: Tag, content_items: Sequence[Tuple[str, str, str]], fallback_model: str = "") -> List[Dict[str, str]]:
    """Extract raw facts, then canonicalise them once.

    Tables and definition lists are treated as structured facts. Narrative
    technical content is read only under an exact technical heading. This stops
    the ordinary Specifications table from being parsed a second time.
    """
    raw_rows: List[Dict[str, str]] = []

    for obj in product_jsonld(soup):
        props = obj.get("additionalProperty", [])
        if isinstance(props, dict):
            props = [props]
        for prop in props if isinstance(props, list) else []:
            if isinstance(prop, dict):
                label = clean_text(prop.get("name"))
                value = clean_text(prop.get("value") or prop.get("description"))
                if label and value:
                    raw_rows.append({"label": label, "source_label": label, "value": value, "section": "Product structured data"})
        sku = clean_text(obj.get("sku") or obj.get("mpn"))
        if sku:
            raw_rows.append({"label": "Model / SKU", "source_label": "Model / SKU", "value": sku, "section": "Product structured data"})

    for table in root.find_all("table"):
        heading = "Specifications"
        previous = table.find_previous(["h1", "h2", "h3", "h4", "h5", "h6"])
        if previous:
            heading = soup_text(previous) or heading
        for tr in table.find_all("tr"):
            cells = [clean_spec_value(soup_text(c)) for c in tr.find_all(["th", "td"], recursive=False)]
            cells = [c for c in cells if c]
            if len(cells) >= 2 and looks_like_spec_label(cells[0]):
                raw_rows.append({"label": cells[0], "source_label": cells[0], "value": " / ".join(cells[1:]), "section": heading})

    for dl in root.find_all("dl"):
        heading_node = dl.find_previous(["h1", "h2", "h3", "h4", "h5", "h6"])
        heading = soup_text(heading_node) or "Specifications"
        for dt in dl.find_all("dt"):
            dd = dt.find_next_sibling("dd")
            label, value = soup_text(dt), soup_text(dd)
            if label and value and looks_like_spec_label(label):
                raw_rows.append({"label": label, "source_label": label, "value": value, "section": heading})

    active = False
    pending_label = ""
    pending_values: List[str] = []

    def flush_pending() -> None:
        nonlocal pending_label, pending_values
        if pending_label and pending_values:
            raw_rows.append({
                "label": pending_label,
                "source_label": pending_label,
                "value": " / ".join(dedupe_keep_order(pending_values)),
                "section": "Technical Specifications",
            })
        pending_label, pending_values = "", []

    for _kind, heading, text in content_items:
        if _technical_heading_exact(heading):
            active = True
        elif active and clean_text(heading).strip() and not _technical_heading_exact(heading):
            flush_pending()
            active = False
        if not active:
            continue
        raw_line = normalize_content_line(text).lstrip("-•* ").strip()
        if not raw_line:
            continue
        if raw_line.endswith(":") and looks_like_spec_label(raw_line[:-1]):
            flush_pending()
            pending_label = clean_spec_value(raw_line[:-1])
            continue
        if ":" in raw_line:
            label, value = raw_line.split(":", 1)
            label, value = clean_spec_value(label), clean_spec_value(value)
            if looks_like_spec_label(label) and value:
                flush_pending()
                raw_rows.append({"label": label, "source_label": label, "value": value, "section": "Technical Specifications"})
                continue
        if pending_label:
            pending_values.append(clean_spec_value(raw_line))
        # Unlabelled technical lines are intentionally not assigned to a guessed
        # label. They remain in full official content, but should not become a
        # potentially wrong key/value specification row.
    flush_pending()

    model = clean_text(fallback_model)
    if model:
        raw_rows.insert(0, {"label": "Model", "source_label": "Model", "value": model, "section": "Input Excel"})

    # Keep raw source rows here. The canonical pipeline runs in main() after
    # all scrape/fallback sources are combined, so conflicts remain detectable
    # and can be sent to Review Spesifikasi instead of being silently dropped.
    return raw_rows


def extract_model(title: str, specs: Sequence[Dict[str, str]], fallback: str = "") -> str:
    for row in specs:
        label = clean_text(row.get("label")).casefold()
        if any(x in label for x in ["model", "sku", "catalog", "cat.no", "part no", "item no", "mpn"]):
            value = clean_text(row.get("value"))
            if value:
                return value
    return clean_text(fallback)


def is_accessory_product(title: str, description: str, specs: Sequence[Dict[str, str]]) -> Tuple[bool, str]:
    """Classify an item as accessory only with affirmative, high-confidence evidence.

    A product title containing "receiver", "transmitter", "probe", or
    "sensor" is not enough.  Lutron uses those words in independent meters and
    transmitters, so the prior broad rule produced incorrect categories and
    compatibility copy.
    """
    title_clean = clean_text(title)
    desc_clean = clean_text(description)
    blob = " ".join([title_clean, desc_clean] + [f"{r.get('label','')} {r.get('value','')}" for r in specs])
    model_like = " ".join([title_clean] + [clean_text(r.get("value")) for r in specs if re.search(r"\b(model|sku|cat|part|item|mpn)\b", clean_text(r.get("label")), re.I)])

    # A recognised instrument family or an explicit main-instrument noun wins
    # over any component word found in the same title.
    if PRIMARY_PRODUCT_TERMS.search(title_clean) or PRIMARY_MODEL_PREFIX_RE.search(model_like):
        return False, "main_instrument_title_or_model"

    functional_labels = 0
    for row in specs:
        label = clean_text(row.get("label"))
        if re.search(r"\b(range|measurement|accuracy|resolution|display|power|battery|output|interface|data logger|datalogger|channel|temperature|humidity|pressure|frequency|weight|dimension)\b", label, re.I):
            functional_labels += 1
    if functional_labels >= 2:
        return False, "main_instrument_spec_signature"

    # Strong phrases are accepted only when the title does not identify an
    # autonomous instrument. A standalone replacement probe/cable can still be
    # classified correctly here.
    if ACCESSORY_STRONG_KEYWORDS.search(title_clean):
        return True, "strong_accessory_keyword_in_title"

    # A short title that is merely a component noun can be an accessory, but
    # unknown/incomplete pages are intentionally not guessed as accessories.
    if ACCESSORY_COMPONENT_TERMS.fullmatch(title_clean.strip().casefold()):
        return True, "standalone_component_title"
    if ACCESSORY_COMPONENT_TERMS.search(title_clean) and len(desc_clean) < 35 and not specs:
        return True, "component_title_with_no_product_data"

    return False, "not_confirmed_as_accessory"


def source_official_name(cfg: SourceConfig) -> str:
    return "halaman produk Lutron"


HOME_PAGE_TITLE_RE = re.compile(
    r"^lutron electronic taiwan official website\s*-\s*expert in electronic instruments and meters$",
    re.I,
)


def _model_token(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).casefold())


def looks_like_lutron_landing_page(soup: BeautifulSoup, expected_model: str = "") -> bool:
    """Detect the generic landing page returned by the www Lutron host."""
    page_title = clean_text(soup_text(soup.title))
    if not HOME_PAGE_TITLE_RE.search(page_title):
        return False
    token = _model_token(expected_model)
    if not token:
        return True
    visible = _model_token(soup.get_text(" ", strip=True))
    return token not in visible


class LutronBrowserRenderer:
    """Render JavaScript content and record actual image URLs requested by Lutron."""

    def __init__(self, timeout_seconds: int = 30):
        self.timeout_ms = max(15, int(timeout_seconds)) * 1000
        self._playwright = None
        self._browser = None
        self._context = None

    def available(self) -> bool:
        return sync_playwright is not None

    def start(self) -> None:
        if self._context is not None:
            return
        if sync_playwright is None:
            raise RuntimeError("playwright_not_installed")
        self._playwright = sync_playwright().start()
        launch_args = {
            "headless": True,
            "args": ["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        }
        # Prefer Playwright's own browser. If it has not been downloaded, use
        # an installed Chrome/Edge/Chromium browser when available.
        try:
            self._browser = self._playwright.chromium.launch(**launch_args)
        except Exception as first_error:
            candidates = [
                clean_text(os.getenv("LUTRON_BROWSER_EXECUTABLE", "")),
                shutil.which("chrome"), shutil.which("google-chrome"),
                shutil.which("chromium"), shutil.which("chromium-browser"),
                shutil.which("msedge"),
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            ]
            executable = next((x for x in candidates if x and Path(x).exists()), "")
            if not executable:
                self.close()
                raise RuntimeError(
                    "browser_not_found. Install browser Playwright: py -m playwright install chromium"
                ) from first_error
            self._browser = self._playwright.chromium.launch(executable_path=executable, **launch_args)
        self._context = self._browser.new_context(
            user_agent=DEFAULT_HEADERS["User-Agent"],
            locale="en-US",
            viewport={"width": 1440, "height": 1400},
            extra_http_headers={"Accept-Language": DEFAULT_HEADERS["Accept-Language"]},
        )

    def fetch(self, url: str, expected_model: str = "") -> Dict[str, Any]:
        page = None
        images: List[str] = []

        def on_response(response: Any) -> None:
            try:
                rtype = clean_text(response.request.resource_type).casefold()
                ctype = clean_text(response.headers.get("content-type", "")).casefold()
                if rtype == "image" or ctype.startswith("image/"):
                    images.append(clean_text(response.url))
            except Exception:
                pass

        try:
            self.start()
            assert self._context is not None
            page = self._context.new_page()
            page.set_default_timeout(self.timeout_ms)
            page.on("response", on_response)
            page.goto(url, wait_until="domcontentloaded", timeout=self.timeout_ms)
            # Scroll through the page to activate lazy-loaded gallery images.
            for y in (0, 650, 1500, 2800, 4400):
                page.evaluate("y => window.scrollTo(0, y)", y)
                page.wait_for_timeout(400)
            try:
                page.wait_for_load_state("networkidle", timeout=min(8000, self.timeout_ms))
            except Exception:
                pass
            try:
                page.wait_for_function(
                    "model => !model || document.body.innerText.toLowerCase().includes(model.toLowerCase())",
                    arg=expected_model,
                    timeout=min(4000, self.timeout_ms),
                )
            except Exception:
                pass

            # Lutron's product photo is commonly NOT an <img>. It is a Quasar
            # q-img element whose real source is CSS `background-image`, exactly
            # like: <div class="q-img__image ..." style="background-image:url(...)">.
            # Therefore, inspect both inline and computed backgrounds across the
            # rendered DOM, rank the image visually related to the product title,
            # and make it the primary candidate.
            css_background_candidates = page.evaluate(
                r"""
                expectedModel => {
                    const expected = (expectedModel || '').trim().toLowerCase();
                    const textOf = el => (el.innerText || el.textContent || '').trim().toLowerCase();
                    const urlsFrom = value => {
                        const out = [];
                        const text = String(value || '');
                        const regex = /url\(\s*["']?([^"')\s]+)["']?\s*\)/gi;
                        let match;
                        while ((match = regex.exec(text)) !== null) {
                            if (match[1]) out.push(match[1].trim());
                        }
                        return out;
                    };
                    const headingNodes = [...document.querySelectorAll(
                        'h1, h2, [class*="product-title"], [class*="product_name"], [class*="product-name"], [class*="model"], [class*="sku"]'
                    )];
                    const title = headingNodes.find(el => expected && textOf(el).includes(expected))
                        || document.querySelector('h1')
                        || headingNodes[0]
                        || null;
                    const tb = title ? title.getBoundingClientRect() : null;
                    const candidates = [];
                    for (const el of [...document.querySelectorAll('*')]) {
                        const rect = el.getBoundingClientRect();
                        if (rect.width < 100 || rect.height < 100) continue;
                        const inlineStyle = el.getAttribute('style') || '';
                        const computedStyle = getComputedStyle(el);
                        const background = computedStyle ? computedStyle.backgroundImage : '';
                        const urls = [...urlsFrom(inlineStyle), ...urlsFrom(background)];
                        if (!urls.length) continue;
                        const classText = String(el.className || '').toLowerCase();
                        for (const u of urls) {
                            const low = u.toLowerCase();
                            if (!u || /(?:logo|icon|sprite|flag|banner(?:s)?|hero|carousel|slider|homepage|landing|corporate|company|placeholder|facebook|instagram|youtube|linkedin)/.test(low)) continue;
                            let score = Math.min(rect.width * rect.height, 800000);
                            if (/q-img(?:__image|__content)?/.test(classText)) score += 600000;
                            if (/product|products|catalog|item|goods|photo|image/.test(low)) score += 150000;
                            if (tb) {
                                const verticallyRelated = rect.bottom >= tb.top - 260 && rect.top <= tb.bottom + 750;
                                const leftOfTitle = rect.left < tb.left && rect.right <= tb.left + 180;
                                const centerGap = Math.abs((rect.top + rect.bottom) / 2 - (tb.top + tb.bottom) / 2);
                                if (verticallyRelated) score += 500000;
                                if (leftOfTitle) score += 800000;
                                score += Math.max(0, 180000 - centerGap * 900);
                            }
                            candidates.push({url: u, score, classText, rect: [rect.left, rect.top, rect.width, rect.height]});
                        }
                    }
                    const deduped = new Map();
                    for (const item of candidates) {
                        const existing = deduped.get(item.url);
                        if (!existing || item.score > existing.score) deduped.set(item.url, item);
                    }
                    return [...deduped.values()].sort((a, b) => b.score - a.score).slice(0, 20);
                }
                """,
                expected_model,
            )
            if isinstance(css_background_candidates, list):
                for candidate in css_background_candidates:
                    if isinstance(candidate, dict) and clean_text(candidate.get("url")):
                        images.insert(0, "__LUTRON_PRIMARY__" + clean_text(candidate.get("url")))

            dom_images = page.eval_on_selector_all(
                "img, source, [style*='background-image']",
                r"""els => els.flatMap(el => {
                    const out = [];
                    if (el.currentSrc) out.push(el.currentSrc);
                    if (el.src) out.push(el.src);
                    for (const key of ['data-src','data-original','data-lazy-src','data-image','data-zoom-image']) {
                      const v = el.getAttribute(key); if (v) out.push(v);
                    }
                    const st = el.getAttribute('style') || '';
                    const m = st.match(/url\((?:['\"])?([^'\")]+)(?:['\"])?\)/i);
                    if (m && m[1]) out.push(m[1]);
                    return out;
                })""",
            )
            images.extend(clean_text(x) for x in dom_images if clean_text(x))
            return {
                "ok": True,
                "html": page.content(),
                "url": page.url,
                "images": dedupe_keep_order(images),
                "background_candidates": css_background_candidates if isinstance(css_background_candidates, list) else [],
                "error": "",
            }
        except Exception as exc:
            return {"ok": False, "html": "", "url": url, "images": [], "error": f"playwright_{type(exc).__name__}: {exc}"}
        finally:
            try:
                if page is not None:
                    page.close()
            except Exception:
                pass

    def close(self) -> None:
        for obj, method in ((self._context, "close"), (self._browser, "close"), (self._playwright, "stop")):
            try:
                if obj is not None:
                    getattr(obj, method)()
            except Exception:
                pass
        self._context = self._browser = self._playwright = None


def page_content_quality(title: str, description: str, specs: Sequence[Dict[str, str]], image_url: str, content_html: str) -> int:
    plain = re.sub(r"<[^>]+>", " ", clean_text(content_html))
    plain = re.sub(r"\s+", " ", plain).strip()
    score = 1 if title else 0
    score += 3 if len(clean_text(description)) >= 80 else 0
    score += 4 if len(specs) >= 4 else (2 if len(specs) >= 2 else 0)
    score += 3 if image_url else 0
    score += 3 if len(plain) >= 250 else 0
    return score


def fetch_lutron_product_page(
    url: str,
    timeout: int,
    session: requests.Session,
    expected_model: str = "",
) -> Tuple[Optional[requests.Response], str]:
    """Fetch a canonical product page and avoid the generic www landing page."""
    canonical_url = normalize_url(url)

    # The first visit establishes normal site cookies once per run. It is a
    # lightweight bootstrap request and avoids a product URL being handled as
    # a generic landing page by some Lutron site configurations.
    if not getattr(session, "_lutron_bootstrapped", False):
        try:
            session.get(
                "https://lutron.com.tw/en-us",
                headers=DEFAULT_HEADERS,
                timeout=timeout,
                allow_redirects=True,
            )
        except Exception:
            pass
        setattr(session, "_lutron_bootstrapped", True)

    headers = dict(DEFAULT_HEADERS)
    headers["Referer"] = "https://lutron.com.tw/en-us/"
    try:
        response = session.get(
            canonical_url,
            headers=headers,
            timeout=timeout,
            allow_redirects=True,
        )
        response.raise_for_status()
    except Exception as exc:
        return None, f"{type(exc).__name__}: {exc}"

    # Defensive retry. This covers old cached www routes and rare cases where
    # a routing layer serves the generic landing page on the first request.
    soup = BeautifulSoup(response.text or "", "html.parser")
    if looks_like_lutron_landing_page(soup, expected_model):
        retry_headers = dict(DEFAULT_HEADERS)
        retry_headers["Referer"] = "https://lutron.com.tw/en-us/categories/2/environmental"
        try:
            retry = session.get(
                canonical_url,
                headers=retry_headers,
                timeout=timeout,
                allow_redirects=True,
            )
            retry.raise_for_status()
            retry_soup = BeautifulSoup(retry.text or "", "html.parser")
            if not looks_like_lutron_landing_page(retry_soup, expected_model):
                response = retry
            else:
                return response, "generic_landing_page_returned"
        except Exception:
            return response, "generic_landing_page_returned"

    return response, ""


def scrape_product_page(record: Dict[str, str], timeout: int, session: requests.Session, renderer: Optional[LutronBrowserRenderer] = None) -> Dict[str, Any]:
    url = normalize_url(record.get("url"))
    cfg = get_source_config(url)
    result: Dict[str, Any] = {
        "url": url,
        "status": "failed",
        "source": cfg.key if cfg else "unknown",
        "brand": LUTRON_CONFIG.brand if cfg else "",
        "title": clean_text(record.get("input_title")),
        "input_title": clean_text(record.get("input_title")),
        "input_model": clean_text(record.get("input_model")),
        "description": "",
        "official_content_html": "",
        "spec_rows": [],
        "image_url": "",
        "source_category": "",
        "model": clean_text(record.get("input_model")),
        "is_accessory": False,
        "accessory_reason": "",
        "error": "",
    }
    if not url:
        result.update(status="no_url", error="URL kosong")
        return result
    if cfg is None:
        result.update(status="unsupported_domain", error="URL wajib berasal dari domain lutron.com.tw")
        return result

    response, fetch_note = fetch_lutron_product_page(
        url,
        timeout=timeout,
        session=session,
        expected_model=record.get("input_model", ""),
    )
    if response is None:
        result.update(status="request_error", error=fetch_note)
        return result

    def parse_document(html_text: str, base_url: str, network_images: Optional[Sequence[str]] = None) -> Dict[str, Any]:
        soup = BeautifulSoup(html_text or "", "html.parser")
        is_landing = looks_like_lutron_landing_page(soup, record.get("input_model", ""))
        products = product_jsonld(soup)
        title = extract_title(soup, record.get("input_title", ""))
        if is_landing:
            title = clean_text(record.get("input_title")) or clean_text(record.get("input_model"))
        category = extract_category(soup, products)
        root = choose_main_container(soup)
        strip_noise_nodes(root)
        description = "" if is_landing else extract_description(soup, products, title)
        content_items = [] if is_landing else extract_all_content(root, title)
        full_content = all_content_html(content_items, title)
        specs = extract_specs(soup, root, content_items, record.get("input_model", ""))
        model = extract_model(title, specs, record.get("input_model", ""))
        if model and not any(re.search(r"\bmodel\b", row.get("label", ""), re.I) and clean_text(row.get("value")) == model for row in specs):
            specs.insert(0, {"label": "Model", "source_label": "Model", "value": model, "section": "Product page"})
        image = "" if is_landing else extract_best_image_url(
            soup, base_url, model=model or record.get("input_model", ""), network_images=network_images
        )
        return {"landing": is_landing, "title": title, "category": category, "description": description,
                "content": full_content, "specs": specs, "model": model, "image": image}

    parsed = parse_document(response.text or "", response.url or url)
    browser_note = ""
    # Browser fallback is essential when the site injects product data and
    # images after load. Keep rendered result only when it is at least as rich.
    initial_quality = page_content_quality(parsed["title"], parsed["description"], parsed["specs"], parsed["image"], parsed["content"])
    if (parsed["landing"] or initial_quality < 8 or not parsed["image"]) and renderer is not None and renderer.available():
        rendered = renderer.fetch(response.url or url, record.get("input_model", ""))
        if rendered.get("ok"):
            browser_parsed = parse_document(rendered.get("html", ""), rendered.get("url", url), rendered.get("images", []))
            browser_quality = page_content_quality(browser_parsed["title"], browser_parsed["description"], browser_parsed["specs"], browser_parsed["image"], browser_parsed["content"])
            # Prefer the rendered DOM if it provides the real product image.
            # A CSS-backed q-img is authoritative even when its text extraction
            # is slightly shorter than the initial HTML shell.
            if browser_parsed["image"] or browser_quality >= initial_quality:
                parsed = browser_parsed
                fetch_note = _merge_source_notes(fetch_note, "browser_rendered")
        else:
            browser_note = clean_text(rendered.get("error", ""))

    is_landing_page = parsed["landing"]
    title = parsed["title"]
    source_category = parsed["category"]
    description = parsed["description"]
    official_content = parsed["content"]
    specs = parsed["specs"]
    model = parsed["model"]
    image_url = parsed["image"]
    is_accessory, reason = is_accessory_product(title, description, specs)

    # Never hide the Playwright error. Earlier versions overwrote it when the
    # HTTP fetch had returned the generic landing shell, making image failures
    # impossible to diagnose from the Excel log.
    correction = _merge_source_notes(
        fetch_note,
        browser_note,
        "generic_landing_page_returned" if is_landing_page else "",
    )

    result.update({
        "url": response.url or url,
        "status": "landing_page" if is_landing_page else ("ok" if title else "partial"),
        "title": title,
        "description": description,
        "official_content_html": official_content,
        "spec_rows": specs,
        "image_url": image_url,
        "source_category": source_category,
        "model": model,
        "is_accessory": is_accessory,
        "accessory_reason": reason,
        "error": correction,
    })
    return result



# ---------------------------------------------------------------------------
# Verified fallbacks for incomplete Lutron product pages
# ---------------------------------------------------------------------------

# Lutron's public product routes sometimes return only the site shell or a
# short product card.  This is why the older versions produced a title/model
# but no image, official content, or specification.  The fallback below never
# invents technical values: it retrieves a model-specific datasheet, extracts
# its selectable text, then records the fallback URL in the output log.
#
# The list is intentionally explicit and configurable.  Any response is
# accepted only if it is a real PDF and contains the requested model token.
DATASHEET_URL_TEMPLATES: Tuple[str, ...] = (
    "https://www.sunwe.com.tw/lutron/{model}.pdf",
    "https://www.sunwe.com.tw/lutron/{model}eop.pdf",
    "https://www.sunwe.com.tw/lutron/{model}_eop.PDF",
    "https://www.instrumentsgroup.co.za/index_files/Lutron/database/pdf/{model}.pdf",
    "https://jin-hua.com.tw/upload/product/{model}.pdf",
    "https://www.100y.com.tw/pdf_file/{model}.pdf",
    "https://siliconinstrument.com.sg/wp-content/uploads/2022/01/{model}.pdf",
)

PDF_SPEC_LABELS: Tuple[str, ...] = (
    "Operating Temperature", "Operating Temp.", "Operating Temp", "Operating Humidity",
    "Power Supply", "Power Consumption", "Power Input", "Power Current",
    "Dimension", "Dimensions", "Weight", "Display", "Circuit", "Measurement",
    "Range", "Resolution", "Accuracy", "Frequency Accuracy", "Width/Duty Accuracy",
    "Sensitivity", "Input Connector", "Sensor Type", "Data Output", "Data Logger",
    "Datalogger", "Memory Card", "Sampling Time", "Channels", "Channel",
    "Interface", "Protection", "Battery", "Accessories", "Standard Accessories",
    "Optional Accessories", "Case", "Time Base", "Output", "Function",
)

_PDF_LABEL_RE = re.compile(
    r"(?P<label>" + "|".join(re.escape(x) for x in sorted(PDF_SPEC_LABELS, key=len, reverse=True)) + r")"
    r"\s*[:.]?\s*",
    re.I,
)


def _is_real_product_data(data: Dict[str, Any]) -> bool:
    """Return True only when the Lutron page supplied substantive product data."""
    title = clean_text(data.get("title"))
    description = clean_text(data.get("description"))
    specs = data.get("spec_rows") if isinstance(data.get("spec_rows"), list) else []
    if not title or HOME_PAGE_TITLE_RE.search(title):
        return False
    non_identity = [
        row for row in specs
        if not re.search(r"\b(model|sku|part|item|mpn|catalog)\b", clean_text(row.get("label")), re.I)
    ]
    return len(description) >= 80 or len(non_identity) >= 3 or len(clean_text(data.get("official_content_html"))) >= 350


def _model_file_candidates(value: Any) -> List[str]:
    """Build safe filename candidates from a model or multi-model set."""
    raw = clean_text(value)
    if not raw:
        return []
    pieces = re.split(r"\s*(?:/|,|;|\+|\band\b|\bwith\b)\s*", raw, flags=re.I)
    candidates: List[str] = []
    for piece in [raw] + pieces:
        p = clean_text(piece)
        p = re.sub(r"\b(?:set|kit|package)\b", "", p, flags=re.I).strip(" -_/")
        if not p:
            continue
        variants = [
            p,
            p.upper(),
            p.replace(" ", ""),
            p.replace("/", "-").replace(" ", ""),
            re.sub(r"[^A-Za-z0-9._-]+", "", p),
        ]
        candidates.extend(v for v in variants if len(v) >= 3)
    return dedupe_keep_order(candidates)


def _find_tesseract_executable() -> str:
    candidates = [
        clean_text(os.getenv("TESSERACT_CMD", "")),
        shutil.which("tesseract") or "",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    return next((candidate for candidate in candidates if candidate and Path(candidate).exists()), "")


def _ocr_pdf_bytes(pdf_bytes: bytes, max_pages: int = 3) -> str:
    """OCR a scanned PDF only when text extraction produced no usable text."""
    executable = _find_tesseract_executable()
    if not pdf_bytes or fitz is None or not executable:
        return ""
    try:
        with tempfile.TemporaryDirectory(prefix="lutron_pdf_ocr_") as tmp:
            pdf_path = Path(tmp) / "datasheet.pdf"
            pdf_path.write_bytes(pdf_bytes)
            document = fitz.open(pdf_path)
            pieces: List[str] = []
            for index, page in enumerate(document[:max_pages]):
                pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
                image_path = Path(tmp) / f"page_{index + 1}.png"
                pix.save(image_path)
                process = subprocess.run(
                    [executable, str(image_path), "stdout", "-l", "eng", "--psm", "6"],
                    capture_output=True,
                    text=True,
                    timeout=75,
                    check=False,
                )
                page_text = clean_text(process.stdout)
                if page_text:
                    pieces.append(page_text)
            return "\n".join(pieces)
    except Exception:
        return ""


def _pdf_response_text(pdf_bytes: bytes) -> Tuple[str, str]:
    """Return extracted datasheet text plus method: text, ocr, or empty."""
    if not pdf_bytes:
        return "", "empty"
    extracted = ""
    if PdfReader is not None:
        try:
            reader = PdfReader(BytesIO(pdf_bytes))
            chunks: List[str] = []
            # Datasheets are usually one or two pages. Capping avoids excessive
            # memory usage when a full operation manual is encountered.
            for page in reader.pages[:16]:
                page_text = clean_text(page.extract_text() or "")
                if page_text:
                    chunks.append(page_text)
            extracted = "\n".join(chunks)
        except Exception:
            extracted = ""
    if len(re.sub(r"\s+", "", extracted)) >= 80:
        return extracted, "text"
    ocr_text = _ocr_pdf_bytes(pdf_bytes)
    if len(re.sub(r"\s+", "", ocr_text)) >= 80:
        return ocr_text, "ocr"
    return extracted, "empty"


def _find_labeled_pdf_specs(pdf_text: str, model: str) -> List[Dict[str, str]]:
    """Extract conservative Label: Value rows from text-based Lutron PDFs.

    Two-column PDF layouts frequently interleave the left and right columns in
    one text line. To avoid false specifications, this parser begins only after
    a real specification heading and reads labels from the start of a line.
    """
    lines = [clean_spec_value(line) for line in clean_text(pdf_text).splitlines()]
    lines = [line for line in lines if line and len(line) <= 1200]
    rows: List[Dict[str, str]] = []
    if model:
        rows.append({"label": "Model", "value": model, "section": "Model-specific datasheet"})

    spec_mode = False
    pending_label = ""
    seen_headings = 0
    for line in lines:
        if re.search(r"\b(?:GENERAL\s*&?\s*ELECTRICAL\s*SPECIFICATIONS?|TECHNICAL\s+(?:SPECIFICATIONS?|DATA)|SPECIFICATIONS?)\b", line, re.I):
            spec_mode = True
            seen_headings += 1
            pending_label = ""
            continue
        if not spec_mode:
            continue
        if re.search(r"^(?:TABLE\s+FOR|OPERATION\s+MANUAL|FEATURES)\b", line, re.I):
            # Table rows may follow; keep parsing only known labels.
            pending_label = ""
            continue

        match = _PDF_LABEL_RE.match(line)
        if match:
            label = normalize_label(match.group("label"))
            value = clean_spec_value(line[match.end():])
            # Trim an interleaved right-column label. It will be picked up if
            # it appears at the start of a later extracted line.
            later = _PDF_LABEL_RE.search(value)
            # pypdf can combine left/right table columns. The right column
            # often begins with a generic technical label that is not in the
            # primary specification-label list.
            interleaved = re.search(
                r"\b(?:Operating(?:\s+Humidity)?|Power(?:\s+(?:Supply|Consumption|Input))?|"
                r"Memory\s+Card|Data\s+Output|Dimensions?|Weight|Accessories?|Standard|Optional|TABLE\s+FOR)\b",
                value,
                flags=re.I,
            )
            cut_at = min(
                [m.start() for m in (later, interleaved) if m is not None] or [len(value)]
            )
            if cut_at < len(value):
                value = clean_spec_value(value[:cut_at])
            value = re.sub(r"^[,:;.\-–—\s]+", "", value)
            if label.casefold() == "model":
                pending_label = ""
                continue
            if value and len(value) >= 4:
                # Ignore clearly fragmented leftovers from a PDF table.
                if value.casefold() not in {"format", "reading", "specifications"}:
                    rows.append({"label": label, "value": value, "section": "Model-specific datasheet"})
                pending_label = ""
            else:
                pending_label = label
            continue

        if pending_label and len(line) <= 500 and len(line) >= 4:
            rows.append({"label": pending_label, "value": line, "section": "Model-specific datasheet"})
            pending_label = ""

    # A datasheet can use non-standard labels. Capture a small number of
    # technical statements only when the clean labelled extraction is scarce.
    if len(rows) < 5:
        for line in lines:
            if re.search(r"\b(?:range|resolution|accuracy|power|dimension|weight|temperature|humidity)\b", line, re.I):
                rows.append({"label": "Technical detail", "value": line, "section": "Model-specific datasheet"})
                if len(rows) >= 12:
                    break
    return rows


def _pdf_feature_description(pdf_text: str, max_chars: int = 1400) -> str:
    lines = [clean_spec_value(line) for line in clean_text(pdf_text).splitlines()]
    lines = [line for line in lines if line and not re.fullmatch(r"(?:LUTRON\s*ELECTRONIC|ISO[- ]?9001.*)", line, flags=re.I)]
    start = next((i for i, line in enumerate(lines) if re.fullmatch(r"FEATURES", line, flags=re.I)), -1)
    if start < 0:
        return ""
    collected: List[str] = []
    for line in lines[start + 1:]:
        if re.search(r"^(?:GENERAL\s*&?\s*ELECTRICAL\s*SPECIFICATIONS?|SPECIFICATIONS?|TECHNICAL\s*DATA|OPERATION\s+MANUAL)", line, re.I):
            break
        line = re.sub(r"^[*•\-]+\s*", "", line).strip()
        if len(line) < 15:
            continue
        collected.append(line)
        if len(" ".join(collected)) >= max_chars:
            break
    return compact(" ".join(dedupe_keep_order(collected)), max_chars)


def _pdf_content_html(pdf_text: str, title: str, max_chars: int = 12000) -> str:
    """Make a clean, limited official-content tab from the datasheet text."""
    lines = [clean_spec_value(line) for line in clean_text(pdf_text).splitlines()]
    lines = [line for line in lines if line and not re.fullmatch(r"(?:LUTRON\s*ELECTRONIC|ISO[- ]?9001.*)", line, flags=re.I)]
    blocks: List[str] = []
    used = 0
    section = "Informasi Produk"
    for line in lines:
        if line.casefold() == clean_text(title).casefold():
            continue
        if len(line) <= 100 and (
            line.isupper() or re.fullmatch(r"(?:FEATURES|SPECIFICATIONS?|GENERAL\s*&?\s*ELECTRICAL\s*SPECIFICATIONS?|ACCESSORIES|DATA\s+OUTPUT)", line, re.I)
        ):
            section = line.title() if line.isupper() else line
            addition = len(section) + 9
            if used + addition > max_chars:
                break
            blocks.append(f"<h3>{esc(section)}</h3>")
            used += addition
            continue
        # Skip repeated product name boilerplate, but retain technical facts.
        if len(line) < 2:
            continue
        addition = len(line) + 7
        if used + addition > max_chars:
            break
        blocks.append(f"<p>{esc(line)}</p>")
        used += addition
    return "".join(blocks)


def _pdf_title(pdf_text: str, model: str, fallback: str) -> str:
    """Create a model-led catalog title from the datasheet heading."""
    lines = [clean_spec_value(line) for line in clean_text(pdf_text).splitlines()]
    candidate = ""
    for idx, line in enumerate(lines[:18]):
        if re.search(r"\b(?:COUNTER|THERMOMETER|ANALYZER|METER|TESTER|TRANSMITTER|RECORDER|LOGGER|GAUGE|HARDNESS)\b", line, re.I):
            candidate = line
            # Datasheets frequently put a qualifier on line 1 and product type
            # on line 2. Join only when it stays short and readable.
            if idx + 1 < len(lines) and len(candidate) < 110 and lines[idx + 1].isupper():
                candidate = clean_text(candidate + " " + lines[idx + 1])
            break
    candidate = re.sub(r"\bModel\s*:\s*.*$", "", candidate, flags=re.I).strip(" -:|")
    candidate = re.sub(r"\s+", " ", candidate)
    if candidate and model:
        return f"Lutron {model} - {candidate.title()}"
    if model and fallback:
        return f"Lutron {model} - {fallback}"
    return fallback or (f"Lutron {model}" if model else "")


def _fetch_datasheet_for_model(
    session: requests.Session,
    model: str,
    timeout: int,
    attempt_limit: int = 20,
) -> Dict[str, Any]:
    """Return a verified model-specific datasheet, or an empty dict."""
    if PdfReader is None:
        return {"ok": False, "reason": "pypdf_not_installed"}
    tried = 0
    model_tokens = [_model_token(x) for x in _model_file_candidates(model)]
    for candidate_model in _model_file_candidates(model):
        safe_model = quote(candidate_model, safe="-_.")
        for template in DATASHEET_URL_TEMPLATES:
            if tried >= attempt_limit:
                return {"ok": False, "reason": "datasheet_attempt_limit_reached"}
            tried += 1
            url = template.format(model=safe_model)
            try:
                response = session.get(
                    url,
                    headers=DEFAULT_HEADERS,
                    timeout=timeout,
                    allow_redirects=True,
                )
            except Exception:
                continue
            if response.status_code != 200:
                continue
            content_type = clean_text(response.headers.get("content-type", "")).casefold()
            payload = response.content or b""
            if not (payload.startswith(b"%PDF") or "pdf" in content_type):
                continue
            if len(payload) < 3000 or len(payload) > 30_000_000:
                continue
            pdf_text, extraction_method = _pdf_response_text(payload)
            compact_token = _model_token(model)
            if not pdf_text or (compact_token and compact_token not in _model_token(pdf_text)):
                # A set can use two models. Accept matching a component as long
                # as one explicit filename candidate matched the document.
                if not any(token and token in _model_token(pdf_text) for token in model_tokens):
                    continue
            return {
                "ok": True,
                "url": response.url or url,
                "text": pdf_text,
                "source_host": (urlparse(response.url or url).hostname or "").casefold(),
                "extraction_method": extraction_method,
            }
    return {"ok": False, "reason": "datasheet_not_found"}


def _merge_source_notes(*parts: Any) -> str:
    """Join semicolon-delimited status notes without repeating them."""
    values: List[str] = []
    for part in parts:
        for item in clean_text(part).split(";"):
            item = clean_text(item)
            if item:
                values.append(item)
    return "; ".join(dedupe_keep_order(values))


def _image_bad_url(url: str) -> bool:
    # Share the same strict gate used for official-page images so optional
    # external fallback cannot reintroduce generic banners or corporate art.
    return not valid_image_url(url)


def _image_candidate_score(url: str, model: str, title: str, engine: str) -> int:
    low = clean_text(url).casefold()
    token = _model_token(model)
    title_token = _model_token(title)
    score = 0
    if token and token in _model_token(low):
        score += 90
    if title_token and len(title_token) > 8 and title_token[:8] in _model_token(low):
        score += 10
    if re.search(r"\.(?:jpe?g|png|webp|avif)(?:[?#]|$)", low, re.I):
        score += 12
    if any(host in low for host in ("directindustry", "lutroninstruments", "sunwe", "instrumentsgroup", "legatool", "reviznipristroje", "advancom")):
        score += 10
    if engine == "bing":
        score += 3
    if _image_bad_url(low):
        score -= 1000
    return score


def _verify_image_candidate(session: requests.Session, url: str, timeout: int) -> bool:
    try:
        response = session.get(
            url,
            headers={"User-Agent": DEFAULT_HEADERS["User-Agent"], "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8"},
            timeout=timeout,
            stream=True,
            allow_redirects=True,
        )
        if response.status_code != 200:
            return False
        content_type = clean_text(response.headers.get("content-type", "")).casefold()
        if content_type.startswith("image/"):
            return True
        # Some CDNs omit the correct content type. A recognizable image
        # extension is enough if the response body is not an HTML page.
        final_url = clean_text(response.url)
        return bool(re.search(r"\.(?:jpe?g|png|webp|avif|gif)(?:[?#]|$)", final_url, re.I)) and "text/html" not in content_type
    except Exception:
        return False


def _bing_image_candidates(session: requests.Session, query: str, timeout: int) -> List[str]:
    try:
        response = session.get(
            "https://www.bing.com/images/search",
            params={"q": query, "form": "HDRSC3", "first": "1"},
            headers=DEFAULT_HEADERS,
            timeout=timeout,
        )
        if response.status_code != 200:
            return []
        html_text = response.text or ""
        candidates: List[str] = []
        # Bing serializes original image URLs in an `m` JSON attribute.
        for raw in re.findall(r'\bm=(["\'])(.*?)\1', html_text, flags=re.I | re.S):
            payload = html.unescape(raw[1])
            try:
                obj = json.loads(payload)
            except Exception:
                continue
            for key in ("murl", "turl"):
                value = clean_text(obj.get(key, ""))
                if value:
                    candidates.append(value)
        # Newer Bing layouts can include the JSON as escaped data-m attributes.
        for match in re.finditer(r'"murl"\s*:\s*"([^"]+)"', html_text, flags=re.I):
            candidates.append(clean_text(match.group(1)).replace("\\/", "/"))
        return dedupe_keep_order(candidates)
    except Exception:
        return []


def _duckduckgo_image_candidates(session: requests.Session, query: str, timeout: int) -> List[str]:
    try:
        landing = session.get(
            "https://duckduckgo.com/",
            params={"q": query, "iax": "images", "ia": "images"},
            headers=DEFAULT_HEADERS,
            timeout=timeout,
        )
        if landing.status_code != 200:
            return []
        raw = landing.text or ""
        token_match = re.search(r'vqd=["\']([^"\']+)', raw, flags=re.I)
        if not token_match:
            token_match = re.search(r'vqd\\?["\']?\s*[:=]\s*["\']([^"\']+)', raw, flags=re.I)
        if not token_match:
            return []
        vqd = token_match.group(1)
        api = session.get(
            "https://duckduckgo.com/i.js",
            params={"l": "us-en", "o": "json", "q": query, "vqd": vqd, "f": ",,,,,", "p": "1"},
            headers={**DEFAULT_HEADERS, "Referer": "https://duckduckgo.com/"},
            timeout=timeout,
        )
        if api.status_code != 200:
            return []
        data = api.json()
        rows = data.get("results", []) if isinstance(data, dict) else []
        out = []
        for row in rows if isinstance(rows, list) else []:
            if isinstance(row, dict):
                out.append(clean_text(row.get("image", "")))
                out.append(clean_text(row.get("thumbnail", "")))
        return dedupe_keep_order(out)
    except Exception:
        return []


def search_product_image_url(
    session: requests.Session,
    model: str,
    title: str,
    timeout: int,
    engines: Sequence[str] = ("bing", "duckduckgo"),
) -> Tuple[str, str]:
    """Find a reachable product-photo URL only when Lutron itself has none."""
    query = " ".join(x for x in ("Lutron", clean_text(model), clean_text(title)) if x)
    if not query:
        return "", ""
    all_candidates: List[Tuple[str, str]] = []
    for engine in engines:
        if engine == "bing":
            urls = _bing_image_candidates(session, query, timeout)
        elif engine == "duckduckgo":
            urls = _duckduckgo_image_candidates(session, query, timeout)
        else:
            urls = []
        all_candidates.extend((engine, url) for url in urls if not _image_bad_url(url))

    seen = set()
    ranked: List[Tuple[int, str, str]] = []
    for engine, url in all_candidates:
        key = url.casefold()
        if key in seen:
            continue
        seen.add(key)
        ranked.append((_image_candidate_score(url, model, title, engine), engine, url))
    ranked.sort(reverse=True)

    # Test only a small ranked set. This makes it practical for hundreds of
    # products while keeping `image_url` usable by WooCommerce.
    for _score, engine, url in ranked[:12]:
        if _verify_image_candidate(session, url, timeout):
            return url, engine
    return "", ""


def _apply_verified_fallbacks(
    data: Dict[str, Any],
    record: Dict[str, str],
    session: requests.Session,
    timeout: int,
    enable_datasheet: bool = True,
    enable_image_search: bool = True,
    datasheet_attempt_limit: int = 20,
) -> Dict[str, Any]:
    """Fill incomplete link scrape with verified datasheet/image fallbacks."""
    model = clean_text(data.get("model") or record.get("input_model"))
    page_was_complete = _is_real_product_data(data)
    notes: List[str] = []
    if not page_was_complete:
        notes.append("lutron_page_incomplete")

    if not page_was_complete and enable_datasheet and model:
        pdf = _fetch_datasheet_for_model(session, model, timeout, datasheet_attempt_limit)
        if pdf.get("ok"):
            pdf_text = clean_text(pdf.get("text"))
            parsed_specs = _find_labeled_pdf_specs(pdf_text, model)
            description = _pdf_feature_description(pdf_text)
            fallback_title = _pdf_title(pdf_text, model, record.get("input_title", ""))
            data["title"] = fallback_title or clean_text(data.get("title")) or clean_text(record.get("input_title")) or model
            data["model"] = model
            data["description"] = description or clean_text(data.get("description"))
            data["official_content_html"] = _pdf_content_html(pdf_text, data["title"]) or clean_text(data.get("official_content_html"))
            data["spec_rows"] = parsed_specs or data.get("spec_rows") or []
            data["datasheet_url"] = clean_text(pdf.get("url"))
            data["datasheet_source_host"] = clean_text(pdf.get("source_host"))
            data["datasheet_extraction_method"] = clean_text(pdf.get("extraction_method"))
            data["status"] = "datasheet_fallback"
            data["source"] = "lutron_link_plus_verified_datasheet"
            notes.append("datasheet_fallback=" + clean_text(pdf.get("url")))
            if clean_text(pdf.get("extraction_method")) == "ocr":
                notes.append("datasheet_ocr_used")
            data["error"] = _merge_source_notes(data.get("error"), *notes)
        else:
            notes.append("datasheet_fallback_failed_" + clean_text(pdf.get("reason", "unknown")))
            data["error"] = _merge_source_notes(data.get("error"), *notes)

    # A link scrape can have useful specs but no image, and an image fallback
    # is only needed in that case. The source is recorded transparently.
    if not clean_text(data.get("image_url")) and enable_image_search:
        image_url, engine = search_product_image_url(
            session,
            model=model or record.get("input_model", ""),
            title=clean_text(data.get("title")) or record.get("input_title", ""),
            timeout=timeout,
        )
        if image_url:
            data["image_url"] = image_url
            data["image_source"] = engine + "_image_search"
            notes.append("image_fallback=" + data["image_source"])
        else:
            notes.append("image_fallback_not_found")
        data["error"] = _merge_source_notes(data.get("error"), *notes)
    elif clean_text(data.get("image_url")):
        data["image_source"] = "lutron_product_page_or_rendered_header"

    # Classification can change after a verified datasheet supplied the real
    # title/specifications. Re-evaluate it here so fallback pages do not retain
    # a stale or guessed accessory status.
    accessory, accessory_reason = is_accessory_product(
        clean_text(data.get("title")),
        clean_text(data.get("description")),
        data.get("spec_rows") if isinstance(data.get("spec_rows"), list) else [],
    )
    data["is_accessory"] = accessory
    data["accessory_reason"] = accessory_reason

    # Do not silently create a seemingly successful catalog item from only the
    # input Excel. A record needs substantive source data, then it may be sent
    # to WooCommerce.
    if not _is_real_product_data(data) and data.get("status") != "datasheet_fallback":
        data["status"] = "incomplete_source"
    data["quality_flags"] = collect_source_quality_flags(data)
    return data


def _temperature_conversion_consistent(value: Any) -> bool:
    """Detect obvious Celsius/Fahrenheit conversion errors in product text."""
    text = clean_text(value)
    if not text:
        return True
    number = r"(-?\d+(?:[.,]\d+)?)"
    c_ranges = re.findall(number + r"\s*(?:to|hingga|sampai|[-–~])\s*" + number + r"\s*°?\s*C", text, flags=re.I)
    f_ranges = re.findall(number + r"\s*(?:to|hingga|sampai|[-–~])\s*" + number + r"\s*°?\s*F", text, flags=re.I)
    def n(v: str) -> float:
        return float(v.replace(",", "."))
    if c_ranges and f_ranges:
        for c_pair, f_pair in zip(c_ranges, f_ranges):
            expected = (n(c_pair[0]) * 9 / 5 + 32, n(c_pair[1]) * 9 / 5 + 32)
            actual = (n(f_pair[0]), n(f_pair[1]))
            if any(abs(a - e) > 4.0 for a, e in zip(actual, expected)):
                return False
    # Single-point notation such as 0°C (32°F).
    singles = re.findall(number + r"\s*°?\s*C[^0-9-]{0,30}" + number + r"\s*°?\s*F", text, flags=re.I)
    for c, f in singles:
        if abs((n(c) * 9 / 5 + 32) - n(f)) > 4.0:
            return False
    return True


def collect_source_quality_flags(data: Dict[str, Any]) -> List[str]:
    flags: List[str] = []
    # Do not use a global Celsius/Fahrenheit pairing heuristic as a publication
    # blocker. Lutron pages can repeat, reorder, or separately format ranges.
    # The translated spec validator preserves every numeric source token instead.
    image = clean_text(data.get("image_url"))
    if image and not valid_image_url(image):
        flags.append("rejected_generic_image")
    return dedupe_keep_order(flags)


def _canonical_image_key(url: Any) -> str:
    parts = urlsplit(clean_text(url))
    return urlunsplit((parts.scheme.casefold(), parts.netloc.casefold(), parts.path, "", ""))


def apply_batch_image_quality_gate(
    products: Sequence[Dict[str, Any]],
    skipped: List[Dict[str, Any]],
    require_verified_image: bool = False,
) -> List[Dict[str, Any]]:
    """Reject generic or cross-model duplicate images before Excel generation."""
    by_url: Dict[str, List[Dict[str, Any]]] = {}
    for data in products:
        image = clean_text(data.get("image_url"))
        if not image:
            continue
        by_url.setdefault(_canonical_image_key(image), []).append(data)

    invalid_ids = set()
    for key, group in by_url.items():
        models = {clean_text(item.get("model")) for item in group if clean_text(item.get("model"))}
        if len(group) > 1 and len(models) > 1:
            for item in group:
                invalid_ids.add(id(item))
                item["image_url"] = ""
                item["image_source"] = ""
                item["quality_flags"] = dedupe_keep_order(list(item.get("quality_flags") or []) + ["duplicate_image_across_models"])
                item["error"] = _merge_source_notes(item.get("error"), "duplicate_image_across_models")

    approved: List[Dict[str, Any]] = []
    for data in products:
        image = clean_text(data.get("image_url"))
        if image and not valid_image_url(image):
            data["image_url"] = ""
            data["image_source"] = ""
            data["quality_flags"] = dedupe_keep_order(list(data.get("quality_flags") or []) + ["rejected_generic_image"])
            data["error"] = _merge_source_notes(data.get("error"), "rejected_generic_image")
            image = ""
        if require_verified_image and not image:
            skipped.append({
                "url": data.get("url", ""),
                "title": data.get("title", ""),
                "status": data.get("status", ""),
                "reason": "missing_or_rejected_product_image",
                "error": _merge_source_notes(data.get("error"), "; ".join(data.get("quality_flags") or [])),
            })
            continue
        approved.append(data)
    return approved


def scrape_product_with_fallbacks(
    record: Dict[str, str],
    timeout: int,
    session: requests.Session,
    renderer: Optional[LutronBrowserRenderer] = None,
    enable_datasheet: bool = True,
    enable_image_search: bool = True,
    datasheet_attempt_limit: int = 20,
) -> Dict[str, Any]:
    data = scrape_product_page(record, timeout, session, renderer)
    return _apply_verified_fallbacks(
        data=data,
        record=record,
        session=session,
        timeout=timeout,
        enable_datasheet=enable_datasheet,
        enable_image_search=enable_image_search,
        datasheet_attempt_limit=datasheet_attempt_limit,
    )


# ---------------------------------------------------------------------------
# Copy and WooCommerce row generation
# ---------------------------------------------------------------------------

def normalize_lutron_category(value: Any, product_title: str = "") -> str:
    """Gunakan kategori breadcrumb hanya jika benar-benar kategori, bukan breadcrumb Products > judul."""
    s = clean_catalog_source_text(value)
    if not s:
        return ""
    s = s.replace("chevron_right", ">")
    parts = [clean_text(x).strip(" >|") for x in re.split(r"\s*>\s*|\s*›\s*", s) if clean_text(x).strip(" >|")]
    filtered: List[str] = []
    title_key = clean_text(product_title).casefold()
    for part in parts:
        key = part.casefold()
        if key in {"lutron", "products", "product", "all products"}:
            continue
        if title_key and key == title_key:
            continue
        filtered.append(part)
    # Breadcrumb halaman Lutron yang hanya "Products > nama produk" tidak boleh menjadi kategori.
    if not filtered:
        return ""
    return " > ".join(filtered)


def detect_family(data: Dict[str, Any]) -> Dict[str, Any]:
    title = clean_text(data.get("title"))
    desc = clean_catalog_source_text(data.get("description"))
    specs = data.get("spec_rows") if isinstance(data.get("spec_rows"), list) else []
    if data.get("is_accessory"):
        family = dict(ACCESSORY_FAMILY)
    else:
        # Model dan title diprioritaskan. Hindari salah baca "Operating Temperature"
        # sebagai kategori thermometer untuk produk yang sebenarnya berbeda.
        core_blob = " ".join([title, clean_text(data.get("model")), clean_catalog_source_text(data.get("source_category"))])
        full_blob = " ".join([core_blob, desc, " ".join(f"{x.get('label','')} {x.get('value','')}" for x in specs)])
        family = dict(DEFAULT_FAMILY)
        for pattern, rule in FAMILY_RULES:
            if pattern.search(core_blob) or pattern.search(full_blob):
                family = dict(rule)
                break
    source_category = normalize_lutron_category(data.get("source_category"), title)
    # Model/title family rules are authoritative. A generic breadcrumb must not
    # override a recognised product class such as TR-PS pressure transmitters.
    if clean_text(family.get("category")) == clean_text(DEFAULT_FAMILY.get("category")) and source_category:
        family["category"] = source_category
    elif not clean_text(family.get("category")).casefold().startswith("lutron >"):
        family["category"] = clean_text(family.get("category"))
    return family


def spec_value(specs: Sequence[Dict[str, str]], labels: Sequence[str]) -> str:
    for row in specs:
        l = clean_text(row.get("label")).casefold()
        if any(term.casefold() in l for term in labels):
            return clean_text(row.get("value"))
    return ""


def feature_bullets(data: Dict[str, Any], family: Dict[str, Any]) -> List[str]:
    title = clean_text(data.get("title"))
    brand = clean_text(data.get("brand"))
    specs = data.get("spec_rows") or []
    if data.get("is_accessory"):
        bullets = [f"Produk ini merupakan aksesori atau komponen pendukung {brand} yang diproses sebagai produk katalog tersendiri."]
        model = clean_text(data.get("model")) or spec_value(specs, ["model", "sku", "cat", "part"])
        if model:
            bullets.append(f"Model atau referensi produk yang tercantum: {model}.")
        bullets.append("Cocokkan kompatibilitas produk dengan unit utama sebelum pemesanan atau penggunaan.")
        bullets.append(f"Data teknis dirujuk dari {source_official_name(get_source_config(data.get('url','')) or LUTRON_CONFIG)}.")
        return dedupe_keep_order(bullets)[:5]

    bullets: List[str] = []
    candidates = [
        ("Rentang pengukuran", ["range", "measurement range", "measuring range"]),
        ("Akurasi", ["accuracy"]),
        ("Resolusi", ["resolution", "minimum scale"]),
        ("Suhu operasi", ["operating temperature", "temperature range", "ambient temperature"]),
        ("Output atau antarmuka", ["output", "interface", "communication"]),
        ("Catu daya", ["power supply", "battery"]),
    ]
    for nice, terms in candidates:
        value = spec_value(specs, terms)
        if value:
            bullets.append(f"{nice}: {value}.")
    bullets.append(f"Dirancang untuk kebutuhan pengukuran, pemeriksaan, atau kontrol kualitas menggunakan {family['term']}.")
    if title and brand:
        bullets.append(f"Merek produk: {brand}.")
    return dedupe_keep_order(bullets)[:6]


def source_description_block(description: str, cfg: SourceConfig) -> str:
    # Jangan pernah meneruskan teks landing page Lutron atau karakter Chinese ke output.
    description = compact(clean_catalog_source_text(description), 1600)
    if not description:
        return ""
    label = "Ringkasan dari halaman Lutron"
    return f"<p><strong>{label}:</strong> {esc(description)}</p>"


def build_product_description(data: Dict[str, Any], family: Dict[str, Any]) -> str:
    title = clean_text(data.get("title"))
    brand = clean_text(data.get("brand"))
    cfg = get_source_config(data.get("url", "")) or LUTRON_CONFIG
    desc = source_description_block(clean_text(data.get("description")), cfg)
    bullets = "".join(f"<li>{esc(x)}</li>" for x in feature_bullets(data, family))
    if data.get("is_accessory"):
        intro = f"<p><strong>{esc(title)}</strong> adalah {esc(family['term'])} dari {esc(brand)} untuk melengkapi unit utama yang kompatibel.</p>"
        selection = "<p>Pastikan model, referensi produk, dan kecocokan teknis dengan unit utama sebelum digunakan.</p>"
    else:
        intro = f"<p><strong>{esc(title)}</strong> adalah {esc(family['term'])} dari {esc(brand)} yang digunakan untuk {esc(family['function'])}.</p>"
        selection = "<p>Gunakan informasi spesifikasi untuk memastikan kesesuaian produk dengan kebutuhan pengukuran dan kondisi penggunaan.</p>"
    return "".join([
        intro, desc,
        "<h2>Fungsi dan Keunggulan</h2>",
        f"<ul>{bullets}</ul>" if bullets else "",
        "<h2>Informasi Pemilihan Produk</h2>", selection,
        UJI_CTA,
    ])


def build_short_description(data: Dict[str, Any], family: Dict[str, Any]) -> str:
    title, brand = clean_text(data.get("title")), clean_text(data.get("brand"))
    specs = data.get("spec_rows") or []
    model = clean_text(data.get("model")) or spec_value(specs, ["model", "sku", "part"])
    if data.get("is_accessory"):
        text = f"{title} adalah aksesori atau komponen pendukung {brand} untuk unit utama yang kompatibel."
        if model:
            text += f" Referensi: {model}."
        return compact(text, 320)
    text = f"{title} adalah {family['term']} dari {brand} untuk {family['function']}."
    range_value = spec_value(specs, ["range", "measurement range"])
    accuracy = spec_value(specs, ["accuracy"])
    if range_value:
        text += f" Rentang: {range_value}."
    if accuracy:
        text += f" Akurasi: {accuracy}."
    return compact(text, 320)


def build_application_tab(data: Dict[str, Any], family: Dict[str, Any]) -> str:
    if data.get("is_accessory"):
        apps = [
            "kelengkapan unit utama yang kompatibel",
            "penggantian atau penambahan komponen sesuai referensi model",
            "pemeliharaan peralatan ukur",
            "kebutuhan laboratorium, inspeksi, atau quality control",
        ]
    else:
        apps = [
            "kontrol kualitas produksi",
            "pemeriksaan sampel atau parameter proses",
            "pengujian di laboratorium atau area inspeksi",
            "validasi parameter sebelum proses lanjutan",
        ]
    return "<h2>Aplikasi Penggunaan</h2><ul>" + "".join(f"<li>{esc(x)}</li>" for x in apps) + "</ul>"


def build_faq(data: Dict[str, Any], family: Dict[str, Any]) -> str:
    title, brand = clean_text(data.get("title")), clean_text(data.get("brand"))
    cfg = get_source_config(data.get("url", "")) or LUTRON_CONFIG
    model = clean_text(data.get("model"))
    if data.get("is_accessory"):
        return (
            f"<h3>Apa fungsi {esc(title)}?</h3><p>Produk ini merupakan aksesori atau komponen pendukung {esc(brand)} untuk unit utama yang kompatibel.</p>"
            f"<h3>Bagaimana memastikan kompatibilitasnya?</h3><p>Periksa model, nomor referensi, dan kebutuhan teknis unit utama sebelum pemesanan atau penggunaan.</p>"
            f"<h3>Apakah produk ini alat ukur utama?</h3><p>Tidak selalu. Produk ini diklasifikasikan sebagai aksesori atau komponen karena informasi halaman produk menunjukkan fungsi pendukung.</p>"
            f"<h3>Dari mana data teknis disusun?</h3><p>Data produk dirujuk dari {esc(source_official_name(cfg))}.</p>"
        )
    return (
        f"<h3>Apa fungsi {esc(title)}?</h3><p>{esc(title)} digunakan sebagai {esc(family['term'])} untuk {esc(family['function'])}.</p>"
        f"<h3>Apakah spesifikasi mengikuti halaman merek?</h3><p>Ya. Tab spesifikasi dan informasi produk dirujuk dari {esc(source_official_name(cfg))}.</p>"
        f"<h3>Bagaimana memilih model yang sesuai?</h3><p>Bandingkan rentang, akurasi, resolusi, kondisi penggunaan, dan kebutuhan antarmuka pada spesifikasi produk.</p>"
        f"<h3>Apakah model produk tercantum?</h3><p>{esc('Model yang terdeteksi: ' + model + '.' if model else 'Periksa tab Spesifikasi dan Informasi Produk Resmi untuk referensi model yang tersedia.')}</p>"
    )


def build_meta_description(data: Dict[str, Any], family: Dict[str, Any]) -> str:
    title, brand = clean_text(data.get("title")), clean_text(data.get("brand"))
    if data.get("is_accessory"):
        return compact(f"{title} adalah aksesori atau komponen {brand}. Periksa kompatibilitas model dan spesifikasinya di nama-website-kalian.", 155)
    range_value = spec_value(data.get("spec_rows") or [], ["range", "measurement range"])
    text = f"{title} dari {brand} untuk {family['function']}."
    if range_value:
        text += f" Rentang: {range_value}."
    text += " Cek spesifikasi di nama-website-kalian."
    return compact(text, 155)


# ---------------------------------------------------------------------------
# Optional DeepSeek content generation and Indonesian localization
# ---------------------------------------------------------------------------

AI_SETTINGS: Dict[str, Any] = {
    "enabled": False,
    "api_key": "",
    "base_url": "https://api.deepseek.com",
    "model": "deepseek-chat",
    "cache": {},
    "cache_path": "",
    "limit": 0,
    "timeout": 90,
    "delay": 0.0,
    "temperature": 0.15,
    "strict_indonesian": True,
    "spec_limit": 120,
}
AI_STATS: Dict[str, int] = {"used": 0, "cache_hit": 0, "fallback": 0, "skipped": 0}
AI_PROMPT_VERSION = "lutron_uji_deepseek_indonesian_v18_ai_semantic_spec_dedup_2026_06_29"

# Fallback hanya untuk label. Dalam mode AI bersih, nilai spesifikasi
# diterjemahkan oleh DeepSeek agar tidak ada teks Inggris/Chinese yang masuk.
SPEC_LABEL_ID_MAP = {
    "model": "Model",
    "circuit": "Sirkuit",
    "display": "Tampilan",
    "displaylcd size": "Ukuran layar LCD",
    "display direction": "Arah tampilan",
    "measurement unit": "Satuan pengukuran",
    "measurement range": "Rentang pengukuran",
    "measuring range": "Rentang pengukuran",
    "range": "Rentang",
    "accuracy": "Akurasi",
    "resolution": "Resolusi",
    "operating temperature": "Suhu operasi",
    "operating humidity": "Kelembapan operasi",
    "power supply": "Catu daya",
    "power current": "Arus daya",
    "power consumption": "Konsumsi daya",
    "weight": "Berat",
    "dimension": "Dimensi",
    "dimensions": "Dimensi",
    "function": "Fungsi",
    "data hold": "Penahan data",
    "data output": "Keluaran data",
    "interface": "Antarmuka",
    "sampling time of display": "Waktu sampling tampilan",
    "data logger sampling time setting range": "Rentang pengaturan waktu sampling pencatat data",
    "accessories included": "Aksesori yang disertakan",
    "accessories": "Aksesori",
    "zero adjust": "Penyetelan nol",
    "span adjust": "Penyetelan rentang",
    "principal": "Prinsip kerja",
    "input": "Masukan",
    "output": "Keluaran",
    "battery": "Baterai",
}


def set_ai_settings(**kwargs: Any) -> None:
    AI_SETTINGS.update(kwargs)
    AI_SETTINGS["base_url"] = clean_text(AI_SETTINGS.get("base_url")).rstrip("/") or "https://api.deepseek.com"
    AI_SETTINGS["cache"] = AI_SETTINGS.get("cache") if isinstance(AI_SETTINGS.get("cache"), dict) else {}
    AI_SETTINGS["limit"] = max(0, int(AI_SETTINGS.get("limit") or 0))
    AI_SETTINGS["timeout"] = max(20, int(AI_SETTINGS.get("timeout") or 90))
    AI_SETTINGS["delay"] = max(0.0, float(AI_SETTINGS.get("delay") or 0.0))
    AI_SETTINGS["temperature"] = max(0.0, min(1.0, float(AI_SETTINGS.get("temperature") or 0.15)))
    AI_SETTINGS["spec_limit"] = max(1, int(AI_SETTINGS.get("spec_limit") or 120))
    AI_SETTINGS["strict_indonesian"] = bool(AI_SETTINGS.get("strict_indonesian", True))


def _html_text(value: Any) -> str:
    s = clean_text(value)
    s = re.sub(r"<\s*br\s*/?>", "\n", s, flags=re.I)
    s = re.sub(r"</\s*(?:p|li|h2|h3|tr)\s*>", "\n", s, flags=re.I)
    s = re.sub(r"<[^>]+>", " ", s)
    return re.sub(r"\s+", " ", html.unescape(s)).strip()


def _sanitize_ai_html(value: Any) -> str:
    s = clean_text(value)
    s = re.sub(r"^```(?:json|html)?\s*|\s*```$", "", s, flags=re.I)
    s = re.sub(r"<\s*(script|style|iframe|object|embed)[^>]*>.*?<\s*/\s*\1\s*>", "", s, flags=re.I | re.S)
    s = re.sub(r"\s+on[a-z]+\s*=\s*(['\"]).*?\1", "", s, flags=re.I | re.S)
    allowed = {"p", "strong", "b", "em", "ul", "ol", "li", "h2", "h3", "br"}
    def tag_repl(m: re.Match) -> str:
        slash = "/" if m.group(1) else ""
        tag = (m.group(2) or "").casefold()
        return f"<{slash}{tag}>" if tag in allowed else ""
    s = re.sub(r"<\s*(/?)\s*([a-zA-Z0-9]+)(?:\s+[^>]*)?>", tag_repl, s)
    return s.strip()


def _sanitize_ai_text(value: Any, max_chars: int) -> str:
    s = compact(_html_text(value), max_chars)
    if s and s[-1] not in ".!?":
        s += "."
    return s


def _parse_json(text: str) -> Dict[str, Any]:
    s = clean_text(text)
    s = re.sub(r"^```(?:json)?\s*|\s*```$", "", s, flags=re.I)
    try:
        value = json.loads(s)
        return value if isinstance(value, dict) else {}
    except Exception:
        start, end = s.find("{"), s.rfind("}")
        if start >= 0 and end > start:
            try:
                value = json.loads(s[start:end + 1])
                return value if isinstance(value, dict) else {}
            except Exception:
                return {}
    return {}


def _spec_rows_for_ai(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Send DeepSeek the untouched source rows, including both spec sections.

    `prepare_canonical_specs_for_product()` creates `_raw_spec_rows` before it
    applies fallback rules. AI must inspect that raw list, otherwise it cannot
    identify duplicates that the rule-based parser did not recognise.
    """
    raw_rows = data.get("_raw_spec_rows")
    rows = raw_rows if isinstance(raw_rows, list) else (
        data.get("spec_rows") if isinstance(data.get("spec_rows"), list) else []
    )
    limit = int(AI_SETTINGS.get("spec_limit") or 120)
    clean_rows: List[Dict[str, Any]] = []
    for index, item in enumerate(rows[:limit], 1):
        if not isinstance(item, dict):
            continue
        source_label = clean_catalog_source_text(item.get("source_label") or item.get("label"))
        value = clean_catalog_source_text(item.get("value"))
        if source_label and value:
            clean_rows.append({
                "id": index,
                "source_label": source_label,
                "source_value": compact(value, 700),
                "section": clean_text(item.get("section")) or "Specifications",
            })
    return clean_rows


def _ai_payload(data: Dict[str, Any], family: Dict[str, Any]) -> Dict[str, Any]:
    source_full = html_to_source_text(data.get("official_content_html"), 8500)
    source_desc = compact(clean_catalog_source_text(data.get("description")), 2200)
    return {
        "source_product_name": clean_catalog_source_text(data.get("title")),
        "brand": clean_text(data.get("brand")) or "Lutron",
        "model": clean_text(data.get("model")),
        "source_url": clean_text(data.get("url")),
        "source_category": clean_catalog_source_text(data.get("source_category")),
        "product_type": "accessory_or_component" if data.get("is_accessory") else "main_product",
        "suggested_category": clean_catalog_source_text(family.get("category")).replace("", "").strip(),
        "family_function": clean_catalog_source_text(family.get("function")),
        "source_description": source_desc,
        "source_full_information": source_full,
        "specifications": _spec_rows_for_ai(data),
    }


def _ai_cache_key(payload: Dict[str, Any]) -> str:
    raw = json.dumps(
        {"version": AI_PROMPT_VERSION, "model": AI_SETTINGS.get("model"), "payload": payload},
        ensure_ascii=False,
        sort_keys=True,
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _numeric_tokens(value: Any) -> List[str]:
    return re.findall(r"(?<![A-Za-z])[-+]?\d+(?:[.,]\d+)?", clean_text(value))


def _localize_source_label(source_label: Any, localized_label: Any) -> str:
    """Return a safe Indonesian display label without changing source identity.

    The original v16 called this helper from both the AI and fallback paths,
    but the function was missing. That NameError made every AI-enabled product
    fail validation and enter the Skipped sheet.
    """
    localized = clean_catalog_source_text(localized_label)
    source = clean_catalog_source_text(source_label)
    return normalize_label(localized or source)


def _coerce_ai_spec_id(value: Any) -> Optional[int]:
    try:
        parsed = int(str(value).strip())
        return parsed if parsed > 0 else None
    except Exception:
        return None


def _validate_all_translated_specs(
    parsed: Dict[str, Any],
    source_specs: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], Dict[int, Dict[str, Any]]]:
    """Validate translation row-by-row before applying AI duplicate selection."""
    raw_items = parsed.get("translated_specifications")
    if not isinstance(raw_items, list) or not source_specs:
        return [], {}

    translated_by_id: Dict[int, Dict[str, Any]] = {}
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        item_id = _coerce_ai_spec_id(item.get("id"))
        if item_id is not None:
            translated_by_id[item_id] = item

    expected_ids = [int(spec["id"]) for spec in source_specs]
    if set(translated_by_id) != set(expected_ids):
        return [], {}

    translated_rows: List[Dict[str, Any]] = []
    mapped: Dict[int, Dict[str, Any]] = {}
    for spec in source_specs:
        source_id = int(spec["id"])
        item = translated_by_id.get(source_id)
        if not item:
            return [], {}
        label = clean_catalog_source_text(item.get("label_id"))
        value = clean_catalog_source_text(item.get("value_id"))
        if not label or not value or has_cjk(label) or has_cjk(value):
            return [], {}
        normalized_value = value.replace(",", ".")
        if any(token.replace(",", ".") not in normalized_value for token in _numeric_tokens(spec["source_value"])):
            return [], {}
        translated = {
            "id": source_id,
            "label": _localize_source_label(spec.get("source_label"), label),
            "source_label": clean_text(spec.get("source_label")),
            "value": value,
            "section": clean_text(spec.get("section")) or "Specifications",
        }
        translated_rows.append(translated)
        mapped[source_id] = translated
    return translated_rows, mapped


def _source_label_duplicate_family(left: Dict[str, Any], right: Dict[str, Any]) -> bool:
    """Reject unsafe AI removals where labels clearly represent different facts."""
    left_key, _left_std, left_source = _canonical_label_info({
        "source_label": left.get("source_label"),
        "label": left.get("source_label"),
    })
    right_key, _right_std, right_source = _canonical_label_info({
        "source_label": right.get("source_label"),
        "label": right.get("source_label"),
    })
    if _canonical_keys_can_compare(left_key, right_key):
        return True
    a, b = _spec_text_key(left_source), _spec_text_key(right_source)
    if not a or not b:
        return False
    if a == b:
        return True
    return SequenceMatcher(None, a, b).ratio() >= 0.92


def _source_values_plausibly_identical(left: Dict[str, Any], right: Dict[str, Any]) -> bool:
    """Allow AI to merge only cosmetic variants of the same source fact.

    This guard keeps AI from deleting valid values such as Meter versus Probe
    dimensions, multiple frequency ranges, or different operating modes.
    """
    a = clean_text(left.get("source_value"))
    b = clean_text(right.get("source_value"))
    if _spec_values_equivalent(a, b):
        return True

    a_numbers = [v.replace(",", ".") for v in _numeric_tokens(a)]
    b_numbers = [v.replace(",", ".") for v in _numeric_tokens(b)]
    if a_numbers != b_numbers:
        return False

    # Values with numerical data may differ only in punctuation, spacing, or
    # equivalent wording such as "to", "-", and "hingga".
    if a_numbers:
        a_units = _spec_unit_signature(a)
        b_units = _spec_unit_signature(b)
        if a_units and b_units and a_units != b_units:
            return False
        return True

    a_key, b_key = _spec_text_key(a), _spec_text_key(b)
    if not a_key or not b_key:
        return False
    if a_key == b_key or a_key in b_key or b_key in a_key:
        return True
    return SequenceMatcher(None, a_key, b_key).ratio() >= 0.88


def _validate_ai_spec_deduplication(
    parsed: Dict[str, Any],
    source_specs: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, str]], int]:
    """Use DeepSeek's keep/remove decision after conservative source checks."""
    translated_rows, translations = _validate_all_translated_specs(parsed, source_specs)
    if not translated_rows:
        return [], 0

    raw_keep_ids = parsed.get("keep_specification_ids")
    raw_removed = parsed.get("removed_duplicate_specifications")
    if not isinstance(raw_keep_ids, list) or not isinstance(raw_removed, list):
        return [], 0

    expected_ids = {int(spec["id"]) for spec in source_specs}
    keep_ids = [_coerce_ai_spec_id(value) for value in raw_keep_ids]
    if any(value is None for value in keep_ids):
        return [], 0
    keep_ids = [int(value) for value in keep_ids if value is not None]
    if len(keep_ids) != len(set(keep_ids)) or not set(keep_ids).issubset(expected_ids):
        return [], 0

    removed_map: Dict[int, int] = {}
    for item in raw_removed:
        if not isinstance(item, dict):
            return [], 0
        source_id = _coerce_ai_spec_id(item.get("source_id", item.get("id")))
        duplicate_of_id = _coerce_ai_spec_id(item.get("duplicate_of_id", item.get("keep_id")))
        if source_id is None or duplicate_of_id is None or source_id == duplicate_of_id:
            return [], 0
        if source_id in removed_map:
            return [], 0
        removed_map[source_id] = duplicate_of_id

    removed_ids = set(removed_map)
    if not removed_ids.issubset(expected_ids):
        return [], 0
    if not set(removed_map.values()).issubset(set(keep_ids)):
        return [], 0
    if set(keep_ids) & removed_ids:
        return [], 0
    # Every source row must be accounted for. This prevents AI from silently
    # losing a unique specification while it selects duplicates.
    if set(keep_ids) | removed_ids != expected_ids:
        return [], 0

    source_by_id = {int(item["id"]): item for item in source_specs}
    for source_id, target_id in removed_map.items():
        source_row = source_by_id[source_id]
        target_row = source_by_id[target_id]
        if not _source_label_duplicate_family(source_row, target_row):
            return [], 0
        if not _source_values_plausibly_identical(source_row, target_row):
            return [], 0

        # If the source sections are a direct Technical-vs-Specifications pair,
        # Technical Specifications may never be removed in favor of the generic
        # Specifications section.
        source_is_technical = _is_technical_spec_section(source_row.get("section"))
        target_is_technical = _is_technical_spec_section(target_row.get("section"))
        if source_is_technical and not target_is_technical:
            return [], 0

    output: List[Dict[str, str]] = []
    for source_id in keep_ids:
        row = translations[source_id]
        output.append({
            "label": clean_text(row.get("label")),
            "source_label": clean_text(row.get("source_label")),
            "value": clean_text(row.get("value")),
            "section": clean_text(row.get("section")),
        })
    return output, len(removed_ids)


def _validate_ai_fields(parsed: Dict[str, Any], source_specs: List[Dict[str, Any]]) -> Tuple[bool, Dict[str, Any], str]:
    fields: Dict[str, Any] = {
        "product_name_id": _sanitize_ai_text(parsed.get("product_name_id"), 180),
        "category_id": _sanitize_ai_text(parsed.get("category_id"), 100).rstrip("."),
        "product_description_html": _sanitize_ai_html(parsed.get("product_description_html")),
        "short_description": _sanitize_ai_text(parsed.get("short_description"), 260),
        "feature_tab_html": _sanitize_ai_html(parsed.get("feature_tab_html")),
        "faq_html": _sanitize_ai_html(parsed.get("faq_html")),
        "meta_description": _sanitize_ai_text(parsed.get("meta_description"), 158),
        "official_information_html": _sanitize_ai_html(parsed.get("official_information_html")),
    }
    translated_rows, removed_count = _validate_ai_spec_deduplication(parsed, source_specs)
    fields["translated_spec_rows"] = translated_rows
    fields["ai_spec_dedup_applied"] = bool(source_specs)
    fields["ai_spec_dedup_removed_count"] = removed_count

    required_text = [
        fields["product_name_id"],
        fields["product_description_html"],
        fields["short_description"],
        fields["feature_tab_html"],
        fields["faq_html"],
        fields["meta_description"],
        fields["official_information_html"],
    ]
    if any(not item for item in required_text):
        return False, {}, "missing_indonesian_ai_field"
    if any(has_cjk(item) for item in required_text):
        return False, {}, "chinese_character_detected_in_ai_output"
    if len(_html_text(fields["product_description_html"])) < 160 or "<h2" not in fields["product_description_html"].lower():
        return False, {}, "invalid_product_description"
    if len(_html_text(fields["official_information_html"])) < 100 or "<h2" not in fields["official_information_html"].lower():
        return False, {}, "invalid_official_information"
    if "<li" not in fields["feature_tab_html"].lower() or "<h3" not in fields["faq_html"].lower():
        return False, {}, "invalid_ai_tabs"
    if source_specs and not fields["translated_spec_rows"]:
        return False, {}, "invalid_ai_specification_deduplication"
    forbidden = re.compile(r"(harga|diskon|promo|garansi resmi|ready stock|stok tersedia|gratis ongkir|termurah)", re.I)
    all_public_text = " ".join(str(x) for x in required_text)
    if forbidden.search(all_public_text):
        return False, {}, "forbidden_sales_claim"
    unsafe_claim = re.compile(r"(prudent avoidance|epa|who|risiko kesehatan|risiko kanker|bahaya kesehatan|diagnosis medis)", re.I)
    if unsafe_claim.search(all_public_text):
        return False, {}, "unsafe_health_or_regulatory_claim"
    return True, fields, "ok"


def _call_ai(payload: Dict[str, Any]) -> Tuple[bool, Dict[str, Any], str]:
    api_key = clean_text(AI_SETTINGS.get("api_key"))
    if not api_key:
        return False, {}, "missing_api_key"

    user = {
        "task": "Buat semua konten katalog WooCommerce dalam bahasa Indonesia dari data produk Lutron berikut.",
        "critical_rules": [
            "SEMUA teks untuk pembaca harus berbahasa Indonesia yang rapi dan profesional.",
            "Jangan tulis karakter Mandarin, Chinese, atau kutipan bahasa sumber.",
            "Pertahankan merek Lutron, model, kode, angka, satuan, lambang, serta singkatan teknis seperti LCD, True RMS, SD card, RS232, USB, dan IP apa adanya.",
            "Gunakan hanya fakta yang didukung data sumber. Jangan menambah klaim, harga, stok, garansi, sertifikasi, atau fitur.",
            "Terjemahkan SETIAP source_label dan source_value dalam specifications. Jangan menghapus, mengganti, atau mengonversi angka maupun satuan.",
            "Setelah menerjemahkan seluruh specifications, lakukan penyaringan duplikat SEMANTIK hanya untuk tab Spesifikasi.",
            "Dua spesifikasi boleh dianggap duplikat hanya jika labelnya menyatakan fakta teknis yang sama DAN semua angka, satuan, rentang, komponen, mode, kanal, sensor/probe, serta kondisi pengujiannya sama. Perbedaan bahasa, kapitalisasi, spasi, tanda baca, atau kata 'to', '-', dan 'hingga' boleh dianggap kosmetik.",
            "Jika fakta yang sama muncul pada Technical Specifications dan Specifications, WAJIB simpan ID dari Technical Specifications dan masukkan ID dari Specifications ke removed_duplicate_specifications.",
            "Jangan menghapus dua baris hanya karena labelnya sama. Pertahankan keduanya jika nilai berbeda, misalnya Meter dan Probe, 1 GHz dan 6 GHz, rentang dan akurasi, operasi normal dan pencatatan data, atau ukuran komponen yang berbeda.",
            "Jangan menggabungkan atau mengonversi nilai. Tampilkan nilai yang disimpan sesuai source_value terpilih, hanya diterjemahkan ke Bahasa Indonesia.",
            "Status product_type sudah ditentukan oleh validasi rule. Jangan mengubah status produk utama menjadi aksesori dan jangan menyebut kompatibilitas unit utama kecuali data sumber secara eksplisit menyatakan produk adalah aksesori atau komponen pengganti.",
            "Jangan membuat klaim kesehatan, risiko paparan, EPA, WHO, Prudent Avoidance, kanker, atau klaim keselamatan medis karena data sumber produk tidak digunakan untuk klaim tersebut.",
            "Jangan menyatakan USB atau RS-232 sebagai kelengkapan bawaan kecuali sumber menyebut included atau standard accessories. Bila sumber hanya menyebut interface atau optional accessory, gunakan kata mendukung atau opsional.",
            "Jangan menyalin tabel spesifikasi ke product_description_html karena tabel dibuat terpisah.",
            "Balas JSON valid saja, tanpa markdown dan tanpa code fence.",
        ],
        "output_schema": {
            "product_name_id": "Nama produk dalam bahasa Indonesia, maksimal 180 karakter. Pertahankan model/kode bila ada.",
            "category_id": "Satu kategori WooCommerce dalam bahasa Indonesia tanpa kata Lutron dan tanpa tanda >.",
            "product_description_html": "HTML aman: 2 paragraf pembuka, <h2>Keunggulan Produk</h2> dengan <ul> 4-6 poin, <h2>Panduan Pemilihan</h2>.",
            "short_description": "1-2 kalimat Indonesia maksimal 260 karakter.",
            "feature_tab_html": "HTML <h2>Keunggulan Produk</h2><ul> dengan 4-6 poin faktual dalam bahasa Indonesia.",
            "faq_html": "HTML dengan 4 pasangan <h3>pertanyaan Indonesia</h3><p>jawaban Indonesia</p>.",
            "meta_description": "Maksimal 158 karakter, bahasa Indonesia.",
            "official_information_html": "HTML Indonesia berisi <h2>Informasi Produk</h2>, ringkasan faktual halaman/dokumen sumber, dan <h2>Catatan Teknis</h2>. Jangan menyalin teks sumber bahasa Inggris atau Chinese.",
            "translated_specifications": [
                {"id": "ID setiap baris specifications dari input", "label_id": "Label spesifikasi Bahasa Indonesia", "value_id": "Nilai spesifikasi Bahasa Indonesia dengan angka/satuan persis dipertahankan"}
            ],
            "keep_specification_ids": ["Semua ID spesifikasi yang dipertahankan di tab Spesifikasi, urut mengikuti sumber. Jika tidak duplikat, ID tetap harus masuk daftar ini."],
            "removed_duplicate_specifications": [
                {"source_id": "ID yang dibuang karena benar-benar duplikat", "duplicate_of_id": "ID yang dipertahankan sebagai versi yang sama", "reason_id": "Alasan singkat Bahasa Indonesia, misalnya Duplikat identik. Tidak boleh menghapus fakta yang nilainya berbeda."}
            ],
        },
        "product_data": payload,
    }
    body = {
        "model": clean_text(AI_SETTINGS.get("model")) or "deepseek-chat",
        "messages": [
            {
                "role": "system",
                "content": (
                    "Anda adalah penerjemah teknis dan editor data spesifikasi untuk nama-website-kalian. "
                    "Tulis dalam bahasa Indonesia baku. Semua karakter Chinese dilarang. "
                    "Jaga fakta teknis tetap akurat. Terapkan deduplikasi spesifikasi sangat konservatif. "
                    "Balas JSON valid saja."
                ),
            },
            {"role": "user", "content": json.dumps(user, ensure_ascii=False)},
        ],
        "temperature": AI_SETTINGS.get("temperature", 0.15),
        "response_format": {"type": "json_object"},
    }
    endpoint = AI_SETTINGS["base_url"] + "/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    try:
        response = requests.post(endpoint, headers=headers, json=body, timeout=AI_SETTINGS["timeout"])
        if response.status_code in {400, 422}:
            body.pop("response_format", None)
            response = requests.post(endpoint, headers=headers, json=body, timeout=AI_SETTINGS["timeout"])
        response.raise_for_status()
        raw = response.json()
        content = clean_text(raw.get("choices", [{}])[0].get("message", {}).get("content", ""))
        parsed = _parse_json(content)
        valid, fields, reason = _validate_ai_fields(parsed, payload.get("specifications") or [])
        return valid, fields if valid else {}, reason
    except Exception as exc:
        return False, {}, f"request_error_{type(exc).__name__}"


def get_ai_content(data: Dict[str, Any], family: Dict[str, Any]) -> Dict[str, Any]:
    existing = data.get("_ai_fields")
    if isinstance(existing, dict):
        return existing
    if not AI_SETTINGS.get("enabled"):
        return {}
    cache = AI_SETTINGS.get("cache", {})
    payload = _ai_payload(data, family)
    key = _ai_cache_key(payload)
    cached = cache.get(key) if isinstance(cache, dict) else None
    if isinstance(cached, dict) and cached.get("ok") and isinstance(cached.get("fields"), dict):
        AI_STATS["cache_hit"] += 1
        data["_ai_status"] = "cache_hit"
        data["_ai_fields"] = cached["fields"]
        return cached["fields"]

    if AI_SETTINGS.get("limit") and AI_STATS["used"] >= AI_SETTINGS["limit"]:
        AI_STATS["skipped"] += 1
        data["_ai_status"] = "limit_reached"
        return {}

    ok, fields, reason = _call_ai(payload)
    cache[key] = {
        "ok": ok,
        "fields": fields,
        "reason": reason,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    AI_SETTINGS["cache"] = cache
    data["_ai_status"] = reason
    if ok:
        AI_STATS["used"] += 1
        data["_ai_fields"] = fields
        if AI_SETTINGS.get("delay"):
            time.sleep(AI_SETTINGS["delay"])
        return fields
    AI_STATS["fallback"] += 1
    return {}


def localize_spec_rows_fallback(rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    """Fallback label Indonesia tanpa menyisipkan teks sumber Chinese."""
    out: List[Dict[str, str]] = []
    for row in rows:
        raw_label = clean_catalog_source_text(row.get("label"))
        raw_value = clean_catalog_source_text(row.get("value"))
        if not raw_label or not raw_value:
            continue
        key = re.sub(r"\s+", " ", raw_label).casefold()
        label = SPEC_LABEL_ID_MAP.get(key, raw_label)
        out.append({
            "label": _localize_source_label(row.get("source_label") or raw_label, label),
            "source_label": clean_text(row.get("source_label")) or raw_label,
            "value": raw_value,
            "section": clean_text(row.get("section")) or "Specifications",
        })
    return dedupe_specs(out)


def fallback_official_information(data: Dict[str, Any]) -> str:
    """Fallback Indonesia yang aman. Tidak memublikasikan teks mentah Inggris/Chinese."""
    model = clean_text(data.get("model"))
    title = clean_catalog_source_text(data.get("title"))
    parts = ["<h2>Informasi Produk</h2>"]
    if title:
        parts.append(f"<p>Informasi teknis produk {esc(title)} dirangkum dari halaman resmi Lutron dan dokumen yang terdeteksi.</p>")
    else:
        parts.append("<p>Informasi teknis dirangkum dari halaman resmi Lutron dan dokumen yang terdeteksi.</p>")
    if model:
        parts.append(f"<p><strong>Model:</strong> {esc(model)}</p>")
    parts.append("<h2>Catatan Teknis</h2><p>Gunakan tab Spesifikasi untuk memeriksa rentang, akurasi, kondisi operasi, dan kebutuhan penggunaan produk sebelum pemesanan.</p>")
    return "".join(parts)


def build_output_row(data: Dict[str, Any], pub_date: datetime) -> Dict[str, Any]:
    brand = LUTRON_CONFIG.brand
    data["brand"] = brand
    family = detect_family(data)
    specs = data.get("spec_rows") or []
    ai = get_ai_content(data, family)

    # Dalam mode AI bersih, seluruh teks publik berasal dari hasil terjemahan/penulisan Indonesia.
    if ai:
        title = clean_text(ai.get("product_name_id")) or clean_catalog_source_text(data.get("title"))
        # Category from deterministic family rules is authoritative. AI may
        # suggest a label but never overrides the validated product class.
        category_leaf = clean_text(family.get("category")).strip()
        # DeepSeek has already selected the retained specification rows from
        # the raw Technical Specifications + Specifications source list.
        specs_out = ai.get("translated_spec_rows") or localize_spec_rows_fallback(specs)
        desc = ai.get("product_description_html")
        short = ai.get("short_description")
        features = ai.get("feature_tab_html")
        faq = ai.get("faq_html")
        meta = ai.get("meta_description")
        official = ai.get("official_information_html")
    else:
        title = clean_catalog_source_text(data.get("title"))
        category_leaf = clean_catalog_source_text(family.get("category")).replace("", "").strip()
        specs_out, _output_spec_issues = canonicalize_specifications(localize_spec_rows_fallback(specs))
        desc = build_product_description(data, family)
        short = build_short_description(data, family)
        features = "<h2>Keunggulan Produk</h2><ul>" + "".join(f"<li>{esc(x)}</li>" for x in feature_bullets(data, family)) + "</ul>"
        faq = build_faq(data, family)
        meta = build_meta_description(data, family)
        official = fallback_official_information(data)

    # Pastikan hasil akhir bebas karakter Chinese, bahkan bila ada respons AI yang tidak sesuai.
    if has_cjk(title) or any(has_cjk(value) for value in [desc, short, features, faq, meta, official, category_leaf]):
        raise ValueError("chinese_character_detected_in_output")

    if UJI_CTA not in desc:
        desc = desc.rstrip() + UJI_CTA
    category = f"{category_leaf}" if category_leaf else "Alat Ukur dan Instrumen"
    model = clean_text(data.get("model"))
    tags = dedupe_keep_order([brand, model, category.replace("", ""), "alat ukur Lutron"])

    corrections: List[str] = []
    if not data.get("description"):
        corrections.append("missing_description")
    if not specs:
        corrections.append("missing_specs")
    if not data.get("official_content_html"):
        corrections.append("missing_full_content")
    if not data.get("image_url"):
        corrections.append("missing_image")
    if data.get("datasheet_url"):
        corrections.append("datasheet_source=" + clean_text(data.get("datasheet_url")))
    if data.get("image_source"):
        corrections.append("image_source=" + clean_text(data.get("image_source")))
    if data.get("error"):
        corrections.append("source_note=" + clean_text(data.get("error")))
    if data.get("is_accessory"):
        corrections.append("accessory_included_" + clean_text(data.get("accessory_reason")))
    if clean_text(family.get("classification_basis")):
        corrections.append("category_basis=" + clean_text(family.get("classification_basis")))
    for flag in data.get("quality_flags") if isinstance(data.get("quality_flags"), list) else []:
        corrections.append("quality_flag=" + clean_text(flag))
    if AI_SETTINGS.get("enabled"):
        if ai:
            removed_count = int(ai.get("ai_spec_dedup_removed_count") or 0)
            corrections.append("deepseek_ai_indonesia_spec_dedup_removed=" + str(removed_count))
        else:
            corrections.append("deepseek_ai_fallback_" + clean_text(data.get("_ai_status", "unknown")))

    row = {col: "" for col in OUTPUT_COLS}
    row.update({
        "Name": title,
        "Brand": brand,
        "Product Description": desc,
        "Product Short Description": short,
        "custom_tab_1_title": "Spesifikasi",
        "custom_tab_1_content": specifications_tab_html(
            specs_out,
            already_deduplicated=bool(ai and ai.get("ai_spec_dedup_applied")),
        ),
        "custom_tab_1_priority": 10,
        "custom_tab_2_title": "Keunggulan Produk",
        "custom_tab_2_content": features,
        "custom_tab_2_priority": 20,
        "custom_tab_3_title": "Aplikasi Penggunaan",
        "custom_tab_3_content": build_application_tab({"title": title, "brand": brand, "is_accessory": data.get("is_accessory")}, family),
        "custom_tab_3_priority": 30,
        "custom_tab_4_title": "FAQ",
        "custom_tab_4_content": faq,
        "custom_tab_4_priority": 40,
        "custom_tab_5_title": "Informasi Produk",
        "custom_tab_5_content": official,
        "custom_tab_5_priority": 50,
        "product categories": category,
        "product tags": ", ".join(tags),
        "focus keyphrase": title[:120],
        "meta description": meta,
        "publication_date": pub_date.strftime("%m/%d/%Y %H:%M"),
        "Processed": "Yes",
        "Processing Time": "lutron_uji_import_deepseek_indonesia" if ai else "lutron_uji_import_rule_fallback",
        "Content Quality": f"{clean_text(data.get('source'))}_live_scrape_{'deepseek_indonesia' if ai else 'rule_fallback'}",
        "image_url": clean_text(data.get("image_url")),
        "Source URL": clean_text(data.get("url")),
        "Website Validation": clean_text(data.get("status")) or f"scraped_{clean_text(data.get('source'))}_live",
        "Website Correction Log": "; ".join(corrections) if corrections else "ok",
    })
    return row



REVIEW_SPEC_COLUMNS = [
    "URL", "Model", "Product", "Status", "Issue Code", "Severity", "Detail",
    "Canonical Key", "Display Label", "Source Label", "Value", "Source Section",
]


def prepare_canonical_specs_for_product(data: Dict[str, Any]) -> List[Dict[str, str]]:
    """Canonicalise from immutable raw rows and attach review issues to data."""
    raw = data.get("_raw_spec_rows")
    if not isinstance(raw, list):
        raw = data.get("spec_rows") if isinstance(data.get("spec_rows"), list) else []
        data["_raw_spec_rows"] = [dict(item) for item in raw if isinstance(item, dict)]
    canonical, issues = canonicalize_specifications(raw)
    data["spec_rows"] = canonical
    data["spec_review_issues"] = issues
    return issues


def build_spec_review_records(data: Dict[str, Any]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for issue in data.get("spec_review_issues") if isinstance(data.get("spec_review_issues"), list) else []:
        if not isinstance(issue, dict):
            continue
        row = {col: "" for col in REVIEW_SPEC_COLUMNS}
        row.update({
            "URL": clean_text(data.get("url")),
            "Model": clean_text(data.get("model")),
            "Product": clean_text(data.get("title")),
            "Status": clean_text(data.get("status")),
        })
        for col in REVIEW_SPEC_COLUMNS:
            if col in issue:
                row[col] = clean_text(issue.get(col))
        rows.append(row)
    return rows


def _append_spec_review_log(data: Dict[str, Any]) -> None:
    issues = data.get("spec_review_issues") if isinstance(data.get("spec_review_issues"), list) else []
    if not issues:
        return
    codes = dedupe_keep_order(clean_text(item.get("Issue Code")) for item in issues if isinstance(item, dict))
    if codes:
        data["error"] = _merge_source_notes(data.get("error"), "spec_review=" + ",".join(codes))


# ---------------------------------------------------------------------------
# Cache / output / CLI
# ---------------------------------------------------------------------------

def load_cache(path: str) -> Dict[str, Any]:
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_cache(path: str, data: Dict[str, Any]) -> None:
    if not path:
        return
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def style_output_xlsx(path: str, sheet_name: str = "UJI Products") -> None:
    if load_workbook is None:
        return
    try:
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill, cell.font, cell.border = header_fill, header_font, border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = {"A": 34, "B": 16, "C": 62, "D": 44, "E": 22, "F": 60, "H": 24, "I": 60,
                  "K": 22, "L": 60, "N": 22, "O": 60, "Q": 28, "R": 70, "AC": 34, "AD": 44,
                  "AE": 30, "AF": 44, "AG": 20, "AH": 16, "AI": 32, "AJ": 30, "AK": 44, "AL": 44,
                  "AM": 32, "AN": 54}
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(get_column_letter(i), 18)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
            ws.row_dimensions[row[0].row].height = 84
        ws.row_dimensions[1].height = 34
        if ws.max_row >= 2 and not ws.tables:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tab = Table(displayName="LUTRON_UJI_Products", ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(tab)
        wb.save(path)
    except Exception:
        pass


def write_outputs(
    rows: Sequence[Dict[str, Any]],
    skipped: Sequence[Dict[str, Any]],
    reviews: Sequence[Dict[str, Any]],
    output_path: str,
    skipped_output: str = "",
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=OUTPUT_COLS).to_excel(writer, index=False, sheet_name="UJI Products")
        if skipped:
            pd.DataFrame(skipped).to_excel(writer, index=False, sheet_name="Skipped")
        if reviews:
            pd.DataFrame(reviews, columns=REVIEW_SPEC_COLUMNS).to_excel(writer, index=False, sheet_name="Review Spesifikasi")
    style_output_xlsx(output_path)
    if skipped_output:
        pd.DataFrame(skipped).to_excel(skipped_output, index=False)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Scrape katalog produk LUTRON ke format WooCommerce nama-website-kalian dengan penghapusan spesifikasi duplikat")
    p.add_argument("--input", required=True, help="Excel berisi URL produk")
    p.add_argument("--output", default="uji_catalog_import_ready.xlsx", help="Output Excel WooCommerce")
    p.add_argument("--sheet", default=0, help="Nama atau indeks sheet input")
    p.add_argument("--url-column", default="", help="Nama kolom URL jika ingin ditentukan manual")
    p.add_argument("--cache", default="lutron_scrape_cache_v7.json", help="Cache JSON hasil scrape")
    p.add_argument("--timeout", type=int, default=25, help="Timeout tiap halaman dalam detik")
    p.add_argument("--delay", type=float, default=0.6, help="Jeda antar request dalam detik")
    p.add_argument("--max-rows", type=int, default=0, help="Batasi URL. 0 = semua")
    p.add_argument("--start-date", default="08/02/2025 10:00", help="Format MM/DD/YYYY HH:MM")
    p.add_argument("--pub-interval-hours", type=int, default=1, help="Jarak jam publication date")
    p.add_argument("--skipped-output", default="", help="Opsional output log produk diskip")
    p.add_argument("--no-cache", action="store_true", help="Abaikan cache dan scrape ulang")
    p.add_argument("--no-render-js", action="store_true", help="Matikan browser Playwright. Default: browser dipakai untuk menangkap gambar, spesifikasi, dan isi Lutron yang dimuat JavaScript.")
    p.add_argument("--no-datasheet-fallback", action="store_true", help="Matikan fallback datasheet model ketika halaman Lutron hanya mengembalikan shell/konten minim.")
    p.add_argument("--allow-external-image-fallback", action="store_true", help="Opsional dan tidak disarankan: izinkan fallback gambar dari mesin pencari. Default hanya memakai gambar resmi yang dimuat halaman Lutron.")
    p.add_argument("--datasheet-attempt-limit", type=int, default=20, help="Maksimum URL datasheet yang dicek per produk sebelum dinyatakan tidak ditemukan.")
    p.add_argument("--strict-source", action="store_true", help="Masukkan hanya produk yang memiliki data substansial dari halaman Lutron atau datasheet model yang terverifikasi. Default aktif secara efektif untuk mencegah output kosong.")
    p.add_argument("--strict-publish", action="store_true", help="Lewati produk yang gambar produknya tidak terverifikasi, gambar duplikat lintas model, atau memiliki flag kualitas sumber. Direkomendasikan untuk batch siap publish.")
    p.add_argument("--allow-spec-review-publish", action="store_true", help="Kompatibilitas v16. Tidak diperlukan karena spesifikasi tidak pernah menahan produk.")
    p.add_argument("--hold-all-spec-reviews", action="store_true", help="Kompatibilitas v16. Diabaikan karena v16 hanya menghapus duplikasi spesifikasi.")
    p.add_argument("--category-from-description", action="store_true", help="Kompatibilitas: V10 sudah selalu menentukan kategori dari model, judul input, judul resmi, dan deskripsi resmi. Spesifikasi tidak dipakai untuk kategori.")
    p.add_argument("--skip-accessories", action="store_true", help="Lewati aksesori/komponen yang sudah terkonfirmasi. Default: tetap dimasukkan")
    p.add_argument("--use-ai", action="store_true", help="Aktifkan generator konten DeepSeek")
    p.add_argument("--deepseek-api-key", default=os.getenv("DEEPSEEK_API_KEY", ""))
    p.add_argument("--deepseek-base-url", default=os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com"))
    p.add_argument("--deepseek-model", default=os.getenv("DEEPSEEK_MODEL", "deepseek-chat"))
    p.add_argument("--ai-cache", default="", help="Cache JSON hasil AI")
    p.add_argument("--ai-limit", type=int, default=0, help="Batas request AI baru. 0 = semua")
    p.add_argument("--ai-timeout", type=int, default=60)
    p.add_argument("--ai-delay", type=float, default=0.0)
    p.add_argument("--ai-temperature", type=float, default=0.15)
    p.add_argument("--ai-spec-limit", type=int, default=120, help="Maksimum baris spesifikasi yang diterjemahkan AI per produk. Default: 120.")
    p.add_argument(
        "--allow-source-language-fallback",
        action="store_true",
        help="Izinkan output fallback tanpa terjemahan AI. Default saat --use-ai: produk yang gagal diterjemahkan masuk sheet Skipped agar output WordPress tetap bersih berbahasa Indonesia.",
    )
    p.add_argument("--catalog-audit-only", action="store_true", help="Audit seluruh input Excel tanpa scrape web. Membuat laporan klasifikasi model, kategori, dan status aksesori untuk semua produk.")
    p.add_argument("--catalog-audit-output", default="", help="Lokasi file Excel audit katalog. Digunakan bersama --catalog-audit-only.")
    return p.parse_args()


def coerce_sheet(value: Any) -> Any:
    return int(value) if isinstance(value, str) and re.fullmatch(r"\d+", value) else value


def main() -> None:
    args = parse_args()
    records = read_input_records(args.input, coerce_sheet(args.sheet), args.url_column)
    if args.max_rows and args.max_rows > 0:
        records = records[:args.max_rows]
    if not records:
        raise SystemExit("Tidak ada URL yang dapat diproses.")

    if getattr(args, "catalog_audit_only", False):
        audit_path = clean_text(getattr(args, "catalog_audit_output", "")) or str(Path(args.input).with_name("lutron_catalog_preflight_audit.xlsx"))
        write_catalog_preflight(records, audit_path)
        print(f"Preflight katalog selesai. Produk diaudit: {len(records)}")
        print(f"Output audit: {audit_path}")
        return

    cache = {} if args.no_cache else load_cache(args.cache)
    out = Path(args.output)
    ai_cache_path = args.ai_cache or str(out.with_suffix(".deepseek_ai_cache.json"))
    set_ai_settings(
        enabled=bool(args.use_ai),
        api_key=args.deepseek_api_key,
        base_url=args.deepseek_base_url,
        model=args.deepseek_model,
        cache={} if args.no_cache else load_cache(ai_cache_path),
        cache_path=ai_cache_path,
        limit=args.ai_limit,
        timeout=args.ai_timeout,
        delay=args.ai_delay,
        temperature=args.ai_temperature,
        spec_limit=args.ai_spec_limit,
        strict_indonesian=bool(args.use_ai and not args.allow_source_language_fallback),
    )
    try:
        start_date = datetime.strptime(args.start_date, "%m/%d/%Y %H:%M")
    except Exception:
        start_date = datetime(2025, 8, 2, 10, 0)

    session = requests.Session()
    if PdfReader is None and not args.no_datasheet_fallback:
        print("PERINGATAN: pypdf belum tersedia. Fallback datasheet dimatikan. Instal: py -m pip install pypdf")
    if not args.no_datasheet_fallback and (fitz is None or not _find_tesseract_executable()):
        print("INFO: OCR datasheet scan tidak aktif. PDF teks tetap bisa dibaca. Untuk PDF scan, instal PyMuPDF dan Tesseract OCR.")
    renderer: Optional[LutronBrowserRenderer] = None if args.no_render_js else LutronBrowserRenderer(max(args.timeout, 25))
    if renderer is not None and not renderer.available():
        print("PERINGATAN: Playwright belum tersedia. Instal: py -m pip install playwright lalu py -m playwright install chromium")
        renderer = None

    accepted: List[Dict[str, Any]] = []
    rows: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    review_specs: List[Dict[str, Any]] = []
    included_accessories = 0

    for idx, record in enumerate(records, 1):
        url = record["url"]
        print(f"[{idx}/{len(records)}] scrape {url}")
        if not args.no_cache and url in cache:
            data = cache[url]
            for key in ("input_brand", "input_title", "input_model"):
                if not clean_text(data.get(key)):
                    data[key] = record.get(key, "")
            if not data.get("brand"):
                data["brand"] = LUTRON_CONFIG.brand
            if not data.get("title"):
                data["title"] = record.get("input_title", "")
            if not data.get("model"):
                data["model"] = record.get("input_model", "")
            data["is_accessory"], data["accessory_reason"] = is_accessory_product(
                clean_text(data.get("title")), clean_text(data.get("description")),
                data.get("spec_rows") if isinstance(data.get("spec_rows"), list) else [],
            )
            data["quality_flags"] = collect_source_quality_flags(data)
            print("    cache hit")
        else:
            data = scrape_product_with_fallbacks(
                record,
                args.timeout,
                session,
                renderer,
                enable_datasheet=not args.no_datasheet_fallback,
                enable_image_search=bool(args.allow_external_image_fallback),
                datasheet_attempt_limit=max(1, int(args.datasheet_attempt_limit or 20)),
            )
            cache[url] = data
            if args.delay > 0:
                time.sleep(args.delay)

        # Canonical pipeline runs on every product, including cache hits.
        # It combines Technical Specifications and Specifications before AI.
        # v16 hanya menghapus baris spesifikasi yang memang duplikat.
        # Tidak membuat review, tidak memperbaiki label, dan tidak menahan produk.
        prepare_canonical_specs_for_product(data)
        spec_issues: List[Dict[str, str]] = []

        image_state = "OK" if clean_text(data.get("image_url")) else "KOSONG"
        flags = ", ".join(data.get("quality_flags") or [])
        print(
            f"    status={clean_text(data.get('status')) or '-'} | "
            f"model={clean_text(data.get('model')) or '-'} | image={image_state}" +
            (f" | flags={flags}" if flags else "")
        )
        if image_state == "KOSONG":
            print(f"    image_reason={clean_text(data.get('error')) or 'source image tidak ditemukan'}")

        if data.get("is_accessory"):
            if args.skip_accessories:
                skipped.append({"url": url, "title": data.get("title", ""), "status": data.get("status", ""), "reason": data.get("accessory_reason", ""), "error": data.get("error", "")})
                print("    skip confirmed accessory")
                continue
            included_accessories += 1
            print("    include confirmed accessory")

        if not clean_text(data.get("title")):
            skipped.append({"url": url, "title": "", "status": data.get("status", ""), "reason": "missing_title_or_scrape_failed", "error": data.get("error", "")})
            print(f"    skip failed: {data.get('error') or data.get('status')}")
            continue
        if not _is_real_product_data(data):
            skipped.append({
                "url": url,
                "title": data.get("title", ""),
                "status": data.get("status", ""),
                "reason": "incomplete_source_data",
                "error": data.get("error", ""),
            })
            print(f"    skip incomplete source: {data.get('error') or data.get('status')}")
            continue

        # v16: spesifikasi tidak pernah menjadi alasan produk ditahan.
        # Script hanya membuang baris yang memang sama dan meneruskan sisanya.

        if AI_SETTINGS.get("enabled") and AI_SETTINGS.get("strict_indonesian"):
            ai_fields = get_ai_content(data, detect_family(data))
            if not ai_fields:
                # AI is optional content enrichment, not a publication gate.
                # Keep the source product and use the deterministic fallback so
                # a transient API error cannot empty the complete output file.
                data["_ai_status"] = clean_text(data.get("_ai_status")) or "ai_translation_failed_rule_fallback_used"
                data["error"] = _merge_source_notes(
                    data.get("error"),
                    "ai_translation_failed_rule_fallback_used",
                )
                print(f"    AI fallback: {data.get('_ai_status')}")

        # Strict publish is intentionally conservative: questionable source data
        # is sent to Skipped instead of being silently published.
        if args.strict_publish and data.get("quality_flags"):
            skipped.append({
                "url": url,
                "title": data.get("title", ""),
                "status": data.get("status", ""),
                "reason": "source_quality_review_required",
                "error": _merge_source_notes(data.get("error"), "; ".join(data.get("quality_flags") or [])),
            })
            print("    skip strict publish: source quality flag")
            continue
        accepted.append(data)

    # Batch-level image validation detects generic images reused across many
    # models. This cannot be done safely while each row is emitted immediately.
    accepted = apply_batch_image_quality_gate(accepted, skipped, require_verified_image=bool(args.strict_publish))

    for data in accepted:
        pub_date = start_date + timedelta(hours=len(rows) * args.pub_interval_hours)
        try:
            rows.append(build_output_row(data, pub_date))
        except ValueError as exc:
            skipped.append({
                "url": data.get("url", ""),
                "title": data.get("title", ""),
                "status": data.get("status", ""),
                "reason": "output_language_validation_failed",
                "error": str(exc),
            })
            print(f"    skip validasi output: {exc}")

    write_outputs(rows, skipped, review_specs, args.output, args.skipped_output)
    save_cache(args.cache, cache)
    if renderer is not None:
        renderer.close()
    if AI_SETTINGS.get("enabled"):
        save_cache(AI_SETTINGS.get("cache_path", ai_cache_path), AI_SETTINGS.get("cache", {}))
    print(f"Selesai. Produk masuk output: {len(rows)} | Diskip: {len(skipped)} | Review spesifikasi: {len(review_specs)} | Accessories masuk: {included_accessories}")
    print(f"Output: {args.output}")
    if AI_SETTINGS.get("enabled"):
        print(f"deepseek_ai_used={AI_STATS['used']} cache_hit={AI_STATS['cache_hit']} fallback={AI_STATS['fallback']} skipped={AI_STATS['skipped']}")



# ---------------------------------------------------------------------------
# V10 category routing: title + official description only
# ---------------------------------------------------------------------------
# V8 fixed the specific false positives in the first test batch, but the
# complete source workbook contains probes, electrodes, cases, software,
# current clamps, buffer solutions, interface cables, and other item families.
# These V9 rules classify all 559 catalog input rows before a live scrape and
# are intentionally conservative: an uncertain component is marked for review
# rather than being described as a main instrument.

ACCESSORY_MODEL_PREFIX_RE_V9 = re.compile(
    r"^(?:TP-|PE-|OXPB-|EP-|FRTP-|CDPB-|LN-|TL-|TLSM-|SMDC-|UPCB-|USB-|GMCB-|EXCB-|"
    r"CA-(?:0[3-8]|52A|203)|AP-|CP-|LS-|KV-|WG-|WH-|WP-|WT-|OXHD-|OXEL-|EA-|PB-|NN-|NB-|AT-|"
    r"ALLIGATOR|PL-|PI-|PWCB-|FS-|IP-|PX-|MP-|MS-71P$|VB-8[34]$|EMF-824$|"
    r"ORP-(?:14|15|400)$|PH-(?:04|04A|07|07A)$|ST-50\b|SB-01\b|CD-14\b|USBP-)",
    re.I,
)

ACCESSORY_TITLE_RE_V9 = re.compile(
    r"\b(?:replacement|spare(?:\s+part)?|accessor(?:y|ies)|optional\s+accessor|"
    r"carrying\s+case|protective\s+cover|extension\s+cable|interface\s+cable|usb\s+cable|rs\s*232\s+cable|"
    r"test\s+lead|alligator\s+clip|test\s+clip|test\s+probe|thermocouple\s+adapter|"
    r"calibration\s+(?:weight|solution)|buffer\s+solution|probe[-\s]?filling\s+electrolyte|"
    r"penetr?ometer\s+tips?|electrode|conductivity\s+solution|standard\s+solution|\bprobe\b|sensor|ear\s*phone|wedge\s+grip|"
    r"wall\s+holder|water\s+resistant\s+bag|current\s+shunt|power\s+adapter|power\s+converter|"
    r"power\s+interface\s+cable|adapter)\b",
    re.I,
)

# Product group rules cover the full supplied Lutron workbook, not only the
# initial eleven products.  The first matching rule wins.  Rules are built on
# model and title because source breadcrumbs are often generic or absent.
V9_FAMILY_RULES: List[Tuple[re.Pattern[str], Dict[str, Any]]] = [
    (re.compile(r"\b(?:software|data\s+acquisition|data\s+transmission|wifi\s+converter|rs\s*232\s+converter|i/o\s+controller\s+to\s+wifi)\b|\b(?:SW-|RSW-|RSC-|RHW-|IOW-)", re.I), {
        "category": "Software dan Interface Instrumentasi", "term": "perangkat lunak atau antarmuka instrumentasi",
        "function": "mendukung akuisisi data, komunikasi, atau integrasi instrumen sesuai spesifikasi produk",
        "tags": ["Lutron", "software instrumentasi", "interface instrumentasi"],
    }),
    (re.compile(r"\b(?:pressure\s+transmitter|humidity\s+transmitter|wind\s+(?:speed|direction)\s+transmitter|ph\s+transmitter|light\s+transmitter|sound\s+transmitter|vibration\s+transmitter|temperature\s+transmitter|dissolved\s+oxygen\s+transmitter|pressure\s+transducer|pressure\s+sensor)\b|\b(?:TR-|TRDO-|TRLX-|TRPH-|TRRH-|TRSL-|TRVB-|PS93|PS100)", re.I), {
        "category": "Transmitter dan Sensor Industri", "term": "transmitter atau sensor industri",
        "function": "mengirimkan atau menyediakan sinyal pengukuran untuk pemantauan, kontrol, atau otomasi",
        "tags": ["Lutron", "transmitter industri", "sensor industri"],
    }),
    (re.compile(r"\b(?:pressure\s+meter|manometer|vacuum\s+meter|pressure\s+data\s+recorder)\b|\b(?:PM-91|PS-93|MPS-)", re.I), {
        "category": "Manometer dan Pressure Meter", "term": "manometer atau pressure meter",
        "function": "mengukur, memantau, atau merekam tekanan sesuai spesifikasi produk",
        "tags": ["Lutron", "manometer", "pressure meter"],
    }),
    (re.compile(r"\b(?:air\s+quality|co2\s*meter|co\s*meter|o2\s*meter|pm\s*2\.5|negative\s+ion)\b|\b(?:AQ-|CO2-|COH-|GC-|GCH-|GCO-|MCH-|PM-10|O2H-|PO2-|PCO-)", re.I), {
        "category": "Air Quality dan Gas Meter", "term": "air quality atau gas meter",
        "function": "mengukur parameter kualitas udara atau gas sesuai spesifikasi produk",
        "tags": ["Lutron", "air quality meter", "gas meter"],
    }),
    (re.compile(r"\b(?:conductivity|tds|salt\s+meter|dissolved\s+oxygen|\bdo\s+meter|\bph\s*(?:meter|tester|monitor)|orp\s+meter|chlorine\s+meter|turbidity|pure\s+water|water\s+hardness|water\s+quality|ion\s+water)\b|\b(?:BCT-|BDO-|BPH-|BWA-|CD-|CL-|DO-|NI-|ORP-(?:2|4)|PCD-|PDO-|PH-(?:2|3)|PSA-|PWA-|PWH-|TU-|WA-|WAC-|YK-(?:21PH|22DOA|22CTA|23RP|30WA|31SA|200[15]))", re.I), {
        "category": "Water Quality Meter", "term": "alat ukur kualitas air",
        "function": "mengukur parameter kualitas air atau larutan sesuai spesifikasi produk",
        "tags": ["Lutron", "water quality meter", "pH meter", "conductivity meter"],
    }),
    (re.compile(r"\b(?:temperature|thermometer|thermocouple|infrared\s+thermometer)\b|\b(?:TM[-_]|PTM-|BTM-|MTM-|PST-|TC-|FT-967)", re.I), {
        "category": "Thermometer dan Temperature Meter", "term": "alat ukur suhu",
        "function": "mengukur suhu sampel, lingkungan, atau proses sesuai spesifikasi produk",
        "tags": ["Lutron", "thermometer", "temperature meter"],
    }),
    (re.compile(r"\b(?:humidity|hygrometer|wet\s+bulb|dew\s+point|barometer)\b|\b(?:HR-|HT-|MHB-|MHT-|PHB-|PHT-)", re.I), {
        "category": "Humidity dan Environmental Meter", "term": "humidity atau environmental meter",
        "function": "mengukur kelembapan dan parameter lingkungan terkait sesuai spesifikasi produk",
        "tags": ["Lutron", "humidity meter", "hygrometer", "environment meter"],
    }),
    (re.compile(r"\b(?:anemometer|air\s+flow|air\s+velocity|wind\s+speed|pitot\s+tube)\b|\b(?:AM-|ABH-|AH-|PAM-|MY-|YK-2004|YK-80A|LM-81A|SP-82A)", re.I), {
        "category": "Anemometer dan Air Flow Meter", "term": "anemometer atau air flow meter",
        "function": "mengukur kecepatan udara, aliran udara, atau parameter lingkungan terkait sesuai spesifikasi produk",
        "tags": ["Lutron", "anemometer", "air flow meter"],
    }),
    (re.compile(r"\b(?:light\s+meter|lux\s+meter|uv\s+light|solar\s+power)\b|\b(?:LX-|UVA?-|UVC-|SPM-|YK-(?:10LX|35UV|37UV)|SP-82LX|SP-82UV|LM-81LX)", re.I), {
        "category": "Light Meter dan Lux Meter", "term": "light meter atau lux meter",
        "function": "mengukur intensitas cahaya atau parameter radiasi terkait sesuai spesifikasi produk",
        "tags": ["Lutron", "light meter", "lux meter", "uv meter"],
    }),
    (re.compile(r"\b(?:sound\s+level|noise\s+dosimeter|sound\s+calibrator|decibel)\b|\b(?:SL-|SC-|DS-)", re.I), {
        "category": "Sound Level Meter dan Kalibrator", "term": "sound level meter atau kalibrator suara",
        "function": "mengukur atau mengkalibrasi tingkat kebisingan dan tekanan suara sesuai spesifikasi produk",
        "tags": ["Lutron", "sound level meter", "sound calibrator"],
    }),
    (re.compile(r"\b(?:electromagnetic\s+field|emf\s+tester|rf\s+field|gauss\s+meter|magnetic\s+meter|micro\s+wave\s+leakage|electrostatic\s+field)\b|\b(?:EMF-|ESF-|GU-|MG-|PMG-|BMG-)", re.I), {
        "category": "Pengukur Medan Elektromagnetik", "term": "pengukur medan elektromagnetik",
        "function": "mengukur medan elektromagnetik sesuai spesifikasi produk",
        "tags": ["Lutron", "EMF tester", "gauss meter"],
    }),
    (re.compile(r"\b(?:power\s+analyzer|power\s+meter|watt\s*meter|energy\s+meter|power\s+quality|three\s+phase|fork\s+current|clamp\s+power|apparent\s+power|power\s+factor)\b|\b(?:DW-|CWF-|PC-|PVA-|PPF-|FT-9950)", re.I), {
        "category": "Power Analyzer dan Electrical Tester", "term": "power analyzer atau electrical tester",
        "function": "menganalisis parameter daya, energi, dan kelistrikan sesuai spesifikasi produk",
        "tags": ["Lutron", "power analyzer", "power meter"],
    }),
    (re.compile(r"\b(?:clamp\s+meter|multimeter|digital\s+multimeter|voltmeter|ammeter|bench\s+meter|leakage\s+tester|insulation\s+tester|rccb|rotation\s+tester)\b|\b(?:CM-|CMF-|DM-|AA-|AV-|DV-|DL-6054|DI-|RCB-|RT-)", re.I), {
        "category": "Multimeter dan Electrical Meter", "term": "multimeter atau electrical meter",
        "function": "mengukur atau memeriksa parameter kelistrikan sesuai spesifikasi produk",
        "tags": ["Lutron", "multimeter", "electrical meter"],
    }),
    (re.compile(r"\b(?:frequency\s+counter|frequency\s+meter|rf\s+detector|line\s+frequency)\b|\b(?:FC-|LF-)", re.I), {
        "category": "Frequency Counter", "term": "frequency counter",
        "function": "mengukur frekuensi dan parameter sinyal elektronik sesuai spesifikasi produk",
        "tags": ["Lutron", "frequency counter"],
    }),
    (re.compile(r"\b(?:vibration|stroboscope|tachometer|rpm)\b|\b(?:VB-|BVB-|MVB-|PVB-|VT-|DT-|PDT-)", re.I), {
        "category": "Vibration dan Tachometer", "term": "vibration meter atau tachometer",
        "function": "mengukur getaran, putaran, atau parameter mekanis terkait sesuai spesifikasi produk",
        "tags": ["Lutron", "vibration meter", "tachometer"],
    }),
    (re.compile(r"\b(?:force\s+gauge|fruit\s+hardness|torque|material\s+tester)\b|\b(?:FG-|FR-|TQ-)", re.I), {
        "category": "Force Gauge dan Material Tester", "term": "alat uji gaya atau material",
        "function": "mendukung pengukuran gaya, kekerasan, torsi, atau parameter material sesuai spesifikasi produk",
        "tags": ["Lutron", "force gauge", "material tester"],
    }),
    (re.compile(r"\b(?:digital\s+balance|digital\s+scale|calibration\s+weight)\b|\b(?:GM-|WT-)", re.I), {
        "category": "Timbangan Digital dan Balance", "term": "timbangan digital atau balance",
        "function": "mengukur massa atau berat sesuai spesifikasi produk",
        "tags": ["Lutron", "digital balance", "digital scale"],
    }),
    (re.compile(r"\b(?:moisture\s+meter|wood\s+moisture|soil\s+moisture|humidity\s+content)\b|\b(?:MS-(?!71P)|PMS-)", re.I), {
        "category": "Moisture Meter", "term": "moisture meter",
        "function": "mengukur kadar kelembapan material atau tanah sesuai spesifikasi produk",
        "tags": ["Lutron", "moisture meter"],
    }),
    (re.compile(r"\b(?:calibrator|decade\s+box|lcr\s+meter|milliohm|micro[-\s]?ohm|smd\s+tester|photo\s+interrupter)\b|\b(?:CC-|CBOX-|LBOX-|RBOX-|LCR-|MO-|SMDA-)", re.I), {
        "category": "Kalibrator dan Component Tester", "term": "kalibrator atau component tester",
        "function": "mendukung kalibrasi atau pengujian komponen elektronik sesuai spesifikasi produk",
        "tags": ["Lutron", "calibrator", "component tester"],
    }),
    (re.compile(r"\b(?:panel\s+meters?|process\s+indicator)\b|\bDR-", re.I), {
        "category": "Panel Meter dan Process Indicator", "term": "panel meter atau process indicator",
        "function": "menampilkan atau memantau parameter proses listrik dan industri sesuai spesifikasi produk",
        "tags": ["Lutron", "panel meter", "process indicator"],
    }),
    (re.compile(r"\b(?:ultrasonic\s+leakage\s+detector|ultrasonic\s+transmitter)\b|\bGS-", re.I), {
        "category": "Ultrasonic Leak Detector", "term": "ultrasonic leak detector",
        "function": "mendukung deteksi kebocoran atau pengujian ultrasonik sesuai spesifikasi produk",
        "tags": ["Lutron", "ultrasonic leak detector"],
    }),
    (re.compile(r"\b(?:current\s+recorder|voltage\s+recorder|sd\s*card\s+data\s+logger)\b|\b(?:MMA-|MMV-|DL-9602)", re.I), {
        "category": "Data Logger dan Recorder", "term": "data logger atau recorder",
        "function": "merekam parameter pengukuran sesuai spesifikasi produk",
        "tags": ["Lutron", "data logger", "recorder"],
    }),
    (re.compile(r"\b(?:controller|monitor|indicator|alarm)\b|\b(?:CT-|PAA-|PAV-|PDA-|PDV-|PIR-|PPH-|PPS-|PTM-9957|PVA-|PVB-8219)", re.I), {
        "category": "Industrial Controller dan Monitor", "term": "controller atau monitor industri",
        "function": "memantau atau mengendalikan parameter proses sesuai spesifikasi produk",
        "tags": ["Lutron", "industrial controller", "process monitor"],
    }),
    (re.compile(r"\b(?:environment\s+meter|weather\s+meter|wbgt|heat\s+index|4\s+in\s+1\s+meter|5\s+in\s+1\s+meter)\b|\b(?:EM-|EMC-|LM-|SP-|WBGT-)", re.I), {
        "category": "Environmental dan Weather Meter", "term": "environmental atau weather meter",
        "function": "mengukur parameter lingkungan sesuai spesifikasi produk",
        "tags": ["Lutron", "environment meter", "weather meter"],
    }),
]


def _v9_model_from_product(title: str, specs: Sequence[Dict[str, str]]) -> str:
    for row in specs:
        if re.search(r"\b(model|sku|cat|part|item|mpn)\b", clean_text(row.get("label")), re.I):
            value = clean_text(row.get("value"))
            if value:
                return value
    # Input data normally includes title and the actual model is passed in specs.
    return ""


def is_accessory_product(title: str, description: str, specs: Sequence[Dict[str, str]]) -> Tuple[bool, str]:
    """V9 catalog-wide accessory classification with model-first overrides."""
    title_clean = clean_text(title)
    desc_clean = clean_text(description)
    model = _v9_model_from_product(title_clean, specs)
    composite = f"{model} {title_clean}".strip()

    # Model-specific rules are strongest because titles can mention a meter or
    # tester only to describe the instrument the component connects to.
    if ACCESSORY_MODEL_PREFIX_RE_V9.search(composite):
        return True, "accessory_model_or_consumable_pattern"

    # Independent instruments retain main-product status even when supplied
    # with a probe or sensor.
    if PRIMARY_MODEL_PREFIX_RE.search(composite):
        return False, "main_instrument_model_pattern"

    functional_labels = sum(
        1 for row in specs
        if re.search(r"\b(range|measurement|accuracy|resolution|display|power|battery|output|interface|data\s*logger|datalogger|channel|temperature|humidity|pressure|frequency|weight|dimension)\b", clean_text(row.get("label")), re.I)
    )
    if functional_labels >= 2:
        return False, "main_instrument_spec_signature"

    if ACCESSORY_TITLE_RE_V9.search(title_clean):
        # A title that begins or ends as a recognised autonomous instrument is
        # not downgraded merely because it has a separate probe.
        if re.search(r"\b(?:meter|tester|analyzer|counter|recorder|logger|transmitter|controller|monitor|calibrator|balance|scale|detector|thermometer)\b|\b(?:co2|co|o2)\s*meter\b|\bco2meter\b", title_clean, re.I):
            return False, "main_instrument_title_pattern"
        return True, "accessory_title_pattern"

    # Incomplete source pages are left as main/unknown, never guessed as an
    # accessory. Strict publish will gate missing source data separately.
    return False, "not_confirmed_as_accessory"


def _v10_category_description(value: Any, max_chars: int = 1600) -> str:
    """Return only the product-description text used for category routing.

    Technical tables, full product HTML, optional accessories, and specification
    rows are deliberately excluded. They often contain words such as USB,
    RS-232, software, interface, probe, or temperature that describe a feature
    or accessory rather than the product's primary category.
    """
    text = compact(clean_catalog_source_text(value), max_chars)
    if not text:
        return ""
    # Some Lutron pages concatenate the summary with a technical section.
    # Keep the description portion and stop before common specification labels.
    stop = re.search(
        r"\b(?:specifications?|technical\s+data|technical\s+specification|"
        r"standard\s+accessories|optional\s+accessories|accessories\s+included)\b",
        text,
        flags=re.I,
    )
    return clean_text(text[:stop.start()] if stop else text)


def _v10_match_rule(text: str, rules: Sequence[Tuple[re.Pattern[str], Dict[str, Any]]]) -> Optional[Dict[str, Any]]:
    """Return the first catalog rule matched by a controlled category text."""
    for pattern, rule in rules:
        if pattern.search(text):
            return dict(rule)
    return None


def _v10_source_category_family(source_category: str, title: str) -> Optional[Dict[str, Any]]:
    normalized = normalize_lutron_category(source_category, title)
    if not normalized:
        return None
    family = dict(DEFAULT_FAMILY)
    family["category"] = normalized
    family["classification_basis"] = "source_category_fallback"
    return family


def detect_family(data: Dict[str, Any]) -> Dict[str, Any]:
    """Classify from product identity and official description, never specs.

    Priority:
    1. Confirmed accessory/consumable state.
    2. Model + input title + official product title.
    3. Official product description only, when title cannot identify the type.
    4. Official source category as a fallback.
    5. Generic category plus an audit flag.

    `spec_rows` and `official_content_html` are intentionally ignored. This
    prevents an EMF meter, pH meter, or humidity meter from becoming a software,
    thermometer, or interface product merely because its specifications mention
    optional USB/RS-232/software/accessories.
    """
    input_title = clean_catalog_source_text(data.get("input_title"))
    official_title = clean_catalog_source_text(data.get("title"))
    model = clean_text(data.get("model"))

    if data.get("is_accessory"):
        family = dict(ACCESSORY_FAMILY)
        family["classification_basis"] = "confirmed_accessory_title_model"
        return family

    title_model_text = clean_text(" ".join(x for x in (model, input_title, official_title) if x))
    description_text = _v10_category_description(data.get("description"))

    # Software/interface is deliberately only allowed from model or title.
    # A complete instrument can mention optional software in its description.
    software_rule = V9_FAMILY_RULES[0]
    if software_rule[0].search(title_model_text):
        family = dict(software_rule[1])
        family["classification_basis"] = "model_title"
        return family

    # The remaining rules represent physical product families. First use
    # model/title. This is the strongest evidence and prevents broad words in
    # prose such as temperature, monitor, or alarm from overriding the product.
    non_software_rules = V9_FAMILY_RULES[1:]
    family = _v10_match_rule(title_model_text, non_software_rules)
    if family:
        family["classification_basis"] = "model_title"
        return family

    # Description becomes the classifier only when title/model is insufficient.
    # Generic controller/monitor text is excluded here because it appears in
    # many ordinary descriptions and is too broad to be trusted alone.
    description_rules = [
        rule for rule in non_software_rules
        if clean_text(rule[1].get("category")) != "Industrial Controller dan Monitor"
    ]
    family = _v10_match_rule(description_text, description_rules)
    if family:
        family["classification_basis"] = "official_description"
        return family

    source_family = _v10_source_category_family(data.get("source_category", ""), official_title or input_title)
    if source_family:
        return source_family

    family = dict(DEFAULT_FAMILY)
    family["classification_basis"] = "generic_review_required"
    return family

def _v9_preflight_review_flag(record: Dict[str, str], accessory: bool, family: Dict[str, Any]) -> str:
    title = clean_text(record.get("input_title"))
    model = clean_text(record.get("input_model"))
    # Explicitly flag component-looking items that have no matching model rule
    # so they can be reviewed once, not silently published as instruments.
    component_words = re.compile(r"\b(probe|sensor|electrode|cable|adapter|clip|lead|case|cover|solution|electrolyte|tips?|holder|bag|weight|shunt|plug|grip|ear\s*phone)\b", re.I)
    if not accessory and component_words.search(title) and not PRIMARY_MODEL_PREFIX_RE.search(model):
        # Terms such as "separate probe" often describe a complete meter.
        # Only route to review when the title itself lacks clear main-product evidence.
        if not re.search(r"\b(?:meter|tester|analyzer|counter|recorder|logger|transmitter|controller|monitor|calibrator|balance|scale|detector|thermometer)\b|\b(?:co2|co|o2)\s*meter\b|\bco2meter\b", title, re.I):
            return "manual_product_type_review"
    if family.get("category") == DEFAULT_FAMILY["category"]:
        return "category_rule_review"
    return ""


def write_catalog_preflight(records: Sequence[Dict[str, str]], output_path: str) -> None:
    """Audit every input record without network calls and save a review-ready Excel file."""
    audit_rows: List[Dict[str, Any]] = []
    for idx, record in enumerate(records, 1):
        model = clean_text(record.get("input_model"))
        title = clean_text(record.get("input_title"))
        specs = [{"label": "Model", "value": model, "section": "Input Excel"}] if model else []
        is_accessory, reason = is_accessory_product(title, "", specs)
        family = detect_family({
            "input_title": title, "title": title, "model": model, "description": "", "source_category": "",
            "spec_rows": specs, "is_accessory": is_accessory,
        })
        review = _v9_preflight_review_flag(record, is_accessory, family)
        audit_rows.append({
            "No": idx,
            "Model": model,
            "Input Title": title,
            "URL": clean_text(record.get("url")),
            "Product Type": "Aksesori/Komponen" if is_accessory else "Produk Utama",
            "Classification Reason": reason,
            "Suggested Category": clean_text(family.get("category")),
            "Suggested Term": clean_text(family.get("term")),
            "Category Basis": clean_text(family.get("classification_basis")),
            "Preflight Status": "Review" if review else "Ready for Live Scrape",
            "Review Reason": review,
        })
    df = pd.DataFrame(audit_rows)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Catalog Preflight")
        summary = (
            df.groupby(["Product Type", "Suggested Category", "Preflight Status"], dropna=False)
            .size().reset_index(name="Count").sort_values(["Product Type", "Suggested Category"])
        )
        summary.to_excel(writer, index=False, sheet_name="Summary")
    style_output_xlsx(output_path, "Catalog Preflight")

if __name__ == "__main__":
    main()
