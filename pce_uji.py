#!/usr/bin/env python3
"""PCE Instruments -> uji.co.id WooCommerce catalog generator.

V3_PARAM_CLEAN: nama produk resmi dipertahankan apa adanya; pembersihan bahasa difokuskan pada parameter, highlight teknis, kategori, tag, dan narasi publik.

Versi ini dibuat untuk file PCE, bukan AMTAST. Fokus utamanya:
- membaca struktur kolom PCE seperti link, Judul, Tags, Image, Short_Description, Brand, Order No;
- membuat Product Description HTML dengan gaya katalog uji.co.id;
- mengurangi pola kalimat yang terlalu template/AI;
- menangani produk utama dan aksesori/komponen secara berbeda;
- menjaga agar klaim teknis tetap berasal dari data website PCE atau Excel.
- memprioritaskan deskripsi asli dari link website PCE ketika tersedia.

Contoh:
python pce_uji_import_ready_v2.py --input "pce-import.xlsx" --final-output "uji_pce_import_ready.xlsx"
"""
from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import time
from urllib.parse import urljoin
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

try:
    import requests
except Exception:  # website scraping is optional and has a fallback
    requests = None

try:
    from bs4 import BeautifulSoup
except Exception:  # HTML parsing will fall back to regex/text extraction
    BeautifulSoup = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:  # styling is optional
    load_workbook = None

OUTPUT_COLS = [
    "Name", "Brand", "Product Description", "Product Short Description",
    "custom_tab_1_title", "custom_tab_1_content", "custom_tab_1_priority",
    "custom_tab_2_title", "custom_tab_2_content", "custom_tab_2_priority",
    "custom_tab_3_title", "custom_tab_3_content", "custom_tab_3_priority",
    "custom_tab_4_title", "custom_tab_4_content", "custom_tab_4_priority",
    "custom_tab_5_title", "custom_tab_5_content", "custom_tab_5_priority",
    "custom_tab_6_title", "custom_tab_6_content", "custom_tab_6_priority",
    "custom_tab_7_title", "custom_tab_7_content", "custom_tab_7_priority",
    "custom_tab_8_title", "custom_tab_8_content", "custom_tab_8_priority",
    "product categories", "product tags", "focus keyphrase", "meta description",
    "publication_date", "Processed", "Processing Time", "Content Quality",
    "image_url", "Source URL", "Website Validation", "Website Correction Log",
]

UJI_CTA = (
    '<p>Untuk konsultasi produk, silakan kunjungi '
    '<strong><a href="https://uji.co.id/about-us-3/" target="_blank" rel="noopener">uji.co.id</a></strong> '
    'atau <strong><a href="https://uji.co.id/contact-us/" target="_blank" rel="noopener">hubungi kami</a></strong>. '
    'Telepon: +62896-2784-2222.</p>'
)

BAD_EMPTY = {"", "nan", "none", "null", "empty", "-"}

PCE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
}

SCRAPE_STOP_HEADINGS = {
    "information", "help", "company", "social media", "similar products", "accessories",
    "related products", "technical hotline", "product presentation", "downloads", "manuals",
}

NOISY_TEXT_PATTERNS = [
    r"use of cookies", r"only necessary cookies", r"also allow comfort cookies",
    r"all cookies allowed", r"price \(exclude vat\)", r"price \(include vat\)",
    r"add to cart", r"question / callback", r"delivery time", r"technical hotline",
    r"would you like a product presentation", r"send us an email", r"copyright",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = html.unescape(str(value))
    s = s.replace("\xa0", " ").replace("\u3000", " ").replace("_x000D_", "\n")
    s = s.replace("\r", "\n")
    s = re.sub(r"[ \t\f\v]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def is_empty(value: Any) -> bool:
    return clean_text(value).casefold() in BAD_EMPTY


def esc(value: Any) -> str:
    return html.escape(clean_text(value), quote=False)


def strip_jual(title: str) -> str:
    title = clean_text(title)
    title = re.sub(r"^\s*jual\s+", "", title, flags=re.I).strip()
    title = re.sub(r"\s+", " ", title)
    return title


def normalize_brand(brand: str) -> str:
    b = clean_text(brand)
    if not b or re.search(r"\bpce\b", b, flags=re.I):
        return "PCE Instruments"
    return b


def pick(row: Dict[str, Any], *names: str) -> str:
    lower_map = {str(k).strip().casefold(): k for k in row.keys()}
    for name in names:
        key = lower_map.get(name.strip().casefold())
        if key is not None and not is_empty(row.get(key)):
            return clean_text(row.get(key))
    return ""


def deterministic_index(text: str, modulo: int) -> int:
    if modulo <= 1:
        return 0
    digest = hashlib.md5(clean_text(text).encode("utf-8")).hexdigest()
    return int(digest[:8], 16) % modulo


def clean_slug_word(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean_text(s).lower()).strip()


def compact_source_text(text: str, max_chars: int = 900) -> str:
    text = clean_text(text)
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= max_chars:
        return text
    return text[:max_chars].rsplit(" ", 1)[0].rstrip(" ,.;") + "."


def is_noise_text(text: str) -> bool:
    t = clean_text(text).lower()
    if not t:
        return True
    if len(t) < 8:
        return True
    return any(re.search(pat, t, flags=re.I) for pat in NOISY_TEXT_PATTERNS)


def normalize_url(url: str) -> str:
    url = clean_text(url)
    if not url:
        return ""
    if url.startswith("//"):
        return "https:" + url
    if not re.match(r"https?://", url, flags=re.I):
        return "https://" + url.lstrip("/")
    return url


def load_scrape_cache(path: str) -> Dict[str, Any]:
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_scrape_cache(path: str, cache: Dict[str, Any]) -> None:
    if not path:
        return
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    try:
        p.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def soup_text(node: Any) -> str:
    if node is None:
        return ""
    try:
        text = node.get_text(" ", strip=True)
    except Exception:
        text = str(node)
    return clean_text(re.sub(r"\s+", " ", text))


def find_meta_content(soup: Any, *selectors: str) -> str:
    if soup is None:
        return ""
    for sel in selectors:
        try:
            node = soup.select_one(sel)
        except Exception:
            node = None
        if node:
            val = clean_text(node.get("content") or node.get("src") or node.get("href") or "")
            if val and not is_noise_text(val):
                return val
    return ""


def extract_best_image_url(soup: Any, base_url: str) -> str:
    """Extract product image with lazy-load fallbacks and basic logo/icon filtering."""
    if soup is None:
        return ""
    candidates: List[str] = []
    for sel in [
        'meta[property="og:image"]', 'meta[name="twitter:image"]', 'link[rel="image_src"]',
        'img[itemprop="image"]', '.product img', '.product-detail img', '.product-info img',
        'img[data-src]', 'img[data-original]', 'img[data-lazy]', 'img[src]'
    ]:
        try:
            nodes = soup.select(sel)
        except Exception:
            nodes = []
        for node in nodes:
            raw = clean_text(
                node.get("content") or node.get("src") or node.get("data-src") or
                node.get("data-original") or node.get("data-lazy") or node.get("href") or ""
            )
            if not raw:
                srcset = clean_text(node.get("srcset") or node.get("data-srcset") or "")
                if srcset:
                    raw = clean_text(srcset.split(",")[0].strip().split(" ")[0])
            if raw:
                candidates.append(urljoin(base_url, raw))

    bad = re.compile(r"(?:logo|icon|sprite|flag|payment|cookie|blank|placeholder|loading|transparent|tracking)", re.I)
    good = re.compile(r"(?:product|produkte|prod|images?|media|artikel|shop|cache|upload)", re.I)
    unique: List[str] = []
    seen = set()
    for url in candidates:
        if not url or url in seen or bad.search(url):
            continue
        seen.add(url)
        unique.append(url)
    if not unique:
        return ""
    # Prefer likely product/media URLs, then fall back to first non-logo image.
    for url in unique:
        if good.search(url):
            return url
    return unique[0]


def extract_text_after_heading(soup: Any, heading_text: str, max_chars: int = 1400) -> str:
    if soup is None:
        return ""
    heading = None
    for tag in soup.find_all(re.compile(r"^h[1-6]$")):
        txt = soup_text(tag).strip().lower()
        if txt == heading_text.lower():
            heading = tag
            break
    if heading is None:
        # Some pages use tabs/buttons/anchors instead of headings.
        for tag in soup.find_all(True):
            txt = soup_text(tag).strip().lower()
            if txt == heading_text.lower():
                heading = tag
                break
    if heading is None:
        return ""

    parts: List[str] = []
    for node in heading.next_elements:
        if node is heading:
            continue
        name = getattr(node, "name", None)
        if not name:
            continue
        if name in {"script", "style", "noscript", "svg", "img", "form", "button"}:
            continue
        if re.fullmatch(r"h[1-6]", str(name), flags=re.I):
            txt = soup_text(node).strip().lower()
            if txt and txt != heading_text.lower():
                if txt in SCRAPE_STOP_HEADINGS or len(parts) > 0:
                    break
        if name not in {"p", "li", "td", "th", "div", "span", "section", "article"}:
            continue
        txt = soup_text(node)
        if is_noise_text(txt):
            continue
        # Avoid collecting the full body repeatedly from wrapper divs.
        if len(txt) > 600 and name in {"div", "section", "article"}:
            continue
        if txt.lower() == heading_text.lower():
            continue
        if txt not in parts:
            parts.append(txt)
        joined = " ".join(parts)
        if len(joined) >= max_chars:
            break
    return compact_source_text(" ".join(parts), max_chars=max_chars)



def is_probable_spec_label(text: str) -> bool:
    label = clean_text(text).strip().lower()
    if not label or len(label) > 80:
        return False
    return bool(re.search(
        r"\b(measurement|measuring|range|accuracy|resolution|display|dimensions?|weight|power|battery|temperature|humidity|interface|sensor|probe|electrode|memory|data|input|output|supply|life|unit|wood|types?|pressure|capacity|readability|repeatability|linearity|operating|storage|protection|material|diameter)\b",
        label,
        flags=re.I,
    ))


SPEC_LABEL_TRANSLATIONS: List[Tuple[str, str]] = [
    (r"^measurement range$|^measuring range$", "Rentang pengukuran"),
    (r"^measurement range\s+(.+)$|^measuring range\s+(.+)$", r"Rentang pengukuran \1"),
    (r"^accuracy$", "Akurasi"),
    (r"^resolution$", "Resolusi"),
    (r"^number of different wood types$", "Jumlah jenis kayu"),
    (r"^wood temperature range$", "Rentang suhu kayu"),
    (r"^temperature range$", "Rentang suhu"),
    (r"^operating temperature$|^operating temperature range$", "Suhu operasional"),
    (r"^storage temperature$|^storage temperature range$", "Suhu penyimpanan"),
    (r"^display$", "Tampilan"),
    (r"^electrode dimensions?$", "Dimensi elektroda"),
    (r"^probe dimensions?$", "Dimensi sonda"),
    (r"^sensor dimensions?$", "Dimensi sensor"),
    (r"^power supply$", "Catu daya"),
    (r"^battery life$", "Daya tahan baterai"),
    (r"^unit dimensions?$|^device dimensions?$|^dimensions?$", "Dimensi unit"),
    (r"^unit weight$|^weight$", "Berat unit"),
    (r"^interface$|^interfaces$", "Antarmuka"),
    (r"^data interface$", "Antarmuka data"),
    (r"^memory$|^data memory$", "Memori data"),
    (r"^input$|^inputs$", "Masukan"),
    (r"^output$|^outputs$", "Keluaran"),
    (r"^signal output$", "Keluaran sinyal"),
    (r"^protection class$|^protection rating$", "Rating perlindungan"),
    (r"^material$", "Material"),
    (r"^capacity$|^weighing range$", "Rentang timbang"),
    (r"^readability$", "Keterbacaan"),
    (r"^repeatability$", "Repeatabilitas"),
    (r"^linearity$", "Linearitas"),
]


def translate_spec_label(label: str) -> str:
    raw = clean_text(label).strip(" :-–—")
    low = re.sub(r"\s+", " ", raw.lower())
    for pattern, repl in SPEC_LABEL_TRANSLATIONS:
        if re.search(pattern, low, flags=re.I):
            out = re.sub(pattern, repl, low, flags=re.I).strip()
            return out[:1].upper() + out[1:]
    out = raw
    replacements = [
        (r"\bmeasurement\b|\bmeasuring\b", "pengukuran"),
        (r"\brange\b", "rentang"),
        (r"\btemperature\b", "suhu"),
        (r"\bhumidity\b", "kelembapan"),
        (r"\baccuracy\b", "akurasi"),
        (r"\bresolution\b", "resolusi"),
        (r"\bdisplay\b", "tampilan"),
        (r"\bdimensions?\b", "dimensi"),
        (r"\bweight\b", "berat"),
        (r"\bpower supply\b", "catu daya"),
        (r"\bbattery\b", "baterai"),
        (r"\bwood\b", "kayu"),
        (r"\btypes?\b", "jenis"),
        (r"\belectrode\b", "elektroda"),
        (r"\bprobe\b", "sonda"),
        (r"\bunit\b", "unit"),
        (r"\binterface\b", "antarmuka"),
        (r"\binput\b", "masukan"),
        (r"\boutput\b", "keluaran"),
        (r"\bmemory\b", "memori"),
    ]
    for pat, rep in replacements:
        out = re.sub(pat, rep, out, flags=re.I)
    out = re.sub(r"\s+", " ", out).strip(" :-–—")
    return out[:1].upper() + out[1:] if out else raw


def translate_spec_value(value: str, label: str = "") -> str:
    s = clean_text(value).strip()
    if not s:
        return ""
    replacements = [
        (r"\bApprox\.?|\bapprox\.?|\bapproximately\b", "Sekitar"),
        (r"\bmoisture content\b", "kadar air"),
        (r"\bhumidity measurement\b", "pengukuran kelembapan"),
        (r"\btemperature measurement\b", "pengukuran suhu"),
        (r"\bwithin\b", "pada rentang"),
        (r"\brange\b", "rentang"),
        (r"\bwood types\b", "jenis kayu"),
        (r"\bwood\b", "kayu"),
        (r"\bmeasurements\b", "pengukuran"),
        (r"\bmeasurement\b", "pengukuran"),
        (r"\bbattery powered\b", "menggunakan baterai"),
        (r"\bbattery\b", "baterai"),
        (r"\bbatteries\b", "baterai"),
        (r"\bdiameter\b", "diameter"),
        (r"\bdisplay\b", "layar"),
        (r"\bprobe\b", "sonda"),
        (r"\belectrode\b", "elektroda"),
        (r"\binterface\b", "antarmuka"),
        (r"\boutput\b", "keluaran"),
        (r"\binput\b", "masukan"),
        (r"\bfrom\b", "dari"),
        (r"\bto\b", "hingga"),
        (r"\blbs\b", "lb"),
    ]
    for pat, rep in replacements:
        s = re.sub(pat, rep, s, flags=re.I)
    s = re.sub(r"(\d)\s*x\s*(\d)", r"\1 × \2", s, flags=re.I)
    s = re.sub(r"(%|°C|°F|ppm|ppb|bar|mbar|Pa|psi|dB|lux|rpm|Hz)\s+rentang\b", r"\1", s, flags=re.I)
    s = re.sub(r"\b1\s*×\s*([0-9]+(?:[.,][0-9]+)?\s*V\s*[0-9A-Za-z]*)\s*baterai\b", r"1 × baterai \1", s, flags=re.I)
    s = re.sub(r"\b(\d+(?:[,.]\d+)?)\s*V\b", r"\1 V", s, flags=re.I)
    s = re.sub(r"\b(\d+(?:[,.]\d+)?)\s*A\b", r"\1 A", s, flags=re.I)
    s = re.sub(r"(?<=\d)\.(?=\d)", ",", s)
    s = re.sub(r"\b(\d{5,})\b", lambda m: f"{int(m.group(1)):,}".replace(",", "."), s)
    s = re.sub(r"\s*/\s*", " / ", s)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    if s:
        s = s[:1].upper() + s[1:]
    return s


def dedupe_spec_rows(rows: List[Dict[str, str]], max_rows: int = 40) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    seen = set()
    for row in rows:
        label = clean_text(row.get("label", ""))
        value = clean_text(row.get("value", ""))
        if not label or not value:
            continue
        key = (label.lower(), value.lower())
        if key in seen:
            continue
        seen.add(key)
        out.append({"label": label, "value": value})
        if len(out) >= max_rows:
            break
    return out


def extract_spec_table_from_soup(soup: Any) -> List[Dict[str, str]]:
    """Extract structured specification rows from the official product page.

    Berbeda dari highlight, fungsi ini mengambil tabel spesifikasi asli seperti
    Measurement range, Accuracy, Resolution, Display, dan seterusnya.
    """
    if soup is None:
        return []

    headings: List[Any] = []
    heading_patterns = re.compile(r"^(specification|specifications|technical data|technical specifications|technical details|product specifications)$", re.I)
    for tag in soup.find_all(True):
        txt = soup_text(tag).strip()
        if not txt or len(txt) > 80:
            continue
        if heading_patterns.match(txt):
            headings.append(tag)

    search_roots: List[Any] = []
    for heading in headings:
        parent = heading.parent
        if parent:
            search_roots.append(parent)
        for sib in heading.find_all_next(True, limit=12):
            if getattr(sib, "name", "") in {"table", "tbody", "dl", "div", "section", "article"}:
                search_roots.append(sib)
    # Fallback: scan tables globally and keep those that look like specification tables.
    try:
        search_roots.extend(soup.find_all("table"))
    except Exception:
        pass

    rows: List[Dict[str, str]] = []
    checked = set()
    for root in search_roots:
        if root is None or id(root) in checked:
            continue
        checked.add(id(root))

        # Normal table rows.
        try:
            trs = root.find_all("tr") if getattr(root, "name", None) != "tr" else [root]
        except Exception:
            trs = []
        for tr in trs:
            cells = tr.find_all(["th", "td"])
            if len(cells) < 2:
                continue
            label_raw = soup_text(cells[0])
            value_raw = clean_text(" ".join(soup_text(c) for c in cells[1:]))
            if not is_probable_spec_label(label_raw):
                continue
            label = translate_spec_label(label_raw)
            value = translate_spec_value(value_raw, label)
            if label and value:
                rows.append({"label": label, "value": value})

        # Definition lists.
        try:
            dts = root.find_all("dt")
        except Exception:
            dts = []
        for dt in dts:
            dd = dt.find_next_sibling("dd")
            if not dd:
                continue
            label_raw = soup_text(dt)
            value_raw = soup_text(dd)
            if not is_probable_spec_label(label_raw):
                continue
            label = translate_spec_label(label_raw)
            value = translate_spec_value(value_raw, label)
            if label and value:
                rows.append({"label": label, "value": value})

    return dedupe_spec_rows(rows)


def spec_table_html(rows: List[Dict[str, str]]) -> str:
    clean_rows = dedupe_spec_rows(rows)
    if not clean_rows:
        return ""
    body = []
    for row in clean_rows:
        label = esc(row.get("label", ""))
        value = esc(row.get("value", ""))
        if label and value:
            body.append(f"<tr><th>{label}</th><td>{value}</td></tr>")
    if not body:
        return ""
    return '<table class="uji-spec-table"><tbody>' + ''.join(body) + '</tbody></table>'

def extract_description_from_plain_text(text: str, max_chars: int = 1000) -> str:
    text = clean_text(text)
    if not text:
        return ""
    # Best fallback for simple PCE pages: capture text between Description and footer headings.
    m = re.search(r"\bDescription\b\s*(.+?)(?:\bInformation\b|\bHelp\b|\bCompany\b|\bSocial Media\b|Copyright|$)", text, flags=re.I | re.S)
    if m:
        candidate = clean_text(m.group(1))
        lines = [clean_text(x) for x in re.split(r"\n+|\s{2,}", candidate) if not is_noise_text(x)]
        return compact_source_text(" ".join(lines), max_chars=max_chars)
    return ""


def scrape_pce_product_page(url: str, timeout: int = 15) -> Dict[str, str]:
    """Fetch a PCE product page and extract source fields used by the catalog generator.

    The function is intentionally conservative: when a field cannot be read confidently,
    it returns an empty string and lets the Excel/template fallback handle the row.
    """
    url = normalize_url(url)
    result: Dict[str, str] = {
        "url": url,
        "status": "not_requested" if not url else "failed",
        "title": "",
        "description": "",
        "specs": "",
        "spec_table": [],
        "image_url": "",
        "order_no": "",
        "gtin": "",
        "error": "",
    }
    if not url:
        result["status"] = "no_url"
        return result
    if requests is None:
        result["error"] = "requests_not_installed"
        return result
    try:
        response = requests.get(url, headers=PCE_HEADERS, timeout=timeout)
        response.raise_for_status()
    except Exception as exc:
        result["error"] = f"request_error:{type(exc).__name__}"
        return result

    html_text = response.text or ""
    if not html_text.strip():
        result["error"] = "empty_html"
        return result

    if BeautifulSoup is None:
        result["description"] = extract_description_from_plain_text(html_text)
        result["status"] = "ok" if result["description"] else "no_description_found"
        return result

    soup = BeautifulSoup(html_text, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    h1 = soup.find("h1")
    title = soup_text(h1)
    if title.lower() in {"description", "information"} or not title:
        title = find_meta_content(soup, 'meta[property="og:title"]', 'meta[name="twitter:title"]')
    if not title and soup.title:
        title = soup_text(soup.title)
    result["title"] = clean_text(re.sub(r"\s*[-|].*$", "", title))

    desc = extract_text_after_heading(soup, "Description", max_chars=1200)
    if not desc:
        desc = find_meta_content(soup, 'meta[name="description"]', 'meta[property="og:description"]', 'meta[name="twitter:description"]')
    if not desc:
        desc = extract_description_from_plain_text(soup_text(soup), max_chars=1200)
    result["description"] = compact_source_text(desc, max_chars=1200)

    spec_table = extract_spec_table_from_soup(soup)
    result["spec_table"] = spec_table
    specs = extract_text_after_heading(soup, "Technical Data", max_chars=2200)
    if not specs:
        specs = extract_text_after_heading(soup, "Specifications", max_chars=2200)
    if not specs:
        specs = extract_text_after_heading(soup, "Specification", max_chars=2200)
    result["specs"] = compact_source_text(specs, max_chars=2200)

    result["image_url"] = extract_best_image_url(soup, url)

    page_text = soup_text(soup)
    m_order = re.search(r"Order\s*no\.?\s*:?\s*([A-Za-z0-9._/-]+)", page_text, flags=re.I)
    if m_order:
        result["order_no"] = clean_text(m_order.group(1))
    m_gtin = re.search(r"GTIN\s*\(?EAN\)?\s*:?\s*([0-9]{6,14})", page_text, flags=re.I)
    if m_gtin:
        result["gtin"] = clean_text(m_gtin.group(1))

    result["status"] = "ok" if result["description"] else "no_description_found"
    return result


def get_scraped_product_data(url: str, cache: Optional[Dict[str, Any]], timeout: int = 15, delay: float = 0.0) -> Dict[str, str]:
    url = normalize_url(url)
    if not url:
        return {"status": "no_url"}
    if cache is not None and url in cache:
        cached = dict(cache.get(url) or {})
        cached["from_cache"] = "yes"
        return cached
    if delay > 0:
        time.sleep(delay)
    data = scrape_pce_product_page(url, timeout=timeout)
    if cache is not None:
        cache[url] = data
    return data


def detect_accessory_reason(name: str, tags: str = "", category: str = "", description: str = "") -> Tuple[bool, str]:
    """Classify product role conservatively.

    Prinsip utama:
    - jangan memakai deskripsi website sebagai pemicu utama aksesori, karena halaman alat utama
      sering menyebut kata sensor/probe/cable di bagian fitur;
    - nama produk dengan kata Meter/Detector/Thermometer/Analyzer/dll diprioritaskan sebagai
      alat utama kecuali judulnya jelas berupa certificate, cable, case, stand, probe, dan sejenisnya;
    - kata "sensor" tidak otomatis aksesori jika nama produk jelas menunjukkan alat utama.
    """
    title = clean_slug_word(name)
    meta = clean_slug_word(" ".join([tags, category]))
    blob = clean_slug_word(" ".join([name, tags, category]))

    if not title:
        return False, "empty_name_default_main"

    def has(pattern: str, target: str = "title") -> bool:
        source = title if target == "title" else blob
        return re.search(pattern, source, flags=re.I) is not None

    # Frasa ini hampir selalu bukan unit utama katalog. Sengaja dicek lebih dulu.
    hard_accessory_patterns = [
        (r"^(?:iso )?calibration certificate\b", "calibration_certificate"),
        (r"^certificate\b", "certificate"),
        (r"\bcalibration kit\b", "calibration_kit"),
        (r"\busb kit\b|\bkit\b.*\bfor\s+pce", "kit_accessory"),
        (r"\bclamping table\b", "clamping_table"),
        (r"\btest bladder\b|\bgas ducts?\b", "test_bladder_accessory"),
        (r"\bcurrent clamp\b(?!\s+meter)", "current_clamp_accessory"),
        (r"\bcarrying case\b|\btransport case\b|\bprotective case\b|\bcase for\b", "case"),
        (r"\bsoftware\b|\blicen[cs]e\b", "software_license"),
        (r"\b(?:usb|interface|data|connection|connecting|measuring|sensor|extension|power) cable\b", "cable"),
        (r"\btest leads?\b", "test_lead"),
        (r"\b(?:power supply|charger|battery pack|battery charger)\b", "power_accessory"),
        (r"\b(?:replacement|spare)\b", "replacement_spare"),
        (r"\b(?:adapter|adaptor|holder|bracket|mounting kit|mounting set|tripod)\b", "mount_adapter_holder"),
        (r"\b(?:spindle|rotor|paddle|beaker|measuring system|measurement system)\b", "viscometer_accessory"),
        (r"\b(?:filter|paper roll|printer paper)\b", "consumable"),
    ]
    for pattern, reason in hard_accessory_patterns:
        if has(pattern):
            return True, reason

    # Pola utama: kalau judul jelas menyebut instrumen, jangan dikalahkan oleh kata sensor/probe di tag/deskripsi.
    main_patterns = [
        (r"\b(?:meter|detector|thermometer|controller|regulator|scale|balance|analy[sz]er|tester|logger|gauge|camera|borescope|endoscope|viscometer|refractometer|tachometer|stroboscope|microscope|centrifuge|counter|monitor|transmitter|manometer|barometer|calibrator)\b", "main_instrument_keyword"),
        (r"\b(?:gasman|tcgard|gasmaster|gasflag)\b", "pce_gas_unit_family"),
        (r"\bpce[- ]?[a-z]{1,6}[- ]?\d", "pce_model_in_title"),
    ]
    for pattern, reason in main_patterns:
        if has(pattern):
            # Pengecualian: judul model PCE bisa muncul pada aksesori seperti probe/cable for PCE-xxx.
            if has(r"\b(?:probe|sensor|electrode|transducer|module|cable|adapter|adaptor|case|holder|stand|tripod|clamp)\b\s+(?:for|to)\b"):
                return True, "accessory_noun_for_main_unit"
            if has(r"\bfor\s+(?:pce|gasman|tcgard|gasmaster|pce[- ])") and has(r"\b(?:probe|sensor|electrode|transducer|module|cable|adapter|adaptor|case|holder|stand|tripod|clamp)\b"):
                return True, "accessory_for_named_unit"
            return False, reason

    # Soft accessory: dipakai setelah cek alat utama.
    soft_accessory_patterns = [
        (r"\b(?:probe|electrode|transducer)\b", "probe_electrode_transducer"),
        (r"\bsensor\b", "sensor_component"),
        (r"\bmodule\b", "module_component"),
        (r"\b(?:cable|adapter|adaptor|case|holder|stand|tripod|clamp)\b", "generic_accessory_noun"),
        (r"\bfor\s+(?:pce|gasman|tcgard|gasmaster|rm|pce[- ])", "for_compatible_unit"),
        (r"\baccessor(?:y|ies)\b", "accessory_tag"),
    ]
    for pattern, reason in soft_accessory_patterns:
        if has(pattern):
            return True, reason

    # Jika kategori/tag eksplisit accessories, gunakan sebagai sinyal pendukung, bukan deskripsi.
    if re.search(r"\baccessor(?:y|ies)|spare parts?\b", meta, flags=re.I):
        return True, "accessory_category_or_tag"

    return False, "default_main_product"


def detect_accessory(name: str, tags: str = "", category: str = "", description: str = "") -> bool:
    return detect_accessory_reason(name, tags, category, description)[0]


FAMILY_CONFIGS: List[Dict[str, Any]] = [
    {
        "id": "salt_solution",
        "match": ["salt meter", "salinity", "brix", "refractometer", "sugar meter"],
        "category": "Salt Meter / Refractometer",
        "term": "alat ukur konsentrasi larutan",
        "function": "memeriksa kadar garam, salinitas, Brix, atau konsentrasi larutan pada sampel cair",
        "object": "sampel cair, air proses, makanan-minuman, atau larutan produksi",
        "criteria": "parameter ukur, rentang salinitas/Brix, akurasi, resolusi, kompensasi suhu, dan kebutuhan kalibrasi",
        "features": [
            "Memeriksa konsentrasi larutan sesuai parameter yang tersedia pada produk.",
            "Praktis untuk pengecekan sampel cair pada proses produksi, laboratorium, atau inspeksi lapangan.",
            "Memberikan acuan pembacaan digital sehingga evaluasi sampel lebih mudah dibanding pembacaan manual.",
        ],
        "apps": [
            "pengukuran salinitas, Brix, atau konsentrasi larutan",
            "analisis sampel cair di laboratorium",
            "kontrol kualitas makanan, minuman, atau air proses",
            "pemeriksaan cairan pada produksi dan inspeksi lapangan",
        ],
        "tags": ["salt meter", "salinity meter", "brix meter", "refractometer", "alat ukur salinitas"],
    },
    {
        "id": "gas_air",
        "match": ["gas detector", "gasman", "gasflag", "gasmaster", "tcgard", "air quality", "co2", "h2s", "oxygen", "ozone", "voc", "combustible gas"],
        "category": "Gas Detector & Air Quality Meter",
        "term": "alat deteksi gas dan kualitas udara",
        "function": "memantau konsentrasi gas atau kondisi kualitas udara sesuai sensor dan konfigurasi produk",
        "object": "area kerja, ruang proses, fasilitas produksi, atau lokasi dengan potensi paparan gas",
        "criteria": "jenis gas, rentang deteksi, tipe sensor, sistem alarm, metode pemasangan, sertifikasi area, dan kebutuhan pencatatan data",
        "features": [
            "Mengidentifikasi gas target sesuai sensor atau konfigurasi yang tersedia.",
            "Cocok untuk pemantauan keselamatan kerja, area industri, dan inspeksi teknis.",
            "Dapat menjadi acuan ketika pengguna perlu memilih perangkat portabel atau sistem deteksi tetap.",
        ],
        "apps": [
            "deteksi gas berbahaya pada area kerja",
            "pemantauan keselamatan di fasilitas industri",
            "kontrol kualitas udara atau ventilasi ruangan",
            "pemasangan sistem monitoring gas secara tetap maupun portabel",
        ],
        "tags": ["gas detector", "air quality meter", "CO2 detector", "H2S detector", "alat deteksi gas"],
    },
    {
        "id": "electrical",
        "match": ["multimeter", "clamp meter", "voltage", "current", "power analyzer", "power meter", "insulation tester", "installation tester", "earth tester", "electrical tester", "phase", "socket tester"],
        "category": "Electrical Tester",
        "term": "alat uji dan ukur kelistrikan",
        "function": "mengecek parameter kelistrikan seperti tegangan, arus, daya, kontinuitas, atau kondisi instalasi sesuai spesifikasi produk",
        "object": "panel listrik, instalasi, mesin, kabel, atau peralatan elektronik",
        "criteria": "parameter listrik yang diuji, rentang pengukuran, kategori keamanan, akurasi, jenis sonda/tang, dan kompatibilitas aksesori",
        "features": [
            "Digunakan untuk pengecekan parameter kelistrikan pada instalasi atau peralatan teknis.",
            "Memudahkan teknisi memperoleh data dasar sebelum analisis lanjutan atau perbaikan.",
            "Pilihan produk dapat disesuaikan dengan kebutuhan pengukuran arus, tegangan, daya, atau inspeksi instalasi.",
        ],
        "apps": [
            "pengujian instalasi listrik",
            "troubleshooting panel dan peralatan listrik",
            "pemantauan arus, tegangan, daya, atau kontinuitas",
            "inspeksi teknis pada pemeliharaan gedung atau industri",
        ],
        "tags": ["electrical tester", "multimeter", "clamp meter", "power analyzer", "alat ukur listrik"],
    },
    {
        "id": "process_control",
        "match": ["process controller", "temperature controller", "pid controller", "regulator", "controller pce", "control systems"],
        "category": "Process Controller & Regulator",
        "term": "alat kontrol proses",
        "function": "mengatur atau memantau parameter proses seperti suhu atau sinyal kontrol sesuai konfigurasi produk",
        "object": "panel kontrol, mesin proses, sensor, aktuator, atau sistem produksi",
        "criteria": "jenis input sensor, mode kontrol, keluaran kontrol, rentang kerja, catu daya, dimensi panel, dan kompatibilitas instalasi",
        "features": [
            "Mendukung pengaturan parameter proses sesuai konfigurasi produk.",
            "Cocok untuk panel kontrol, mesin produksi, dan aplikasi otomasi sederhana.",
            "Pemilihan produk perlu disesuaikan dengan jenis input sensor dan kebutuhan keluaran kontrol.",
        ],
        "apps": ["kontrol suhu proses", "integrasi panel kontrol", "pengaturan mesin produksi", "aplikasi otomasi dan monitoring proses"],
        "tags": ["process controller", "temperature controller", "PID controller", "regulator", "alat kontrol proses"],
    },
    {
        "id": "temperature",
        "match": ["thermometer", "temperature", "thermal camera", "infrared", "thermocouple", "temperature logger"],
        "category": "Temperature Meter",
        "term": "alat ukur suhu",
        "function": "mengukur suhu objek, sampel, permukaan, atau lingkungan kerja sesuai metode pengukuran yang tersedia",
        "object": "proses produksi, ruang penyimpanan, sampel laboratorium, atau titik pemeriksaan lapangan",
        "criteria": "rentang suhu, akurasi, tipe sensor/sonda, waktu respons, resolusi, dan kondisi lingkungan pengukuran",
        "features": [
            "Mengukur suhu untuk pemeriksaan proses, lingkungan, atau sampel teknis.",
            "Cocok untuk pemantauan rutin di laboratorium, produksi, dan inspeksi lapangan.",
            "Dapat dipilih berdasarkan kebutuhan kontak, non-kontak, pencatatan data, atau inspeksi termal.",
        ],
        "apps": ["pemantauan suhu proses", "kontrol kualitas produksi", "pemeriksaan suhu lingkungan", "inspeksi titik panas pada objek atau peralatan"],
        "tags": ["temperature meter", "thermometer", "thermal camera", "infrared thermometer", "alat ukur suhu"],
    },
    {
        "id": "moisture_humidity",
        "match": ["moisture", "humidity", "hygrometer", "dew point", "wood", "grain"],
        "category": "Moisture & Humidity Meter",
        "term": "alat ukur kelembapan dan kadar air",
        "function": "mengecek kelembapan udara atau kadar air material sesuai jenis sampel dan metode pengukuran produk",
        "object": "bahan baku, material bangunan, ruang penyimpanan, komoditas, atau lingkungan kerja",
        "criteria": "jenis material, rentang kelembapan, akurasi, metode sensor, kompensasi suhu, dan kebutuhan pencatatan data",
        "features": [
            "Mengevaluasi kadar air atau kelembapan pada material dan lingkungan kerja.",
            "Cocok untuk kontrol kualitas bahan, penyimpanan, dan inspeksi teknis.",
            "Membantu pengguna memilih metode pengujian yang sesuai dengan jenis sampel.",
        ],
        "apps": ["pengukuran kadar air material", "pemantauan kelembapan ruang", "kontrol kualitas bahan baku", "inspeksi penyimpanan atau proses produksi"],
        "tags": ["moisture meter", "humidity meter", "hygrometer", "alat ukur kelembapan", "kadar air"],
    },
    {
        "id": "water_quality",
        "match": ["ph meter", "pH", "conductivity", "tds", "turbidity", "dissolved oxygen", "orp", "redox", "water quality", "chlorine", "fluoride"],
        "category": "Water Quality Meter",
        "term": "alat analisis kualitas air",
        "function": "menguji parameter air atau sampel cair seperti pH, konduktivitas, TDS, ORP, turbiditas, dissolved oxygen, atau parameter lain sesuai produk",
        "object": "air proses, air limbah, air minum, sampel laboratorium, atau cairan produksi",
        "criteria": "parameter ukur, rentang kerja, akurasi, resolusi, tipe elektroda/sensor, kompensasi suhu, dan metode kalibrasi",
        "features": [
            "Menganalisis parameter kualitas air sesuai kemampuan produk.",
            "Cocok untuk laboratorium, pengolahan air, kontrol kualitas produksi, dan inspeksi lapangan.",
            "Memberikan data pengujian yang dapat dijadikan acuan evaluasi sampel cair.",
        ],
        "apps": ["analisis kualitas air", "pengujian pH, TDS, konduktivitas, atau ORP", "kontrol kualitas air proses", "pemeriksaan sampel cair di laboratorium"],
        "tags": ["water quality meter", "pH meter", "conductivity meter", "TDS meter", "alat uji kualitas air"],
    },
    {
        "id": "pressure_flow",
        "match": ["pressure", "manometer", "barometer", "differential pressure", "flow meter", "flowmeter", "air flow", "anemometer", "wind speed"],
        "category": "Pressure & Flow Meter",
        "term": "alat ukur tekanan dan aliran",
        "function": "mengukur tekanan, perbedaan tekanan, aliran udara, atau kecepatan angin sesuai spesifikasi produk",
        "object": "sistem HVAC, saluran udara, proses industri, ruang kerja, atau titik inspeksi teknis",
        "criteria": "rentang tekanan/aliran, satuan pengukuran, akurasi, jenis sonda, kondisi media, dan kebutuhan pencatatan data",
        "features": [
            "Mengukur tekanan atau aliran sebagai acuan evaluasi sistem teknis.",
            "Cocok untuk HVAC, pemeliharaan, inspeksi proses, dan pemantauan lingkungan kerja.",
            "Memudahkan pemilihan alat berdasarkan rentang kerja dan jenis media yang diuji.",
        ],
        "apps": ["pengujian tekanan sistem", "pemeriksaan aliran udara HVAC", "inspeksi proses industri", "pemantauan ventilasi atau lingkungan kerja"],
        "tags": ["pressure meter", "manometer", "anemometer", "flow meter", "alat ukur tekanan"],
    },
    {
        "id": "light_sound_env",
        "match": ["sound", "noise", "decibel", "lux", "light meter", "luminance", "uv", "particle", "dust", "environmental"],
        "category": "Environmental Meter",
        "term": "alat ukur parameter lingkungan",
        "function": "memantau parameter lingkungan seperti kebisingan, intensitas cahaya, UV, partikel, atau debu sesuai jenis produk",
        "object": "area kerja, ruang produksi, fasilitas umum, laboratorium, atau lokasi inspeksi lapangan",
        "criteria": "parameter yang diuji, rentang pengukuran, akurasi, resolusi, metode sensor, dan kebutuhan pencatatan data",
        "features": [
            "Mendukung pemantauan parameter lingkungan sesuai kategori produk.",
            "Cocok untuk inspeksi keselamatan, audit lingkungan, dan kontrol kondisi ruang kerja.",
            "Data pengukuran dapat menjadi acuan evaluasi sebelum tindakan koreksi dilakukan.",
        ],
        "apps": ["pemantauan kebisingan atau pencahayaan", "inspeksi lingkungan kerja", "pengukuran UV, debu, atau partikel", "audit kondisi ruang produksi atau laboratorium"],
        "tags": ["environmental meter", "sound level meter", "lux meter", "particle counter", "alat ukur lingkungan"],
    },
    {
        "id": "viscosity",
        "match": ["viscometer", "viscosity", "rheometer", "measuring system ms-", "spindle", "rotor"],
        "category": "Viscometer & Rheology",
        "term": "alat ukur viskositas",
        "function": "mengukur viskositas atau kekentalan sampel cair, semi-cair, maupun bahan proses sesuai sistem pengukuran produk",
        "object": "sampel cair, bahan formulasi, produk proses, atau material yang perlu diuji kekentalannya",
        "criteria": "rentang viskositas, tipe spindle atau measuring system, kompatibilitas unit utama, suhu kerja, volume sampel, dan metode pengukuran",
        "features": [
            "Digunakan untuk evaluasi viskositas pada sampel cair atau bahan proses.",
            "Cocok untuk formulasi, kontrol kualitas produksi, laboratorium, dan pengujian material cair/semi-cair.",
            "Pemilihan perlu disesuaikan dengan rentang viskositas, spindle, dan sistem pengukuran yang digunakan.",
        ],
        "apps": ["pengujian viskositas sampel cair", "kontrol kualitas formulasi", "evaluasi kekentalan bahan proses", "pengujian laboratorium untuk sampel semi-cair"],
        "tags": ["viscometer", "viscosity meter", "rheology", "alat ukur viskositas", "measuring system"],
    },
    {
        "id": "vibration_rotation",
        "match": ["accelerometer", "vibration", "vibrometer", "tachometer", "stroboscope", "rpm", "balancing"],
        "category": "Vibration & Rotation Meter",
        "term": "alat ukur getaran dan putaran",
        "function": "memantau getaran, kecepatan putaran, RPM, atau kondisi mesin sesuai sensor dan spesifikasi produk",
        "object": "mesin, motor, bearing, poros, kipas, atau sistem mekanik yang berputar",
        "criteria": "parameter getaran/RPM, rentang pengukuran, tipe sensor, metode pemasangan, resolusi, dan kebutuhan pencatatan data",
        "features": [
            "Memantau kondisi mesin melalui data getaran atau putaran.",
            "Cocok untuk pemeliharaan prediktif, balancing, dan inspeksi mesin berputar.",
            "Data pengukuran dapat menjadi acuan awal untuk menganalisis perubahan kondisi mekanik.",
        ],
        "apps": ["pemantauan getaran mesin", "pengukuran RPM atau putaran", "pemeliharaan motor, bearing, dan poros", "inspeksi kondisi mekanik di area produksi"],
        "tags": ["vibration meter", "accelerometer", "tachometer", "RPM meter", "alat ukur getaran"],
    },
    {
        "id": "ndt_flaw",
        "match": ["defectoscope", "flaw detector", "ndt", "ultrasonic flaw", "crack detector", "borescope", "inspection camera", "endoscope"],
        "category": "NDT & Inspection Instrument",
        "term": "alat inspeksi dan pengujian non-destruktif",
        "function": "memeriksa cacat, kondisi internal, atau area yang sulit dijangkau tanpa merusak objek uji",
        "object": "material, sambungan las, komponen mesin, pipa, rongga, atau objek inspeksi teknis",
        "criteria": "metode inspeksi, rentang kerja, tipe sonda/kamera, resolusi tampilan, material uji, dan kebutuhan dokumentasi hasil",
        "features": [
            "Digunakan untuk inspeksi teknis tanpa perlu merusak objek uji.",
            "Cocok untuk kontrol kualitas material, pemeliharaan, inspeksi sambungan, dan pemeriksaan komponen industri.",
            "Membantu pengguna mengevaluasi kondisi objek berdasarkan metode inspeksi yang tersedia.",
        ],
        "apps": ["inspeksi cacat material", "pemeriksaan sambungan las atau komponen", "pengujian non-destruktif di lapangan", "pemeliharaan pipa, mesin, atau struktur teknis"],
        "tags": ["NDT", "flaw detector", "defectoscope", "inspection camera", "alat inspeksi material"],
    },
    {
        "id": "material_testing",
        "match": ["thickness", "coating", "hardness", "roughness", "gloss", "colorimeter", "spectrophotometer", "force gauge", "torque"],
        "category": "Material Testing & Inspection",
        "term": "alat uji dan inspeksi material",
        "function": "mengevaluasi karakteristik material, permukaan, gaya, torsi, atau kondisi objek sesuai jenis pengujian produk",
        "object": "material, komponen, permukaan, mesin, sambungan, atau objek inspeksi teknis",
        "criteria": "metode uji, rentang pengukuran, resolusi, tipe sonda/sensor, jenis material, dan kondisi objek uji",
        "features": [
            "Digunakan untuk inspeksi kualitas material atau kondisi objek teknis.",
            "Cocok untuk kontrol kualitas produksi, pemeliharaan, laboratorium material, dan pemeriksaan lapangan.",
            "Membantu menentukan kesesuaian produk berdasarkan metode pengujian yang dibutuhkan.",
        ],
        "apps": ["kontrol kualitas material", "inspeksi komponen produksi", "pengujian permukaan, ketebalan, gaya, atau torsi", "pemeliharaan dan pemeriksaan teknis lapangan"],
        "tags": ["material testing", "thickness gauge", "hardness tester", "force gauge", "alat uji material"],
    },
    {
        "id": "scale_lab",
        "match": ["balance", "scale", "weighing", "microscope", "centrifuge", "pipette", "colony counter", "stirrer", "laboratory"],
        "category": "Laboratory Equipment",
        "term": "peralatan laboratorium",
        "function": "mendukung proses penimbangan, pengamatan, pemisahan, preparasi, atau pengujian sampel di laboratorium",
        "object": "sampel laboratorium, bahan uji, media preparasi, atau proses analisis",
        "criteria": "kapasitas, resolusi, metode kerja, kompatibilitas sampel, kebutuhan kalibrasi, dan kondisi penggunaan laboratorium",
        "features": [
            "Mendukung pekerjaan laboratorium yang membutuhkan alat bantu analisis atau preparasi sampel.",
            "Cocok untuk riset, kontrol kualitas, pendidikan, dan pemeriksaan teknis di laboratorium.",
            "Pemilihan produk dapat disesuaikan dengan kapasitas, resolusi, dan jenis sampel yang digunakan.",
        ],
        "apps": ["analisis dan preparasi sampel", "kontrol kualitas laboratorium", "penimbangan, pengamatan, atau pemisahan sampel", "kebutuhan riset dan pendidikan"],
        "tags": ["laboratory equipment", "balance", "microscope", "centrifuge", "peralatan laboratorium"],
    },
]

DEFAULT_FAMILY = {
    "id": "general",
    "category": "Test and Measurement Instrument",
    "term": "perangkat pengujian teknis",
    "function": "mendukung proses pengukuran, pengujian, inspeksi, atau kontrol kualitas sesuai spesifikasi produk",
    "object": "sampel, material, proses produksi, atau objek inspeksi teknis",
    "criteria": "parameter yang diuji, rentang kerja, akurasi, resolusi, metode penggunaan, dan kompatibilitas aksesori",
    "features": [
        "Mendukung kebutuhan pengujian teknis sesuai data produk.",
        "Cocok untuk laboratorium, kontrol kualitas, produksi, atau inspeksi lapangan.",
        "Spesifikasi produk dapat dijadikan acuan sebelum menentukan pilihan alat.",
    ],
    "apps": ["kontrol kualitas", "pengujian teknis", "inspeksi proses produksi", "penggunaan laboratorium atau lapangan"],
    "tags": ["test instrument", "measurement instrument", "alat ukur", "alat uji", "PCE Instruments"],
}


def detect_family(name: str, tags: str = "", category: str = "", description: str = "") -> Dict[str, Any]:
    blob = " ".join([name, tags, category, description]).lower()
    best = None
    best_score = 0
    for fam in FAMILY_CONFIGS:
        score = 0
        for kw in fam["match"]:
            if kw.lower() in blob:
                score += 2 if kw.lower() in name.lower() else 1
        if score > best_score:
            best_score = score
            best = fam
    return best or DEFAULT_FAMILY


def extract_model(name: str, order_no: str = "") -> str:
    if order_no:
        return clean_text(order_no)
    clean_name = clean_text(name)
    m_for = re.search(r"\bfor\s+([A-Za-z]{1,8}(?:[- ]?[A-Za-z0-9]{1,12}){0,4})\b", clean_name, flags=re.I)
    if m_for:
        return clean_text(m_for.group(1)).strip(" .;,()")
    tokens = clean_name.split()
    candidates = []
    for tok in tokens[-6:]:
        t = tok.strip('"\'(),;:/[]')
        if re.search(r"\d", t) and re.search(r"[A-Za-z]", t):
            candidates.append(t)
    return " ".join(candidates[-2:])


def extract_params(text: str, limit: int = 4) -> List[str]:
    text = clean_text(text)
    if not text:
        return []
    # Unit regex sengaja dibuat lebih ketat agar model angka produk seperti "PCE-SM 11 adalah"
    # tidak terbaca sebagai parameter "11 a".
    units_i = r"(?:%\s*RH|%RH|%|ppm|ppb|mg/L|g/L|pH|NTU|°C|℃|K|bar|mbar|Pa|hPa|psi|dB|lux|lx|rpm|m/s|m³/h|m3/h|mm|cm|µm|um|μm|kg|g|kN|Nm|Hz)"
    patterns = [
        (rf"(?:hingga|sampai|up to|range|rentang|nilai|measurement range|measuring range)\s*[:\-]?\s*[^.;,]{{0,45}}?\d[\d.,\s\-–~]*\s*{units_i}", re.I),
        (rf"\d[\d.,\s\-–~]*\s*{units_i}\b", re.I),
        (r"\b\d[\d.,\s\-–~]*\s*(?:A|mA|V|W|kW|N)\b", 0),
        (r"\b(?:CO2|CO₂|H2S|H₂S|O2|O₂|CO|NO2|NO₂|SO2|SO₂|CH4|CH₄|O3|O₃|VOC|Cl2|NH3|PH3)\b", re.I),
        (r"\b(?:ATEX|UL|DIN ISO|ISO)\b", re.I),
    ]
    found: List[str] = []
    for pat, flags in patterns:
        for m in re.finditer(pat, text, flags=flags):
            val = clean_text(m.group(0)).strip(" .;,")
            # Skip artifacts from Indonesian prose or product name fragments.
            if re.search(r"\b(adalah|alat|air)\b", val, flags=re.I):
                continue
            if re.fullmatch(r"\d+(?:[.,]\d+)?\s*[a-z]", val):
                continue
            if len(val) > 90:
                val = val[:90].rsplit(" ", 1)[0]
            if val and val.lower() not in [x.lower() for x in found]:
                found.append(val)
            if len(found) >= limit:
                return found
    return found


def sentence_from_source(text: str) -> str:
    """Return a clean sentence from source text for internal checks only.

    Output katalog tidak lagi menempelkan kalimat sumber mentah, karena mayoritas halaman
    PCE berbahasa Inggris dan beberapa hasil scrape bisa terpotong. Fungsi ini tetap
    disediakan untuk kebutuhan audit internal / pengembangan lanjutan.
    """
    text = clean_text(re.sub(r"^\s*Deskripsi Produk\s*:\s*", "", text, flags=re.I))
    parts = re.split(r"(?<=[.!?])\s+", text)
    for p in parts:
        p = clean_text(p)
        if 60 <= len(p) <= 230 and not re.search(r"harga|promo|garansi|stok", p, flags=re.I):
            return p
    return ""


def source_feature_notes(desc: str, params: List[str], accessory: bool = False) -> List[str]:
    """Create short Indonesian feature notes from explicit source keywords only.

    Fungsi ini sengaja konservatif: hanya membuat poin ketika ada kata/fitur yang
    benar-benar muncul pada teks sumber. Tidak menyalin kalimat Inggris mentah.
    """
    text = clean_text(desc)
    blob = clean_slug_word(text)
    notes: List[str] = []

    def add_if(pattern: str, note: str) -> None:
        if len(notes) >= 3:
            return
        if re.search(pattern, blob, flags=re.I) and note not in notes:
            notes.append(note)

    if accessory:
        add_if(r"\bcompatible\b|\bfor use with\b|\bsuitable for\b|\bfor pce\b", "Dirancang untuk digunakan bersama unit utama yang kompatibel sesuai keterangan produk")
        add_if(r"\breplacement\b|\bspare\b", "Dapat dipakai sebagai komponen pengganti ketika sesuai dengan model alat")
        add_if(r"\bmount\b|\bholder\b|\bstand\b|\btripod\b|\bclamp\b", "Membantu pemasangan atau konfigurasi alat pada aplikasi tertentu")
        return notes

    add_if(r"\bdatalogger\b|\bdata logger\b|\blogging\b|\bsave measurements\b|\bmemory\b", "Mencantumkan dukungan pencatatan atau penyimpanan data pada informasi produk")
    add_if(r"\balarm\b|\bwarning\b|\boptical signal\b|\baudio signal\b|\baudible\b", "Mencantumkan fungsi peringatan atau alarm sesuai konfigurasi produk")
    add_if(r"\bautomatic temperature compensation\b|\btemperature compensation\b|\batc\b", "Mencantumkan kompensasi suhu untuk membantu stabilitas pembacaan")
    add_if(r"\bsignal output\b|\banalogue output\b|\banalog output\b|\brelay output\b|\b4 20 ma\b", "Mencantumkan opsi keluaran sinyal untuk integrasi dengan sistem terkait")
    add_if(r"\bat[e]?x\b|\bexplosion proof\b|\bflammable gas\b", "Mencantumkan informasi sertifikasi atau penggunaan pada aplikasi gas tertentu")
    add_if(r"\bip\s?6[4578]\b|\bwater resistant\b|\bwaterproof\b", "Mencantumkan perlindungan terhadap air atau lingkungan sesuai rating produk")
    add_if(r"\bportable\b|\bhandheld\b|\bcompact\b", "Format produk mendukung penggunaan portabel atau pemeriksaan lapangan")
    add_if(r"\bmin\b.*\bmax\b|\bhold\b", "Mencantumkan fungsi bantu pembacaan seperti Min/Max atau Hold")
    return notes


def split_source_highlights(text: str, limit: int = 8) -> List[str]:
    """Extract concise highlight bullets from source text without copying long paragraphs."""
    text = clean_text(text)
    if not text:
        return []
    # Prioritise text after Highlights/Features when available.
    m = re.search(r"\b(?:Highlights|Features)\b\s*[-:]?\s*(.+)", text, flags=re.I | re.S)
    candidate = m.group(1) if m else text
    raw_parts = re.split(r"\s+-\s+|[\n•]+|\s+\*\s+", candidate)
    out: List[str] = []
    for part in raw_parts:
        part = clean_text(part).strip(" -–—.;")
        part = re.sub(r"\s+", " ", part)
        if not part or len(part) < 4 or len(part) > 140:
            continue
        if is_noise_text(part):
            continue
        if re.search(r"\b(?:copyright|add to cart|delivery time|question|callback|price|vat)\b", part, flags=re.I):
            continue
        # Long marketing sentences are not useful as specification bullets.
        if len(part.split()) > 18 and not re.search(r"\d|%|°C|ppm|bar|mbar|IP\s?\d|ATEX|HOLD|Min|Max", part, flags=re.I):
            continue
        if part.lower() not in [x.lower() for x in out]:
            out.append(part)
        if len(out) >= limit:
            break
    return out



PUBLIC_CATEGORY_BY_FAMILY_ID = {
    "salt_solution": "Alat Ukur Konsentrasi Larutan",
    "gas_air": "Alat Deteksi Gas dan Kualitas Udara",
    "electrical": "Alat Uji dan Ukur Kelistrikan",
    "process_control": "Alat Kontrol Proses",
    "temperature": "Alat Ukur Suhu",
    "moisture_humidity": "Alat Ukur Kelembapan dan Kadar Air",
    "water_quality": "Alat Uji Kualitas Air",
    "pressure_flow": "Alat Ukur Tekanan dan Aliran",
    "light_sound_env": "Alat Ukur Parameter Lingkungan",
    "viscosity": "Alat Ukur Viskositas",
    "vibration_rotation": "Alat Ukur Getaran dan Putaran",
    "ndt_flaw": "Alat Inspeksi dan Uji Non-Destruktif",
    "material_testing": "Alat Uji dan Inspeksi Material",
    "scale_lab": "Peralatan Laboratorium",
    "general": "Alat Uji dan Ukur Teknis",
}

PUBLIC_TAGS_BY_FAMILY_ID = {
    "salt_solution": ["alat ukur salinitas", "alat ukur brix", "refraktometer", "analisis larutan"],
    "gas_air": ["detektor gas", "alat kualitas udara", "pemantauan gas", "keselamatan kerja"],
    "electrical": ["alat ukur listrik", "penguji kelistrikan", "analisis daya", "inspeksi instalasi"],
    "process_control": ["kontrol proses", "pengatur suhu", "kontrol PID", "otomasi proses"],
    "temperature": ["alat ukur suhu", "termometer", "inspeksi suhu", "pemantauan suhu"],
    "moisture_humidity": ["alat ukur kelembapan", "alat ukur kadar air", "higrometer", "inspeksi material"],
    "water_quality": ["alat uji kualitas air", "alat ukur pH", "alat ukur konduktivitas", "analisis air"],
    "pressure_flow": ["alat ukur tekanan", "alat ukur aliran", "manometer", "inspeksi HVAC"],
    "light_sound_env": ["alat ukur lingkungan", "alat ukur kebisingan", "alat ukur cahaya", "inspeksi lingkungan"],
    "viscosity": ["alat ukur viskositas", "uji kekentalan", "rheologi", "kontrol kualitas cairan"],
    "vibration_rotation": ["alat ukur getaran", "alat ukur putaran", "takometer", "pemantauan mesin"],
    "ndt_flaw": ["uji non-destruktif", "alat inspeksi", "inspeksi material", "kamera inspeksi"],
    "material_testing": ["alat uji material", "inspeksi permukaan", "uji kekerasan", "uji ketebalan"],
    "scale_lab": ["peralatan laboratorium", "timbangan laboratorium", "analisis sampel", "alat preparasi"],
    "general": ["alat uji teknis", "alat ukur", "inspeksi teknis", "kontrol kualitas"],
}

PUBLIC_CATEGORY_KEYWORDS = [
    (r"accessor(?:y|ies)|spare parts?|probe|sensor|adapter|adaptor|cable|case|holder|stand|module|certificate", "Aksesori dan Suku Cadang PCE"),
    (r"salt|salinity|brix|refractometer", "Alat Ukur Konsentrasi Larutan"),
    (r"gas|air quality|co2|h2s|oxygen|ozone|voc", "Alat Deteksi Gas dan Kualitas Udara"),
    (r"multimeter|clamp meter|power|voltage|current|electrical|insulation|earth tester", "Alat Uji dan Ukur Kelistrikan"),
    (r"controller|regulator|process control|pid", "Alat Kontrol Proses"),
    (r"temperature|thermometer|thermal|infrared|thermocouple", "Alat Ukur Suhu"),
    (r"moisture|humidity|hygrometer|dew point", "Alat Ukur Kelembapan dan Kadar Air"),
    (r"ph|conductivity|tds|turbidity|water quality|orp|dissolved oxygen", "Alat Uji Kualitas Air"),
    (r"pressure|manometer|barometer|flow|anemometer|wind", "Alat Ukur Tekanan dan Aliran"),
    (r"sound|noise|lux|light|uv|particle|dust|environment", "Alat Ukur Parameter Lingkungan"),
    (r"viscometer|viscosity|rheometer", "Alat Ukur Viskositas"),
    (r"vibration|accelerometer|tachometer|stroboscope|rpm", "Alat Ukur Getaran dan Putaran"),
    (r"defectoscope|flaw|ndt|borescope|inspection camera|endoscope", "Alat Inspeksi dan Uji Non-Destruktif"),
    (r"thickness|coating|hardness|roughness|force|torque|gloss|colorimeter", "Alat Uji dan Inspeksi Material"),
    (r"balance|scale|weighing|microscope|centrifuge|pipette|laboratory", "Peralatan Laboratorium"),
]

PUBLIC_ENGLISH_BLOCK_RE = re.compile(
    r"\b(?:range|measurement|measuring|weighing|readability|interface|output|input|display|"
    r"without|with|from|up to|accuracy|resolution|power|battery|current clamp|clamp|adapter|module|"
    r"software|accessory|spare|certificate|calibration|automatic|waterproof|water resistant|"
    r"portable|compact|warning|baud rate|sample rate|data rate|memory|stainless steel|"
    r"pressure measurement|humidity measurement|temperature measurement|expansion|plug[- ]?in|"
    r"communication|automatically|detected|graphic|opening|harmonic|distortion|order|"
    r"suitable|compatible|support|supports|probe|electrode|feature|features|function|"
    r"external|internal|included|including|high|low|sampling|rate|screen|phase|phases)\b",
    re.I,
)

TECH_PHRASE_REPLACEMENTS = [
    (r"\bplug[- ]?in communication module\b", "modul komunikasi tambahan"),
    (r"\bmodule is automatically detected\b", "modul terdeteksi otomatis"),
    (r"\bexpansion module\b", "modul ekspansi"),
    (r"\bexpansion\b", "ekspansi"),
    (r"\bcommunication module\b", "modul komunikasi"),
    (r"\bgraphic display\b", "layar grafis"),
    (r"\bgraphic\b", "grafis"),
    (r"\bcurrent clamp opening\b", "bukaan tang arus"),
    (r"\bclamp opening\b", "bukaan tang"),
    (r"\bopening\b", "bukaan"),
    (r"\bharmonic distortion\b", "distorsi harmonik"),
    (r"\bthe\s+(\d+)(?:st|nd|rd|th)\s+order\b", r"orde ke-\1"),
    (r"\b(\d+)(?:st|nd|rd|th)\s+order\b", r"orde ke-\1"),
    (r"\border\b", "orde"),
    (r"\bfrom\b", "dari"),
    (r"\bmeasurement for\b", "pengukuran untuk"),
    (r"\bmeasurements? for\b", "pengukuran untuk"),
    (r"\b1\s+and\s+3\s+phases?\b", "1 dan 3 fase"),
    (r"\b3\s+phases?\b", "3 fase"),
    (r"\b1\s+phase\b", "1 fase"),
    (r"\bphases?\b", "fase"),
    (r"\bfor use with\b", "untuk digunakan bersama"),
    (r"\bcompatible with\b", "kompatibel dengan"),
    (r"\bsuitable for\b", "sesuai untuk"),
    (r"\bsupports?\b", "mendukung"),
    (r"\bfeatures?\b", "fitur"),
    (r"\bfunction\b", "fungsi"),
    (r"\bscreen\b", "layar"),
    (r"\bsampling rate\b", "laju pengambilan sampel"),
    (r"\bincluded\b", "termasuk"),
    (r"\bincluding\b", "termasuk"),
    (r"\bexternal\b", "eksternal"),
    (r"\binternal\b", "internal"),
    (r"\bmeasurement range up to\b", "rentang pengukuran hingga"),
    (r"\bmeasurement range\b", "rentang pengukuran"),
    (r"\bmeasuring range\b", "rentang pengukuran"),
    (r"\bweighing range\b", "rentang timbang"),
    (r"\breadability\b", "keterbacaan"),
    (r"\baccuracy\b", "akurasi"),
    (r"\bresolution\b", "resolusi"),
    (r"\brange of\b", "rentang"),
    (r"\brange\b", "rentang"),
    (r"\bup to\b", "hingga"),
    (r"\bof the weight\b", "dari berat"),
    (r"\bwith the weight\b", "berdasarkan berat"),
    (r"\bautomatic temperature compensation\b", "kompensasi suhu otomatis"),
    (r"\btemperature compensation\b", "kompensasi suhu"),
    (r"\btemperature measurement\b", "pengukuran suhu"),
    (r"\bhumidity measurement\b", "pengukuran kelembapan"),
    (r"\bpressure measurement\b", "pengukuran tekanan"),
    (r"\bdifferential pressure\b", "tekanan diferensial"),
    (r"\bsignal output\b", "keluaran sinyal"),
    (r"\brelay output\b", "keluaran relai"),
    (r"\banalogue output\b", "keluaran analog"),
    (r"\banalog output\b", "keluaran analog"),
    (r"\balarm output\b", "keluaran alarm"),
    (r"\btrigger input\b", "masukan pemicu"),
    (r"\binput\b", "masukan"),
    (r"\boutput\b", "keluaran"),
    (r"\bUSB interface\b", "antarmuka USB"),
    (r"\bRS[- ]?232 interface\b", "antarmuka RS-232"),
    (r"\binterface\b", "antarmuka"),
    (r"\bdatalogger\b", "pencatat data"),
    (r"\bdata logger\b", "pencatat data"),
    (r"\bmemory\b", "memori"),
    (r"\balarm\b", "alarm"),
    (r"\bwarning\b", "peringatan"),
    (r"\bauto shut[- ]?off\b", "mati otomatis"),
    (r"\bwithout display\b", "tanpa layar"),
    (r"\bdigital display\b", "layar digital"),
    (r"\bdisplay\b", "layar"),
    (r"\bwater resistant\b", "tahan air"),
    (r"\bwaterproof\b", "tahan air"),
    (r"\bportable\b", "portabel"),
    (r"\bhandheld\b", "genggam"),
    (r"\bcompact\b", "ringkas"),
    (r"\bmaneuverable\b", "mudah digunakan"),
    (r"\bcurrent clamp\b", "tang arus"),
    (r"\bclamp\b", "penjepit"),
    (r"\bprobe\b", "sonda"),
    (r"\belectrode\b", "elektroda"),
    (r"\bsensor\b", "sensor"),
    (r"\bcable\b", "kabel"),
    (r"\badapter\b", "adaptor"),
    (r"\badaptor\b", "adaptor"),
    (r"\bmodule\b", "modul"),
    (r"\bsoftware\b", "perangkat lunak"),
    (r"\baccessory\b", "aksesori"),
    (r"\baccessories\b", "aksesori"),
    (r"\bspare parts?\b", "suku cadang"),
    (r"\bcalibration certificate\b", "sertifikat kalibrasi"),
    (r"\bcertificate\b", "sertifikat"),
    (r"\bcalibration\b", "kalibrasi"),
    (r"\bstainless steel\b", "baja tahan karat"),
    (r"\bbaud rate\b", "laju baud"),
    (r"\bsample rate\b", "laju sampel"),
    (r"\bdata rate\b", "laju data"),
    (r"\bbattery\b", "baterai"),
    (r"\bpower supply\b", "catu daya"),
    (r"\bpower\b", "daya"),
    (r"\bmin\.?\s*/\s*max\b", "nilai minimum/maksimum"),
    (r"\bhold\b", "tahan baca"),
    (r"\bfor\b", "untuk"),
    (r"\band\b", "dan"),
    (r"\bwith\b", "dengan"),
    (r"\band\b", "dan"),
    (r"\bor\b", "atau"),
    (r"\bwith\b", "dengan"),
    (r"\bwithout\b", "tanpa"),
]


def public_category(category: str, fam: Dict[str, Any], accessory: bool = False) -> str:
    if accessory:
        return "Aksesori dan Suku Cadang PCE"
    fam_id = str((fam or {}).get("id", "general"))
    if fam_id in PUBLIC_CATEGORY_BY_FAMILY_ID:
        return PUBLIC_CATEGORY_BY_FAMILY_ID[fam_id]
    raw = clean_text(category)
    raw_low = raw.lower()
    for pattern, label in PUBLIC_CATEGORY_KEYWORDS:
        if re.search(pattern, raw_low, flags=re.I):
            return label
    return PUBLIC_CATEGORY_BY_FAMILY_ID["general"]


def translate_technical_phrase(text: str) -> str:
    """Translate short technical phrases into Indonesian and drop unfinished fragments."""
    s = clean_text(text).strip(" .;:,-–—")
    if not s:
        return ""
    s = re.sub(r"\s*/\s*s\b", "/detik", s, flags=re.I)
    s = re.sub(r"\s+\.\.\.\s+", " ... ", s)
    s = re.sub(r"\bof\s+(-?\d)", r"\1", s, flags=re.I)
    for pattern, repl in TECH_PHRASE_REPLACEMENTS:
        s = re.sub(pattern, repl, s, flags=re.I)
    # Rapikan kombinasi hasil terjemahan yang sering muncul dari highlight PCE.
    s = re.sub(r"\bdistorsi harmonik hingga\s+(?:the\s+)?(\d+)(?:st|nd|rd|th)?\s+orde\b", r"distorsi harmonik hingga orde ke-\1", s, flags=re.I)
    s = re.sub(r"\bdistorsi harmonik hingga\s+orde\s+ke[- ]?(\d+)\b", r"distorsi harmonik hingga orde ke-\1", s, flags=re.I)
    s = re.sub(r"\bbukaan tang arus\s*(\d+(?:[.,]\d+)?\s*mm)\b", r"bukaan tang arus \1", s, flags=re.I)
    s = re.sub(r"\brentang pengukuran dari\s+(-?\d)", r"rentang pengukuran \1", s, flags=re.I)
    s = re.sub(r"\brentang dari\s+(-?\d)", r"rentang \1", s, flags=re.I)
    s = re.sub(r"\bpengukuran kelembapan pengukuran\b", "pengukuran kelembapan", s, flags=re.I)
    s = re.sub(r"\bpengukuran suhu pengukuran\b", "pengukuran suhu", s, flags=re.I)
    s = re.sub(r"\bmodul komunikasi modul\b", "modul komunikasi", s, flags=re.I)
    s = re.sub(r"\bmodul ekspansi modul\b", "modul ekspansi", s, flags=re.I)
    s = re.sub(r"\bmodul terdeteksi otomatis\b", "modul terdeteksi otomatis", s, flags=re.I)
    s = re.sub(r"\bto\b", "hingga", s, flags=re.I)
    s = re.sub(r"\bup\s+hingga\b", "hingga", s, flags=re.I)
    s = re.sub(r"\brentang pengukuran pengukuran\b", "rentang pengukuran", s, flags=re.I)
    s = re.sub(r"\brentang timbang pengukuran\b", "rentang timbang", s, flags=re.I)
    s = re.sub(r"\bmin\b", "minimum", s, flags=re.I)
    s = re.sub(r"\bmax\b", "maksimum", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    # Use Indonesian decimal comma only in simple baud-rate style values; keep model/ranges unchanged.
    s = re.sub(r"(?<=\d)\.(?=\d\s*(?:kbit|Mbit|bit))", ",", s)
    if not s:
        return ""
    # Capitalise first letter when the phrase starts with a word.
    if re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


def has_untranslated_public_english(text: str) -> bool:
    s = clean_text(text)
    if not s:
        return False
    # Product names are allowed elsewhere, but technical bullets/parameters must not keep English fragments.
    return PUBLIC_ENGLISH_BLOCK_RE.search(s) is not None


def is_weak_public_parameter(label: str, original: str = "") -> bool:
    """Drop isolated numeric fragments that are not useful as public catalog parameters.

    Nama produk tetap dibiarkan apa adanya. Fungsi ini hanya menyaring parameter
    seperti "10V" atau "2 1.5V" yang biasanya muncul dari catu daya/baterai dan
    kurang jelas bila ditampilkan tanpa konteks.
    """
    s = clean_text(label).strip(" .;:,-–—")
    if not s:
        return True
    # Isolated electrical/mechanical values without a label are usually scrape fragments.
    if re.fullmatch(r"(?:\d+(?:[.,]\d+)?\s*){1,3}(?:×|x)?\s*(?:V|A|mA|W|kW|N)", s, flags=re.I):
        return True
    if re.fullmatch(r"\d+\s+\d+(?:[.,]\d+)?\s*(?:V|A|mA|W|kW|N)", s, flags=re.I):
        return True
    # Keep meaningful ranges such as 0 ... 100 % or -210 ... 1370 °C, but drop very short unlabeled values.
    if re.fullmatch(r"\d+(?:[.,]\d+)?\s*(?:V|A|mA|W|kW|N)", s, flags=re.I):
        return True
    return False


def normalize_public_phrase(label: str) -> str:
    """Final cleanup for public parameter/highlight text, without touching product names."""
    s = clean_text(label)
    if not s:
        return ""
    s = re.sub(r"\bmaintenance\b", "pemeliharaan", s, flags=re.I)
    s = re.sub(r"\bphases?\b", "fase", s, flags=re.I)
    s = re.sub(r"\bprobe\b", "sonda", s, flags=re.I)
    s = re.sub(r"\bclamp\b", "tang", s, flags=re.I)
    s = re.sub(r"\bfrom\b", "dari", s, flags=re.I)
    s = re.sub(r"\bto\b", "hingga", s, flags=re.I)
    s = re.sub(r"\s*/\s*s\b", "/detik", s, flags=re.I)
    s = re.sub(r"\bpengukuran untuk 1 dan 3 fase\b", "Pengukuran untuk 1 dan 3 fase", s, flags=re.I)
    s = re.sub(r"(?<=\d)\.(?=\d\s*(?:V|A|mA|W|kW|N|kbit|Mbit|bit))", ",", s)
    s = re.sub(r"\b(\d+)\s+(\d+(?:[,.]\d+)?)\s*V\b", r"\1 × \2 V", s, flags=re.I)
    if re.fullmatch(r"\d+\s*×\s*\d+(?:[,.]\d+)?\s*V", s, flags=re.I):
        s = "Catu daya " + s
    s = re.sub(r"\b(\d+(?:[,.]\d+)?)\s*([VANW])\b", r"\1 \2", s, flags=re.I)
    s = re.sub(r"\b(\d+(?:[,.]\d+)?)\s*(mA|kW)\b", r"\1 \2", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


def public_highlight_label(line: str) -> str:
    """Convert PCE highlight snippets into Indonesian-only catalog bullets.

    Baris yang masih mengandung frasa Inggris tidak dipaksa masuk ke output publik.
    Ini lebih aman daripada menempel potongan spesifikasi mentah yang campur bahasa.
    """
    original = clean_text(line).strip(" .")
    if not original:
        return ""
    low_norm = re.sub(r"\s+", " ", original.lower())

    if re.search(r"\bATEX\b", original, flags=re.I):
        return "Mencantumkan informasi ATEX sesuai data produk"
    if re.search(r"\bIP\s?\d{2}\b", original, flags=re.I):
        ip = re.search(r"\bIP\s?\d{2}\b", original, flags=re.I).group(0).upper().replace(" ", "")
        return f"Mencantumkan perlindungan lingkungan dengan rating {ip}"
    if "automatic temperature compensation" in low_norm:
        return "Kompensasi suhu otomatis"
    if "temperature compensation" in low_norm:
        return "Kompensasi suhu"
    if "auto shut" in low_norm:
        return "Mati otomatis untuk membantu menghemat baterai"
    if "water resistant" in low_norm or "waterproof" in low_norm:
        return "Tahan air sesuai informasi produk"
    if "without display" in low_norm:
        return "Tanpa layar"
    if re.search(r"\bmin\.?\s*/?\s*max\b", low_norm) or re.search(r"\bhold\b", low_norm):
        return "Fungsi nilai minimum/maksimum dan tahan baca sesuai informasi produk"

    translated = normalize_public_phrase(translate_technical_phrase(original))
    if not translated:
        return ""
    # Reject long prose, weak numeric fragments, and snippets that still contain common English words.
    if len(translated.split()) > 16:
        return ""
    if is_weak_public_parameter(translated, original):
        return ""
    if has_untranslated_public_english(translated):
        return ""
    # Keep useful technical bullets only: values/units or known Indonesian feature words.
    if re.search(r"\d|%|°C|ppm|ppb|bar|mbar|Pa\b|psi|dB|lux|rpm|m/s|Hz|bit|IP\d|ATEX|USB|RS-?232", translated, flags=re.I):
        return translated
    if re.search(r"\b(?:pengukuran|kompensasi|antarmuka|keluaran|masukan|alarm|peringatan|layar|baterai|memori|portabel|ringkas|tahan air|adaptor|kabel|modul|sertifikat|kalibrasi|sonda|elektroda|sensor|tang|fase|distorsi|harmonik)\b", translated, flags=re.I):
        return translated
    return ""


def public_param_text(params: List[str]) -> str:
    labels: List[str] = []
    for param in params:
        p = clean_text(param)
        if not p:
            continue
        label = public_highlight_label(p) or translate_technical_phrase(p)
        label = normalize_public_phrase(label)
        if not label:
            continue
        if is_weak_public_parameter(label, p):
            continue
        if has_untranslated_public_english(label):
            continue
        if label.lower() not in [x.lower() for x in labels]:
            labels.append(label)
    return ", ".join(labels)

def html_ul(items: List[str]) -> str:
    clean_items = [clean_text(x).strip(" .") for x in items if clean_text(x)]
    return "<ul>" + "".join(f"<li>{esc(x)}.</li>" for x in clean_items) + "</ul>"


def unique_items(items: List[str], max_items: int = 10) -> List[str]:
    out: List[str] = []
    seen = set()
    for item in items:
        item = clean_text(item).strip(" .")
        if not item:
            continue
        key = re.sub(r"\s+", " ", item.lower())
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
        if len(out) >= max_items:
            break
    return out


def build_features(name: str, fam: Dict[str, Any], desc: str, params: List[str], accessory: bool) -> List[str]:
    if accessory:
        model = extract_model(name)
        base = [
            f"Dirancang sebagai aksesori atau komponen pendukung untuk konfigurasi {model or 'PCE'} yang kompatibel",
            "Berguna saat pengguna membutuhkan penggantian, perluasan fungsi, atau kelengkapan sistem pengukuran",
            "Pemilihan perlu disesuaikan dengan unit utama, konektor, sensor, dan metode penggunaan yang dipakai",
        ]
    else:
        base = list(fam.get("features", []))
    param_note = public_param_text(params[:3]) if params else ""
    if param_note:
        base.insert(1, f"Data produk mencantumkan parameter penting seperti {param_note}")
    base.extend(source_feature_notes(desc, params, accessory))
    return unique_items(base, 5)


def build_applications(fam: Dict[str, Any], accessory: bool) -> List[str]:
    if accessory:
        return [
            "penggantian atau pelengkap unit utama yang kompatibel",
            "penyesuaian konfigurasi alat dengan aplikasi pengujian tertentu",
            "pemeliharaan perangkat pengukuran yang sudah digunakan",
            "kebutuhan teknis ketika sonda, sensor, tang, kabel, atau modul bawaan belum mencukupi",
        ]
    return list(fam.get("apps", DEFAULT_FAMILY["apps"]))[:5]


def build_opening(name: str, brand: str, fam: Dict[str, Any], desc: str, accessory: bool) -> str:
    term = fam.get("term", DEFAULT_FAMILY["term"])
    function = fam.get("function", DEFAULT_FAMILY["function"])
    obj = fam.get("object", DEFAULT_FAMILY["object"])
    idx = deterministic_index(name, 3)

    if accessory:
        model = extract_model(name)
        templates = [
            f"<p><strong>{esc(name)}</strong> adalah aksesori atau komponen pendukung dari {esc(brand)} untuk melengkapi konfigurasi alat yang kompatibel. Produk ini dipakai ketika pengguna membutuhkan komponen tambahan, pengganti, atau perluasan fungsi pada sistem pengukuran.</p>",
            f"<p><strong>{esc(name)}</strong> merupakan komponen pendukung {esc(brand)} yang digunakan bersama unit utama yang sesuai{(' seperti ' + esc(model)) if model else ''}. Produk ini membantu menjaga konfigurasi alat tetap sesuai dengan kebutuhan pengujian.</p>",
            f"<p><strong>{esc(name)}</strong> adalah aksesori PCE yang perlu dipilih berdasarkan kompatibilitas unit utama, konektor, dan metode pengukuran. Produk ini relevan untuk kebutuhan pemeliharaan, penggantian komponen, atau penyesuaian aplikasi teknis.</p>",
        ]
        return templates[idx]

    templates = [
        f"<p><strong>{esc(name)}</strong> adalah {esc(term)} dari {esc(brand)} yang digunakan untuk {esc(function)}. Produk ini dapat dipakai sebagai acuan saat mengevaluasi {esc(obj)}.</p>",
        f"<p><strong>{esc(name)}</strong> merupakan {esc(term)} PCE Instruments untuk {esc(function)}. Perangkat ini cocok dipertimbangkan ketika pengguna perlu mendapatkan data teknis dari {esc(obj)} secara lebih terarah.</p>",
        f"<p><strong>{esc(name)}</strong> adalah {esc(term)} yang dirancang untuk {esc(function)}. Dalam penggunaan harian, produk ini membantu proses pengecekan pada {esc(obj)} dengan data yang lebih mudah dievaluasi.</p>",
    ]
    return templates[idx]


def source_note_html(row: Dict[str, Any]) -> str:
    # Tidak ditampilkan ke katalog publik. Sumber tetap dicatat pada kolom Source URL,
    # Website Validation, Website Correction Log, dan Content Quality.
    return ""


def generate_product_description(row: Dict[str, str]) -> str:
    name = row["name"]
    brand = row["brand"]
    fam = row["family"]
    desc = row.get("description", "")
    params = row.get("params", [])
    accessory = row.get("is_accessory", False)
    features = build_features(name, fam, desc, params, accessory)
    apps = build_applications(fam, accessory)
    criteria = fam.get("criteria", DEFAULT_FAMILY["criteria"])

    parts: List[str] = [build_opening(name, brand, fam, desc, accessory)]
    param_note = public_param_text(params[:4]) if params else ""
    if param_note and not accessory:
        parts.append(
            f"<p>Informasi dari data produk menunjukkan beberapa parameter yang perlu diperhatikan, seperti {esc(param_note)}. Detail tersebut sebaiknya dicocokkan kembali dengan kebutuhan sampel dan kondisi penggunaan.</p>"
        )

    parts.append(f"<h2>Fungsi dan Keunggulan {esc(name)}</h2>")
    parts.append(html_ul(features))
    parts.append("<h2>Contoh Penggunaan</h2>")
    parts.append(f"<p>{esc(name)} dapat dipertimbangkan untuk kebutuhan seperti:</p>")
    parts.append(html_ul(apps))
    parts.append("<h2>Informasi Pemilihan Produk</h2>")
    if accessory:
        parts.append(
            f"<p>Sebelum digunakan, pastikan kompatibilitas {esc(name)} dengan unit utama, model alat, konektor, jenis sensor, dan konfigurasi pengukuran yang dipakai. Produk aksesori tidak selalu dapat digunakan sebagai alat utama secara mandiri.</p>"
        )
    else:
        parts.append(
            f"<p>Sebelum digunakan, pastikan {esc(criteria)} sudah sesuai dengan sampel, objek uji, atau kondisi kerja yang akan diperiksa.</p>"
        )
    parts.append(UJI_CTA)
    return "".join(parts)


def generate_short_description(row: Dict[str, str]) -> str:
    name = row["name"]
    fam = row["family"]
    accessory = row.get("is_accessory", False)
    if accessory:
        s = f"{name} adalah aksesori PCE untuk melengkapi atau mengganti komponen pada unit utama yang kompatibel."
    else:
        s = f"{name} adalah {fam.get('term', DEFAULT_FAMILY['term'])} PCE Instruments untuk {fam.get('function', DEFAULT_FAMILY['function'])}."
    if len(s) > 220:
        s = s[:220].rsplit(" ", 1)[0].rstrip(".,") + "."
    return s


def has_official_spec_table(row: Dict[str, Any]) -> bool:
    """Return True only when an official specification table was parsed from the PCE page."""
    spec_rows = row.get("spec_table", []) if isinstance(row.get("spec_table", []), list) else []
    return bool(spec_table_html(spec_rows))


def generate_spec_tab(row: Dict[str, str]) -> str:
    """Generate the public SPEC tab only from the official PCE specification table.

    Aturan publish:
    - Jika tabel spesifikasi resmi PCE berhasil dibaca, tampilkan tabel itu saja.
    - Jangan menambahkan Ringkasan Data Produk, Data Produk, Parameter Terdeteksi,
      atau Highlight Teknis ke tab SPESIFIKASI.
    - Jika tabel spesifikasi resmi tidak ada / tidak terbaca, kosongkan custom tab
      SPESIFIKASI agar tidak terlihat seperti membuat spesifikasi sendiri.
    """
    spec_rows = row.get("spec_table", []) if isinstance(row.get("spec_table", []), list) else []
    spec_html = spec_table_html(spec_rows)
    if not spec_html:
        return ""
    return "<h2>Spesifikasi Teknis</h2>" + spec_html


def source_highlights_for_feature_tab(row: Dict[str, Any], limit: int = 6) -> List[str]:
    """Move website Highlights/Features into the Keunggulan tab, not the SPEC tab.

    Output tetap konservatif: hanya highlight singkat yang bisa diterjemahkan menjadi
    bahasa Indonesia katalog. Baris mentah bahasa Inggris yang tidak dikenali dibuang.
    """
    highlights: List[str] = []
    for source_text in [clean_text(row.get("website_specs", "")), clean_text(row.get("description", ""))]:
        for item in split_source_highlights(source_text, limit=limit):
            label = public_highlight_label(item)
            if not label:
                continue
            label = normalize_public_phrase(label)
            if has_untranslated_public_english(label):
                continue
            if label.lower() not in [x.lower() for x in highlights]:
                highlights.append(label)
            if len(highlights) >= limit:
                return highlights
    return highlights


def generate_feature_tab(row: Dict[str, str]) -> str:
    features = build_features(row["name"], row["family"], row.get("description", ""), row.get("params", []), row.get("is_accessory", False))
    features.extend(source_highlights_for_feature_tab(row, limit=6))
    features = unique_items(features, 10)
    return f"<h2>Keunggulan Produk</h2>{html_ul(features)}"


def generate_faq(row: Dict[str, str]) -> str:
    name = row["name"]
    fam = row["family"]
    accessory = row.get("is_accessory", False)
    criteria = fam.get("criteria", DEFAULT_FAMILY["criteria"])
    params = row.get("params", [])
    qas: List[Tuple[str, str]] = []
    if accessory:
        qas = [
            (f"Apa fungsi {name}?", f"{name} digunakan sebagai aksesori atau komponen pendukung untuk unit utama yang kompatibel. Produk ini perlu dipilih berdasarkan model alat dan kebutuhan aplikasi."),
            ("Apakah produk ini bisa digunakan secara mandiri?", "Tidak selalu. Jika produk ini berupa sonda, sensor, tang, kabel, adaptor, atau modul, pengguna harus memastikan kompatibilitasnya dengan unit utama terlebih dahulu."),
            (f"Apa yang perlu dicek sebelum memilih {name}?", "Periksa model alat, konektor, rentang kerja, jenis sensor, dan metode penggunaan agar aksesori sesuai dengan sistem yang sudah digunakan."),
            ("Untuk siapa produk ini cocok?", "Produk ini cocok untuk teknisi, laboratorium, tim kontrol kualitas, atau pengguna alat PCE yang membutuhkan komponen tambahan atau pengganti."),
        ]
    else:
        qas = [
            (f"Apa fungsi {name}?", f"{name} digunakan untuk {fam.get('function', DEFAULT_FAMILY['function'])}. Produk ini membantu pengguna memperoleh data pengujian sesuai parameter yang tersedia."),
            (f"Untuk kebutuhan apa {name} digunakan?", f"Produk ini dapat digunakan untuk {', '.join(build_applications(fam, False)[:3])}. Penggunaan tetap perlu disesuaikan dengan sampel atau objek uji."),
            (f"Apa yang perlu dicek sebelum memilih {name}?", f"Periksa {criteria} agar produk sesuai dengan kebutuhan pengujian dan kondisi kerja."),
            ("Apakah produk ini cocok untuk laboratorium atau lapangan?", "Bisa, selama spesifikasi produk sesuai dengan aplikasi yang dibutuhkan. Beberapa produk lebih cocok untuk laboratorium, sedangkan produk portabel dapat digunakan untuk inspeksi lapangan."),
        ]
        param_note = public_param_text(params[:4]) if params else ""
        if param_note:
            qas.insert(2, ("Parameter apa yang perlu diperhatikan?", f"Data produk mencantumkan informasi seperti {param_note}. Detail ini perlu dicocokkan dengan kebutuhan sampel, rentang kerja, dan standar internal pengguna."))
    return "".join(f"<h3>{esc(q)}</h3><p>{esc(a)}</p>" for q, a in qas[:5])


def meta_description(row: Dict[str, str]) -> str:
    name = row["name"]
    fam = row["family"]
    cat = public_category(row.get("category", ""), fam, row.get("is_accessory", False))
    if row.get("is_accessory", False):
        candidates = [
            f"{name} aksesori PCE untuk unit kompatibel. Cek fungsi, kecocokan model, dan detail produk di uji.co.id.",
            f"{name} aksesori PCE. Cek kompatibilitas dan detail produk di uji.co.id.",
        ]
    else:
        candidates = [
            f"{name} PCE Instruments untuk {cat}. Cek fungsi, spesifikasi, dan konsultasi produk di uji.co.id.",
            f"{name} PCE Instruments. Cek fungsi dan spesifikasi produk di uji.co.id.",
        ]
    for s in candidates:
        if len(s) <= 158:
            return s
    s = candidates[-1]
    return s[:158].rsplit(" ", 1)[0].rstrip(" ,.;") + "."

def generate_tags(row: Dict[str, str]) -> str:
    fam = row["family"]
    fam_id = str((fam or {}).get("id", "general"))
    is_acc = bool(row.get("is_accessory", False))
    tags: List[str] = []
    tags.append(public_category(row.get("category", ""), fam, is_acc))
    tags.extend(PUBLIC_TAGS_BY_FAMILY_ID.get(fam_id, PUBLIC_TAGS_BY_FAMILY_ID["general"]))
    if is_acc:
        tags.extend(["aksesori PCE", "suku cadang PCE", "komponen pendukung"])

    # Tambahkan model/SKU sebagai tag teknis. Nama produk penuh sengaja tidak dimasukkan
    # agar kolom tag tidak dipenuhi frasa Inggris dari judul produk.
    for token in re.findall(r"\b(?:PCE|CAL|IRMAX|RT|FM|PPC|PAC)[- ]?[A-Za-z0-9]+(?:[- ][A-Za-z0-9]+){0,4}\b", row.get("name", ""), flags=re.I):
        token = clean_text(token).strip(" ,.;()[]")
        token = re.sub(r"\b(?:incl|with|for|accessory|software)\b.*$", "", token, flags=re.I).strip(" -_/.,;()[]")
        if token and len(token) <= 40:
            tags.append(token)

    clean_tags: List[str] = []
    seen = set()
    for tag in tags:
        tag = clean_text(tag).strip(" ,.;")
        if not tag or len(tag) > 80:
            continue
        # Hindari tag publik yang masih memuat frasa Inggris umum.
        if has_untranslated_public_english(tag):
            continue
        key = tag.lower()
        if key in seen:
            continue
        seen.add(key)
        clean_tags.append(tag)
    return ", ".join(clean_tags[:10])

def normalize_row(
    row: Dict[str, Any],
    scrape_web: bool = True,
    scrape_cache: Optional[Dict[str, Any]] = None,
    request_timeout: int = 15,
    request_delay: float = 0.0,
) -> Dict[str, Any]:
    source_url = pick(row, "link", "Source URL", "url", "URL")
    scraped: Dict[str, Any] = {}
    if scrape_web and source_url:
        scraped = get_scraped_product_data(source_url, scrape_cache, timeout=request_timeout, delay=request_delay)

    name = strip_jual(pick(row, "Judul", "Name", "Title", "Product Name"))
    scraped_title = strip_jual(clean_text(scraped.get("title", "")))
    if not name and scraped_title:
        name = scraped_title
    if not name:
        # Beberapa baris PCE tidak punya Judul, tetapi masih punya Order No/Internal Reference
        # dan deskripsi yang valid. Gunakan kode produk sebagai nama fallback agar tidak kosong.
        name = strip_jual(pick(row, "Order No", "Internal Reference", "costumiz"))

    brand = normalize_brand(pick(row, "Brand", "Merek"))
    source_tags = pick(row, "Tags", "Product_Tags", "product tags")
    excel_desc = pick(
        row,
        "Website Description", "API Description", "Source Description", "Original Description",
        "Short_Description", "Product Short Description", "description", "Description",
        "description_sale", "description_purchase", "description_ecommerce", "body_html",
    )
    website_desc = clean_text(scraped.get("description", ""))
    if website_desc and len(website_desc) >= 20:
        desc = website_desc
        description_origin = "website_cache" if scraped.get("from_cache") == "yes" else "website"
    elif excel_desc:
        desc = excel_desc
        description_origin = "excel"
    else:
        desc = ""
        description_origin = "fallback_template"

    source_category = pick(row, "Product Category/Name", "product categories", "Category", "Tags")
    # Avoid generic Odoo category if it is not a useful product category.
    if source_category.lower() in {"saleable", "storable product", "units"}:
        source_category = source_tags
    fam = detect_family(name, source_tags, source_category, desc)
    category = source_category if source_category and source_category.lower() != "saleable" else fam.get("category")
    is_acc, accessory_reason = detect_accessory_reason(name, source_tags, category, desc)
    category = public_category(category, fam, is_acc)
    params = extract_params("\n".join([name, source_tags, desc, clean_text(scraped.get("specs", ""))]))
    # Sertifikasi seperti ISO/UL/ATEX hanya dimunculkan ketika relevan dengan produk keselamatan, gas,
    # atau kelistrikan. Untuk produk umum, sertifikasi tunggal sering muncul dari teks sumber yang terlalu luas.
    if fam.get("id") not in {"gas_air", "electrical"}:
        params = [p for p in params if p.upper() not in {"ISO", "UL", "ATEX", "DIN ISO"}]

    image_url = pick(row, "Image", "image_url", "Images", "Gambar") or clean_text(scraped.get("image_url", ""))
    order_no = pick(row, "Order No", "Order No.", "order_no", "Model") or clean_text(scraped.get("order_no", ""))
    gtin = pick(row, "GTIN", "GTIN (EAN)", "EAN") or clean_text(scraped.get("gtin", ""))

    return {
        "name": name,
        "brand": brand,
        "source_tags": source_tags,
        "category": category or fam.get("category"),
        "family": fam,
        "description": desc,
        "description_origin": description_origin,
        "website_specs": clean_text(scraped.get("specs", "")),
        "spec_table": scraped.get("spec_table", []) if isinstance(scraped.get("spec_table", []), list) else [],
        "scrape_status": clean_text(scraped.get("status", "")) if scrape_web else "disabled",
        "scrape_error": clean_text(scraped.get("error", "")),
        "params": params,
        "param_text": ", ".join(params),
        "is_accessory": is_acc,
        "accessory_reason": accessory_reason,
        "source_url": normalize_url(source_url),
        "image_url": image_url,
        "order_no": order_no,
        "internal_reference": pick(row, "Internal Reference", "Internal Reference", "costumiz", "SKU"),
        "gtin": gtin,
    }

def row_to_output(norm: Dict[str, Any], idx: int) -> Dict[str, Any]:
    out = {col: "" for col in OUTPUT_COLS}
    log_notes = []
    if not norm["name"]:
        log_notes.append("missing_name")
    if not norm.get("image_url"):
        log_notes.append("missing_image")
    if not norm.get("description"):
        log_notes.append("missing_short_description")
    if norm.get("is_accessory"):
        log_notes.append("detected_accessory_component")
    else:
        log_notes.append("detected_main_product")
    if norm.get("accessory_reason"):
        safe_reason = re.sub(r"[^a-z0-9_]+", "_", str(norm.get("accessory_reason", "")).lower()).strip("_")
        if safe_reason:
            log_notes.append(f"role_reason_{safe_reason}")
    if norm.get("description_origin"):
        log_notes.append(f"description_source_{norm.get('description_origin')}")
    if norm.get("scrape_status") and norm.get("scrape_status") not in {"ok", "disabled"}:
        log_notes.append(f"scrape_{norm.get('scrape_status')}")
    if norm.get("scrape_error"):
        log_notes.append(norm.get("scrape_error"))

    validation = "scraped_from_source_url" if norm.get("description_origin") in {"website", "website_cache"} else "fallback_from_excel_or_template_not_live_checked"
    quality = "pce_uji_generated_from_website_source" if norm.get("description_origin") in {"website", "website_cache"} else "pce_uji_generated_from_excel_fallback"

    spec_tab_content = generate_spec_tab(norm)
    if spec_tab_content:
        spec_tab_title = "SPESIFIKASI"
        spec_tab_priority = 20
        log_notes.append("official_spec_table_found")
    else:
        spec_tab_title = ""
        spec_tab_priority = ""
        log_notes.append("official_spec_table_not_found_spec_tab_skipped")

    out.update({
        "Name": norm["name"],
        "Brand": norm["brand"],
        "Product Description": generate_product_description(norm),
        "Product Short Description": generate_short_description(norm),
        "custom_tab_1_title": "Keunggulan Produk",
        "custom_tab_1_content": generate_feature_tab(norm),
        "custom_tab_1_priority": 10,
        "custom_tab_2_title": spec_tab_title,
        "custom_tab_2_content": spec_tab_content,
        "custom_tab_2_priority": spec_tab_priority,
        "custom_tab_3_title": "FAQ",
        "custom_tab_3_content": generate_faq(norm),
        "custom_tab_3_priority": 30,
        "product categories": norm["category"],
        "product tags": generate_tags(norm),
        "focus keyphrase": norm["name"],
        "meta description": meta_description(norm),
        "publication_date": (datetime(2025, 2, 8, 10, 0) + timedelta(hours=idx)).strftime("%d/%m/%Y %H:%M"),
        "Processed": "Yes",
        "Processing Time": "pce_uji_full_indonesia_pipeline_v6_spec_only_highlight_to_feature",
        "Content Quality": quality,
        "image_url": norm.get("image_url", ""),
        "Source URL": norm.get("source_url", ""),
        "Website Validation": validation,
        "Website Correction Log": "; ".join(log_notes),
    })
    return out


def style_workbook(path: str) -> None:
    if load_workbook is None:
        return
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "")
        if any(k in header.lower() for k in ["description", "content", "specifikasi", "correction"]):
            width = 58
        elif any(k in header.lower() for k in ["url", "image", "source"]):
            width = 44
        elif header.lower() in {"name", "focus keyphrase"}:
            width = 36
        else:
            width = min(max(len(header) + 4, 14), 30)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    for r in range(2, min(ws.max_row, 250) + 1):
        ws.row_dimensions[r].height = 80
    try:
        ref = ws.dimensions
        tab = Table(displayName="UJIProductsTable", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        pass
    wb.save(path)


def format_duration(seconds: float) -> str:
    seconds = max(0, int(seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def progress_eta(start_ts: float, done: int, total: int) -> str:
    elapsed = time.time() - start_ts
    if done <= 0:
        return f"elapsed={format_duration(elapsed)} eta=--:--"
    avg = elapsed / done
    remaining = max(0, total - done) * avg
    return f"elapsed={format_duration(elapsed)} eta={format_duration(remaining)} avg={avg:.1f}s/row"


def shorten_for_progress(text: Any, limit: int = 90) -> str:
    text = clean_text(text)
    if not text:
        return "-"
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= limit:
        return text
    return text[:limit - 1].rstrip() + "…"


def get_raw_progress_name(raw_row: Dict[str, Any], fallback_index: int) -> str:
    value = pick(raw_row, "Judul", "Title", "Name", "Product Name", "product_name")
    value = strip_jual(value)
    return shorten_for_progress(value or f"row_{fallback_index}")


def print_progress(enabled: bool, message: str) -> None:
    if enabled:
        print(message, flush=True)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Path file Excel input PCE")
    ap.add_argument("--sheet", default="", help="Nama sheet. Kosong = sheet pertama")
    ap.add_argument("--final-output", default="uji_pce_import_ready.xlsx")
    ap.add_argument("--limit", type=int, default=0, help="Batasi jumlah baris untuk testing. 0 = semua")
    ap.add_argument("--no-scrape-web", action="store_true", help="Matikan pengambilan deskripsi dari link website PCE")
    ap.add_argument("--scrape-cache", default="", help="Path cache JSON hasil scraping. Kosong = otomatis di sebelah output")
    ap.add_argument("--request-timeout", type=int, default=15, help="Timeout request website dalam detik")
    ap.add_argument("--request-delay", type=float, default=0.2, help="Jeda antar request website agar tidak terlalu agresif")
    ap.add_argument("--use-ai", action="store_true", help="Aktifkan generator konten AI DeepSeek untuk deskripsi, short description, keunggulan, FAQ, dan meta description")
    ap.add_argument("--deepseek-api-key", default=os.getenv("DEEPSEEK_API_KEY", ""), help="API key DeepSeek. Bisa juga memakai env DEEPSEEK_API_KEY")
    ap.add_argument("--deepseek-base-url", default=os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com"), help="Base URL DeepSeek/OpenAI-compatible API")
    ap.add_argument("--deepseek-model", default=os.getenv("DEEPSEEK_MODEL", "deepseek-chat"), help="Model DeepSeek yang digunakan")
    ap.add_argument("--ai-cache", default="", help="Path cache JSON hasil AI. Kosong = otomatis di sebelah output")
    ap.add_argument("--ai-limit", type=int, default=0, help="Batasi jumlah request AI baru. 0 = semua baris")
    ap.add_argument("--ai-timeout", type=int, default=60, help="Timeout request AI dalam detik")
    ap.add_argument("--ai-delay", type=float, default=0.0, help="Jeda antar request AI agar tidak terlalu agresif")
    ap.add_argument("--ai-temperature", type=float, default=0.25, help="Temperature AI untuk menjaga output tetap stabil")
    ap.add_argument("--progress-every", type=int, default=1, help="Tampilkan progress setiap N baris. Default 1 = setiap baris")
    ap.add_argument("--no-progress", action="store_true", help="Matikan tampilan progress per baris")
    ap.add_argument("--save-cache-every", type=int, default=10, help="Simpan cache tiap N baris agar aman jika proses dihentikan. 0 = hanya di akhir")
    args = ap.parse_args()

    if args.sheet:
        df = pd.read_excel(args.input, sheet_name=args.sheet)
    else:
        df = pd.read_excel(args.input)
    if args.limit and args.limit > 0:
        df = df.head(args.limit).copy()

    out_path = Path(args.final_output)
    cache_path = args.scrape_cache or str(out_path.with_suffix(".scrape_cache.json"))
    scrape_cache = load_scrape_cache(cache_path) if not args.no_scrape_web else None

    ai_cache_path = args.ai_cache or str(out_path.with_suffix(".deepseek_ai_cache.json"))
    ai_cache = load_scrape_cache(ai_cache_path) if args.use_ai else {}
    set_deepseek_ai_settings(
        enabled=args.use_ai,
        api_key=args.deepseek_api_key,
        base_url=args.deepseek_base_url,
        model=args.deepseek_model,
        cache=ai_cache,
        cache_path=ai_cache_path,
        limit=args.ai_limit,
        timeout=args.ai_timeout,
        delay=args.ai_delay,
        temperature=args.ai_temperature,
    )

    rows = []
    scraped_ok = 0
    scraped_failed = 0
    fallback_count = 0
    total_rows = len(df)
    progress_enabled = not args.no_progress
    progress_every = max(1, int(args.progress_every or 1))
    save_cache_every = max(0, int(args.save_cache_every or 0))
    start_ts = time.time()

    print_progress(
        progress_enabled,
        f"START generate | total_rows={total_rows} | use_ai={args.use_ai} | scrape_web={not args.no_scrape_web} | output={out_path}",
    )

    for row_number, (_, r) in enumerate(df.iterrows(), start=1):
        raw_row = r.to_dict()
        should_log = progress_enabled and (row_number == 1 or row_number == total_rows or row_number % progress_every == 0)
        raw_name = get_raw_progress_name(raw_row, row_number)
        print_progress(should_log, f"[{row_number}/{total_rows}] START | {raw_name}")

        norm = normalize_row(
            raw_row,
            scrape_web=not args.no_scrape_web,
            scrape_cache=scrape_cache,
            request_timeout=args.request_timeout,
            request_delay=args.request_delay,
        )
        if norm.get("description_origin") in {"website", "website_cache"}:
            scraped_ok += 1
        else:
            fallback_count += 1
            if norm.get("scrape_status") not in {"ok", "disabled", "no_url", ""}:
                scraped_failed += 1

        rows.append(row_to_output(norm, row_number - 1))

        ai_text = "ai=off"
        if args.use_ai:
            ai_state = norm.get("_deepseek_ai", {}) if isinstance(norm.get("_deepseek_ai", {}), dict) else {}
            if ai_state.get("ok"):
                ai_text = "ai=used"
            elif ai_state.get("from_cache"):
                ai_text = "ai=cache"
            else:
                reason = clean_text(ai_state.get("reason", "fallback")) or "fallback"
                ai_text = f"ai=fallback:{shorten_for_progress(reason, 36)}"

        done_name = shorten_for_progress(norm.get("name") or raw_name)
        print_progress(
            should_log,
            f"[{row_number}/{total_rows}] DONE  | {done_name} | desc={norm.get('description_origin', '-')} | scrape={norm.get('scrape_status', '-')} | {delivery_progress_status_v14(norm)} | {ai_text} | {progress_eta(start_ts, row_number, total_rows)}",
        )

        if save_cache_every and row_number % save_cache_every == 0:
            if scrape_cache is not None:
                save_scrape_cache(cache_path, scrape_cache)
            if args.use_ai:
                save_scrape_cache(ai_cache_path, AI_SETTINGS.get("cache", {}))
            print_progress(progress_enabled, f"[cache] saved after {row_number}/{total_rows}")

    if scrape_cache is not None:
        save_scrape_cache(cache_path, scrape_cache)
    if args.use_ai:
        save_scrape_cache(ai_cache_path, AI_SETTINGS.get("cache", {}))

    print_progress(progress_enabled, "[excel] writing output workbook...")
    out_df = pd.DataFrame(rows, columns=OUTPUT_COLS)
    dup_mask = out_df["Name"].astype(str).duplicated(keep=False) & (out_df["Name"].astype(str).str.strip() != "")
    if dup_mask.any():
        existing = out_df.loc[dup_mask, "Website Correction Log"].fillna("").astype(str)
        out_df.loc[dup_mask, "Website Correction Log"] = existing.apply(lambda x: (x + "; duplicate_name_check").strip("; "))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="UJI Products")
    style_workbook(str(out_path))
    print_progress(progress_enabled, f"[excel] saved | {out_path}")

    print("DONE", flush=True)
    print(f"input_rows={len(df)}", flush=True)
    print(f"output_rows={len(out_df)}", flush=True)
    print(f"final_output={out_path}", flush=True)
    print(f"total_elapsed={format_duration(time.time() - start_ts)}", flush=True)
    print(f"missing_image={int((out_df['image_url'].astype(str).str.strip() == '').sum())}", flush=True)
    accessory_rows = int(out_df['Website Correction Log'].astype(str).str.contains('detected_accessory_component', regex=False).sum())
    print(f"accessory_component_rows={accessory_rows}", flush=True)
    print(f"main_product_rows={len(out_df) - accessory_rows}", flush=True)
    print(f"scraped_from_website={scraped_ok}", flush=True)
    print(f"fallback_used={fallback_count}", flush=True)
    print(f"scrape_failed={scraped_failed}", flush=True)
    if scrape_cache is not None:
        print(f"scrape_cache={cache_path}", flush=True)
    if args.use_ai:
        print(f"deepseek_ai_enabled={AI_SETTINGS.get('enabled')}", flush=True)
        print(f"deepseek_model={AI_SETTINGS.get('model')}", flush=True)
        print(f"deepseek_ai_used={AI_STATS.get('used', 0)}", flush=True)
        print(f"deepseek_ai_cache_hits={AI_STATS.get('cache_hit', 0)}", flush=True)
        print(f"deepseek_ai_fallback={AI_STATS.get('fallback', 0)}", flush=True)
        print(f"deepseek_ai_cache={ai_cache_path}", flush=True)


# ---------------------------------------------------------------------------
# V7 override: stricter Indonesian cleanup for public specification tables and
# feature highlights. Product names/SKU are intentionally not modified.
# ---------------------------------------------------------------------------

PUBLIC_ENGLISH_BLOCK_RE = re.compile(
    r"\b(?:range|measurement|measuring|weighing|readability|interface|output|input|display|"
    r"without|with|from|up to|accuracy|resolution|power|battery|current clamp|clamp|adapter|module|"
    r"software|accessory|spare|certificate|calibration|automatic|waterproof|water resistant|"
    r"portable|compact|warning|baud rate|sample rate|data rate|memory|stainless steel|"
    r"pressure measurement|humidity measurement|temperature measurement|expansion|plug[- ]?in|"
    r"communication|automatically|detected|graphic|opening|harmonic|distortion|order|"
    r"suitable|compatible|support|supports|probe|electrode|feature|features|function|"
    r"external|included|including|high|low|sampling|rate|screen|phase|phases|"
    r"operating|storage|conditions|medium|capacity|separate|summari[sz]ed|records?|"
    r"preset|characteristic|curves?|wood types?|easy[- ]?to[- ]?read|backlit|symbols?)\b",
    re.I,
)

SPEC_LABEL_TRANSLATIONS_V7: List[Tuple[str, str]] = [
    (r"^voltage measurement$", "Pengukuran tegangan"),
    (r"^current measurement$", "Pengukuran arus"),
    (r"^active power(?:\s+p)?$", "Daya aktif P"),
    (r"^apparent power(?:\s+s)?$", "Daya semu S"),
    (r"^reactive power(?:\s+q)?$", "Daya reaktif Q"),
    (r"^power factor(?:\s*\(cos\s*φ\))?$", "Faktor daya (cos φ)"),
    (r"^measurement of real power.*$", "Pengukuran daya aktif"),
    (r"^measurement of apparent power.*$", "Pengukuran daya semu"),
    (r"^measurement range up to$", "Rentang pengukuran"),
    (r"^measurement range$|^measuring range$", "Rentang pengukuran"),
    (r"^measurement range\s+(.+)$|^measuring range\s+(.+)$", r"Rentang pengukuran \1"),
    (r"^moisture measurement range$|^moisture measuring range$", "Rentang pengukuran kadar air"),
    (r"^humidity measurement range$", "Rentang pengukuran kelembapan"),
    (r"^temperature measurement range$", "Rentang pengukuran suhu"),
    (r"^frequency range$", "Rentang frekuensi"),
    (r"^adjustment range$", "Rentang penyesuaian"),
    (r"^accuracy$", "Akurasi"),
    (r"^resolution$", "Resolusi"),
    (r"^repeatability$", "Repeatabilitas"),
    (r"^linearity$", "Linearitas"),
    (r"^readability$", "Keterbacaan"),
    (r"^number of different wood types$", "Jumlah jenis kayu"),
    (r"^wood temperature range$", "Rentang suhu kayu"),
    (r"^temperature range$", "Rentang suhu"),
    (r"^temperature compensation$", "Kompensasi suhu"),
    (r"^operating conditions$|^operating condition$", "Kondisi operasi"),
    (r"^operating temperature$|^operating temperature range$", "Suhu operasional"),
    (r"^storage conditions$|^storage condition$", "Kondisi penyimpanan"),
    (r"^storage temperature$|^storage temperature range$", "Suhu penyimpanan"),
    (r"^display$", "Tampilan"),
    (r"^display type$", "Jenis tampilan"),
    (r"^display size$", "Ukuran tampilan"),
    (r"^display refresh rate$", "Laju refresh tampilan"),
    (r"^electrode dimensions?$", "Dimensi elektroda"),
    (r"^probe dimensions?$", "Dimensi sonda"),
    (r"^sensor dimensions?$", "Dimensi sensor"),
    (r"^power supply$", "Catu daya"),
    (r"^power$", "Daya"),
    (r"^battery life$", "Daya tahan baterai"),
    (r"^\(?rechargeable\)? battery$|^battery$|^batteries$", "Baterai"),
    (r"^unit dimensions?$|^device dimensions?$|^dimensions?$", "Dimensi unit"),
    (r"^unit weight$|^weight$", "Berat unit"),
    (r"^interface$|^interfaces$", "Antarmuka"),
    (r"^data interface$", "Antarmuka data"),
    (r"^memory$|^data memory$", "Memori data"),
    (r"^storage medium$", "Media penyimpanan"),
    (r"^storage capacity$", "Kapasitas penyimpanan"),
    (r"^storage interval of$", "Interval penyimpanan mulai"),
    (r"^storage interval to$", "Interval penyimpanan hingga"),
    (r"^memory capacity\s*\(additional information\)$", "Kapasitas memori (informasi tambahan)"),
    (r"^input$|^inputs$", "Masukan"),
    (r"^input current$", "Arus masukan"),
    (r"^output$|^outputs$", "Keluaran"),
    (r"^signal output$", "Keluaran sinyal"),
    (r"^measurement rate$|^measuring rate$|^sampling rate$", "Laju pengukuran"),
    (r"^measurement time$|^measuring time$", "Waktu pengukuran"),
    (r"^measurement functions?$", "Fungsi pengukuran"),
    (r"^number of measurement channels$", "Jumlah kanal pengukuran"),
    (r"^protection class(?:\s*\(device\))?$|^protection rating$", "Kelas perlindungan (perangkat)"),
    (r"^material$", "Material"),
    (r"^capacity$|^weighing range$", "Rentang timbang"),
]

SPEC_VALUE_REPLACEMENTS_V7: List[Tuple[str, str]] = [
    (r"&gt;", ">"),
    (r"&lt;", "<"),
    (r"\bApprox\.?\b|\bapproximately\b", "Sekitar"),
    (r"\bmax\.\b", "Maks."),
    (r"\bmin\.\b", "Min."),
    (r"\bRMS value\b", "nilai RMS"),
    (r"\breal value of power\b", "nilai daya riil"),
    (r"\bof the measured value\b|\bof measured value\b|\bof the value\b|\bof value\b", "dari nilai"),
    (r"\bof Rd\b|\bof\s*R\.D\.\b|\bof reading\b", "dari pembacaan"),
    (r"\bof Mw\.?\b|\bof Mv\.?\b|\bv\.Mw\.\b", "dari nilai ukur"),
    (r"\bmeasured value\b", "nilai terukur"),
    (r"\bdigits?\b|\bdgt\b", "digit"),
    (r"\bin\s+(\d+)\s+measurement ranges?\b", r"dalam \1 rentang pengukuran"),
    (r"\baccording\s+to\s+the\s+measurement\s+range\b", "sesuai rentang pengukuran"),
    (r"\bmeasurement ranges?\b", "rentang pengukuran"),
    (r"\bmeasurement range\b|\bmeasuring range\b", "rentang pengukuran"),
    (r"\bhumidity measurement\b", "pengukuran kelembapan"),
    (r"\btemperature measurement\b", "pengukuran suhu"),
    (r"\bmoisture content\b", "kadar air"),
    (r"\bwood types\b", "jenis kayu"),
    (r"\bgrain types\b", "jenis biji-bijian"),
    (r"\bstandardized grain types\b", "jenis biji-bijian standar"),
    (r"\bmeasurements\b", "pengukuran"),
    (r"\bmeasurement\b", "pengukuran"),
    (r"\baverage value\b", "nilai rata-rata"),
    (r"\bInternal memory\b", "memori internal"),
    (r"\bExternal memory\b", "memori eksternal"),
    (r"\bSeparate phases\b", "fase terpisah"),
    (r"\bSummari[sz]ed phases\b", "fase gabungan"),
    (r"\bdata records?\b", "rekaman data"),
    (r"\bdata sets?\b", "kumpulan data"),
    (r"\bstorage medium\b", "media penyimpanan"),
    (r"\bstorage capacity\b", "kapasitas penyimpanan"),
    (r"\boperating conditions\b", "kondisi operasi"),
    (r"\bstorage conditions\b", "kondisi penyimpanan"),
    (r"\bprotection class\b", "kelas perlindungan"),
    (r"\btemperature compensation\b", "kompensasi suhu"),
    (r"\bautomatic power[- ]?off\b", "mati otomatis"),
    (r"\bautomatically\b", "otomatis"),
    (r"\bautomatic\b", "otomatis"),
    (r"\bwithin\b", "pada rentang"),
    (r"\babove\b", "di atas"),
    (r"\bbelow\b", "di bawah"),
    (r"\bfor standardized grain types\b", "untuk jenis biji-bijian standar"),
    (r"\bfor\s*<\s*10\s*%", "untuk <10%"),
    (r"\bfor\s*>\s*10\s*%", "untuk >10%"),
    (r"\bfor\b", "untuk"),
    (r"\band\b", "dan"),
    (r"\bwith\b", "dengan"),
    (r"\band\b", "dan"),
    (r"\bor\b", "atau"),
    (r"\bwith\b", "dengan"),
    (r"\bwithout\b", "tanpa"),
    (r"\bfrom\b", "dari"),
    (r"\bto\b", "hingga"),
    (r"\bsecond\b|\bseconds\b", "detik"),
    (r"\bmains adaptor\b", "adaptor listrik"),
    (r"\bblock battery\b", "baterai blok"),
    (r"\bbattery powered\b", "menggunakan baterai"),
    (r"\bbatteries\b", "baterai"),
    (r"\bbattery\b", "baterai"),
    (r"\bAlkali[- ]?manganese\b", "alkali-mangan"),
    (r"\bmulti functional LCD\b|\bmultifunctional LCD\b", "LCD multifungsi"),
    (r"\bdisplay\b", "layar"),
    (r"\bdiameter\b", "diameter"),
    (r"\bprobe\b", "sonda"),
    (r"\belectrode\b", "elektroda"),
    (r"\binterface\b", "antarmuka"),
    (r"\boutput\b", "keluaran"),
    (r"\binput\b", "masukan"),
    (r"\bpower supply\b", "catu daya"),
    (r"\bpower\b", "daya"),
    (r"\btypically between\b|\bis typically between\b", "biasanya antara"),
    (r"\br\.h\.\b", "RH"),
    (r"\binch\b", "inci"),
]


def _apply_replacements_v7(s: str, replacements: List[Tuple[str, str]]) -> str:
    for pat, repl in replacements:
        s = re.sub(pat, repl, s, flags=re.I)
    return s


def normalize_public_phrase(label: str) -> str:
    """Final cleanup for public parameter/highlight/spec text. Product names are not passed here."""
    s = clean_text(label)
    if not s:
        return ""
    s = html.unescape(s)
    s = re.sub(r"\s+/\s*s\b", "/detik", s, flags=re.I)
    s = re.sub(r"\s+\.\.\.\s+", " ... ", s)
    s = _apply_replacements_v7(s, TECH_PHRASE_REPLACEMENTS)
    s = _apply_replacements_v7(s, SPEC_VALUE_REPLACEMENTS_V7)

    cleanup_pairs = [
        (r"\bRentang pengukuran up hingga\b|\bRentang pengukuran up to\b", "Rentang pengukuran"),
        (r"\brentang pengukuran dari\s+(-?\d)", r"rentang pengukuran \1"),
        (r"\brentang dari\s+(-?\d)", r"rentang \1"),
        (r"\bmoisture pengukuran rentang\b", "rentang pengukuran kadar air"),
        (r"\bpengukuran kelembapan pengukuran\b", "pengukuran kelembapan"),
        (r"\bpengukuran suhu pengukuran\b", "pengukuran suhu"),
        (r"\bfrequency rentang\b", "rentang frekuensi"),
        (r"\bpengukuran rate\b", "laju pengukuran"),
        (r"\btampilan size\b", "ukuran tampilan"),
        (r"\btampilan refresh rate\b", "laju refresh tampilan"),
        (r"\bstorage medium\b", "media penyimpanan"),
        (r"\bstorage capacity\b", "kapasitas penyimpanan"),
        (r"\bstorage interval of\b", "interval penyimpanan mulai"),
        (r"\bstorage interval to\b", "interval penyimpanan hingga"),
        (r"\boperating conditions\b", "kondisi operasi"),
        (r"\bstorage conditions\b", "kondisi penyimpanan"),
        (r"\btemperature compensation\b|\bsuhu compensation\b", "kompensasi suhu"),
        (r"\bprotection class\s*\(device\)\b", "kelas perlindungan (perangkat)"),
        (r"\bnumber of pengukuran channels\b", "jumlah kanal pengukuran"),
        (r"\bpengukuran functions\b", "fungsi pengukuran"),
        (r"\bpengukuran time\b", "waktu pengukuran"),
        (r"\bpower factor\b", "faktor daya"),
        (r"\bactive daya\b", "daya aktif"),
        (r"\bapparent daya\b", "daya semu"),
        (r"\breactive daya\b", "daya reaktif"),
        (r"\bpower\b", "daya"),
        (r"\bvoltage\b", "tegangan"),
        (r"\bcurrent\b", "arus"),
        (r"\branges\b", "rentang"),
        (r"\brange\b", "rentang"),
        (r"\bmeasurement\b", "pengukuran"),
        (r"\bmeasuring\b", "pengukuran"),
        (r"\bof Rd\b", "dari pembacaan"),
        (r"\bof Mw\.?\b", "dari nilai ukur"),
        (r"\bof the value\b", "dari nilai"),
        (r"\baccording hingga the pengukuran rentang\b", "sesuai rentang pengukuran"),
        (r"\bin 5 pengukuran ranges\b", "dalam 5 rentang pengukuran"),
        (r"\b(\d+)\s+dan\s+(\d+)\s+fase\b", r"\1 dan \2 fase"),
        (r"\bwood types\b", "jenis kayu"),
        (r"\bgrain types\b", "jenis biji-bijian"),
        (r"\bPreset characteristic curves untuk (\d+) jenis kayu\b", r"Kurva karakteristik bawaan untuk \1 jenis kayu"),
        (r"\bPreset characteristic curves untuk (\d+) wood types\b", r"Kurva karakteristik bawaan untuk \1 jenis kayu"),
        (r"\bEasy[- ]?hingga[- ]?read LCD layar\b|\bEasy[- ]?to[- ]?read LCD layar\b", "Layar LCD mudah dibaca"),
        (r"\bBaterai powered\b", "Menggunakan baterai"),
        (r"\bHigh[- ]?contrast\b", "kontras tinggi"),
        (r"\bbacklit\b", "lampu latar"),
        (r"\bMeasured value memori\b", "memori nilai terukur"),
        (r"\bSymbols for\b", "simbol untuk"),
        (r"\bmaintenance\b", "pemeliharaan"),
        (r"\bphases?\b", "fase"),
        (r"\bprobe\b", "sonda"),
        (r"\bclamp\b", "tang"),
        (r"\bfrom\b", "dari"),
        (r"\bto\b", "hingga"),
    ]
    for pat, repl in cleanup_pairs:
        s = re.sub(pat, repl, s, flags=re.I)

    # Numeric/unit cleanup.
    s = re.sub(r"\b(\d+)\s+(\d+(?:[,.]\d+)?)\s*V\b", r"\1 × \2 V", s, flags=re.I)
    if re.fullmatch(r"\d+\s*×\s*\d+(?:[,.]\d+)?\s*V", s, flags=re.I):
        s = "Catu daya " + s
    s = re.sub(r"(?<=\d)\.(?=\d\s*(?:V|A|mA|W|kW|N|kbit|Mbit|bit))", ",", s)
    s = re.sub(r"(?<!\d)(\d{5,})(?!\d)", lambda m: f"{int(m.group(1)):,}".replace(",", "."), s)
    s = re.sub(r"\s*/\s*", " / ", s)
    s = re.sub(r"\s+([,.;:])", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


def translate_spec_label(label: str) -> str:
    raw = clean_text(label).strip(" :-–—")
    if not raw:
        return ""
    low = re.sub(r"\s+", " ", raw.lower())
    low = re.sub(r"\bup to\b", "up to", low)
    for pattern, repl in SPEC_LABEL_TRANSLATIONS_V7:
        if re.search(pattern, low, flags=re.I):
            out = re.sub(pattern, repl, low, flags=re.I).strip()
            return normalize_public_phrase(out)
    out = raw
    label_replacements = [
        (r"\bvoltage\b", "tegangan"),
        (r"\bcurrent\b", "arus"),
        (r"\bmeasurement\b|\bmeasuring\b", "pengukuran"),
        (r"\brange\b", "rentang"),
        (r"\bfrequency\b", "frekuensi"),
        (r"\btemperature\b", "suhu"),
        (r"\bhumidity\b", "kelembapan"),
        (r"\bmoisture\b", "kadar air"),
        (r"\baccuracy\b", "akurasi"),
        (r"\bresolution\b", "resolusi"),
        (r"\bdisplay\b", "tampilan"),
        (r"\btype\b", "jenis"),
        (r"\bsize\b", "ukuran"),
        (r"\brefresh rate\b", "laju refresh"),
        (r"\bdimensions?\b", "dimensi"),
        (r"\bweight\b", "berat"),
        (r"\bpower supply\b", "catu daya"),
        (r"\bpower\b", "daya"),
        (r"\bbattery\b", "baterai"),
        (r"\bwood\b", "kayu"),
        (r"\btypes?\b", "jenis"),
        (r"\belectrode\b", "elektroda"),
        (r"\bprobe\b", "sonda"),
        (r"\bunit\b|\bdevice\b", "unit"),
        (r"\binterface\b", "antarmuka"),
        (r"\binput\b", "masukan"),
        (r"\boutput\b", "keluaran"),
        (r"\bmemory\b", "memori"),
        (r"\bstorage\b", "penyimpanan"),
        (r"\bmedium\b", "media"),
        (r"\bcapacity\b", "kapasitas"),
        (r"\bconditions?\b", "kondisi"),
        (r"\boperating\b", "operasi"),
        (r"\bprotection class\b", "kelas perlindungan"),
        (r"\bchannels?\b", "kanal"),
        (r"\bfunctions?\b", "fungsi"),
    ]
    for pat, rep in label_replacements:
        out = re.sub(pat, rep, out, flags=re.I)
    return normalize_public_phrase(out)


def translate_spec_value(value: str, label: str = "") -> str:
    s = clean_text(value).strip()
    if not s:
        return ""
    s = normalize_public_phrase(s)
    # Improve common mixed fragments after the generic cleanup.
    extra = [
        (r"\b±\s*([0-9,.]+)\s*%\s+dari\s+nilai\s*\+\s*([0-9,.]+)\s*digit\b", r"±\1 % dari nilai + \2 digit"),
        (r"\b([><])\s*([0-9,.]+)\s*V\s+y\s*([><])\s*([0-9,.]+)\s*A\b", r"\1\2 V dan \3\4 A"),
        (r"\b([><])\s*([0-9,.]+)\s*V\s+dan\s*([><])\s*([0-9,.]+)\s*A\b", r"\1\2 V dan \3\4 A"),
        (r"\b0,1\s*W\s*\.\.\.\s*1\s*kW\s+sesuai\s+rentang\s+pengukuran\b", "0,1 W ... 1 kW sesuai rentang pengukuran"),
        (r"\b0,1\s*VA\s*\.\.\.\s*1\s*kVA\s+sesuai\s+rentang\s+pengukuran\b", "0,1 VA ... 1 kVA sesuai rentang pengukuran"),
        (r"\b0,1\s*VAr\s*\.\.\.\s*1\s*kVAr\s+sesuai\s+rentang\s+pengukuran\b", "0,1 VAr ... 1 kVAr sesuai rentang pengukuran"),
        (r"\bdi atas\s+(-?\d)", r"di atas \1"),
        (r"\bdi bawah\s+(-?\d)", r"di bawah \1"),
    ]
    for pat, repl in extra:
        s = re.sub(pat, repl, s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


def translate_technical_phrase(text: str) -> str:
    """Translate short technical phrases into Indonesian and drop unfinished fragments."""
    s = clean_text(text).strip(" .;:,-–—")
    if not s:
        return ""
    return normalize_public_phrase(s)


def has_untranslated_public_english(text: str) -> bool:
    s = clean_text(text)
    if not s:
        return False
    # Allowed technical abbreviations/acronyms. Remove them before detecting leftover English words.
    s = re.sub(r"\b(?:PCE|USB|RS[- ]?232|LCD|LED|AC|DC|RMS|IP\d{2}|ATEX|ISO|DIN|VAR|VAr|kVAR|kVAr|VA|kVA|Hz|kHz|Mbit|kbit|RH|MIN|MAX|HOLD)\b", " ", s, flags=re.I)
    return PUBLIC_ENGLISH_BLOCK_RE.search(s) is not None


def public_highlight_label(line: str) -> str:
    """Convert PCE highlight snippets into Indonesian-only catalog bullets."""
    original = clean_text(line).strip(" .")
    if not original:
        return ""
    low_norm = re.sub(r"\s+", " ", original.lower())

    special = [
        (r"preset characteristic curves.*?(\d+)\s+wood types", r"Kurva karakteristik bawaan untuk \1 jenis kayu"),
        (r"easy[- ]?to[- ]?read.*lcd", "Layar LCD mudah dibaca"),
        (r"battery powered", "Menggunakan baterai"),
        (r"high[- ]?contrast.*backlit.*display", "Layar kontras tinggi dengan lampu latar"),
        (r"measured value memory", "Memori nilai terukur"),
        (r"symbols? for", "Mencantumkan simbol status pada tampilan"),
        (r"automatic temperature compensation", "Kompensasi suhu otomatis"),
        (r"temperature compensation", "Kompensasi suhu"),
        (r"auto shut", "Mati otomatis untuk membantu menghemat baterai"),
        (r"water resistant|waterproof", "Tahan air sesuai informasi produk"),
        (r"without display", "Tanpa layar"),
        (r"min\.?\s*/?\s*max|hold", "Fungsi nilai minimum/maksimum dan tahan baca sesuai informasi produk"),
    ]
    for pat, repl in special:
        m = re.search(pat, low_norm, flags=re.I)
        if m:
            try:
                out = m.expand(repl)
            except Exception:
                out = repl
            return normalize_public_phrase(out)

    if re.search(r"\bATEX\b", original, flags=re.I):
        return "Mencantumkan informasi ATEX sesuai data produk"
    if re.search(r"\bIP\s?\d{2}\b", original, flags=re.I):
        ip = re.search(r"\bIP\s?\d{2}\b", original, flags=re.I).group(0).upper().replace(" ", "")
        return f"Mencantumkan perlindungan lingkungan dengan rating {ip}"

    translated = normalize_public_phrase(original)
    if not translated:
        return ""
    if len(translated.split()) > 18:
        return ""
    if is_weak_public_parameter(translated, original):
        return ""
    if has_untranslated_public_english(translated):
        return ""
    if re.search(r"\d|%|°C|ppm|ppb|bar|mbar|Pa\b|psi|dB|lux|rpm|m/s|Hz|bit|IP\d|ATEX|USB|RS-?232", translated, flags=re.I):
        return translated
    if re.search(r"\b(?:pengukuran|kompensasi|antarmuka|keluaran|masukan|alarm|peringatan|layar|baterai|memori|portabel|ringkas|tahan air|adaptor|kabel|modul|sertifikat|kalibrasi|sonda|elektroda|sensor|tang|fase|distorsi|harmonik|kurva|karakteristik|kayu)\b", translated, flags=re.I):
        return translated
    return ""


def public_param_text(params: List[str]) -> str:
    labels: List[str] = []
    for param in params:
        p = clean_text(param)
        if not p:
            continue
        label = public_highlight_label(p) or translate_technical_phrase(p)
        label = normalize_public_phrase(label)
        if not label:
            continue
        if is_weak_public_parameter(label, p):
            continue
        if has_untranslated_public_english(label):
            continue
        if label.lower() not in [x.lower() for x in labels]:
            labels.append(label)
    return ", ".join(labels)


def spec_table_html(rows: List[Dict[str, str]]) -> str:
    """Render only official PCE spec rows, with final Indonesian cleanup."""
    clean_rows = dedupe_spec_rows(rows)
    if not clean_rows:
        return ""
    body: List[str] = []
    for row in clean_rows:
        label = translate_spec_label(row.get("label", ""))
        value = translate_spec_value(row.get("value", ""), label)
        if not label or not value:
            continue
        # Do not display rows that are obviously navigation/marketing fragments.
        if re.search(r"\b(?:add to cart|delivery time|question|callback|price|copyright)\b", label + " " + value, flags=re.I):
            continue
        body.append(f"<tr><th>{esc(label)}</th><td>{esc(value)}</td></tr>")
    if not body:
        return ""
    return '<table class="uji-spec-table"><tbody>' + ''.join(body) + '</tbody></table>'

# V7.1 final cleanup pass for fragments created by mixed English/Indonesian replacements.
_normalize_public_phrase_v7 = normalize_public_phrase

def normalize_public_phrase(label: str) -> str:
    s = _normalize_public_phrase_v7(label)
    if not s:
        return ""
    fixes = [
        (r"\bInternal memori\b", "Memori internal"),
        (r"\bExternal memori\b", "Memori eksternal"),
        (r"\bSeparate fase\b", "Fase terpisah"),
        (r"\bSummari[sz]ed fase\b", "Fase gabungan"),
        (r"\bMoisture rentang pengukuran\b", "Rentang pengukuran kadar air"),
        (r"\bMoisture pengukuran rentang\b", "Rentang pengukuran kadar air"),
        (r"\bSuhu compensation\b", "Kompensasi suhu"),
        (r"\bAdjustment rentang\b", "Rentang penyesuaian"),
        (r"\bFrequency rentang\b", "Rentang frekuensi"),
        (r"\bPengukuran rate\b", "Laju pengukuran"),
        (r"\bStorage medium\b", "Media penyimpanan"),
        (r"\bStorage capacity\b", "Kapasitas penyimpanan"),
        (r"\bOperating conditions\b", "Kondisi operasi"),
        (r"\bStorage conditions\b", "Kondisi penyimpanan"),
        (r"\bAutomatic power-off\b", "Mati otomatis"),
        (r"\bProtection class\b", "Kelas perlindungan"),
        (r"\bstandardized jenis biji-bijian\b", "jenis biji-bijian standar"),
        (r"\bstandardized grain jenis\b", "jenis biji-bijian standar"),
        (r"\br\.h\.\b|\brh\.\b", "RH"),
        (r"\bMax\.\s*", "Maks. "),
        (r"\bMin\.\s*", "Min. "),
        (r"\bEasy[- ]?hingga[- ]?read LCD tampilan\b", "Layar LCD mudah dibaca"),
        (r"\bEasy[- ]?hingga[- ]?read LCD layar\b", "Layar LCD mudah dibaca"),
        (r"\bPreset characteristic curves untuk (\d+) kayu jenis\b", r"Kurva karakteristik bawaan untuk \1 jenis kayu"),
        (r"\bPreset characteristic curves untuk (\d+) jenis kayu\b", r"Kurva karakteristik bawaan untuk \1 jenis kayu"),
        (r"\bMeasured value memori\b", "Memori nilai terukur"),
        (r"\bHigh[- ]?contrast\b", "Kontras tinggi"),
        (r"\bbacklit\b", "lampu latar"),
        (r"\bAverage value\b", "Nilai rata-rata"),
        (r"\bMIN, MAX, tahan baca, nilai rata-rata\b", "MIN, MAX, HOLD, nilai rata-rata"),
        (r"\bdi atas\s+(-?\d)", r"di atas \1"),
        (r"\bdi bawah\s+(-?\d)", r"di bawah \1"),
        (r"\bnilai ukur\.\+", "nilai ukur +"),
        (r"\bnilai\.\+", "nilai +"),
    ]
    for pat, repl in fixes:
        s = re.sub(pat, repl, s, flags=re.I)
    s = re.sub(r"\s+\.\.\.\s+", " ... ", s)
    s = re.sub(r"\.{3}", "...", s)
    s = re.sub(r"\s+([,.;:])", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s

# V7.2 polishing: decimal comma, RH, and ellipsis spacing for public spec text.
_normalize_public_phrase_v71 = normalize_public_phrase

def normalize_public_phrase(label: str) -> str:
    s = _normalize_public_phrase_v71(label)
    if not s:
        return ""
    s = re.sub(r"r\.h\.", "RH", s, flags=re.I)
    s = re.sub(r"(?<=\d)\.(?=\d{1,2}(?:\D|$))", ",", s)
    s = re.sub(r"(\d)\.\.\.\s*(\d)", r"\1 ... \2", s)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s




# V10: Strict Delivery Scope support + dedupe delivery items + FAQ last.
# Mapping:
# - PCE Highlights/Features -> custom_tab_1 Keunggulan Produk
# - PCE Specification/Technical Data -> custom_tab_2 SPESIFIKASI, only when official table exists
# - PCE Delivery Scope/Scope of Delivery -> custom_tab_4 Kelengkapan Pengiriman, only when official content exists
#
# Penting: Delivery Scope dibuat ketat. Kalau yang ditemukan ternyata area Downloads,
# tabel spesifikasi, atau tab navigation, hasil dikosongkan. Lebih aman tidak membuat
# tab Kelengkapan Pengiriman daripada mengisi tab itu dengan data yang salah.

DELIVERY_SCOPE_STOP_HEADINGS = {
    "description", "highlights", "features", "specification", "specifications", "technical data",
    "technical specifications", "accessories", "similar products", "related products", "downloads",
    "manuals", "information", "help", "company", "social media", "reviews", "shipping and delivery",
    "product description", "technical specification", "technical specifications", "manual", "data sheet",
}

DELIVERY_SCOPE_HEADING_TEXTS = {
    "delivery scope", "scope of delivery", "delivery contents", "delivery content",
    "standard delivery", "package contents", "package content", "included in delivery",
    "what is included", "contents of delivery", "supplied with", "items supplied",
}

# Jangan pakai kata "included" sendirian sebagai heading. Di halaman PCE kata itu sering
# muncul dalam tabel spesifikasi seperti "battery included" lalu extractor salah lompat
# ke bagian spesifikasi/download.

DELIVERY_BAD_ITEM_RE = re.compile(
    r"\b(?:downloads?|technical specifications?|technical data|specifications?|description|highlights?|features?|"
    r"measurement range|measuring range|accuracy|resolution|display type|display size|storage capacity|"
    r"storage medium|storage interval|power factor|phase angle|active power|reactive power|apparent power|"
    r"frequency|ac voltage|dc voltage|alternating current|temperature range|operating conditions|"
    r"protection class|response time|emissivity|wavelength|input current|weight\s*-|dimensions\s*-|"
    r"signal output|alarm/control|norms|sensor type|gas sensor type|type of measurement|"
    r"general technical data|cereal\s*-|repeatability|adjustment range)\b",
    re.I,
)

DELIVERY_GOOD_ITEM_RE = re.compile(
    r"(?:\b\d+\s*(?:x|×)\b|manual|instructions?|battery|batteries|case|bag|carrying|transport|"
    r"power supply|adapter|charger|usb|cable|software|certificate|probe|sensor|electrode|test lead|"
    r"holder|tripod|clamp|screwdriver|key|tool|set|pce[-\s]?[a-z0-9])",
    re.I,
)


def _normalise_heading_key(text: str) -> str:
    t = clean_text(text).strip().lower()
    t = re.sub(r"\s+", " ", t)
    t = t.strip(" :;-–—")
    return t


def _is_delivery_heading_text(text: str) -> bool:
    return _normalise_heading_key(text) in DELIVERY_SCOPE_HEADING_TEXTS


def _is_stop_heading_text(text: str) -> bool:
    return _normalise_heading_key(text) in DELIVERY_SCOPE_STOP_HEADINGS


def _looks_like_delivery_heading_tag(tag: Any) -> bool:
    if tag is None:
        return False
    txt = soup_text(tag).strip()
    if not txt or len(txt) > 90 or not _is_delivery_heading_text(txt):
        return False
    name = str(getattr(tag, "name", "") or "").lower()
    attrs = " ".join([
        str(tag.get("id") or ""), str(tag.get("class") or ""), str(tag.get("role") or ""),
        str(tag.get("href") or ""), str(tag.get("data-bs-target") or ""), str(tag.get("data-target") or ""),
        str(tag.get("aria-controls") or ""),
    ]).lower()
    # Heading asli atau tab link/button. Hindari span/div biasa di dalam konten teknis.
    if re.fullmatch(r"h[1-6]", name):
        return True
    if name in {"a", "button"}:
        return True
    if name == "li" and "tab" in attrs:
        return True
    if name in {"div", "span"} and re.search(r"\btab\b|delivery|scope", attrs):
        return True
    return False


def _panel_id_from_heading(tag: Any) -> str:
    if tag is None:
        return ""
    vals = [
        clean_text(tag.get("href") or ""),
        clean_text(tag.get("data-bs-target") or ""),
        clean_text(tag.get("data-target") or ""),
        clean_text(tag.get("aria-controls") or ""),
    ]
    # Kadang href ada di child anchor ketika heading tag-nya li.
    try:
        a = tag.find("a")
        if a:
            vals.extend([
                clean_text(a.get("href") or ""), clean_text(a.get("data-bs-target") or ""),
                clean_text(a.get("data-target") or ""), clean_text(a.get("aria-controls") or ""),
            ])
    except Exception:
        pass
    for val in vals:
        if not val:
            continue
        if val.startswith("#") and len(val) > 1:
            return val[1:]
        if re.fullmatch(r"[A-Za-z][\w\-:.]*", val):
            return val
    return ""


def _find_linked_delivery_panel(soup: Any, heading: Any) -> Any:
    panel_id = _panel_id_from_heading(heading)
    if panel_id:
        try:
            node = soup.find(id=panel_id)
            if node is not None and node is not heading:
                return node
        except Exception:
            pass
    return None


def _delivery_item_cleanup(text: str) -> str:
    """Clean one Delivery Scope item without changing official product/model names."""
    s = clean_text(text).strip(" -–—•.;:")
    if not s:
        return ""
    # Drop obvious navigation / spec / footer fragments.
    if DELIVERY_BAD_ITEM_RE.search(s):
        return ""
    if re.search(r"\b(?:add to cart|delivery time|question|callback|price|vat|copyright|technical hotline|similar products|related products)\b", s, flags=re.I):
        return ""
    replacements = [
        (r"\bincl\.\b", "termasuk"),
        (r"\bincluded in delivery\b", "termasuk dalam pengiriman"),
        (r"\bincluded\b", "termasuk"),
        (r"\boperating instructions?\b", "petunjuk penggunaan"),
        (r"\buser manual\b", "manual pengguna"),
        (r"\bmanual\b", "manual"),
        (r"\bcarrying case\b", "tas pembawa"),
        (r"\btransport case\b", "koper pembawa"),
        (r"\bprotective case\b", "kotak pelindung"),
        (r"\bpower supply\b", "catu daya"),
        (r"\bpower adapter\b", "adaptor daya"),
        (r"\bmains adapter\b", "adaptor listrik"),
        (r"\badapter\b", "adaptor"),
        (r"\bbattery charger\b", "pengisi daya baterai"),
        (r"\bbatteries\b", "baterai"),
        (r"\bbattery\b", "baterai"),
        (r"\bUSB cable\b", "kabel USB"),
        (r"\bdata cable\b", "kabel data"),
        (r"\binterface cable\b", "kabel antarmuka"),
        (r"\bcable\b", "kabel"),
        (r"\bsoftware\b", "perangkat lunak"),
        (r"\bISO calibration certificate\b", "sertifikat kalibrasi ISO"),
        (r"\bcalibration certificate\b", "sertifikat kalibrasi"),
        (r"\bprobe\b", "sonda"),
        (r"\belectrode\b", "elektroda"),
        (r"\btest leads?\b", "kabel uji"),
        (r"\btest cable\b", "kabel uji"),
        (r"\bclamp\b", "tang"),
        (r"\bholder\b", "dudukan"),
        (r"\btripod\b", "tripod"),
        (r"\bcase\b", "kotak"),
        (r"\bset\b", "set"),
        (r"\bpieces?\b", "unit"),
    ]
    for pat, rep in replacements:
        s = re.sub(pat, rep, s, flags=re.I)
    # Common quantity formatting: 1 x -> 1 ×
    s = re.sub(r"\b(\d+)\s*x\s*", r"\1 × ", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -–—•.;:")
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


def _strip_delivery_prefix(text: str) -> str:
    """Remove heading-like prefixes that sometimes get included in the first delivery item."""
    s = clean_text(text).strip(" -–—•.;:")
    s = re.sub(r"^(?:delivery\s+(?:scope|contents?|content)|scope\s+of\s+delivery|package\s+contents?|standard\s+delivery)\s*[:\-–—]?\s*", "", s, flags=re.I)
    return s.strip(" -–—•.;:")


def _split_bundled_delivery_item(item: str) -> List[str]:
    """Split one long delivery line containing repeated quantity markers into item lines.

    PCE often renders Delivery Scope as both one combined sentence and individual list items.
    Example: "1 × Meter 4 × batteries 1 × manual". Splitting before each quantity marker
    lets the later dedupe step remove the combined duplicate without losing data.
    """
    s = _strip_delivery_prefix(clean_text(item))
    if not s:
        return []
    qty_markers = list(re.finditer(r"(?<![A-Za-z0-9.,])\d+\s*(?:×|x)\s+", s, flags=re.I))
    if len(qty_markers) < 2:
        return [s]
    parts: List[str] = []
    for idx, marker in enumerate(qty_markers):
        start = marker.start()
        end = qty_markers[idx + 1].start() if idx + 1 < len(qty_markers) else len(s)
        part = s[start:end].strip(" -–—•.;:")
        if part:
            parts.append(part)
    return parts or [s]


def _delivery_item_key(item: str) -> str:
    key = clean_text(item).lower()
    key = re.sub(r"<[^>]+>", " ", key)
    key = re.sub(r"\b(?:delivery contents?|delivery scope|scope of delivery|package contents?)\b", " ", key, flags=re.I)
    key = key.replace("×", "x")
    key = re.sub(r"\s+", " ", key)
    key = re.sub(r"[^a-z0-9°%.,/+\- ]+", "", key)
    return key.strip()


def _dedupe_delivery_items(items: List[str], max_items: int = 30) -> List[str]:
    """Clean and de-duplicate Delivery Scope items.

    Besides exact duplicates, this removes bundled parent lines when the same items are
    already available as separate list entries. If only a bundled line exists, it is split
    into individual quantity-based items.
    """
    expanded: List[str] = []
    for raw_item in items:
        raw = clean_text(raw_item)
        if not raw or _is_delivery_heading_text(raw) or _is_stop_heading_text(raw):
            continue
        for part in _split_bundled_delivery_item(raw):
            item = _delivery_item_cleanup(part)
            item = _strip_delivery_prefix(item)
            if not item or len(item) < 2:
                continue
            # Delivery Scope umumnya berupa item paket pendek. Buang baris tabel panjang.
            if len(item) > 260:
                continue
            expanded.append(item)

    # First pass: exact-normalized unique items.
    unique: List[str] = []
    seen = set()
    for item in expanded:
        key = _delivery_item_key(item)
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(item)

    # Second pass: remove a long bundled/container item when several shorter items are contained in it.
    filtered: List[str] = []
    keys = [_delivery_item_key(x) for x in unique]
    for idx, item in enumerate(unique):
        item_key = keys[idx]
        contained_short_items = 0
        for j, other_key in enumerate(keys):
            if j == idx or not other_key:
                continue
            if len(other_key) < len(item_key) and other_key in item_key:
                contained_short_items += 1
        if contained_short_items >= 2 and len(item) > 120:
            continue
        filtered.append(item)
        if len(filtered) >= max_items:
            break
    return filtered


def _collect_delivery_items_from_container(container: Any) -> List[str]:
    items: List[str] = []
    if container is None:
        return items

    # Kalau container masih berisi tab lain/spec/download, jangan ambil semua wrapper.
    text_all = soup_text(container)
    if len(text_all) > 2500 and re.search(r"\b(?:technical specification|measurement range|accuracy|resolution|downloads?)\b", text_all, flags=re.I):
        return []

    # Prefer explicit list items. Delivery scope PCE biasanya berupa list atau paragraf pendek.
    try:
        for li in container.find_all("li"):
            txt = soup_text(li)
            if txt:
                items.append(txt)
    except Exception:
        pass

    # Table rows only allowed if they look like shipping/package items, not spec rows.
    try:
        for tr in container.find_all("tr"):
            cells = [soup_text(c) for c in tr.find_all(["th", "td"])]
            cells = [c for c in cells if c]
            joined = " - ".join(cells)
            if joined and DELIVERY_GOOD_ITEM_RE.search(joined) and not DELIVERY_BAD_ITEM_RE.search(joined):
                items.append(joined)
    except Exception:
        pass

    # Paragraph fallback.
    if not items:
        try:
            for p in container.find_all(["p", "div"]):
                txt = soup_text(p)
                if txt and len(txt) <= 300 and not DELIVERY_BAD_ITEM_RE.search(txt):
                    items.append(txt)
        except Exception:
            pass
    return items


def _collect_after_heading_until_stop(heading: Any) -> List[str]:
    items: List[str] = []
    try:
        for node in heading.next_elements:
            if node is heading:
                continue
            name = str(getattr(node, "name", "") or "").lower()
            if not name:
                continue
            if name in {"script", "style", "noscript", "svg", "img", "form", "button"}:
                continue
            txt = soup_text(node)
            if not txt:
                continue
            short = len(txt) <= 110
            if node is not heading and short and (_is_stop_heading_text(txt) or (re.fullmatch(r"h[1-6]", name) and not _is_delivery_heading_text(txt))):
                if items:
                    break
                # Kalau langsung ketemu heading lain, berarti delivery kosong / tab nav saja.
                return []
            if DELIVERY_BAD_ITEM_RE.search(txt):
                # Spec/download muncul sebelum item delivery; jangan teruskan, agar tidak salah isi.
                if not items:
                    return []
                break
            if name == "li":
                items.append(txt)
            elif name == "tr":
                cells = [soup_text(c) for c in node.find_all(["th", "td"])]
                cells = [c for c in cells if c]
                joined = " - ".join(cells)
                if joined:
                    items.append(joined)
            elif name in {"p", "span", "div"} and len(txt) <= 300:
                items.append(txt)
            if len(items) >= 30:
                break
    except Exception:
        pass
    return items


def _delivery_scope_is_valid(items: List[str]) -> bool:
    if not items:
        return False
    joined = " ".join(items)
    # Reject if technical/download language remains in the extracted result.
    if DELIVERY_BAD_ITEM_RE.search(joined):
        return False
    # At least one package/delivery signal is required.
    if not DELIVERY_GOOD_ITEM_RE.search(joined):
        return False
    # Reject unusually large extracts, usually wrapper/panel capture.
    if len(items) > 18 or len(joined) > 1600:
        return False
    return True


def extract_delivery_scope_from_soup(soup: Any) -> List[str]:
    """Extract official Delivery Scope / Scope of Delivery content from PCE page.

    Returns [] when there is no reliable Delivery Scope section. This prevents generating
    a custom tab from assumptions, specs, or downloads.
    """
    if soup is None:
        return []

    headings: List[Any] = []
    try:
        all_tags = soup.find_all(True)
    except Exception:
        all_tags = []

    for tag in all_tags:
        if _looks_like_delivery_heading_tag(tag):
            headings.append(tag)

    for heading in headings:
        candidates: List[str] = []

        # Best case: tab link points to the matching tab-pane. Only collect that pane.
        panel = _find_linked_delivery_panel(soup, heading)
        if panel is not None:
            candidates.extend(_collect_delivery_items_from_container(panel))

        # Heading section fallback. Do not collect parent wrappers; that caused specs/downloads
        # to be inserted into Delivery Scope in V8.
        if not candidates and re.fullmatch(r"h[1-6]", str(getattr(heading, "name", "") or ""), flags=re.I):
            candidates.extend(_collect_after_heading_until_stop(heading))

        # Direct next sibling fallback for simple markup.
        if not candidates:
            try:
                sib = heading.find_next_sibling()
                if sib is not None:
                    candidates.extend(_collect_delivery_items_from_container(sib))
            except Exception:
                pass

        cleaned = _dedupe_delivery_items(candidates)
        if _delivery_scope_is_valid(cleaned):
            return cleaned

    # Conservative fallback with heading extractor. This is only accepted when the cleaned
    # result has delivery-package signals and no spec/download terms.
    for heading_text in ["Delivery Scope", "Scope of Delivery", "Delivery contents", "Included in delivery", "Standard delivery", "Package contents"]:
        text = extract_text_after_heading(soup, heading_text, max_chars=900)
        if not text:
            continue
        parts = re.split(r"\s+-\s+|\n+|\s+•\s+|;\s+", text)
        cleaned = _dedupe_delivery_items(parts)
        if _delivery_scope_is_valid(cleaned):
            return cleaned

    return []


def delivery_scope_html(items: List[str]) -> str:
    clean_items = _dedupe_delivery_items(items)
    if not _delivery_scope_is_valid(clean_items):
        return ""
    return "<h2>Kelengkapan Pengiriman</h2><ul>" + "".join(f"<li>{esc(item)}</li>" for item in clean_items) + "</ul>"


def generate_delivery_scope_tab(row: Dict[str, Any]) -> str:
    items = row.get("delivery_scope", [])
    if not isinstance(items, list):
        return ""
    return delivery_scope_html(items)


# Override the scraper so cached/new scrape data can include delivery_scope.
def scrape_pce_product_page(url: str, timeout: int = 15) -> Dict[str, Any]:
    url = normalize_url(url)
    result: Dict[str, Any] = {
        "url": url,
        "status": "not_requested" if not url else "failed",
        "title": "",
        "description": "",
        "specs": "",
        "spec_table": [],
        "delivery_scope": [],
        "image_url": "",
        "order_no": "",
        "gtin": "",
        "error": "",
    }
    if not url:
        result["status"] = "no_url"
        return result
    if requests is None:
        result["error"] = "requests_not_installed"
        return result
    try:
        response = requests.get(url, headers=PCE_HEADERS, timeout=timeout)
        response.raise_for_status()
    except Exception as exc:
        result["error"] = f"request_error:{type(exc).__name__}"
        return result

    html_text = response.text or ""
    if not html_text.strip():
        result["error"] = "empty_html"
        return result

    if BeautifulSoup is None:
        result["description"] = extract_description_from_plain_text(html_text)
        result["status"] = "ok" if result["description"] else "no_description_found"
        return result

    soup = BeautifulSoup(html_text, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    h1 = soup.find("h1")
    title = soup_text(h1)
    if title.lower() in {"description", "information"} or not title:
        title = find_meta_content(soup, 'meta[property="og:title"]', 'meta[name="twitter:title"]')
    if not title and soup.title:
        title = soup_text(soup.title)
    result["title"] = clean_text(re.sub(r"\s*[-|].*$", "", title))

    desc = extract_text_after_heading(soup, "Description", max_chars=1200)
    if not desc:
        desc = find_meta_content(soup, 'meta[name="description"]', 'meta[property="og:description"]', 'meta[name="twitter:description"]')
    if not desc:
        desc = extract_description_from_plain_text(soup_text(soup), max_chars=1200)
    result["description"] = compact_source_text(desc, max_chars=1200)

    spec_table = extract_spec_table_from_soup(soup)
    result["spec_table"] = spec_table

    specs = extract_text_after_heading(soup, "Technical Data", max_chars=2200)
    if not specs:
        specs = extract_text_after_heading(soup, "Specifications", max_chars=2200)
    if not specs:
        specs = extract_text_after_heading(soup, "Specification", max_chars=2200)
    result["specs"] = compact_source_text(specs, max_chars=2200)

    result["delivery_scope"] = extract_delivery_scope_from_soup(soup)
    result["image_url"] = extract_best_image_url(soup, url)

    page_text = soup_text(soup)
    m_order = re.search(r"Order\s*no\.?\s*:?\s*([A-Za-z0-9._/-]+)", page_text, flags=re.I)
    if m_order:
        result["order_no"] = clean_text(m_order.group(1))
    m_gtin = re.search(r"GTIN\s*\(?EAN\)?\s*:?\s*([0-9]{6,14})", page_text, flags=re.I)
    if m_gtin:
        result["gtin"] = clean_text(m_gtin.group(1))

    result["status"] = "ok" if result["description"] else "no_description_found"
    return result


def normalize_row(
    row: Dict[str, Any],
    scrape_web: bool = True,
    scrape_cache: Optional[Dict[str, Any]] = None,
    request_timeout: int = 15,
    request_delay: float = 0.0,
) -> Dict[str, Any]:
    source_url = pick(row, "link", "Source URL", "url", "URL")
    scraped: Dict[str, Any] = {}
    if scrape_web and source_url:
        scraped = get_scraped_product_data(source_url, scrape_cache, timeout=request_timeout, delay=request_delay)

    name = strip_jual(pick(row, "Judul", "Name", "Title", "Product Name"))
    scraped_title = strip_jual(clean_text(scraped.get("title", "")))
    if not name and scraped_title:
        name = scraped_title
    if not name:
        name = strip_jual(pick(row, "Order No", "Internal Reference", "costumiz"))

    brand = normalize_brand(pick(row, "Brand", "Merek"))
    source_tags = pick(row, "Tags", "Product_Tags", "product tags")
    excel_desc = pick(
        row,
        "Website Description", "API Description", "Source Description", "Original Description",
        "Short_Description", "Product Short Description", "description", "Description",
        "description_sale", "description_purchase", "description_ecommerce", "body_html",
    )
    website_desc = clean_text(scraped.get("description", ""))
    if website_desc and len(website_desc) >= 20:
        desc = website_desc
        description_origin = "website_cache" if scraped.get("from_cache") == "yes" else "website"
    elif excel_desc:
        desc = excel_desc
        description_origin = "excel"
    else:
        desc = ""
        description_origin = "fallback_template"

    source_category = pick(row, "Product Category/Name", "product categories", "Category", "Tags")
    if source_category.lower() in {"saleable", "storable product", "units"}:
        source_category = source_tags
    fam = detect_family(name, source_tags, source_category, desc)
    category = source_category if source_category and source_category.lower() != "saleable" else fam.get("category")
    is_acc, accessory_reason = detect_accessory_reason(name, source_tags, category, desc)
    category = public_category(category, fam, is_acc)

    params = extract_params("\n".join([name, source_tags, desc, clean_text(scraped.get("specs", ""))]))
    if fam.get("id") not in {"gas_air", "electrical"}:
        params = [p for p in params if p.upper() not in {"ISO", "UL", "ATEX", "DIN ISO"}]

    image_url = pick(row, "Image", "image_url", "Images", "Gambar") or clean_text(scraped.get("image_url", ""))
    order_no = pick(row, "Order No", "Order No.", "order_no", "Model") or clean_text(scraped.get("order_no", ""))
    gtin = pick(row, "GTIN", "GTIN (EAN)", "EAN") or clean_text(scraped.get("gtin", ""))

    return {
        "name": name,
        "brand": brand,
        "source_tags": source_tags,
        "category": category or fam.get("category"),
        "family": fam,
        "description": desc,
        "description_origin": description_origin,
        "website_specs": clean_text(scraped.get("specs", "")),
        "spec_table": scraped.get("spec_table", []) if isinstance(scraped.get("spec_table", []), list) else [],
        "delivery_scope": scraped.get("delivery_scope", []) if isinstance(scraped.get("delivery_scope", []), list) else [],
        "scrape_status": clean_text(scraped.get("status", "")) if scrape_web else "disabled",
        "scrape_error": clean_text(scraped.get("error", "")),
        "params": params,
        "param_text": ", ".join(params),
        "is_accessory": is_acc,
        "accessory_reason": accessory_reason,
        "source_url": normalize_url(source_url),
        "image_url": image_url,
        "order_no": order_no,
        "internal_reference": pick(row, "Internal Reference", "Internal Reference", "costumiz", "SKU"),
        "gtin": gtin,
    }


def row_to_output(norm: Dict[str, Any], idx: int) -> Dict[str, Any]:
    out = {col: "" for col in OUTPUT_COLS}
    log_notes = []
    if not norm["name"]:
        log_notes.append("missing_name")
    if not norm.get("image_url"):
        log_notes.append("missing_image")
    if not norm.get("description"):
        log_notes.append("missing_short_description")
    if norm.get("is_accessory"):
        log_notes.append("detected_accessory_component")
    else:
        log_notes.append("detected_main_product")
    if norm.get("accessory_reason"):
        safe_reason = re.sub(r"[^a-z0-9_]+", "_", str(norm.get("accessory_reason", "")).lower()).strip("_")
        if safe_reason:
            log_notes.append(f"role_reason_{safe_reason}")
    if norm.get("description_origin"):
        log_notes.append(f"description_source_{norm.get('description_origin')}")
    if norm.get("scrape_status") and norm.get("scrape_status") not in {"ok", "disabled"}:
        log_notes.append(f"scrape_{norm.get('scrape_status')}")
    if norm.get("scrape_error"):
        log_notes.append(norm.get("scrape_error"))

    validation = "scraped_from_source_url" if norm.get("description_origin") in {"website", "website_cache"} else "fallback_from_excel_or_template_not_live_checked"
    quality = "pce_uji_generated_from_website_source" if norm.get("description_origin") in {"website", "website_cache"} else "pce_uji_generated_from_excel_fallback"

    spec_tab_content = generate_spec_tab(norm)
    if spec_tab_content:
        spec_tab_title = "SPESIFIKASI"
        spec_tab_priority = 20
        log_notes.append("official_spec_table_found")
    else:
        spec_tab_title = ""
        spec_tab_priority = ""
        log_notes.append("official_spec_table_not_found_spec_tab_skipped")

    delivery_tab_content = generate_delivery_scope_tab(norm)
    if delivery_tab_content:
        delivery_tab_title = "Kelengkapan Pengiriman"
        delivery_tab_priority = 30
        log_notes.append("delivery_scope_found")
    else:
        delivery_tab_title = ""
        delivery_tab_priority = ""
        log_notes.append("delivery_scope_not_found_tab_skipped")

    out.update({
        "Name": norm["name"],
        "Brand": norm["brand"],
        "Product Description": generate_product_description(norm),
        "Product Short Description": generate_short_description(norm),
        "custom_tab_1_title": "Keunggulan Produk",
        "custom_tab_1_content": generate_feature_tab(norm),
        "custom_tab_1_priority": 10,
        "custom_tab_2_title": spec_tab_title,
        "custom_tab_2_content": spec_tab_content,
        "custom_tab_2_priority": spec_tab_priority,
        # Urutan tab publik: Keunggulan -> Spesifikasi -> Kelengkapan Pengiriman -> FAQ.
        # Jika Delivery Scope tidak ada, custom_tab_3 dibiarkan blank. FAQ tetap ditempatkan paling akhir.
        "custom_tab_3_title": delivery_tab_title,
        "custom_tab_3_content": delivery_tab_content,
        "custom_tab_3_priority": delivery_tab_priority,
        "custom_tab_4_title": "FAQ",
        "custom_tab_4_content": generate_faq(norm),
        "custom_tab_4_priority": 90,
        "product categories": norm["category"],
        "product tags": generate_tags(norm),
        "focus keyphrase": norm["name"],
        "meta description": meta_description(norm),
        "publication_date": (datetime(2025, 2, 8, 10, 0) + timedelta(hours=idx)).strftime("%d/%m/%Y %H:%M"),
        "Processed": "Yes",
        "Processing Time": "pce_uji_full_indonesia_pipeline_v10_dedup_delivery_faq_last",
        "Content Quality": quality,
        "image_url": norm.get("image_url", ""),
        "Source URL": norm.get("source_url", ""),
        "Website Validation": validation,
        "Website Correction Log": "; ".join(log_notes),
    })
    return out


# ---------------------------------------------------------------------------
# V11 override: DeepSeek AI content generator.
# - Official specification table and delivery scope stay rule/scrape based.
# - AI is used only for public marketing copy: Product Description, Short
#   Description, Keunggulan Produk, FAQ, and meta description.
# - If AI is disabled, API key is missing, request fails, or JSON is invalid,
#   the V10 rule/template generator is used automatically.
# ---------------------------------------------------------------------------

AI_PROMPT_VERSION = "pce_uji_deepseek_v11_2026_05_25"
AI_SETTINGS: Dict[str, Any] = {
    "enabled": False,
    "api_key": "",
    "base_url": "https://api.deepseek.com",
    "model": "deepseek-chat",
    "cache": {},
    "cache_path": "",
    "limit": 0,
    "timeout": 60,
    "delay": 0.0,
    "temperature": 0.25,
}
AI_STATS: Dict[str, int] = {"used": 0, "cache_hit": 0, "fallback": 0, "skipped": 0}


def set_deepseek_ai_settings(
    enabled: bool = False,
    api_key: str = "",
    base_url: str = "https://api.deepseek.com",
    model: str = "deepseek-chat",
    cache: Optional[Dict[str, Any]] = None,
    cache_path: str = "",
    limit: int = 0,
    timeout: int = 60,
    delay: float = 0.0,
    temperature: float = 0.25,
) -> None:
    AI_SETTINGS.update({
        "enabled": bool(enabled),
        "api_key": clean_text(api_key),
        "base_url": clean_text(base_url).rstrip("/") or "https://api.deepseek.com",
        "model": clean_text(model) or "deepseek-chat",
        "cache": cache if isinstance(cache, dict) else {},
        "cache_path": clean_text(cache_path),
        "limit": max(0, int(limit or 0)),
        "timeout": max(10, int(timeout or 60)),
        "delay": max(0.0, float(delay or 0.0)),
        "temperature": max(0.0, min(1.0, float(temperature if temperature is not None else 0.25))),
    })


def _truncate_for_ai(text: Any, max_chars: int) -> str:
    s = clean_text(text)
    if len(s) <= max_chars:
        return s
    return s[:max_chars].rsplit(" ", 1)[0].rstrip(" ,.;:") + "..."


def _html_to_text_for_ai(value: Any, max_chars: int = 1600) -> str:
    s = clean_text(value)
    if not s:
        return ""
    s = re.sub(r"<\s*br\s*/?>", "\n", s, flags=re.I)
    s = re.sub(r"</\s*(p|li|h2|h3|tr)\s*>", "\n", s, flags=re.I)
    s = re.sub(r"<[^>]+>", " ", s)
    s = html.unescape(s)
    s = re.sub(r"\s+", " ", s).strip()
    return _truncate_for_ai(s, max_chars)


def _spec_rows_for_ai(row: Dict[str, Any], limit: int = 14) -> List[Dict[str, str]]:
    rows = row.get("spec_table", []) if isinstance(row.get("spec_table", []), list) else []
    out: List[Dict[str, str]] = []
    for item in rows[:limit]:
        if not isinstance(item, dict):
            continue
        label = _truncate_for_ai(item.get("label", ""), 80)
        value = _truncate_for_ai(item.get("value", ""), 140)
        if label and value:
            out.append({"label": label, "value": value})
    return out


def _delivery_for_ai(row: Dict[str, Any], limit: int = 12) -> List[str]:
    items = row.get("delivery_scope", []) if isinstance(row.get("delivery_scope", []), list) else []
    return [_truncate_for_ai(x, 120) for x in items[:limit] if clean_text(x)]


def _feature_candidates_for_ai(row: Dict[str, Any], limit: int = 10) -> List[str]:
    try:
        items = build_features(
            row.get("name", ""),
            row.get("family", DEFAULT_FAMILY),
            row.get("description", ""),
            row.get("params", []),
            row.get("is_accessory", False),
        )
        items.extend(source_highlights_for_feature_tab(row, limit=6))
        items = unique_items(items, limit)
    except Exception:
        items = []
    return [_truncate_for_ai(x, 160) for x in items if clean_text(x)]


def _row_payload_for_ai(row: Dict[str, Any]) -> Dict[str, Any]:
    fam = row.get("family", DEFAULT_FAMILY) if isinstance(row.get("family", DEFAULT_FAMILY), dict) else DEFAULT_FAMILY
    payload = {
        "name": clean_text(row.get("name", "")),
        "brand": clean_text(row.get("brand", "PCE Instruments")),
        "category": clean_text(row.get("category", "")),
        "family_id": clean_text(fam.get("id", "general")),
        "family_term": clean_text(fam.get("term", DEFAULT_FAMILY["term"])),
        "family_function": clean_text(fam.get("function", DEFAULT_FAMILY["function"])),
        "family_object": clean_text(fam.get("object", DEFAULT_FAMILY["object"])),
        "family_criteria": clean_text(fam.get("criteria", DEFAULT_FAMILY["criteria"])),
        "applications": build_applications(fam, bool(row.get("is_accessory", False)))[:6],
        "is_accessory": bool(row.get("is_accessory", False)),
        "accessory_reason": clean_text(row.get("accessory_reason", "")),
        "source_description": _truncate_for_ai(row.get("description", ""), 1400),
        "website_specs_text": _truncate_for_ai(row.get("website_specs", ""), 1400),
        "official_spec_rows": _spec_rows_for_ai(row),
        "delivery_scope": _delivery_for_ai(row),
        "detected_parameters": row.get("params", [])[:8] if isinstance(row.get("params", []), list) else [],
        "feature_candidates": _feature_candidates_for_ai(row),
        "source_url": clean_text(row.get("source_url", "")),
        "order_no": clean_text(row.get("order_no", "")),
        "gtin": clean_text(row.get("gtin", "")),
    }
    return payload


def _ai_cache_key(payload: Dict[str, Any]) -> str:
    key_material = {
        "prompt_version": AI_PROMPT_VERSION,
        "model": AI_SETTINGS.get("model", ""),
        "payload": payload,
    }
    raw = json.dumps(key_material, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _strip_code_fences(text: str) -> str:
    s = clean_text(text)
    s = re.sub(r"^```(?:json|html)?\s*", "", s, flags=re.I)
    s = re.sub(r"\s*```$", "", s)
    return s.strip()


def _extract_json_object(text: str) -> Dict[str, Any]:
    s = _strip_code_fences(text)
    try:
        data = json.loads(s)
        return data if isinstance(data, dict) else {}
    except Exception:
        pass
    start = s.find("{")
    end = s.rfind("}")
    if start >= 0 and end > start:
        try:
            data = json.loads(s[start:end + 1])
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def _sanitize_ai_html(value: Any, fallback: str = "") -> str:
    s = _strip_code_fences(clean_text(value))
    if not s:
        return fallback
    s = re.sub(r"<\s*(script|style|iframe|object|embed)[^>]*>.*?<\s*/\s*\1\s*>", "", s, flags=re.I | re.S)
    s = re.sub(r"\s+on[a-z]+\s*=\s*(['\"]).*?\1", "", s, flags=re.I | re.S)
    s = re.sub(r"\s+(style|class|id)\s*=\s*(['\"]).*?\2", "", s, flags=re.I | re.S)
    allowed = {"p", "strong", "b", "em", "ul", "ol", "li", "h2", "h3", "br"}

    def repl_tag(m: re.Match) -> str:
        slash = "/" if m.group(1) else ""
        tag = (m.group(2) or "").lower()
        if tag not in allowed:
            return ""
        if tag == "br":
            return "<br>"
        return f"<{slash}{tag}>"

    s = re.sub(r"<\s*(/?)\s*([a-zA-Z0-9]+)(?:\s+[^>]*)?>", repl_tag, s)
    s = re.sub(r"\n{3,}", "\n\n", s).strip()
    return s or fallback


def _sanitize_ai_text(value: Any, max_chars: int = 220, fallback: str = "") -> str:
    s = _html_to_text_for_ai(value, max_chars=max_chars * 2)
    s = re.sub(r"\s+", " ", s).strip(" \n\t.,;")
    if not s:
        return fallback
    if len(s) > max_chars:
        s = s[:max_chars].rsplit(" ", 1)[0].rstrip(" ,.;:") + "."
    if s and s[-1] not in ".!?":
        s += "."
    return s


def _validate_ai_content(data: Dict[str, Any]) -> Tuple[bool, str, Dict[str, str]]:
    fields = {
        "product_description_html": _sanitize_ai_html(data.get("product_description_html", "")),
        "short_description": _sanitize_ai_text(data.get("short_description", ""), max_chars=240),
        "feature_tab_html": _sanitize_ai_html(data.get("feature_tab_html", "")),
        "faq_html": _sanitize_ai_html(data.get("faq_html", "")),
        "meta_description": _sanitize_ai_text(data.get("meta_description", ""), max_chars=158),
    }
    if len(_html_to_text_for_ai(fields["product_description_html"], 4000)) < 220:
        return False, "short_product_description_html", fields
    if "<h2" not in fields["product_description_html"].lower():
        return False, "product_description_missing_h2", fields
    if len(_html_to_text_for_ai(fields["faq_html"], 4000)) < 160 or "<h3" not in fields["faq_html"].lower():
        return False, "invalid_faq_html", fields
    if len(_html_to_text_for_ai(fields["feature_tab_html"], 2000)) < 80 or "<li" not in fields["feature_tab_html"].lower():
        return False, "invalid_feature_tab_html", fields
    if not fields["short_description"] or not fields["meta_description"]:
        return False, "missing_short_or_meta", fields
    forbidden = re.compile(r"\b(harga|diskon|promo|garansi resmi|ready stock|stok tersedia|gratis ongkir)\b", re.I)
    joined = " ".join(fields.values())
    if forbidden.search(joined):
        return False, "forbidden_sales_claim", fields
    return True, "ok", fields


def _build_deepseek_messages(payload: Dict[str, Any]) -> List[Dict[str, str]]:
    system_prompt = (
        "Anda adalah penulis katalog teknis berbahasa Indonesia untuk uji.co.id. "
        "Tulis copy produk yang natural, tidak terdengar seperti template, dan tidak mengarang spesifikasi. "
        "Gunakan hanya data yang diberikan. Jangan membuat klaim harga, garansi, stok, promo, sertifikasi, atau akurasi jika tidak ada pada data. "
        "Nama produk/SKU/model harus dipertahankan apa adanya. Satuan teknis seperti °C, %, ppm, g, Hz boleh tetap. "
        "Hindari campuran Inggris-Indonesia kecuali istilah teknis/nama produk. "
        "Jangan memasukkan tabel spesifikasi atau delivery scope ke product_description_html; itu akan diisi oleh tab resmi terpisah. "
        "Balas JSON valid saja."
    )
    user_prompt = {
        "task": "Buat konten katalog WooCommerce uji.co.id dalam bahasa Indonesia.",
        "output_schema": {
            "product_description_html": "HTML: 2 paragraf pembuka, <h2>Fungsi dan Keunggulan ...</h2> + <ul>, <h2>Contoh Penggunaan</h2> + paragraf/ul, <h2>Informasi Pemilihan Produk</h2> + paragraf. Tanpa CTA kontak.",
            "short_description": "1 kalimat ringkas maksimal 220 karakter.",
            "feature_tab_html": "HTML <h2>Keunggulan Produk</h2><ul> berisi 4-7 poin natural dan faktual.",
            "faq_html": "HTML berisi 4-5 pasang <h3>pertanyaan</h3><p>jawaban</p>. FAQ harus spesifik sesuai jenis produk.",
            "meta_description": "Maksimal 158 karakter, natural untuk SEO.",
        },
        "rules": [
            "Output hanya JSON valid, tanpa markdown dan tanpa code fence.",
            "Jangan tulis kalimat seperti 'Data produk mencantumkan...' kecuali benar-benar perlu.",
            "Jangan memasukkan parameter mentah yang tidak jelas ke deskripsi utama.",
            "Jika data spesifikasi kurang lengkap, arahkan pembaca untuk mencocokkan spesifikasi resmi tanpa mengarang angka.",
            "Untuk aksesori, jelaskan kompatibilitas dengan unit utama dan jangan sebut sebagai alat ukur mandiri.",
            "Gunakan tag HTML yang aman: p, strong, ul, li, h2, h3, em.",
        ],
        "product_data": payload,
    }
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(user_prompt, ensure_ascii=False)},
    ]


def _call_deepseek_chat(messages: List[Dict[str, str]]) -> Tuple[bool, str, str]:
    if requests is None:
        return False, "requests_not_installed", ""
    api_key = clean_text(AI_SETTINGS.get("api_key", ""))
    if not api_key:
        return False, "missing_api_key", ""
    url = clean_text(AI_SETTINGS.get("base_url", "https://api.deepseek.com")).rstrip("/") + "/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    body = {
        "model": AI_SETTINGS.get("model", "deepseek-chat"),
        "messages": messages,
        "temperature": AI_SETTINGS.get("temperature", 0.25),
        "response_format": {"type": "json_object"},
    }
    try:
        response = requests.post(url, headers=headers, json=body, timeout=AI_SETTINGS.get("timeout", 60))
        if response.status_code >= 400:
            # Some OpenAI-compatible gateways do not support response_format. Retry once without it.
            if response.status_code in {400, 422}:
                body.pop("response_format", None)
                response = requests.post(url, headers=headers, json=body, timeout=AI_SETTINGS.get("timeout", 60))
        response.raise_for_status()
        data = response.json()
        choices = data.get("choices", []) if isinstance(data, dict) else []
        if not choices:
            return False, "empty_choices", ""
        message = choices[0].get("message", {}) if isinstance(choices[0], dict) else {}
        content = clean_text(message.get("content", ""))
        if not content:
            return False, "empty_content", ""
        return True, "ok", content
    except Exception as exc:
        return False, f"request_error_{type(exc).__name__}", ""


def get_deepseek_ai_content(row: Dict[str, Any]) -> Dict[str, str]:
    if not AI_SETTINGS.get("enabled"):
        return {}
    existing = row.get("_deepseek_ai")
    if isinstance(existing, dict):
        return existing.get("fields", {}) if existing.get("ok") else {}

    payload = _row_payload_for_ai(row)
    cache_key = _ai_cache_key(payload)
    cache = AI_SETTINGS.get("cache", {}) if isinstance(AI_SETTINGS.get("cache", {}), dict) else {}
    cached = cache.get(cache_key)
    if isinstance(cached, dict) and cached.get("ok") and isinstance(cached.get("fields"), dict):
        row["_deepseek_ai"] = cached
        AI_STATS["cache_hit"] = AI_STATS.get("cache_hit", 0) + 1
        return cached.get("fields", {})

    if not clean_text(AI_SETTINGS.get("api_key", "")):
        row["_deepseek_ai"] = {"ok": False, "reason": "missing_api_key", "fields": {}}
        AI_STATS["fallback"] = AI_STATS.get("fallback", 0) + 1
        return {}

    limit = int(AI_SETTINGS.get("limit", 0) or 0)
    if limit > 0 and AI_STATS.get("used", 0) >= limit:
        row["_deepseek_ai"] = {"ok": False, "reason": "ai_limit_reached", "fields": {}}
        AI_STATS["skipped"] = AI_STATS.get("skipped", 0) + 1
        return {}

    messages = _build_deepseek_messages(payload)
    ok, reason, raw = _call_deepseek_chat(messages)
    if ok:
        data = _extract_json_object(raw)
        valid, validation_reason, fields = _validate_ai_content(data)
        if valid:
            record = {"ok": True, "reason": "ok", "fields": fields, "created_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")}
            cache[cache_key] = record
            AI_SETTINGS["cache"] = cache
            row["_deepseek_ai"] = record
            AI_STATS["used"] = AI_STATS.get("used", 0) + 1
            delay = float(AI_SETTINGS.get("delay", 0.0) or 0.0)
            if delay > 0:
                time.sleep(delay)
            return fields
        reason = validation_reason

    row["_deepseek_ai"] = {"ok": False, "reason": reason, "fields": {}}
    cache[cache_key] = row["_deepseek_ai"]
    AI_SETTINGS["cache"] = cache
    AI_STATS["fallback"] = AI_STATS.get("fallback", 0) + 1
    delay = float(AI_SETTINGS.get("delay", 0.0) or 0.0)
    if delay > 0:
        time.sleep(delay)
    return {}


# Keep references to the V10 rule generator for safe fallback.
RULE_generate_product_description = generate_product_description
RULE_generate_short_description = generate_short_description
RULE_generate_feature_tab = generate_feature_tab
RULE_generate_faq = generate_faq
RULE_meta_description = meta_description
RULE_row_to_output = row_to_output


def generate_product_description(row: Dict[str, Any]) -> str:
    fields = get_deepseek_ai_content(row)
    html_value = fields.get("product_description_html", "") if fields else ""
    if html_value:
        return html_value + UJI_CTA
    return RULE_generate_product_description(row)


def generate_short_description(row: Dict[str, Any]) -> str:
    fields = get_deepseek_ai_content(row)
    value = fields.get("short_description", "") if fields else ""
    return value or RULE_generate_short_description(row)


def generate_feature_tab(row: Dict[str, Any]) -> str:
    fields = get_deepseek_ai_content(row)
    value = fields.get("feature_tab_html", "") if fields else ""
    return value or RULE_generate_feature_tab(row)


def generate_faq(row: Dict[str, Any]) -> str:
    fields = get_deepseek_ai_content(row)
    value = fields.get("faq_html", "") if fields else ""
    return value or RULE_generate_faq(row)


def meta_description(row: Dict[str, Any]) -> str:
    fields = get_deepseek_ai_content(row)
    value = fields.get("meta_description", "") if fields else ""
    return value or RULE_meta_description(row)


def row_to_output(norm: Dict[str, Any], idx: int) -> Dict[str, Any]:
    out = RULE_row_to_output(norm, idx)
    if AI_SETTINGS.get("enabled"):
        ai_state = norm.get("_deepseek_ai", {}) if isinstance(norm.get("_deepseek_ai", {}), dict) else {}
        log = clean_text(out.get("Website Correction Log", ""))
        extra: List[str] = []
        if ai_state.get("ok"):
            extra.append("deepseek_ai_used")
            out["Content Quality"] = "pce_uji_generated_with_deepseek_ai"
            out["Processing Time"] = "pce_uji_full_indonesia_pipeline_v11_deepseek_ai"
        else:
            reason = re.sub(r"[^a-z0-9_]+", "_", clean_text(ai_state.get("reason", "unknown")).lower()).strip("_") or "unknown"
            extra.append(f"deepseek_ai_fallback_{reason}")
            out["Processing Time"] = "pce_uji_full_indonesia_pipeline_v11_deepseek_ai_fallback_rule"
        out["Website Correction Log"] = "; ".join([x for x in [log] + extra if x])
    return out


# ---------------------------------------------------------------------------
# V13 cleanup layer: bahasa Delivery Scope, spesifikasi, kategori moisture, dan
# meta description. Layer ini sengaja ditempatkan di akhir file agar override
# diterapkan tanpa mengubah fondasi pipeline V10/V11/V12.
# ---------------------------------------------------------------------------

AI_PROMPT_VERSION = "pce_uji_deepseek_v13_2026_05_25"

_V12_detect_family = detect_family
_V12_public_category = public_category
_V12_normalize_public_phrase = normalize_public_phrase
_V12_delivery_item_cleanup = _delivery_item_cleanup
_V12_delivery_scope_html = delivery_scope_html
_V12_sanitize_ai_text = _sanitize_ai_text
_V12_build_deepseek_messages = _build_deepseek_messages
_V12_meta_description = meta_description
_V12_row_to_output = row_to_output


def detect_family(name: str, tags: str = "", category: str = "", description: str = "") -> Dict[str, Any]:
    """V13: prioritise moisture family for Absolute Moisture Meter / moisture analyser items.

    Pada versi sebelumnya, beberapa Absolute Moisture Meter bisa masuk ke
    Peralatan Laboratorium karena teks sumber menyebut balance/weighing/lab.
    Untuk katalog publik, keluarga utamanya tetap kadar air/kelembapan.
    """
    title = clean_slug_word(name)
    blob = clean_slug_word(" ".join([name, tags, category, description]))
    moisture_title = re.search(r"\b(?:absolute\s+)?moisture\s+(?:meter|analy[sz]er|balance)\b", title, flags=re.I)
    moisture_context = ("moisture" in title and re.search(r"\bpce[- ]?ma\b|\bma[- ]?\d", title, flags=re.I))
    if moisture_title or moisture_context:
        for fam in FAMILY_CONFIGS:
            if fam.get("id") == "moisture_humidity":
                return fam
    # Kalau nama produk jelas moisture tetapi kategori lama/lab lebih kuat karena teks
    # spesifikasi, tetap pilih moisture_humidity.
    if "moisture" in title and re.search(r"\b(?:meter|analyzer|analyser|tester)\b", title, flags=re.I):
        for fam in FAMILY_CONFIGS:
            if fam.get("id") == "moisture_humidity":
                return fam
    return _V12_detect_family(name, tags, category, description)


def public_category(category: str, fam: Dict[str, Any], accessory: bool = False) -> str:
    if not accessory and isinstance(fam, dict) and fam.get("id") == "moisture_humidity":
        return "Alat Ukur Kelembapan dan Kadar Air"
    return _V12_public_category(category, fam, accessory)


def normalize_public_phrase(label: str) -> str:
    s = _V12_normalize_public_phrase(label)
    if not s:
        return ""
    fixes = [
        (r"\bW\s+eighing\b", "Weighing"),
        (r"\bWeighing\s+rentang\b", "Rentang timbang"),
        (r"\bW eighing\s+rentang\b", "Rentang timbang"),
        (r"\bweighing\s+range\b", "Rentang timbang"),
        (r"\bweighing\b", "penimbangan"),
        (r"\bLC\s+Touch\s+(?:layar|tampilan|display)\b", "Layar sentuh"),
        (r"\bLCD\s+Touch\s+(?:layar|tampilan|display)\b", "Layar sentuh LCD"),
        (r"\bTouch\s+(?:layar|tampilan|display)\b", "Layar sentuh"),
        (r"\b([0-9]+(?:[,.][0-9]+)?)\s+inches\b", r"\1 inci"),
        (r"\b([0-9]+(?:[,.][0-9]+)?)\s+inch\b", r"\1 inci"),
        (r"\b([0-9]+(?:[,.][0-9]+)?)\s+times\s+per\s+detik\b", r"\1 kali/detik"),
        (r"\b([0-9]+(?:[,.][0-9]+)?)\s+time\s+per\s+detik\b", r"\1 kali/detik"),
        (r"\b([0-9]+(?:[,.][0-9]+)?)\s+times\s*/\s*detik\b", r"\1 kali/detik"),
        (r"\bUnit\s+berat\s+dengan\s+Package\b", "Berat unit dengan kemasan"),
        (r"\bUnit\s+weight\s+with\s+package\b", "Berat unit dengan kemasan"),
        (r"\bberat\s+unit\s+dengan\s+Package\b", "Berat unit dengan kemasan"),
        (r"\bAdditional\s+dimensi\b", "Dimensi tambahan"),
        (r"\bAdditional\s+dimensions?\b", "Dimensi tambahan"),
        (r"\bPackage\b", "kemasan"),
        (r"\bDisplay\b", "Layar"),
        (r"\bScreen\b", "Layar"),
        (r"\bReadability\b", "Keterbacaan"),
        (r"\bRepeatability\b", "Repeatabilitas"),
        (r"\bLinearity\b", "Linearitas"),
        (r"\bResolution\b", "Resolusi"),
        (r"\bAccuracy\b", "Akurasi"),
        (r"\bDimensions?\b", "Dimensi"),
        (r"\bWeight\b", "Berat"),
        (r"\bØ\s*([0-9])", r"diameter \1"),
    ]
    for pat, repl in fixes:
        s = re.sub(pat, repl, s, flags=re.I)
    # Rapikan frasa yang muncul akibat penerjemahan bertahap.
    s = re.sub(r"\bRentang\s+timbang\s+rentang\b", "Rentang timbang", s, flags=re.I)
    s = re.sub(r"\bLayar\s+layar\b", "Layar", s, flags=re.I)
    s = re.sub(r"\bKemasan\b", "kemasan", s) if len(s.split()) > 1 else s
    s = re.sub(r"\s*/\s*detik\b", "/detik", s, flags=re.I)
    s = re.sub(r"\s+([,.;:])", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


DELIVERY_REPLACEMENTS_V13: List[Tuple[str, str]] = [
    (r"\bincl\.\b", "termasuk"),
    (r"\bincluded\s+in\s+delivery\b", "termasuk dalam pengiriman"),
    (r"\bscope\s+of\s+delivery\b|\bdelivery\s+scope\b", "kelengkapan pengiriman"),
    (r"\bpackage\s+contents?\b", "isi paket"),
    (r"\boperating\s+instructions?\b", "petunjuk penggunaan"),
    (r"\buser'?s\s+manual\b|\buser\s+manual\b", "buku manual"),
    (r"\binstruction\s+manual\b", "buku manual"),
    (r"\bmanual\b", "buku manual"),
    (r"\bsample\s+pans?\b", "wadah sampel"),
    (r"\bsample\s+pan\s+holder\b", "dudukan wadah sampel"),
    (r"\bpan\s+holder\b", "dudukan wadah"),
    (r"\bcalibration\s+weight\b", "anak timbangan kalibrasi"),
    (r"\bweight\s+for\s+calibration\b", "anak timbangan kalibrasi"),
    (r"\bwind\s+protection\s+shield\b", "pelindung angin"),
    (r"\bwindshield\b", "pelindung angin"),
    (r"\bpower\s+kabel\b", "kabel daya"),
    (r"\bpower\s+cable\b", "kabel daya"),
    (r"\bmains\s+cable\b", "kabel listrik"),
    (r"\bpower\s+supply\b", "catu daya"),
    (r"\bpower\s+adapter\b", "adaptor daya"),
    (r"\bmains\s+adap[dt]er\b", "adaptor listrik"),
    (r"\badap[dt]er\b", "adaptor"),
    (r"\bbattery\s+charger\b", "pengisi daya baterai"),
    (r"\brechargeable\s+batteries\b", "baterai isi ulang"),
    (r"\bbatteries\b", "baterai"),
    (r"\bbattery\b", "baterai"),
    (r"\bUSB\s+cable\b", "kabel USB"),
    (r"\bdata\s+cable\b", "kabel data"),
    (r"\binterface\s+cable\b", "kabel antarmuka"),
    (r"\bconnection\s+cable\b", "kabel koneksi"),
    (r"\bcable\b", "kabel"),
    (r"\bsoftware\b", "perangkat lunak"),
    (r"\bISO\s+calibration\s+certificate\b", "sertifikat kalibrasi ISO"),
    (r"\bcalibration\s+certificate\b", "sertifikat kalibrasi"),
    (r"\btest\s+leads?\b", "kabel uji"),
    (r"\bmeasuring\s+leads?\b", "kabel ukur"),
    (r"\btest\s+cable\b", "kabel uji"),
    (r"\bcrocodile\s+clips?\b", "klip buaya"),
    (r"\balligator\s+clips?\b", "klip buaya"),
    (r"\bcleaning\s+brush\b", "sikat pembersih"),
    (r"\bcarrying\s+case\b", "tas pembawa"),
    (r"\btransport\s+case\b", "koper transportasi"),
    (r"\bprotective\s+case\b", "kotak pelindung"),
    (r"\bcase\b", "kotak"),
    (r"\bprobe\b", "sonda"),
    (r"\belectrode\b", "elektroda"),
    (r"\bsensor\b", "sensor"),
    (r"\bholder\b", "dudukan"),
    (r"\bstand\b", "dudukan"),
    (r"\bmounting\s+bracket\b", "braket pemasangan"),
    (r"\btripod\b", "tripod"),
    (r"\bclamp\b", "tang"),
    (r"\bscrewdriver\b", "obeng"),
    (r"\bthermocouple\b", "termokopel"),
    (r"\bset\b", "set"),
    (r"\bpieces?\b", "unit"),
]


def normalize_delivery_item_v13(text: str) -> str:
    s = clean_text(text).strip(" -–—•.;:")
    if not s:
        return ""
    if DELIVERY_BAD_ITEM_RE.search(s):
        return ""
    if re.search(r"\b(?:add to cart|delivery time|question|callback|price|vat|copyright|technical hotline|similar products|related products)\b", s, flags=re.I):
        return ""
    s = _strip_delivery_prefix(s)
    s = html.unescape(s)
    s = re.sub(r"\b(\d+)\s*x\s*", r"\1 × ", s, flags=re.I)
    for pat, repl in DELIVERY_REPLACEMENTS_V13:
        s = re.sub(pat, repl, s, flags=re.I)
    # Common mixed-output fixes from previous versions.
    mixed_fixes = [
        (r"\bPower\s+kabel\b", "kabel daya"),
        (r"\bCalibration\s+berat\b", "anak timbangan kalibrasi"),
        (r"\bUser'?s\s+buku\s+manual\b", "buku manual"),
        (r"\bbuku\s+buku\s+manual\b", "buku manual"),
        (r"\bwadah\s+sampels\b", "wadah sampel"),
        (r"\bkabel\s+kabel\b", "kabel"),
        (r"\bkotak\s+pembawa\b", "tas pembawa"),
    ]
    for pat, repl in mixed_fixes:
        s = re.sub(pat, repl, s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -–—•.;:")
    if not s:
        return ""
    # Jangan ubah kapitalisasi SKU/model; cukup kapitalisasi huruf pertama item.
    if re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


def _delivery_item_cleanup(text: str) -> str:
    return normalize_delivery_item_v13(text)


def _delivery_key_v13(item: str) -> str:
    key = _delivery_item_key(item)
    key = re.sub(r"\b(?:buku manual|manual pengguna|petunjuk penggunaan)\b", "manual", key, flags=re.I)
    key = re.sub(r"\b(?:wadah sampels?|sample pans?)\b", "wadah sampel", key, flags=re.I)
    key = re.sub(r"\b(?:anak timbangan kalibrasi|calibration weight)\b", "kalibrasi weight", key, flags=re.I)
    key = re.sub(r"\s+", " ", key).strip()
    return key


def delivery_scope_html(items: List[str]) -> str:
    raw_items = _dedupe_delivery_items(items)
    cleaned: List[str] = []
    seen = set()
    for item in raw_items:
        fixed = normalize_delivery_item_v13(item)
        if not fixed:
            continue
        key = _delivery_key_v13(fixed)
        if not key or key in seen:
            continue
        seen.add(key)
        cleaned.append(fixed)
    if not _delivery_scope_is_valid(cleaned):
        return ""
    return "<h2>Kelengkapan Pengiriman</h2><ul>" + "".join(f"<li>{esc(item)}</li>" for item in cleaned) + "</ul>"


WEAK_META_END_RE = re.compile(r"\b(?:dan|atau|untuk|dengan|yang|serta|seperti|pada|di|ke|dalam|sebagai|meliputi|mencakup|cocok)\.$", re.I)


def _clip_complete_sentence_v13(text: str, max_chars: int = 158) -> str:
    s = _html_to_text_for_ai(text, max_chars=max_chars * 3)
    s = re.sub(r"\s+", " ", s).strip(" \t\n\r,;:")
    if not s:
        return ""
    # Potong di batas kalimat jika ada sebelum limit.
    if len(s) > max_chars:
        candidate = s[:max_chars].rstrip(" ,;:")
        sentence_matches = list(re.finditer(r"[.!?]", candidate))
        if sentence_matches and sentence_matches[-1].end() >= 80:
            s = candidate[:sentence_matches[-1].end()].strip()
        else:
            s = candidate.rsplit(" ", 1)[0].rstrip(" ,;:") + "."
    if s and s[-1] not in ".!?":
        s += "."
    s = re.sub(r"\s+([.!?])", r"\1", s)
    return s


def _compose_meta_v13(row: Dict[str, Any]) -> str:
    name = clean_text(row.get("name", "Produk PCE"))
    fam = row.get("family", DEFAULT_FAMILY) if isinstance(row.get("family", DEFAULT_FAMILY), dict) else DEFAULT_FAMILY
    is_acc = bool(row.get("is_accessory", False))
    if is_acc:
        base = f"{name} adalah aksesori PCE untuk mendukung penggunaan unit utama yang kompatibel sesuai kebutuhan pengujian teknis."
    else:
        category = public_category(row.get("category", ""), fam, False)
        if fam.get("id") == "moisture_humidity":
            base = f"{name} untuk pengukuran kadar air atau kelembapan material pada kebutuhan laboratorium, produksi, dan kontrol kualitas."
        elif fam.get("id") == "scale_lab":
            base = f"{name} untuk kebutuhan laboratorium, preparasi sampel, penimbangan, atau kontrol kualitas sesuai spesifikasi produk."
        else:
            base = f"{name} untuk {category.lower()} pada kebutuhan laboratorium, produksi, inspeksi, dan kontrol kualitas."
    return _clip_complete_sentence_v13(base, 158)


def clean_meta_description_v13(value: Any, row: Optional[Dict[str, Any]] = None, max_chars: int = 158) -> str:
    s = _clip_complete_sentence_v13(value, max_chars=max_chars)
    if not s:
        return _compose_meta_v13(row or {}) if row else ""
    if WEAK_META_END_RE.search(s) or len(s) < 70:
        return _compose_meta_v13(row or {}) if row else re.sub(WEAK_META_END_RE, ".", s)
    return s


def _sanitize_ai_text(value: Any, max_chars: int = 220, fallback: str = "") -> str:
    s = _V12_sanitize_ai_text(value, max_chars=max_chars, fallback=fallback)
    if max_chars <= 170:
        s = clean_meta_description_v13(s, None, max_chars=max_chars) or fallback
    return s


def _build_deepseek_messages(payload: Dict[str, Any]) -> List[Dict[str, str]]:
    system_prompt = (
        "Anda adalah penulis katalog teknis berbahasa Indonesia untuk uji.co.id. "
        "Tulis copy produk yang natural, rapi, dan faktual. Gunakan hanya data yang diberikan. "
        "Jangan mengarang spesifikasi, isi paket, harga, stok, promo, garansi, sertifikasi, atau klaim akurasi. "
        "Nama produk/SKU/model harus dipertahankan apa adanya. Satuan teknis seperti °C, %, ppm, g, Hz boleh tetap. "
        "Hindari campuran Inggris-Indonesia kecuali istilah teknis, nama produk, satuan, atau model. "
        "Jangan memasukkan tabel spesifikasi atau delivery scope ke product_description_html karena keduanya punya tab resmi terpisah. "
        "Meta description wajib berupa satu kalimat utuh, tidak boleh berakhir dengan kata sambung seperti 'dan', 'untuk', 'dengan', atau 'yang'. "
        "Balas JSON valid saja."
    )
    user_prompt = {
        "task": "Buat konten katalog WooCommerce uji.co.id dalam bahasa Indonesia.",
        "output_schema": {
            "product_description_html": "HTML: 2 paragraf pembuka, <h2>Fungsi dan Keunggulan ...</h2> + <ul>, <h2>Contoh Penggunaan</h2> + paragraf/ul, <h2>Informasi Pemilihan Produk</h2> + paragraf. Tanpa CTA kontak.",
            "short_description": "1 kalimat ringkas maksimal 220 karakter.",
            "feature_tab_html": "HTML <h2>Keunggulan Produk</h2><ul> berisi 4-7 poin natural dan faktual. Jangan masukkan delivery scope.",
            "faq_html": "HTML berisi 4-5 pasang <h3>pertanyaan</h3><p>jawaban</p>. FAQ harus spesifik sesuai jenis produk.",
            "meta_description": "Maksimal 158 karakter, satu kalimat utuh, natural untuk SEO, tidak terpotong di tengah frasa.",
        },
        "rules": [
            "Output hanya JSON valid, tanpa markdown dan tanpa code fence.",
            "Jangan tulis kalimat seperti 'Data produk mencantumkan...' kecuali benar-benar perlu.",
            "Jangan memasukkan parameter mentah yang tidak jelas ke deskripsi utama.",
            "Jika data spesifikasi kurang lengkap, arahkan pembaca untuk mencocokkan spesifikasi resmi tanpa mengarang angka.",
            "Untuk aksesori, jelaskan kompatibilitas dengan unit utama dan jangan sebut sebagai alat ukur mandiri.",
            "Gunakan tag HTML yang aman: p, strong, ul, li, h2, h3, em.",
            "Jangan menulis daftar kelengkapan pengiriman pada product_description_html atau feature_tab_html.",
            "Meta description harus selesai sebagai kalimat normal; hindari akhiran: dan, atau, untuk, dengan, yang, seperti, pada.",
        ],
        "product_data": payload,
    }
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(user_prompt, ensure_ascii=False)},
    ]


def meta_description(row: Dict[str, Any]) -> str:
    fields = get_deepseek_ai_content(row)
    value = fields.get("meta_description", "") if fields else ""
    if value:
        return clean_meta_description_v13(value, row, 158)
    return clean_meta_description_v13(RULE_meta_description(row), row, 158)


def row_to_output(norm: Dict[str, Any], idx: int) -> Dict[str, Any]:
    out = _V12_row_to_output(norm, idx)
    # Final post-processing guard, including values that may come from AI cache/fallback.
    out["meta description"] = clean_meta_description_v13(out.get("meta description", ""), norm, 158)
    if clean_text(out.get("custom_tab_3_content", "")):
        out["custom_tab_3_content"] = generate_delivery_scope_tab(norm)
    log = clean_text(out.get("Website Correction Log", ""))
    extra = "v13_cleanup_meta_delivery_spec_category"
    if extra not in log:
        out["Website Correction Log"] = "; ".join([x for x in [log, extra] if x])
    if "v11" in clean_text(out.get("Processing Time", "")) or "v10" in clean_text(out.get("Processing Time", "")):
        out["Processing Time"] = clean_text(out.get("Processing Time", "")).replace("v11_deepseek_ai", "v13_deepseek_ai_cleaner").replace("v10_dedup_delivery_faq_last", "v13_deepseek_ai_cleaner")
    return out


# ---------------------------------------------------------------------------
# V14 override: strict official Delivery Scope handling.
# - Kelengkapan Pengiriman is created only when a valid official Delivery Scope
#   section was scraped from the product page.
# - If the official page has no delivery scope, scraping is disabled, or scraping
#   fails, custom_tab_3_title/content/priority are left blank.
# - AI is never allowed to invent delivery contents.
# ---------------------------------------------------------------------------

_V13_row_to_output = row_to_output


def delivery_skip_reason_v14(norm: Dict[str, Any]) -> str:
    source_url = clean_text(norm.get("source_url", ""))
    scrape_status = clean_text(norm.get("scrape_status", ""))
    scrape_error = clean_text(norm.get("scrape_error", ""))
    if not source_url:
        return "delivery_scope_skipped_no_source_url"
    if scrape_status == "disabled":
        return "delivery_scope_skipped_scrape_disabled"
    if scrape_error or scrape_status in {"failed", "no_description_found"}:
        return "delivery_scope_skipped_scrape_failed"
    if scrape_status and scrape_status not in {"ok", "website_cache"} and norm.get("description_origin") not in {"website", "website_cache"}:
        return f"delivery_scope_skipped_scrape_{re.sub(r'[^a-zA-Z0-9_]+', '_', scrape_status)[:40]}"
    return "delivery_scope_skipped_official_empty"


def delivery_progress_status_v14(norm: Dict[str, Any]) -> str:
    content = generate_delivery_scope_tab(norm)
    if clean_text(content):
        return "delivery=found"
    reason = delivery_skip_reason_v14(norm)
    if reason == "delivery_scope_skipped_official_empty":
        return "delivery=skipped:official_empty"
    if reason == "delivery_scope_skipped_scrape_failed":
        return "delivery=skipped:scrape_failed"
    if reason == "delivery_scope_skipped_scrape_disabled":
        return "delivery=skipped:scrape_disabled"
    if reason == "delivery_scope_skipped_no_source_url":
        return "delivery=skipped:no_url"
    return "delivery=skipped"


def row_to_output(norm: Dict[str, Any], idx: int) -> Dict[str, Any]:
    out = _V13_row_to_output(norm, idx)

    # Final strict guard: Delivery Scope must come from valid scraped official items.
    delivery_content = generate_delivery_scope_tab(norm)
    log = clean_text(out.get("Website Correction Log", ""))
    log_parts = [x.strip() for x in log.split(";") if x.strip()]

    # Remove older/ambiguous delivery log markers so the reason is clear.
    log_parts = [x for x in log_parts if not x.startswith("delivery_scope_")]

    if clean_text(delivery_content):
        out["custom_tab_3_title"] = "Kelengkapan Pengiriman"
        out["custom_tab_3_content"] = delivery_content
        out["custom_tab_3_priority"] = 30
        log_parts.append("delivery_scope_found_official")
    else:
        out["custom_tab_3_title"] = ""
        out["custom_tab_3_content"] = ""
        out["custom_tab_3_priority"] = ""
        log_parts.append(delivery_skip_reason_v14(norm))

    extra = "v14_strict_official_delivery_only"
    if extra not in log_parts:
        log_parts.append(extra)
    out["Website Correction Log"] = "; ".join(log_parts)
    out["Processing Time"] = clean_text(out.get("Processing Time", "")).replace("v13_deepseek_ai_cleaner", "v14_deepseek_ai_strict_delivery")
    if "v14_deepseek_ai_strict_delivery" not in out["Processing Time"]:
        out["Processing Time"] = (out["Processing Time"] + "_v14_deepseek_ai_strict_delivery").strip("_")
    return out



# ---------------------------------------------------------------------------
# V15 override: publish cleaner for Delivery Scope and Specification tables.
# Fokus:
# - menghapus catatan non-delivery seperti kalimat sertifikat/custom return;
# - membuang item delivery yang jelas berasal dari produk lain/tab lain;
# - memperbaiki sisa campuran Inggris-Indonesia pada Delivery Scope;
# - memperbaiki frasa spesifikasi campuran seperti "according hingga the", "of unit",
#   "Drying suhu rentang", "sample cup", dan sejenisnya;
# - tetap tidak membuat Kelengkapan Pengiriman jika official Delivery Scope tidak valid.
# ---------------------------------------------------------------------------

V15_NON_DELIVERY_NOTE_RE = re.compile(
    r"\b(?:calibration certificates? are issued|excluded from the right of return|right of return|"
    r"available in the accessories tab|available as an accessory|soft[- ]?lut[- ]?usb|"
    r"all gas sensors are calibrated|certification included|download|data sheet)\b",
    re.I,
)

V15_ALLOWED_DELIVERY_WORD_RE = re.compile(
    r"\b(?:kabel|cable|manual|petunjuk|panduan|baterai|battery|tas|bag|case|koper|kotak|"
    r"sertifikat|certificate|kalibrasi|calibration|sensor|sonda|probe|elektroda|electrode|"
    r"dudukan|holder|pegangan|handle|adaptor|adapter|clip|klip|tang|clamp|coil|koil|"
    r"software|perangkat lunak|flashdisk|usb|catu daya|power supply|pelindung|wadah|sample|"
    r"sampel|anak timbangan|weight|obeng|screwdriver|sikat|brush|pin|tip|lead|test|uji)\b",
    re.I,
)

V15_OTHER_PRODUCT_WORDS = [
    "gasflag", "gasmaster", "xgard", "gasman", "tcgard", "pce-ma", "pce-pa", "pce-gpa", "pce-t", "pce-ir", "pce-sm", "pce-gmm", "pce-wmh", "pce-smm",
]


def _product_model_tokens_v15(name: str) -> List[str]:
    s = clean_text(name).lower()
    tokens = set()
    for m in re.finditer(r"\b(?:pce[-\s]?[a-z0-9]+(?:[-\s]?[a-z0-9]+){0,3}|gasflag|gasmaster|gasman|tcgard|xgard)\b", s, flags=re.I):
        tok = re.sub(r"\s+", "-", m.group(0).lower())
        tokens.add(tok)
    # Also add family words that are part of official name.
    for word in ["gasflag", "gasmaster", "gasman", "tcgard", "xgard"]:
        if word in s:
            tokens.add(word)
    return sorted(tokens)


def _delivery_item_matches_product_v15(item: str, product_name: str) -> bool:
    """Reject obvious product-cross contamination, but keep generic accessories."""
    s = clean_text(item).lower()
    name = clean_text(product_name).lower()
    if not s:
        return False
    if V15_NON_DELIVERY_NOTE_RE.search(s):
        return False

    # If an item mentions another main product family that is not in the current product name,
    # keep it only when it is clearly an accessory/manual for that product in the current bundle.
    for word in V15_OTHER_PRODUCT_WORDS:
        if word in s and word not in name:
            # Manual/standalone product names from other tabs are common false positives on PCE pages.
            if re.search(r"\b(?:manual|buku manual|petunjuk|instructions?)\b", s, flags=re.I):
                return False
            # A bare product item such as "1 × Gasflag" / "1 × Gasmaster" is not a delivery
            # item for TCgard and similar pages.
            if re.fullmatch(r"\s*\d+\s*(?:×|x)\s*" + re.escape(word) + r"(?:\s*\([^)]*\))?\s*", s, flags=re.I):
                return False
            # If it has no accessory terms, treat it as cross-product contamination.
            if not V15_ALLOWED_DELIVERY_WORD_RE.search(s):
                return False
    return True


_V14_normalize_public_phrase = normalize_public_phrase


def normalize_public_phrase(label: str) -> str:
    s = _V14_normalize_public_phrase(label)
    if not s:
        return ""
    fixes = [
        (r"\baccording\s+hingga\s+the\s+rentang\s+pengukuran\b", "sesuai rentang pengukuran"),
        (r"\baccording\s+hingga\s+rentang\s+pengukuran\b", "sesuai rentang pengukuran"),
        (r"\bin\s+(\d+)\s+pengukuran\s+rentang\b", r"dalam \1 rentang pengukuran"),
        (r"\b(>\s*\d+(?:[,.]\d+)?)\s*V\s+y\s*(>\s*\d+(?:[,.]\d+)?)\s*A\b", r"\1 V dan \2 A"),
        (r"\b(<\s*\d+(?:[,.]\d+)?)\s*V\s+y\s*(<\s*\d+(?:[,.]\d+)?)\s*A\b", r"\1 V dan \2 A"),
        (r"\bPada rentang the akurasi of the device\b", "Sesuai akurasi perangkat pada rentang pengukuran"),
        (r"\bRepeatabilitas\s+\|?\s*Pada rentang the akurasi of the device\b", "Repeatabilitas | Sesuai akurasi perangkat pada rentang pengukuran"),
        (r"\bMinimum diameter of the pengukuran point\b", "Diameter minimum titik pengukuran"),
        (r"\bDiameter:\s*", "Diameter: "),
        (r"\bLength:\s*", "Panjang: "),
        (r"\bThermal element\b", "elemen termal"),
        (r"\bSensor jenis\b", "Jenis sensor"),
        (r"\bGas sensor jenis\b", "Jenis sensor gas"),
        (r"\bKeluaran signal\b", "Keluaran sinyal"),
        (r"\bProtection\b", "Perlindungan"),
        (r"\bwith cap resistant hingga inclement weather\b", "dengan penutup yang tahan cuaca buruk"),
        (r"\bconductors?\b", "konduktor"),
        (r"\bLine height\b", "Tinggi baris"),
        (r"\bSymbols? untuk\b", "Simbol untuk"),
        (r"\bMax-tahan baca\b", "Max-Hold"),
        (r"\bBatt-Low\b", "baterai lemah"),
        (r"\bzero kalibrasi\b", "kalibrasi nol"),
        (r"\bthe\s+8\s+h\s+TWA\s+alarm\b", "alarm TWA 8 jam"),
        (r"\b15\s+min\s+dan\s+8\s+h\s+TWA\s+alarm\b", "alarm TWA 15 menit dan 8 jam"),
        (r"\bAdjustable acquisition interval,? maximum capacity of 900 hours of continuous pengukuran at a laju pengukuran of one minute \(depending on the gas concentration\)\b",
         "Interval akuisisi dapat diatur, dengan kapasitas hingga 900 jam pengukuran kontinu pada laju pengukuran 1 menit tergantung konsentrasi gas"),
        (r"\bStandard perangkat lunak is available as an aksesori untuk configuration dan kalibrasi as well\b",
         "Perangkat lunak standar tersedia sebagai aksesori untuk konfigurasi dan kalibrasi"),
        (r"\bToxic gases dan oxygen lithium baterai \(replaceable\) \(24 months intermittent use\)\b",
         "Baterai litium untuk gas toksik dan oksigen, dapat diganti, hingga 24 bulan untuk penggunaan intermiten"),
        (r"\bDaya factor\b", "Faktor daya"),
        (r"\bAC\s+V\s+masukan\s+impedance\b", "Impedansi masukan AC V"),
        (r"\bOverload protection\b", "Perlindungan beban berlebih"),
        (r"\bData penyimpanan\b", "Penyimpanan data"),
        (r"\bPenyimpanan rate\b", "Laju penyimpanan"),
        (r"\bPenyimpanan format\b", "Format penyimpanan"),
        (r"\blive presentation hingga PC\b", "tampilan langsung ke PC"),
        (r"\bJaw kapasitas\b", "Kapasitas rahang"),
        (r"\bAbout\b", "Sekitar"),
        (r"\bBerat percentage\b", "persentase berat"),
        (r"\bDrying suhu rentang\b", "Rentang suhu pengeringan"),
        (r"\bHeating suhu\b", "Suhu pemanasan"),
        (r"\bKadar air content pengukuran rentang\b", "Rentang pengukuran kadar air"),
        (r"\bKadar air content resolusi atau keterbacaan\b", "Resolusi kadar air"),
        (r"\bPenimbangan resolusi atau keterbacaan\b", "Resolusi timbang"),
        (r"\bStores hingga (\d+) drying profiles atau programs\b", r"Menyimpan hingga \1 profil atau program pengeringan"),
        (r"\bTampilan layar\b", "Tampilan"),
        (r"\bdigit height\b", "tinggi digit"),
        (r"\bDimensi of unit\b", "Dimensi unit"),
        (r"\bBerat of unit\b", "Berat unit"),
        (r"\bMemori kapasitas\b", "Kapasitas memori"),
        (r"\bOuter Packaging\b", "kemasan luar"),
        (r"\bAdditional dimensi\b", "Dimensi tambahan"),
        (r"\bSample pan\b", "Wadah sampel"),
        (r"\bSchuko Plug\b", "steker Schuko"),
        (r"\bLC Touch layar\b", "layar sentuh LC"),
        (r"\bTouch layar\b", "layar sentuh"),
        (r"\binches\b", "inci"),
        (r"\btimes per detik\b", "kali/detik"),
        (r"\bper detik\b", "/detik"),
        (r"\bPackage\b", "kemasan"),
        (r"\bconfiguration\b", "konfigurasi"),
        (r"\bkalibrasi as well\b", "kalibrasi"),
        (r"\bdepending on the gas concentration\b", "tergantung konsentrasi gas"),
        (r"\bdepending\b", "tergantung"),
        (r"\bthe\b", ""),
        (r"\bwith\b", "dengan"),
        (r"\bbetween\b", "antara"),
        (r"\bdegree\b", "derajat"),
        (r"\bof unit\b", "unit"),
    ]
    for pat, repl in fixes:
        s = re.sub(pat, repl, s, flags=re.I)
    # Clean unit formatting and accidental leftovers.
    s = re.sub(r"\b(\d+(?:[,.]\d+)?)\s*\.\.\.\s*(\d+(?:[,.]\d+)?)\b", r"\1... \2", s)
    s = re.sub(r"\s+(/|-)\s+", r" \1 ", s)
    s = re.sub(r"\s+([,.;:])", r"\1", s)
    s = re.sub(r"\s{2,}", " ", s).strip(" .;:")
    # Restore ellipsis spacing after generic cleanup.
    s = re.sub(r"(\d)\.\.\.\s*(\d)", r"\1... \2", s)
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


_V14_translate_spec_label = translate_spec_label


def translate_spec_label(label: str) -> str:
    raw = clean_text(label)
    low = raw.lower().strip(" :-–—")
    label_map = {
        "gas sensor type": "Jenis sensor gas",
        "sensor type": "Jenis sensor",
        "minimum diameter of the measurement point": "Diameter minimum titik pengukuran",
        "drying temperature range": "Rentang suhu pengeringan",
        "heating temperature": "Suhu pemanasan",
        "moisture content measurement range": "Rentang pengukuran kadar air",
        "moisture content resolution or readability": "Resolusi kadar air",
        "weighing resolution or readability": "Resolusi timbang",
        "data storage": "Penyimpanan data",
        "storage rate": "Laju penyimpanan",
        "storage format": "Format penyimpanan",
        "jaw capacity": "Kapasitas rahang",
        "ac v input impedance": "Impedansi masukan AC V",
        "overload protection": "Perlindungan beban berlebih",
    }
    if low in label_map:
        return normalize_public_phrase(label_map[low])
    return normalize_public_phrase(_V14_translate_spec_label(label))


_V14_translate_spec_value = translate_spec_value


def translate_spec_value(value: str, label: str = "") -> str:
    s = _V14_translate_spec_value(value, label)
    s = normalize_public_phrase(s)
    # Extra cleanup specifically for long values.
    extras = [
        (r"\b±\s*([0-9,.]+)%\s+dari\s+nilai\s*\+\s*([0-9,.]+)\s*digit\b", r"± \1% dari nilai + \2 digit"),
        (r"\b(>\s*20\s*V)\s+y\s*(>\s*20\s*A)\b", r"\1 dan \2"),
        (r"\b0,001\b", "0,001"),
        (r"\b2\s*x\s*1,5V\b", "2 × 1,5 V"),
        (r"\b1\.5V\b", "1,5 V"),
        (r"\b4-mA\b", "4 mA"),
        (r"\b20-mA\b", "20 mA"),
        (r"\bRS232\b", "RS-232"),
    ]
    for pat, repl in extras:
        s = re.sub(pat, repl, s, flags=re.I)
    s = re.sub(r"\s+([,.;:])", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip(" .;:")
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


V15_DELIVERY_EXTRA_REPLACEMENTS: List[Tuple[str, str]] = [
    (r"\buser[’']?s\s+buku\s+manual\b", "buku manual"),
    (r"\bInstructions?\b", "petunjuk penggunaan"),
    (r"\bQuickstart\s+Guide\b", "panduan mulai cepat"),
    (r"\bService\s+bag\b", "tas servis"),
    (r"\bPocket\s+clip\b", "klip saku"),
    (r"\bUSB\s+stick\s+dengan\s+petunjuk\s+penggunaan\s+dan\s+PC\s+perangkat\s+lunak\b", "flashdisk USB berisi petunjuk penggunaan dan perangkat lunak PC"),
    (r"\bUSB\s+stick\s+with\s+instructions?\s+and\s+PC\s+perangkat\s+lunak\b", "flashdisk USB berisi petunjuk penggunaan dan perangkat lunak PC"),
    (r"\bUSB\s+stick\s+with\s+petunjuk\s+penggunaan\s+and\s+perangkat\s+lunak\s+PC\b", "flashdisk USB berisi petunjuk penggunaan dan perangkat lunak PC"),
    (r"\bUSB\s+stick\b", "flashdisk USB"),
    (r"\bPC\s+perangkat\s+lunak\b", "perangkat lunak PC"),
    (r"\bperangkat\s+lunak\s+\(in\s+English\)", "perangkat lunak (bahasa Inggris)"),
    (r"\bRS-232\s+kabel\b", "kabel RS-232"),
    (r"\bkabel\s+USB\s+for\s+a\s+PC\b", "kabel USB untuk PC"),
    (r"\bUSB\s+C\s+kabel\b", "kabel USB-C"),
    (r"\bMicro-kabel\s+USB\b", "kabel Micro-USB"),
    (r"\bMicro\s+SD\s+memory\s+card\b", "kartu memori Micro SD"),
    (r"\bCarrying\s+bag\b", "tas pembawa"),
    (r"\bwith\s+core\s+sonda\b", "dengan sonda inti"),
    (r"\bK-type\s+wire\s+sonda\b", "sonda kawat tipe-K"),
    (r"\bwire\s+sonda\b", "sonda kawat"),
    (r"\bIsolated\s+measuring\s+pins?\s+with\s+2\s*m\s*/\s*6[,.]6\s*ft\s+of\s+kabel\s+each\b", "pin ukur berisolasi dengan kabel 2 m masing-masing"),
    (r"\bIsolated\s+measuring\s+pins?\s+with\s+2\s*m\s+of\s+kabel\s+each\b", "pin ukur berisolasi dengan kabel 2 m masing-masing"),
    (r"\bMeasuring\s+sensor\b", "sensor ukur"),
    (r"\bMagnetic\s+sensor\b", "sensor magnetik"),
    (r"\bMeasuring\s+tip\b", "ujung ukur"),
    (r"\bmagnetic\s+measuring\s+adaptors?\b", "adaptor ukur magnetik"),
    (r"\bRogowski\s+coils?\b", "koil Rogowski"),
    (r"\bset\s+of\s+test\s+cables\b", "set kabel uji"),
    (r"\btest\s+cables\b", "kabel uji"),
    (r"\bblue\b", "biru"),
    (r"\bbrown\b", "cokelat"),
    (r"\bblack\b", "hitam"),
    (r"\bgrey\b", "abu-abu"),
    (r"\bblock\s+baterai\b", "baterai blok"),
    (r"\bpair\s+of\s+elektroda\b", "sepasang elektroda"),
    (r"\breplacement\s+sepasang\s+elektroda\b", "sepasang elektroda pengganti"),
    (r"\breplacement\s+pair\s+of\s+elektroda\b", "sepasang elektroda pengganti"),
    (r"\bcatu\s+daya\s+cord\b", "kabel daya"),
    (r"\bMains\s+kabel\s+daya\b", "kabel daya"),
    (r"\bsample\s+cup\s+dudukan\s*/\s*handle\b", "dudukan/pegangan wadah sampel"),
    (r"\bSample\s+cup\s+dudukan\s*/\s*handle\b", "dudukan/pegangan wadah sampel"),
    (r"\bwadah\s+sampel\s+handle\b", "pegangan wadah sampel"),
    (r"\bwadah\s+sampel\s+dudukan\b", "dudukan wadah sampel"),
    (r"\b100\s*g\s+weight\s+in\s+plastic\s+box\b", "anak timbangan 100 g dalam kotak plastik"),
    (r"\b100\s*g\s+anak\s+timbangan\s+kalibrasi\s+in\s+plastic\s+box\b", "anak timbangan kalibrasi 100 g dalam kotak plastik"),
    (r"\b100\s*g\s+test\s+weight\b", "anak timbangan uji 100 g"),
    (r"\btest\s+weight\b", "anak timbangan uji"),
    (r"\bISO-sertifikat\s+kalibrasi\b", "sertifikat kalibrasi ISO"),
    (r"\bISO\s+kalibrasi\s+sertifikat\b", "sertifikat kalibrasi ISO"),
    (r"\bwith\b", "dengan"),
    (r"\band\b", "dan"),
    (r"\bor\b", "atau"),
    (r"\bdepending\s+on\s+gas\s+type\b", "tergantung jenis gas"),
    (r"\bdepending\s+on\s+the\s+gas\s+type\b", "tergantung jenis gas"),
]


_V14_normalize_delivery_item_v13 = normalize_delivery_item_v13


def normalize_delivery_item_v15(text: str, product_name: str = "") -> str:
    s = clean_text(text).strip(" -–—•.;:")
    if not s:
        return ""
    if V15_NON_DELIVERY_NOTE_RE.search(s):
        return ""
    if not _delivery_item_matches_product_v15(s, product_name):
        return ""
    s = _V14_normalize_delivery_item_v13(s)
    if not s or V15_NON_DELIVERY_NOTE_RE.search(s):
        return ""
    for pat, repl in V15_DELIVERY_EXTRA_REPLACEMENTS:
        s = re.sub(pat, repl, s, flags=re.I)
    # Mixed cleanup after replacements.
    mixed = [
        (r"\bbuku\s+manual\s+Gasflag\b", ""),
        (r"\bbuku\s+manual\s+Gasmaster\b", ""),
        (r"\b1\s*×\s*Gasflag\b", ""),
        (r"\b1\s*×\s*Gasmaster\b", ""),
        (r"\b1\s*×\s*buku\s+manual\b", "1 × buku manual"),
        (r"\bmanual\s+pengguna\b", "buku manual"),
        (r"\bbuku\s+buku\s+manual\b", "buku manual"),
        (r"\bkabel\s+kabel\b", "kabel"),
        (r"\bperangkat\s+lunak\s+PC\s+perangkat\s+lunak\b", "perangkat lunak PC"),
        (r"\s*/\s*", "/"),
        (r"\b1\.5\s*V\b", "1,5 V"),
        (r"\b9V\b", "9 V"),
    ]
    for pat, repl in mixed:
        s = re.sub(pat, repl, s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -–—•.;:")
    if not s or V15_NON_DELIVERY_NOTE_RE.search(s):
        return ""
    if product_name and not _delivery_item_matches_product_v15(s, product_name):
        return ""
    if s and re.match(r"[a-zà-ÿ]", s, flags=re.I):
        s = s[:1].upper() + s[1:]
    return s


def _delivery_item_cleanup(text: str) -> str:
    return normalize_delivery_item_v15(text)


def _dedupe_delivery_items_v15(items: List[str], product_name: str = "", max_items: int = 24) -> List[str]:
    expanded: List[str] = []
    for raw_item in items:
        raw = clean_text(raw_item)
        if not raw or _is_delivery_heading_text(raw) or _is_stop_heading_text(raw):
            continue
        for part in _split_bundled_delivery_item(raw):
            item = normalize_delivery_item_v15(part, product_name)
            item = _strip_delivery_prefix(item)
            if not item or len(item) < 2 or len(item) > 240:
                continue
            expanded.append(item)

    unique: List[str] = []
    seen = set()
    for item in expanded:
        key = _delivery_key_v13(item)
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(item)

    filtered: List[str] = []
    keys = [_delivery_item_key(x) for x in unique]
    for idx, item in enumerate(unique):
        item_key = keys[idx]
        contained_short_items = 0
        for j, other_key in enumerate(keys):
            if j == idx or not other_key:
                continue
            if len(other_key) < len(item_key) and other_key in item_key:
                contained_short_items += 1
        if contained_short_items >= 2 and len(item) > 120:
            continue
        filtered.append(item)
        if len(filtered) >= max_items:
            break
    return filtered


def delivery_scope_html(items: List[str], product_name: str = "") -> str:
    cleaned = _dedupe_delivery_items_v15(items, product_name=product_name)
    if not _delivery_scope_is_valid(cleaned):
        return ""
    return "<h2>Kelengkapan Pengiriman</h2><ul>" + "".join(f"<li>{esc(item)}</li>" for item in cleaned) + "</ul>"


def generate_delivery_scope_tab(row: Dict[str, Any]) -> str:
    items = row.get("delivery_scope", [])
    if not isinstance(items, list):
        return ""
    return delivery_scope_html(items, clean_text(row.get("name", "")))


_V14_spec_table_html = spec_table_html


def spec_table_html(rows: List[Dict[str, str]]) -> str:
    clean_rows = dedupe_spec_rows(rows)
    if not clean_rows:
        return ""
    body: List[str] = []
    seen = set()
    for row in clean_rows:
        label = translate_spec_label(row.get("label", ""))
        value = translate_spec_value(row.get("value", ""), label)
        if not label or not value:
            continue
        combined = label + " " + value
        if re.search(r"\b(?:add to cart|delivery time|question|callback|price|copyright|soft-lut-usb|right of return)\b", combined, flags=re.I):
            continue
        if re.fullmatch(r"[-–—\s]+", value):
            continue
        if len(value) > 900:
            value = value[:900].rsplit(" ", 1)[0].rstrip(" ,.;:")
        key = (label.lower(), value.lower())
        if key in seen:
            continue
        seen.add(key)
        body.append(f"<tr><th>{esc(label)}</th><td>{esc(value)}</td></tr>")
    if not body:
        return ""
    return '<table class="uji-spec-table"><tbody>' + ''.join(body) + '</tbody></table>'


_V14_row_to_output = row_to_output


def row_to_output(norm: Dict[str, Any], idx: int) -> Dict[str, Any]:
    out = _V14_row_to_output(norm, idx)
    # Re-render official spec and delivery tabs through V15 cleaners.
    spec_tab_content = generate_spec_tab(norm)
    if clean_text(spec_tab_content):
        out["custom_tab_2_title"] = "SPESIFIKASI"
        out["custom_tab_2_content"] = spec_tab_content
        out["custom_tab_2_priority"] = 20
    else:
        out["custom_tab_2_title"] = ""
        out["custom_tab_2_content"] = ""
        out["custom_tab_2_priority"] = ""

    delivery_content = generate_delivery_scope_tab(norm)
    if clean_text(delivery_content):
        out["custom_tab_3_title"] = "Kelengkapan Pengiriman"
        out["custom_tab_3_content"] = delivery_content
        out["custom_tab_3_priority"] = 30
    else:
        out["custom_tab_3_title"] = ""
        out["custom_tab_3_content"] = ""
        out["custom_tab_3_priority"] = ""

    log = clean_text(out.get("Website Correction Log", ""))
    parts = [x.strip() for x in log.split(";") if x.strip()]
    extra = "v15_publish_clean_spec_delivery"
    if extra not in parts:
        parts.append(extra)
    out["Website Correction Log"] = "; ".join(parts)
    out["Processing Time"] = clean_text(out.get("Processing Time", "")).replace("v14_deepseek_ai_strict_delivery", "v15_deepseek_ai_publish_clean")
    if "v15_deepseek_ai_publish_clean" not in out["Processing Time"]:
        out["Processing Time"] = (out["Processing Time"] + "_v15_deepseek_ai_publish_clean").strip("_")
    return out

if __name__ == "__main__":
    main()
