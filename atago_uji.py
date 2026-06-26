#!/usr/bin/env python3
"""ATAGO -> uji.co.id WooCommerce catalog generator.

Script khusus ATAGO.

Fokus:
- input Excel cukup berisi link produk ATAGO;
- scrape dari halaman produk ATAGO DIRECT;
- ambil title, description, Product information, For more detail, dan gambar;
- gabungkan Product information + For more detail menjadi tab Spesifikasi;
- produk aksesori/komponen ATAGO tetap masuk output dan diberi kategori khusus;
- output memakai format kolom WooCommerce seperti script PCE.

Contoh pakai:
python atago_uji_import_ready.py --input "atago.xlsx" --output "uji_atago_import_ready.xlsx"
python atago_uji_import_ready.py --input "atago.xlsx" --sheet atago_with_price --output "uji_atago_import_ready.xlsx" --delay 0.5
"""
from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urljoin

try:
    import pandas as pd
except Exception as exc:  # pragma: no cover
    raise SystemExit("pandas wajib tersedia untuk membaca/menulis Excel. Install: pip install pandas openpyxl") from exc

try:
    import requests
except Exception as exc:  # pragma: no cover
    raise SystemExit("requests wajib tersedia untuk scraping. Install: pip install requests") from exc

try:
    from bs4 import BeautifulSoup
except Exception as exc:  # pragma: no cover
    raise SystemExit("beautifulsoup4 wajib tersedia untuk parsing HTML. Install: pip install beautifulsoup4") from exc

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:  # styling optional
    load_workbook = None


OUTPUT_COLS = [
    "Name", "Brand", "Product Description", "Product Short Description",
    "custom_tab_1_title", "custom_tab_1_content", "custom_tab_1_priority",
    "custom_tab_2_title", "custom_tab_2_content", "custom_tab_2_priority",
    "custom_tab_3_title", "custom_tab_3_content", "custom_tab_3_priority",
    "custom_tab_4_title", "custom_tab_4_content", "custom_tab_4_priority",
    "custom_tab_5_title", "custom_tab_5_content", "custom_tab_5_priority",
    "custom_tab_6_title", "custom_tab_6_content", "custom_tab_6_priority",
    "custom_tab_7_title", "custom_tab_7_content", "custom_tab_7_priority",
    "custom_tab_8_title", "custom_tab_8_content", "custom_tab_8_priority",
    "product categories", "product tags", "focus keyphrase", "meta description",
    "publication_date", "Processed", "Processing Time", "Content Quality",
    "image_url", "Source URL", "Website Validation", "Website Correction Log",
]

BRAND = "ATAGO"

ATAGO_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
}

STOP_AFTER_PRODUCT_INFO = {
    "accessories", "what's new", "products", "user support", "e-commerce", "data book",
    "company", "privacypolicy＆terms of use", "privacypolicy&terms of use",
}

NOISE_LINES = {
    "price", "please contact us", "delivery", "inquiry", "quotation", "search by products",
    "search by standard solution", "shopping guide", "atago hp", "new products", "exhibition",
    "news", "contact us", "warranty registration", "world service center", "amazon uk",
    "ebay", "lazada", "shopee sg", "shopee my", "message", "about us", "philosophy",
    "sdgs", "creditworthiness", "history", "hear from our employees", "location",
    "company introduction", "all rights reserved.", "(c) 2003- atago co.,ltd.",
}

SPEC_LABEL_PATTERNS: List[Tuple[re.Pattern[str], str]] = [
    (re.compile(r"^model\b\s*[:\-]?\s*", re.I), "Model"),
    (re.compile(r"^cat\.?\s*no\.?\s*[:\-]?\s*", re.I), "Cat.No."),
    (re.compile(r"^range\b\s*[:\-]?\s*", re.I), "Range"),
    (re.compile(r"^accuracy\b\s*[:\-]?\s*", re.I), "Accuracy"),
    (re.compile(r"^minimum\s+scale\b\s*[:\-]?\s*", re.I), "Minimum Scale"),
    (re.compile(r"^measurement\s+temp\.?\s*[:\-]?\s*", re.I), "Measurement Temp."),
    (re.compile(r"^measurement\s+temperature\b\s*[:\-]?\s*", re.I), "Measurement Temp."),
    (re.compile(r"^temperature\s+range\b\s*[:\-]?\s*", re.I), "Temperature Range"),
    (re.compile(r"^ambient\s+temperature\b\s*[:\-]?\s*", re.I), "Ambient Temperature"),
    (re.compile(r"^light\s+source\b\s*[:\-]?\s*", re.I), "Light source"),
    (re.compile(r"^output\b\s*[:\-]?\s*", re.I), "Output"),
    (re.compile(r"^interface\b\s*[:\-]?\s*", re.I), "Interface"),
    (re.compile(r"^power\s+supply\b\s*[:\-]?\s*", re.I), "Power Supply"),
    (re.compile(r"^battery\s+life\b\s*[:\-]?\s*", re.I), "Battery Life"),
    (re.compile(r"^dimensions?\s*&\s*weight\b\s*[:\-]?\s*", re.I), "Dimensions & Weight"),
    (re.compile(r"^dimensions?\b\s*[:\-]?\s*", re.I), "Dimensions"),
    (re.compile(r"^weight\b\s*[:\-]?\s*", re.I), "Weight"),
    (re.compile(r"^sample\s+volume\b\s*[:\-]?\s*", re.I), "Sample Volume"),
    (re.compile(r"^measurement\s+time\b\s*[:\-]?\s*", re.I), "Measurement Time"),
    (re.compile(r"^display\s+range\b\s*[:\-]?\s*", re.I), "Display Range"),
    (re.compile(r"^display\b\s*[:\-]?\s*", re.I), "Display"),
    (re.compile(r"^resolution\b\s*[:\-]?\s*", re.I), "Resolution"),
    (re.compile(r"^repeatability\b\s*[:\-]?\s*", re.I), "Repeatability"),
    (re.compile(r"^international\s+protection\s+class\b\s*[:\-]?\s*", re.I), "International Protection Class"),
    (re.compile(r"^protection\s+class\b\s*[:\-]?\s*", re.I), "Protection Class"),
    (re.compile(r"^materials?\b\s*[:\-]?\s*", re.I), "Material"),
    (re.compile(r"^accessories\b\s*[:\-]?\s*", re.I), "Accessories"),
]

SPEC_LABEL_ID = {
    "Model": "Model",
    "Cat.No.": "Cat.No.",
    "Range": "Rentang pengukuran",
    "Accuracy": "Akurasi",
    "Minimum Scale": "Skala minimum",
    "Measurement Temp.": "Suhu pengukuran",
    "Temperature Range": "Rentang suhu",
    "Ambient Temperature": "Suhu lingkungan",
    "Light source": "Sumber cahaya",
    "Output": "Output",
    "Interface": "Antarmuka",
    "Power Supply": "Catu daya",
    "Battery Life": "Daya tahan baterai",
    "Dimensions & Weight": "Dimensi & berat",
    "Dimensions": "Dimensi",
    "Weight": "Berat",
    "Sample Volume": "Volume sampel",
    "Measurement Time": "Waktu pengukuran",
    "Display Range": "Rentang tampilan",
    "Display": "Tampilan",
    "Resolution": "Resolusi",
    "Repeatability": "Repeatabilitas",
    "International Protection Class": "Kelas perlindungan internasional",
    "Protection Class": "Kelas perlindungan",
    "Material": "Material",
    "Accessories": "Aksesori",
}

ACCESSORY_KEYWORDS = re.compile(
    r"\b(spindle|solution|standard solution|sucrose solution|cable|printer|test piece|eyepiece|"
    r"desiccant|adapter|adaptor|case|cover|battery|charger|bottle|cup|beaker|tube|cap|"
    r"stand|holder|paper|filter|sensor replacement|replacement|spare|accessor(?:y|ies)|"
    r"pouch|strap|syringe|funnel|nozzle|needle|tube|hose|kit|lid|o-ring|washer)\b",
    re.I,
)

MAIN_PRODUCT_KEYWORDS = re.compile(
    r"\b(refractometer|brix\s*meter|salt\s*meter|ph\s*meter|acidity\s*meter|viscometer|"
    r"polarimeter|saccharimeter|density\s*meter|concentration\s*meter|in-line|inline|"
    r"digital\s*meter|rx-|pal-|master-|pen-|ap-|cm-|rx|dr-|sac-|visco|pocket)\b",
    re.I,
)

FAMILY_RULES: List[Tuple[re.Pattern[str], Dict[str, Any]]] = [
    (re.compile(r"refractometer|brix|rx-|pal-|master-|pen-|dr-", re.I), {
        "category": "Refractometer dan Brix Meter",
        "term": "refractometer ATAGO",
        "function": "mengukur indeks bias, Brix, atau konsentrasi sampel cair sesuai spesifikasi produk",
        "tags": ["ATAGO", "refractometer", "brix meter", "alat ukur brix", "alat ukur konsentrasi"],
    }),
    (re.compile(r"salt|salinity|nacl", re.I), {
        "category": "Salt Meter dan Salinity Meter",
        "term": "salt meter ATAGO",
        "function": "mengukur kadar garam atau salinitas pada sampel cair sesuai rentang produk",
        "tags": ["ATAGO", "salt meter", "salinity meter", "alat ukur garam", "alat ukur salinitas"],
    }),
    (re.compile(r"ph\s*meter|\bph\b", re.I), {
        "category": "pH Meter",
        "term": "pH meter ATAGO",
        "function": "mengukur nilai pH pada sampel cair untuk kebutuhan kontrol kualitas dan pengujian laboratorium",
        "tags": ["ATAGO", "pH meter", "alat ukur pH", "analisis cairan", "laboratorium"],
    }),
    (re.compile(r"acidity|acid", re.I), {
        "category": "Acidity Meter",
        "term": "acidity meter ATAGO",
        "function": "mengukur tingkat keasaman sampel sesuai metode dan rentang produk",
        "tags": ["ATAGO", "acidity meter", "alat ukur keasaman", "analisis makanan", "quality control"],
    }),
    (re.compile(r"viscometer|viscosity|visco", re.I), {
        "category": "Viscometer",
        "term": "viscometer ATAGO",
        "function": "mengukur viskositas atau kekentalan sampel cair sesuai sistem pengukuran produk",
        "tags": ["ATAGO", "viscometer", "alat ukur viskositas", "uji kekentalan", "quality control"],
    }),
    (re.compile(r"polarimeter|saccharimeter|sac-", re.I), {
        "category": "Polarimeter dan Saccharimeter",
        "term": "polarimeter ATAGO",
        "function": "mendukung pengukuran optik sampel untuk analisis laboratorium dan kontrol kualitas",
        "tags": ["ATAGO", "polarimeter", "saccharimeter", "analisis optik", "laboratorium"],
    }),
]

DEFAULT_FAMILY = {
    "category": "Alat Ukur Laboratorium ATAGO",
    "term": "alat ukur ATAGO",
    "function": "mendukung pengukuran sampel cair, bahan proses, atau kebutuhan kontrol kualitas sesuai spesifikasi produk",
    "tags": ["ATAGO", "alat ukur", "alat laboratorium", "quality control", "analisis sampel"],
}

ACCESSORY_FAMILY = {
    "category": "Aksesori dan Komponen ATAGO",
    "term": "aksesori atau komponen pendukung ATAGO",
    "function": "mendukung penggunaan unit utama ATAGO sesuai model, Cat.No., dan keterangan pada halaman produk",
    "tags": ["ATAGO", "aksesori ATAGO", "komponen ATAGO", "spare part ATAGO", "kelengkapan alat ATAGO"],
}

UJI_CTA = (
    '<p>Untuk konsultasi produk, silakan kunjungi '
    '<strong><a href="https://uji.co.id/about-us-3/" target="_blank" rel="noopener">uji.co.id</a></strong> '
    'atau <strong><a href="https://uji.co.id/contact-us/" target="_blank" rel="noopener">hubungi kami</a></strong>. '
    'Telepon: +62896-2784-2222.</p>'
)


# ----------------------------- basic helpers -----------------------------

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = html.unescape(str(value))
    s = s.replace("\xa0", " ").replace("\u3000", " ").replace("_x000D_", "\n")
    s = s.replace("\r", "\n")
    s = re.sub(r"[ \t\f\v]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def esc(value: Any) -> str:
    return html.escape(clean_text(value), quote=False)


def normalize_url(url: str) -> str:
    url = clean_text(url)
    if not url:
        return ""
    if url.startswith("//"):
        return "https:" + url
    if not re.match(r"https?://", url, flags=re.I):
        return "https://" + url.lstrip("/")
    return url


def compact(text: str, max_chars: int = 900) -> str:
    text = clean_text(re.sub(r"\s+", " ", clean_text(text))).strip()
    if len(text) <= max_chars:
        return text
    return text[:max_chars].rsplit(" ", 1)[0].rstrip(" ,.;:") + "."


def deterministic_index(text: str, modulo: int) -> int:
    if modulo <= 1:
        return 0
    digest = hashlib.md5(clean_text(text).encode("utf-8")).hexdigest()
    return int(digest[:8], 16) % modulo


def dedupe_keep_order(items: Iterable[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    for item in items:
        val = clean_text(item)
        if not val:
            continue
        key = val.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(val)
    return out


def sentence_case_id(text: str) -> str:
    text = clean_text(text)
    return text[:1].upper() + text[1:] if text else text


# ----------------------------- Excel IO -----------------------------

def find_url_column(df: pd.DataFrame, requested: str = "") -> str:
    if requested:
        if requested in df.columns:
            return requested
        lower = {str(c).strip().casefold(): c for c in df.columns}
        key = lower.get(requested.strip().casefold())
        if key is not None:
            return key
        raise ValueError(f"Kolom URL '{requested}' tidak ditemukan. Kolom tersedia: {list(df.columns)}")

    preferred = ["url", "link", "source url", "product url", "product_link", "product link"]
    lower = {str(c).strip().casefold(): c for c in df.columns}
    for name in preferred:
        if name in lower:
            return lower[name]

    for col in df.columns:
        values = df[col].dropna().astype(str).head(25).tolist()
        if any("atago.net" in v.lower() or v.lower().startswith("http") for v in values):
            return col

    raise ValueError("Tidak menemukan kolom URL. Pastikan Excel punya kolom url/link/Source URL atau kolom berisi link ATAGO.")


def read_input_urls(input_path: str, sheet: Any = 0, url_column: str = "") -> List[str]:
    df = pd.read_excel(input_path, sheet_name=sheet)
    if df.empty:
        return []
    col = find_url_column(df, url_column)
    urls = []
    for val in df[col].tolist():
        url = normalize_url(clean_text(val))
        if not url:
            continue
        if "atago.net" not in url.lower():
            # Tetap simpan jika pengguna sengaja isi URL non-ATAGO, tapi akan gagal validasi domain.
            pass
        urls.append(url)
    return dedupe_keep_order(urls)


# ----------------------------- scrape/parsing -----------------------------

def soup_text(node: Any) -> str:
    if node is None:
        return ""
    try:
        return clean_text(node.get_text(" ", strip=True))
    except Exception:
        return clean_text(str(node))


def page_lines(soup: BeautifulSoup) -> List[str]:
    text = soup.get_text("\n", strip=True)
    lines = []
    for raw in text.splitlines():
        line = clean_text(raw)
        if not line:
            continue
        lines.append(line)
    return lines


def normalize_line_key(line: str) -> str:
    return re.sub(r"\s+", " ", clean_text(line)).strip().casefold()


def find_line_index(lines: List[str], target: str) -> int:
    target_key = target.casefold()
    for i, line in enumerate(lines):
        if normalize_line_key(line) == target_key:
            return i
    return -1


def find_first_stop_index(lines: List[str], start: int, stops: Iterable[str]) -> int:
    stop_keys = {s.casefold() for s in stops}
    for i in range(start, len(lines)):
        if normalize_line_key(lines[i]) in stop_keys:
            return i
    return len(lines)


def clean_title(title: str) -> str:
    title = clean_text(title)
    title = re.sub(r"\s*\|\s*ATAGO DIRECT.*$", "", title, flags=re.I)
    title = re.sub(r"\s+", " ", title)
    return title.strip(" -|\t\n")


def extract_title(soup: BeautifulSoup) -> str:
    h1 = soup.find("h1")
    title = soup_text(h1)
    if not title and soup.title:
        title = soup_text(soup.title)
    return clean_title(title)


def extract_best_image_url(soup: BeautifulSoup, base_url: str) -> str:
    candidates: List[str] = []
    selectors = [
        'meta[property="og:image"]', 'meta[name="twitter:image"]', 'link[rel="image_src"]',
        'img[itemprop="image"]', '.product img', '.product-detail img', '.detail img',
        'img[data-src]', 'img[data-original]', 'img[src]',
    ]
    for sel in selectors:
        try:
            nodes = soup.select(sel)
        except Exception:
            nodes = []
        for node in nodes:
            raw = clean_text(
                node.get("content") or node.get("src") or node.get("data-src") or
                node.get("data-original") or node.get("href") or ""
            )
            if not raw:
                srcset = clean_text(node.get("srcset") or node.get("data-srcset") or "")
                if srcset:
                    raw = clean_text(srcset.split(",")[0].strip().split(" ")[0])
            if raw:
                candidates.append(urljoin(base_url, raw))

    bad = re.compile(r"(?:logo|icon|sprite|flag|payment|blank|placeholder|tracking|banner|btn)", re.I)
    preferred = re.compile(r"(?:/images/products/|img_l|products|prod|direct|atagodirect)", re.I)
    unique = []
    seen = set()
    for url in candidates:
        if not url or url in seen or bad.search(url):
            continue
        seen.add(url)
        unique.append(url)
    if not unique:
        return ""
    for url in unique:
        if preferred.search(url):
            return url
    return unique[0]


def extract_description(lines: List[str], title: str) -> str:
    info_idx = find_line_index(lines, "Product information")
    if info_idx < 0:
        return ""

    # Mulai setelah H1 jika ketemu, kalau tidak mulai beberapa baris sebelum Product information.
    title_idx = -1
    title_key = normalize_line_key(title)
    for i, line in enumerate(lines[:info_idx]):
        if normalize_line_key(line) == title_key:
            title_idx = i
            break
    start = title_idx + 1 if title_idx >= 0 else max(0, info_idx - 12)

    desc_lines: List[str] = []
    for line in lines[start:info_idx]:
        key = normalize_line_key(line)
        if key in NOISE_LINES:
            continue
        if re.fullmatch(r"\$?\d+[\d,.]*\s*(?:usd)?", line, flags=re.I):
            continue
        if len(line) < 25:
            continue
        desc_lines.append(line)
    return compact(" ".join(desc_lines), max_chars=1300)


def extract_product_info_lines(lines: List[str]) -> List[str]:
    start = find_line_index(lines, "Product information")
    if start < 0:
        return []
    end = find_first_stop_index(lines, start + 1, STOP_AFTER_PRODUCT_INFO)
    out = []
    for line in lines[start + 1:end]:
        key = normalize_line_key(line)
        if not key or key in NOISE_LINES:
            continue
        out.append(line)
    return out


def match_spec_label(line: str) -> Tuple[str, str]:
    line = clean_text(line)
    for pattern, label in SPEC_LABEL_PATTERNS:
        m = pattern.match(line)
        if m:
            rest = clean_text(line[m.end():]).strip(" :-–—")
            return label, rest
    return "", ""


def parse_spec_rows(product_info_lines: List[str]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    current: Optional[Dict[str, str]] = None
    in_more_detail = False

    for raw in product_info_lines:
        line = clean_text(raw)
        if not line:
            continue
        if normalize_line_key(line) == "for more detail":
            in_more_detail = True
            current = None
            continue

        label, value = match_spec_label(line)
        if label:
            row = {
                "label": SPEC_LABEL_ID.get(label, label),
                "value": value,
                "section": "For more detail" if in_more_detail else "Product information",
                "raw_label": label,
            }
            rows.append(row)
            current = row
            continue

        # Baris lanjutan, misalnya "Brix : 0.0 to 100.0%" setelah Range.
        if current is not None:
            add = line
            if add:
                current["value"] = clean_text(((current.get("value") or "") + " / " + add).strip(" /"))
            continue

        # Fallback jika ATAGO menampilkan label tidak dikenal tetapi formatnya masih key:value.
        if ":" in line and len(line) <= 160:
            left, right = line.split(":", 1)
            left = clean_text(left).strip(" :-–—")
            right = clean_text(right).strip(" :-–—")
            if 2 <= len(left) <= 45 and right:
                rows.append({
                    "label": sentence_case_id(left),
                    "value": right,
                    "section": "For more detail" if in_more_detail else "Product information",
                    "raw_label": left,
                })
                current = rows[-1]

    return dedupe_spec_rows(rows)


def dedupe_spec_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    seen = set()
    for row in rows:
        label = clean_text(row.get("label"))
        value = clean_text(row.get("value"))
        if not label or not value:
            continue
        key = (label.casefold(), value.casefold())
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "label": label,
            "value": value,
            "section": clean_text(row.get("section")) or "Product information",
            "raw_label": clean_text(row.get("raw_label")) or label,
        })
    return out


def is_accessory_product(title: str, description: str, spec_rows: List[Dict[str, str]]) -> Tuple[bool, str]:
    title_clean = clean_text(title)
    title_low = title_clean.lower()
    labels = {clean_text(r.get("label")).casefold() for r in spec_rows}
    non_identity_specs = [r for r in spec_rows if clean_text(r.get("label")).casefold() not in {"model", "cat.no."}]

    has_main_keyword = MAIN_PRODUCT_KEYWORDS.search(title_clean) is not None
    has_accessory_keyword = ACCESSORY_KEYWORDS.search(title_clean) is not None
    has_real_specs = bool(non_identity_specs)
    has_desc = len(clean_text(description)) >= 35

    # Aksesori ATAGO biasanya hanya punya Model + Cat.No. dan tidak punya deskripsi/range/accuracy/detail teknis.
    if not has_desc and not has_real_specs:
        return True, "only_model_catno_no_description_or_specs"

    # Kata aksesori eksplisit di title tetap dilewati, kecuali jelas produk utama dan punya spesifikasi nyata.
    if has_accessory_keyword and not (has_main_keyword and has_real_specs):
        return True, "accessory_keyword_in_title"

    # Jika tidak ada parameter penting, jangan dipaksa sebagai produk utama.
    important_labels = {"rentang pengukuran", "akurasi", "skala minimum", "resolusi", "suhu pengukuran", "output"}
    if not has_desc and labels.isdisjoint(important_labels):
        return True, "no_main_product_description_or_technical_specs"

    return False, "main_product_detected"


def scrape_atago_product_page(url: str, timeout: int = 20, session: Optional[requests.Session] = None) -> Dict[str, Any]:
    url = normalize_url(url)
    result: Dict[str, Any] = {
        "url": url,
        "status": "failed",
        "title": "",
        "description": "",
        "spec_rows": [],
        "image_url": "",
        "is_accessory": False,
        "accessory_reason": "",
        "error": "",
    }
    if not url:
        result["status"] = "no_url"
        return result
    if "atago.net" not in url.lower():
        result["status"] = "invalid_domain"
        result["error"] = "URL bukan domain atago.net"
        return result

    sess = session or requests.Session()
    try:
        resp = sess.get(url, headers=ATAGO_HEADERS, timeout=timeout)
        resp.raise_for_status()
    except Exception as exc:
        result["status"] = "request_error"
        result["error"] = f"{type(exc).__name__}: {exc}"
        return result

    soup = BeautifulSoup(resp.text or "", "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    lines = page_lines(soup)
    title = extract_title(soup)
    desc = extract_description(lines, title)
    product_info = extract_product_info_lines(lines)
    spec_rows = parse_spec_rows(product_info)
    image_url = extract_best_image_url(soup, url)
    is_acc, reason = is_accessory_product(title, desc, spec_rows)

    result.update({
        "status": "ok" if title and spec_rows else "partial",
        "title": title,
        "description": desc,
        "product_info_lines": product_info,
        "spec_rows": spec_rows,
        "image_url": image_url,
        "is_accessory": is_acc,
        "accessory_reason": reason,
    })
    return result



# ----------------------------- DeepSeek AI generation -----------------------------

AI_PROMPT_VERSION = "atago_uji_deepseek_v2_include_accessories_2026_06_03"
AI_SETTINGS: Dict[str, Any] = {
    "enabled": False,
    "api_key": "",
    "base_url": "https://api.deepseek.com",
    "model": "deepseek-chat",
    "cache": {},
    "cache_path": "",
    "limit": 0,
    "timeout": 60,
    "delay": 0.0,
    "temperature": 0.25,
}
AI_STATS: Dict[str, int] = {"used": 0, "cache_hit": 0, "fallback": 0, "skipped": 0}


def set_deepseek_ai_settings(
    enabled: bool = False,
    api_key: str = "",
    base_url: str = "https://api.deepseek.com",
    model: str = "deepseek-chat",
    cache: Optional[Dict[str, Any]] = None,
    cache_path: str = "",
    limit: int = 0,
    timeout: int = 60,
    delay: float = 0.0,
    temperature: float = 0.25,
) -> None:
    """Configure DeepSeek/OpenAI-compatible AI generation.

    AI hanya dipakai untuk copy publik: Product Description, Short Description,
    Keunggulan, FAQ, dan meta description. Spesifikasi tetap berasal dari scrape
    Product information + For more detail halaman ATAGO.
    """
    AI_SETTINGS.update({
        "enabled": bool(enabled),
        "api_key": clean_text(api_key),
        "base_url": clean_text(base_url).rstrip("/") or "https://api.deepseek.com",
        "model": clean_text(model) or "deepseek-chat",
        "cache": cache if isinstance(cache, dict) else {},
        "cache_path": clean_text(cache_path),
        "limit": max(0, int(limit or 0)),
        "timeout": max(10, int(timeout or 60)),
        "delay": max(0.0, float(delay or 0.0)),
        "temperature": max(0.0, min(1.0, float(temperature if temperature is not None else 0.25))),
    })


def _truncate_for_ai(text: Any, max_chars: int) -> str:
    s = clean_text(text)
    if len(s) <= max_chars:
        return s
    return s[:max_chars].rsplit(" ", 1)[0].rstrip(" ,.;:") + "..."


def _html_to_text_for_ai(value: Any, max_chars: int = 1600) -> str:
    s = clean_text(value)
    if not s:
        return ""
    s = re.sub(r"<\s*br\s*/?>", "\n", s, flags=re.I)
    s = re.sub(r"</\s*(p|li|h2|h3|tr)\s*>", "\n", s, flags=re.I)
    s = re.sub(r"<[^>]+>", " ", s)
    s = html.unescape(s)
    s = re.sub(r"\s+", " ", s).strip()
    return _truncate_for_ai(s, max_chars)


def _spec_rows_for_ai(spec_rows: List[Dict[str, str]], limit: int = 18) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    for item in (spec_rows or [])[:limit]:
        if not isinstance(item, dict):
            continue
        label = _truncate_for_ai(item.get("label", ""), 80)
        value = _truncate_for_ai(item.get("value", ""), 180)
        section = _truncate_for_ai(item.get("section", ""), 60)
        if label and value:
            out.append({"label": label, "value": value, "section": section})
    return out


def _family_for_ai(family: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "category": clean_text(family.get("category", DEFAULT_FAMILY["category"])),
        "term": clean_text(family.get("term", DEFAULT_FAMILY["term"])),
        "function": clean_text(family.get("function", DEFAULT_FAMILY["function"])),
        "tags": [clean_text(x) for x in family.get("tags", []) if clean_text(x)][:8],
    }


def _data_payload_for_ai(data: Dict[str, Any], family: Dict[str, Any]) -> Dict[str, Any]:
    spec_rows = data.get("spec_rows") if isinstance(data.get("spec_rows"), list) else []
    is_accessory = bool(data.get("is_accessory"))
    return {
        "name": clean_text(data.get("title", "")),
        "brand": BRAND,
        "source_url": clean_text(data.get("url", "")),
        "source_description": _truncate_for_ai(data.get("description", ""), 1600),
        "source_sections": {
            "product_information_and_for_more_detail": _spec_rows_for_ai(spec_rows),
        },
        "family": _family_for_ai(family),
        "product_type": "accessory_or_component" if is_accessory else "main_product",
        "accessory_reason": clean_text(data.get("accessory_reason", "")),
        "image_url_detected": clean_text(data.get("image_url", "")),
        "notes": [
            "Spesifikasi resmi akan ditampilkan di tab Spesifikasi, jadi jangan membuat ulang tabel spesifikasi di Product Description.",
            "Jika product_type adalah accessory_or_component, tulis sebagai aksesori/komponen pendukung ATAGO, bukan sebagai alat ukur utama.",
            "Gunakan hanya klaim yang didukung source_description atau specification rows.",
        ],
    }


def _ai_cache_key(payload: Dict[str, Any]) -> str:
    key_material = {
        "prompt_version": AI_PROMPT_VERSION,
        "model": AI_SETTINGS.get("model", ""),
        "payload": payload,
    }
    raw = json.dumps(key_material, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _strip_code_fences(text: str) -> str:
    s = clean_text(text)
    s = re.sub(r"^```(?:json|html)?\s*", "", s, flags=re.I)
    s = re.sub(r"\s*```$", "", s)
    return s.strip()


def _extract_json_object(text: str) -> Dict[str, Any]:
    s = _strip_code_fences(text)
    try:
        data = json.loads(s)
        return data if isinstance(data, dict) else {}
    except Exception:
        pass
    start = s.find("{")
    end = s.rfind("}")
    if start >= 0 and end > start:
        try:
            data = json.loads(s[start:end + 1])
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def _sanitize_ai_html(value: Any, fallback: str = "") -> str:
    s = _strip_code_fences(clean_text(value))
    if not s:
        return fallback
    s = re.sub(r"<\s*(script|style|iframe|object|embed)[^>]*>.*?<\s*/\s*\1\s*>", "", s, flags=re.I | re.S)
    s = re.sub(r"\s+on[a-z]+\s*=\s*(['\"]).*?\1", "", s, flags=re.I | re.S)
    s = re.sub(r"\s+(style|class|id)\s*=\s*(['\"]).*?\2", "", s, flags=re.I | re.S)
    allowed = {"p", "strong", "b", "em", "ul", "ol", "li", "h2", "h3", "br"}

    def repl_tag(m: re.Match) -> str:
        slash = "/" if m.group(1) else ""
        tag = (m.group(2) or "").lower()
        if tag not in allowed:
            return ""
        if tag == "br":
            return "<br>"
        return f"<{slash}{tag}>"

    s = re.sub(r"<\s*(/?)\s*([a-zA-Z0-9]+)(?:\s+[^>]*)?>", repl_tag, s)
    s = re.sub(r"\n{3,}", "\n\n", s).strip()
    return s or fallback


def _sanitize_ai_text(value: Any, max_chars: int = 220, fallback: str = "") -> str:
    s = _html_to_text_for_ai(value, max_chars=max_chars * 2)
    s = re.sub(r"\s+", " ", s).strip(" \n\t.,;")
    if not s:
        return fallback
    if len(s) > max_chars:
        s = s[:max_chars].rsplit(" ", 1)[0].rstrip(" ,.;:") + "."
    if s and s[-1] not in ".!?":
        s += "."
    return s


def _validate_ai_content(data: Dict[str, Any]) -> Tuple[bool, str, Dict[str, str]]:
    fields = {
        "product_description_html": _sanitize_ai_html(data.get("product_description_html", "")),
        "short_description": _sanitize_ai_text(data.get("short_description", ""), max_chars=260),
        "feature_tab_html": _sanitize_ai_html(data.get("feature_tab_html", "")),
        "faq_html": _sanitize_ai_html(data.get("faq_html", "")),
        "meta_description": _sanitize_ai_text(data.get("meta_description", ""), max_chars=158),
    }
    if len(_html_to_text_for_ai(fields["product_description_html"], 5000)) < 220:
        return False, "short_product_description_html", fields
    if "<h2" not in fields["product_description_html"].lower():
        return False, "product_description_missing_h2", fields
    if len(_html_to_text_for_ai(fields["feature_tab_html"], 2400)) < 80 or "<li" not in fields["feature_tab_html"].lower():
        return False, "invalid_feature_tab_html", fields
    if len(_html_to_text_for_ai(fields["faq_html"], 4000)) < 150 or "<h3" not in fields["faq_html"].lower():
        return False, "invalid_faq_html", fields
    if not fields["short_description"] or not fields["meta_description"]:
        return False, "missing_short_or_meta", fields
    forbidden = re.compile(r"\b(harga|diskon|promo|garansi resmi|ready stock|stok tersedia|gratis ongkir|termurah|terbaik di indonesia)\b", re.I)
    joined = " ".join(fields.values())
    if forbidden.search(joined):
        return False, "forbidden_sales_claim", fields
    return True, "ok", fields


def _build_deepseek_messages(payload: Dict[str, Any]) -> List[Dict[str, str]]:
    system_prompt = (
        "Anda adalah penulis katalog teknis berbahasa Indonesia untuk uji.co.id. "
        "Tulis copy produk ATAGO yang natural, rapi, dan tidak terdengar seperti template. "
        "Gunakan hanya data yang diberikan. Jangan membuat klaim harga, stok, promo, garansi, sertifikasi, akurasi, range, atau fitur yang tidak ada pada data. "
        "Nama produk, model, Cat.No., SKU, dan satuan teknis harus dipertahankan apa adanya. "
        "Jangan mencampur bahasa Inggris-Indonesia kecuali istilah teknis yang memang lazim. "
        "Jangan memasukkan tabel spesifikasi ke product_description_html karena tab Spesifikasi dibuat terpisah dari data scrape. "
        "Balas JSON valid saja."
    )
    user_prompt = {
        "task": "Buat konten WooCommerce uji.co.id untuk produk ATAGO dalam bahasa Indonesia.",
        "output_schema": {
            "product_description_html": "HTML aman: 2 paragraf pembuka, <h2>Fungsi dan Keunggulan</h2> + <ul>, <h2>Contoh Penggunaan</h2> + paragraf/ul, <h2>Informasi Pemilihan Produk</h2> + paragraf. Tanpa CTA kontak.",
            "short_description": "1-2 kalimat ringkas maksimal 260 karakter.",
            "feature_tab_html": "HTML <h2>Keunggulan Produk</h2><ul> berisi 4-6 poin faktual dan natural.",
            "faq_html": "HTML berisi 4 pasang <h3>pertanyaan</h3><p>jawaban</p>. FAQ harus spesifik sesuai jenis produk.",
            "meta_description": "Maksimal 158 karakter, natural untuk SEO, tidak clickbait.",
        },
        "rules": [
            "Output hanya JSON valid, tanpa markdown dan tanpa code fence.",
            "Sebut brand sebagai ATAGO.",
            "Gunakan data Product information dan For more detail hanya sebagai dasar klaim; jangan menambah angka baru.",
            "Jangan menyebut accessories sebagai included kecuali tertulis eksplisit dalam data.",
            "Jika product_type adalah accessory_or_component, jangan memaksa produk menjadi alat utama; jelaskan sebagai aksesori/komponen pendukung dan tekankan pengecekan kompatibilitas model.",
            "Gunakan tag HTML aman: p, strong, ul, li, h2, h3, em.",
        ],
        "product_data": payload,
    }
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(user_prompt, ensure_ascii=False)},
    ]


def _call_deepseek_chat(messages: List[Dict[str, str]]) -> Tuple[bool, str, str]:
    api_key = clean_text(AI_SETTINGS.get("api_key", ""))
    if not api_key:
        return False, "missing_api_key", ""
    url = clean_text(AI_SETTINGS.get("base_url", "https://api.deepseek.com")).rstrip("/") + "/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    body = {
        "model": AI_SETTINGS.get("model", "deepseek-chat"),
        "messages": messages,
        "temperature": AI_SETTINGS.get("temperature", 0.25),
        "response_format": {"type": "json_object"},
    }
    try:
        response = requests.post(url, headers=headers, json=body, timeout=AI_SETTINGS.get("timeout", 60))
        if response.status_code in {400, 422}:
            body.pop("response_format", None)
            response = requests.post(url, headers=headers, json=body, timeout=AI_SETTINGS.get("timeout", 60))
        response.raise_for_status()
        data = response.json()
        choices = data.get("choices", []) if isinstance(data, dict) else []
        if not choices:
            return False, "empty_choices", ""
        message = choices[0].get("message", {}) if isinstance(choices[0], dict) else {}
        content = clean_text(message.get("content", ""))
        if not content:
            return False, "empty_content", ""
        return True, "ok", content
    except Exception as exc:
        return False, f"request_error_{type(exc).__name__}", ""


def get_deepseek_ai_content(data: Dict[str, Any], family: Dict[str, Any]) -> Dict[str, str]:
    if not AI_SETTINGS.get("enabled"):
        return {}
    existing = data.get("_deepseek_ai")
    if isinstance(existing, dict):
        return existing.get("fields", {}) if existing.get("ok") else {}

    payload = _data_payload_for_ai(data, family)
    cache_key = _ai_cache_key(payload)
    cache = AI_SETTINGS.get("cache", {}) if isinstance(AI_SETTINGS.get("cache", {}), dict) else {}
    cached = cache.get(cache_key)
    if isinstance(cached, dict) and cached.get("ok") and isinstance(cached.get("fields"), dict):
        data["_deepseek_ai"] = cached
        AI_STATS["cache_hit"] = AI_STATS.get("cache_hit", 0) + 1
        return cached.get("fields", {})

    if not clean_text(AI_SETTINGS.get("api_key", "")):
        data["_deepseek_ai"] = {"ok": False, "reason": "missing_api_key", "fields": {}}
        AI_STATS["fallback"] = AI_STATS.get("fallback", 0) + 1
        return {}

    limit = int(AI_SETTINGS.get("limit", 0) or 0)
    if limit > 0 and AI_STATS.get("used", 0) >= limit:
        data["_deepseek_ai"] = {"ok": False, "reason": "ai_limit_reached", "fields": {}}
        AI_STATS["skipped"] = AI_STATS.get("skipped", 0) + 1
        return {}

    messages = _build_deepseek_messages(payload)
    ok, reason, raw = _call_deepseek_chat(messages)
    if ok:
        parsed = _extract_json_object(raw)
        valid, validation_reason, fields = _validate_ai_content(parsed)
        if valid:
            record = {
                "ok": True,
                "reason": "ok",
                "fields": fields,
                "created_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            }
            cache[cache_key] = record
            AI_SETTINGS["cache"] = cache
            data["_deepseek_ai"] = record
            AI_STATS["used"] = AI_STATS.get("used", 0) + 1
            delay = float(AI_SETTINGS.get("delay", 0.0) or 0.0)
            if delay > 0:
                time.sleep(delay)
            return fields
        reason = validation_reason

    record = {"ok": False, "reason": reason, "fields": {}}
    data["_deepseek_ai"] = record
    cache[cache_key] = record
    AI_SETTINGS["cache"] = cache
    AI_STATS["fallback"] = AI_STATS.get("fallback", 0) + 1
    delay = float(AI_SETTINGS.get("delay", 0.0) or 0.0)
    if delay > 0:
        time.sleep(delay)
    return {}

# ----------------------------- content generation -----------------------------

def detect_family(title: str, description: str, spec_rows: List[Dict[str, str]]) -> Dict[str, Any]:
    # Produk accessories/komponen tetap masuk output, tetapi kategorinya dibedakan agar copy tidak memaksa menjadi alat utama.
    is_acc, _reason = is_accessory_product(title, description, spec_rows)
    if is_acc:
        return ACCESSORY_FAMILY
    blob = " ".join([title, description, " ".join(f"{r.get('label')} {r.get('value')}" for r in spec_rows)])
    for pattern, family in FAMILY_RULES:
        if pattern.search(blob):
            return family
    return DEFAULT_FAMILY


def get_spec_value(spec_rows: List[Dict[str, str]], label_contains: str) -> str:
    needle = label_contains.casefold()
    for row in spec_rows:
        if needle in clean_text(row.get("label")).casefold():
            return clean_text(row.get("value"))
    return ""


def spec_table_html(rows: List[Dict[str, str]]) -> str:
    if not rows:
        return ""
    body = []
    for row in rows:
        label = esc(row.get("label"))
        value = esc(row.get("value"))
        if label and value:
            body.append(f"<tr><th>{label}</th><td>{value}</td></tr>")
    if not body:
        return ""
    return '<table class="uji-spec-table"><tbody>' + "".join(body) + "</tbody></table>"


def source_description_block(description: str) -> str:
    description = compact(description, max_chars=700)
    if not description:
        return ""
    # Description ATAGO umumnya berbahasa Inggris. Disimpan sebagai ringkasan sumber agar klaim tetap terlacak.
    return f"<p><strong>Ringkasan dari halaman ATAGO:</strong> {esc(description)}</p>"


def feature_bullets(title: str, family: Dict[str, Any], spec_rows: List[Dict[str, str]]) -> List[str]:
    bullets: List[str] = []
    is_accessory = family.get("category") == ACCESSORY_FAMILY["category"]
    model = get_spec_value(spec_rows, "Model")
    cat_no = get_spec_value(spec_rows, "Cat.No.")

    if is_accessory:
        bullets.append("Merupakan aksesori atau komponen pendukung ATAGO yang diproses sebagai produk katalog tersendiri.")
        if model:
            bullets.append(f"Model yang tercantum pada halaman ATAGO: {model}.")
        if cat_no:
            bullets.append(f"Cat.No. yang tercantum pada halaman ATAGO: {cat_no}.")
        bullets.append("Perlu dicocokkan dengan unit utama ATAGO yang digunakan sebelum pemesanan atau penggunaan.")
        bullets.append("Data teknis yang tersedia disusun dari bagian Product information dan For more detail pada halaman produk ATAGO.")
        return dedupe_keep_order(bullets)[:5]

    range_val = get_spec_value(spec_rows, "Rentang")
    accuracy_val = get_spec_value(spec_rows, "Akurasi")
    min_scale = get_spec_value(spec_rows, "Skala minimum") or get_spec_value(spec_rows, "Resolusi")
    temp = get_spec_value(spec_rows, "Suhu")
    output = get_spec_value(spec_rows, "Output") or get_spec_value(spec_rows, "Antarmuka")

    if range_val:
        bullets.append(f"Memiliki rentang pengukuran yang tercantum pada data ATAGO: {range_val}.")
    if accuracy_val:
        bullets.append(f"Mencantumkan akurasi pengukuran: {accuracy_val}.")
    if min_scale:
        bullets.append(f"Menyediakan skala minimum atau resolusi pembacaan: {min_scale}.")
    if temp:
        bullets.append(f"Informasi suhu kerja/pengukuran tersedia pada spesifikasi: {temp}.")
    if output:
        bullets.append(f"Mencantumkan dukungan output atau antarmuka: {output}.")

    bullets.append(f"Cocok dipertimbangkan untuk kebutuhan laboratorium, quality control, dan pemeriksaan sampel yang membutuhkan {family['term']}.")
    return dedupe_keep_order(bullets)[:5]

def build_product_description(title: str, family: Dict[str, Any], description: str, spec_rows: List[Dict[str, str]]) -> str:
    is_accessory = family.get("category") == ACCESSORY_FAMILY["category"]
    if is_accessory:
        intro = f"<p><strong>{esc(title)}</strong> adalah {esc(family['term'])} yang digunakan untuk melengkapi atau mendukung penggunaan unit utama ATAGO yang kompatibel.</p>"
        compat_note = "<p>Karena produk aksesori biasanya memiliki data teknis yang lebih singkat, pemilihan produk perlu memperhatikan model, Cat.No., dan kecocokan dengan alat utama yang digunakan.</p>"
        desc_block = source_description_block(description)
        bullets = "".join(f"<li>{esc(b)}</li>" for b in feature_bullets(title, family, spec_rows))
        spec_note = "<p>Informasi produk disusun dari bagian <em>Product information</em> dan <em>For more detail</em> pada halaman produk ATAGO.</p>"
        return "".join([
            intro,
            desc_block,
            compat_note,
            "<h2>Fungsi dan Keunggulan</h2>",
            f"<ul>{bullets}</ul>" if bullets else "",
            spec_note,
            UJI_CTA,
        ])

    intro_options = [
        f"<p><strong>{esc(title)}</strong> adalah {esc(family['term'])} dari ATAGO yang digunakan untuk {esc(family['function'])}.</p>",
        f"<p><strong>{esc(title)}</strong> merupakan produk ATAGO untuk kebutuhan pengukuran sampel, kontrol kualitas, dan pekerjaan laboratorium yang memerlukan {esc(family['term'])}.</p>",
    ]
    intro = intro_options[deterministic_index(title, len(intro_options))]
    desc_block = source_description_block(description)
    bullets = "".join(f"<li>{esc(b)}</li>" for b in feature_bullets(title, family, spec_rows))
    spec_note = "<p>Data teknis utama produk ini disusun dari bagian <em>Product information</em> dan <em>For more detail</em> pada halaman produk ATAGO.</p>"
    return "".join([
        intro,
        desc_block,
        "<h2>Keunggulan Produk</h2>",
        f"<ul>{bullets}</ul>" if bullets else "",
        spec_note,
        UJI_CTA,
    ])

def build_short_description(title: str, family: Dict[str, Any], spec_rows: List[Dict[str, str]]) -> str:
    is_accessory = family.get("category") == ACCESSORY_FAMILY["category"]
    model = get_spec_value(spec_rows, "Model")
    cat_no = get_spec_value(spec_rows, "Cat.No.")
    if is_accessory:
        parts = [f"{title} adalah aksesori/komponen pendukung ATAGO untuk unit utama yang kompatibel."]
        if model:
            parts.append(f"Model: {model}.")
        if cat_no:
            parts.append(f"Cat.No.: {cat_no}.")
        return compact(" ".join(parts), max_chars=320)

    range_val = get_spec_value(spec_rows, "Rentang")
    accuracy_val = get_spec_value(spec_rows, "Akurasi")
    parts = [f"{title} adalah {family['term']} dari ATAGO untuk {family['function']}."]
    if range_val:
        parts.append(f"Rentang: {range_val}.")
    if accuracy_val:
        parts.append(f"Akurasi: {accuracy_val}.")
    return compact(" ".join(parts), max_chars=320)

def build_application_tab(family: Dict[str, Any]) -> str:
    if family.get("category") == ACCESSORY_FAMILY["category"]:
        apps = [
            "kelengkapan unit utama ATAGO yang kompatibel",
            "penggantian atau penambahan komponen pendukung sesuai model alat",
            "kebutuhan pemeliharaan alat ukur ATAGO",
            "persiapan aksesori pendukung untuk pekerjaan laboratorium atau quality control",
        ]
    else:
        apps = [
            "kontrol kualitas produksi",
            "pengujian sampel di laboratorium",
            "pemeriksaan bahan cair atau larutan",
            "validasi parameter sampel sebelum proses lanjutan",
        ]
        if "refractometer" in family["term"].lower():
            apps = ["pengukuran Brix atau indeks bias", "kontrol kualitas makanan dan minuman", "analisis konsentrasi larutan", "pemeriksaan sampel cair di laboratorium"]
        elif "salt" in family["term"].lower():
            apps = ["pengukuran kadar garam", "pemeriksaan salinitas sampel", "kontrol kualitas makanan/minuman", "pengujian larutan proses"]
        elif "viscometer" in family["term"].lower():
            apps = ["pengukuran viskositas", "kontrol kualitas cairan proses", "formulasi produk", "pengujian kekentalan sampel laboratorium"]
    items = "".join(f"<li>{esc(app)}</li>" for app in apps)
    return f"<h2>Aplikasi Penggunaan</h2><ul>{items}</ul>"

def build_faq(title: str, family: Dict[str, Any]) -> str:
    if family.get("category") == ACCESSORY_FAMILY["category"]:
        return (
            f"<h3>Apa fungsi {esc(title)}?</h3>"
            f"<p>{esc(title)} digunakan sebagai aksesori atau komponen pendukung ATAGO untuk unit utama yang kompatibel.</p>"
            f"<h3>Apakah produk ini alat utama?</h3>"
            f"<p>Produk ini diproses sebagai aksesori/komponen. Pastikan kecocokan model dan Cat.No. dengan alat utama sebelum digunakan.</p>"
            f"<h3>Apakah data produknya berasal dari halaman ATAGO?</h3>"
            f"<p>Ya. Tab spesifikasi disusun dari bagian Product information dan For more detail pada halaman produk ATAGO.</p>"
            f"<h3>Apakah aksesori ATAGO ikut masuk output?</h3>"
            f"<p>Ya. Versi script ini memasukkan produk aksesori/komponen ke output WooCommerce dan membedakan kategorinya dari alat utama.</p>"
        )
    return (
        f"<h3>Apa fungsi {esc(title)}?</h3>"
        f"<p>{esc(title)} digunakan sebagai {esc(family['term'])} untuk {esc(family['function'])}.</p>"
        f"<h3>Apakah spesifikasinya berasal dari halaman produk?</h3>"
        f"<p>Ya. Tab spesifikasi disusun dari bagian Product information dan For more detail pada halaman ATAGO.</p>"
        f"<h3>Apakah aksesori ATAGO ikut diproses?</h3>"
        f"<p>Ya. Produk aksesori/komponen ATAGO juga dimasukkan ke output, tetapi kategorinya dibedakan sebagai aksesori atau komponen pendukung.</p>"
    )

def build_meta_description(title: str, family: Dict[str, Any], spec_rows: List[Dict[str, str]]) -> str:
    if family.get("category") == ACCESSORY_FAMILY["category"]:
        meta = f"{title} adalah aksesori/komponen ATAGO. Cek model, Cat.No., dan konsultasi kebutuhan di uji.co.id."
        return compact(meta, max_chars=155)
    range_val = get_spec_value(spec_rows, "Rentang")
    if range_val:
        meta = f"{title} dari ATAGO untuk {family['function']}. Rentang: {range_val}. Hubungi uji.co.id."
    else:
        meta = f"{title} dari ATAGO untuk pengukuran dan kontrol kualitas sampel. Cek spesifikasi dan konsultasi di uji.co.id."
    return compact(meta, max_chars=155)

def build_output_row(data: Dict[str, Any], pub_date: datetime, processing_label: str) -> Dict[str, Any]:
    title = clean_text(data.get("title"))
    description = clean_text(data.get("description"))
    spec_rows = data.get("spec_rows") or []
    family = detect_family(title, description, spec_rows)

    # Rule/template fallback selalu tersedia. DeepSeek hanya override copy publik jika valid.
    fallback_product_description = build_product_description(title, family, description, spec_rows)
    fallback_short_description = build_short_description(title, family, spec_rows)
    fallback_feature_tab = "<h2>Keunggulan Produk</h2><ul>" + "".join(
        f"<li>{esc(b)}</li>" for b in feature_bullets(title, family, spec_rows)
    ) + "</ul>"
    fallback_faq = build_faq(title, family)
    fallback_meta = build_meta_description(title, family, spec_rows)

    ai_fields = get_deepseek_ai_content(data, family)
    ai_state = data.get("_deepseek_ai", {}) if isinstance(data.get("_deepseek_ai", {}), dict) else {}
    ai_ok = bool(ai_fields)

    product_description = ai_fields.get("product_description_html") or fallback_product_description
    if UJI_CTA not in product_description:
        product_description = product_description.rstrip() + UJI_CTA
    short_description = ai_fields.get("short_description") or fallback_short_description
    feature_tab_html = ai_fields.get("feature_tab_html") or fallback_feature_tab
    faq_html = ai_fields.get("faq_html") or fallback_faq
    meta = ai_fields.get("meta_description") or fallback_meta

    spec_html = "<h2>Spesifikasi</h2>" + spec_table_html(spec_rows)
    category = family["category"]
    tags = dedupe_keep_order([category] + family["tags"] + ["ATAGO"])
    focus_keyphrase = title[:120]

    correction_log = []
    if not description:
        correction_log.append("missing_description")
    if not spec_rows:
        correction_log.append("missing_product_information")
    if not data.get("image_url"):
        correction_log.append("missing_image")
    if data.get("is_accessory"):
        correction_log.append(f"accessory_included_{clean_text(data.get('accessory_reason', 'detected'))}")
    if AI_SETTINGS.get("enabled"):
        if ai_ok:
            correction_log.append("deepseek_ai_used")
        else:
            correction_log.append(f"deepseek_ai_fallback_{clean_text(ai_state.get('reason', 'unknown'))}")

    row = {col: "" for col in OUTPUT_COLS}
    row.update({
        "Name": title,
        "Brand": BRAND,
        "Product Description": product_description,
        "Product Short Description": short_description,
        "custom_tab_1_title": "Spesifikasi",
        "custom_tab_1_content": spec_html,
        "custom_tab_1_priority": 10,
        "custom_tab_2_title": "Keunggulan Produk",
        "custom_tab_2_content": feature_tab_html,
        "custom_tab_2_priority": 20,
        "custom_tab_3_title": "Aplikasi Penggunaan",
        "custom_tab_3_content": build_application_tab(family),
        "custom_tab_3_priority": 30,
        "custom_tab_4_title": "FAQ",
        "custom_tab_4_content": faq_html,
        "custom_tab_4_priority": 90,
        "product categories": category,
        "product tags": ", ".join(tags),
        "focus keyphrase": focus_keyphrase,
        "meta description": meta,
        "publication_date": pub_date.strftime("%m/%d/%Y %H:%M"),
        "Processed": "Yes",
        "Processing Time": processing_label + ("_deepseek_ai" if ai_ok else ("_deepseek_ai_fallback" if AI_SETTINGS.get("enabled") else "")),
        "Content Quality": "atago_live_scrape_deepseek_ai" if ai_ok else "atago_live_scrape_rule_fallback",
        "image_url": data.get("image_url", ""),
        "Source URL": data.get("url", ""),
        "Website Validation": "scraped_atago_live",
        "Website Correction Log": "; ".join(correction_log) if correction_log else "ok",
    })
    return row


# ----------------------------- cache/output -----------------------------

def load_cache(path: str) -> Dict[str, Any]:
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_cache(path: str, cache: Dict[str, Any]) -> None:
    if not path:
        return
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def style_output_xlsx(path: str, sheet_name: str = "UJI Products") -> None:
    if load_workbook is None:
        return
    try:
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        width_map = {
            "A": 34, "B": 16, "C": 60, "D": 44, "E": 22, "F": 60, "H": 24, "I": 60,
            "K": 22, "L": 60, "AC": 34, "AD": 44, "AE": 30, "AF": 44, "AG": 20,
            "AH": 16, "AI": 32, "AJ": 28, "AK": 44, "AL": 44, "AM": 28, "AN": 48,
        }
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = width_map.get(letter, 18)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
            ws.row_dimensions[row[0].row].height = 72
        ws.row_dimensions[1].height = 32

        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        if ws.max_row >= 1 and not ws.tables:
            tab = Table(displayName="ATAGO_UJI_Products", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)
        wb.save(path)
    except Exception:
        # Styling tidak boleh menggagalkan output utama.
        pass


def write_outputs(rows: List[Dict[str, Any]], skipped: List[Dict[str, Any]], output_path: str, skipped_output: str = "") -> None:
    out_df = pd.DataFrame(rows, columns=OUTPUT_COLS)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="UJI Products")
        if skipped:
            pd.DataFrame(skipped).to_excel(writer, index=False, sheet_name="Skipped")
    style_output_xlsx(output_path, "UJI Products")

    if skipped_output:
        pd.DataFrame(skipped).to_excel(skipped_output, index=False)


# ----------------------------- CLI -----------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="ATAGO product/accessory link scraper to WooCommerce import format for uji.co.id")
    parser.add_argument("--input", required=True, help="Path Excel input berisi link produk ATAGO")
    parser.add_argument("--output", default="uji_atago_import_ready.xlsx", help="Path output Excel WooCommerce")
    parser.add_argument("--sheet", default=0, help="Nama sheet atau index sheet. Default: 0 / sheet pertama")
    parser.add_argument("--url-column", default="", help="Nama kolom URL jika ingin ditentukan manual")
    parser.add_argument("--cache", default="atago_scrape_cache.json", help="Cache JSON hasil scrape agar tidak request ulang")
    parser.add_argument("--timeout", type=int, default=20, help="Timeout request per URL, detik")
    parser.add_argument("--delay", type=float, default=0.4, help="Delay antar request, detik")
    parser.add_argument("--max-rows", type=int, default=0, help="Batasi jumlah URL untuk uji coba. 0 = semua")
    parser.add_argument("--start-date", default="08/02/2025 10:00", help="Tanggal publikasi awal format MM/DD/YYYY HH:MM")
    parser.add_argument("--pub-interval-hours", type=int, default=1, help="Jarak jam antar publication_date")
    parser.add_argument("--skipped-output", default="", help="Opsional: file Excel khusus log produk yang diskip")
    parser.add_argument("--no-cache", action="store_true", help="Abaikan cache dan scrape ulang semua URL")
    parser.add_argument("--skip-accessories", action="store_true", help="Lewati produk yang terdeteksi sebagai accessories/komponen. Default: accessories tetap masuk output.")
    parser.add_argument("--use-ai", action="store_true", help="Aktifkan generator konten AI DeepSeek untuk deskripsi, short description, keunggulan, FAQ, dan meta description")
    parser.add_argument("--deepseek-api-key", default=os.getenv("DEEPSEEK_API_KEY", ""), help="API key DeepSeek. Bisa juga memakai env DEEPSEEK_API_KEY")
    parser.add_argument("--deepseek-base-url", default=os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com"), help="Base URL DeepSeek/OpenAI-compatible API")
    parser.add_argument("--deepseek-model", default=os.getenv("DEEPSEEK_MODEL", "deepseek-chat"), help="Model DeepSeek yang digunakan")
    parser.add_argument("--ai-cache", default="", help="Path cache JSON hasil AI. Kosong = otomatis di sebelah output")
    parser.add_argument("--ai-limit", type=int, default=0, help="Batasi jumlah request AI baru. 0 = semua baris")
    parser.add_argument("--ai-timeout", type=int, default=60, help="Timeout request AI dalam detik")
    parser.add_argument("--ai-delay", type=float, default=0.0, help="Jeda antar request AI agar tidak terlalu agresif")
    parser.add_argument("--ai-temperature", type=float, default=0.25, help="Temperature AI untuk menjaga output tetap stabil")
    return parser.parse_args()


def coerce_sheet_arg(value: Any) -> Any:
    if isinstance(value, str) and re.fullmatch(r"\d+", value):
        return int(value)
    return value


def main() -> None:
    args = parse_args()
    sheet = coerce_sheet_arg(args.sheet)
    urls = read_input_urls(args.input, sheet=sheet, url_column=args.url_column)
    if args.max_rows and args.max_rows > 0:
        urls = urls[:args.max_rows]
    if not urls:
        raise SystemExit("Tidak ada URL yang bisa diproses.")

    cache: Dict[str, Any] = {} if args.no_cache else load_cache(args.cache)
    out_path = Path(args.output)
    ai_cache_path = args.ai_cache or str(out_path.with_suffix(".deepseek_ai_cache.json"))
    ai_cache: Dict[str, Any] = {} if args.no_cache else load_cache(ai_cache_path)
    set_deepseek_ai_settings(
        enabled=bool(args.use_ai),
        api_key=args.deepseek_api_key,
        base_url=args.deepseek_base_url,
        model=args.deepseek_model,
        cache=ai_cache,
        cache_path=ai_cache_path,
        limit=args.ai_limit,
        timeout=args.ai_timeout,
        delay=args.ai_delay,
        temperature=args.ai_temperature,
    )
    sess = requests.Session()
    rows: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    accessory_included_count = 0
    try:
        start_date = datetime.strptime(args.start_date, "%m/%d/%Y %H:%M")
    except Exception:
        start_date = datetime(2025, 8, 2, 10, 0)

    for idx, url in enumerate(urls, 1):
        print(f"[{idx}/{len(urls)}] scrape {url}")
        if not args.no_cache and url in cache:
            data = cache[url]
            data["from_cache"] = "yes"
        else:
            data = scrape_atago_product_page(url, timeout=args.timeout, session=sess)
            cache[url] = data
            if args.delay > 0:
                time.sleep(args.delay)

        if data.get("is_accessory"):
            if args.skip_accessories:
                skipped.append({
                    "url": url,
                    "title": data.get("title", ""),
                    "reason": data.get("accessory_reason", "detected_accessory"),
                    "status": data.get("status", ""),
                    "error": data.get("error", ""),
                })
                print(f"    skip accessory: {data.get('title', '')} ({data.get('accessory_reason', '')})")
                continue
            accessory_included_count += 1
            print(f"    include accessory: {data.get('title', '')} ({data.get('accessory_reason', '')})")

        if not data.get("title"):
            skipped.append({
                "url": url,
                "title": "",
                "reason": "missing_title_or_scrape_failed",
                "status": data.get("status", ""),
                "error": data.get("error", ""),
            })
            print(f"    skip failed: {data.get('error', data.get('status', ''))}")
            continue

        pub_date = start_date + timedelta(hours=len(rows) * args.pub_interval_hours)
        rows.append(build_output_row(data, pub_date, "atago_uji_import_ready"))

    write_outputs(rows, skipped, args.output, skipped_output=args.skipped_output)
    save_cache(args.cache, cache)
    if AI_SETTINGS.get("enabled"):
        save_cache(AI_SETTINGS.get("cache_path", ai_cache_path), AI_SETTINGS.get("cache", {}))
    print(f"Selesai. Produk masuk output: {len(rows)} | Produk diskip: {len(skipped)} | Accessories masuk: {accessory_included_count}")
    print(f"Output: {args.output}")
    if AI_SETTINGS.get("enabled"):
        print(f"deepseek_ai_enabled={AI_SETTINGS.get('enabled')}")
        print(f"deepseek_model={AI_SETTINGS.get('model')}")
        print(f"deepseek_ai_used={AI_STATS.get('used', 0)}")
        print(f"deepseek_ai_cache_hits={AI_STATS.get('cache_hit', 0)}")
        print(f"deepseek_ai_fallback={AI_STATS.get('fallback', 0)}")
        print(f"deepseek_ai_cache={AI_SETTINGS.get('cache_path', ai_cache_path)}")
    if skipped and args.skip_accessories:
        print("Catatan: produk accessories/komponen ATAGO disimpan di sheet 'Skipped' karena --skip-accessories aktif.")
    elif accessory_included_count:
        print("Catatan: produk accessories/komponen ATAGO ikut masuk output dan diberi kategori Aksesori dan Komponen ATAGO.")


if __name__ == "__main__":
    main()
